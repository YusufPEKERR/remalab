def style_excel_file(filepath: str):
    """Excel dosyasını openpyxl kullanarak premium ve estetik bir tasarıma kavuşturur."""
    try:
        import re
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.worksheet.table import Table, TableStyleInfo

        wb = openpyxl.load_workbook(filepath)
        used_table_names = set()
        for sheet_idx, sheet in enumerate(wb.worksheets):
            max_row = sheet.max_row
            max_col = sheet.max_column
            if max_row < 1 or max_col < 1:
                continue

            # Başlık (satır 1) rengi/fontu: özel marka paleti, tek satır olduğu için ucuz.
            header_fill = PatternFill(start_color="212B36", end_color="212B36", fill_type="solid") # Koyu şık gri/lacivert
            header_font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=11)
            center_align = Alignment(horizontal="center", vertical="center")
            header_border = Border(bottom=Side(style='medium', color='1F6FEB')) # Başlığın altına mavi bir vurgu

            sheet.row_dimensions[1].height = 28 # Başlık daha ferah

            flow_col_letter = None
            max_lengths = [0] * (max_col + 1)  # 1-index kullanılacak

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = header_border
                if cell.value and str(cell.value).strip().lower() == 'flow':
                    flow_col_letter = openpyxl.utils.get_column_letter(cell.column)
                val_str = str(cell.value) if cell.value is not None else ""
                max_lengths[cell.column] = len(val_str)

            # Flow sütununa dropdown validation ekle
            if flow_col_letter and max_row > 1:
                flow_options = '"Refurbish,Repair,RMA,Battery Replacement"'
                dv = DataValidation(
                    type="list",
                    formula1=flow_options,
                    allow_blank=False,
                    showDropDown=False,
                    showErrorMessage=True,
                    errorTitle="Geçersiz Değer",
                    error="Lütfen listeden bir seçim yapın.",
                    showInputMessage=True,
                    promptTitle="Akış Durumu",
                    prompt="Refurbish, Repair, RMA veya Battery Replacement seçin."
                )
                dv.sqref = f"{flow_col_letter}2:{flow_col_letter}{max(max_row, 1000)}"
                sheet.add_data_validation(dv)

            # Veri satırlarının biçimlendirmesi (zebra + kenarlık): hücre başına yazmak yerine
            # Excel'in yerleşik "Table" (banded) stiline devredilir - O(satır*sütun) hücre
            # yazma yerine O(1) tablo tanımı, büyük tablolarda kat kat hızlı.
            # Sütun genişliği için yine de değerleri satır bazlı gezip ölçmemiz gerekiyor.
            for row_cells in sheet.iter_rows(min_row=2, max_row=max_row):
                for cell in row_cells:
                    val_str = str(cell.value) if cell.value is not None else ""
                    if len(val_str) > max_lengths[cell.column]:
                        max_lengths[cell.column] = len(val_str)

            if max_row > 1:
                last_col_letter = openpyxl.utils.get_column_letter(max_col)
                table_name = re.sub(r'[^A-Za-z0-9_]', '_', f"Tablo_{sheet.title}_{sheet_idx}")
                if not table_name or not table_name[0].isalpha():
                    table_name = f"T_{table_name}"
                base_name = table_name
                suffix = 1
                while table_name in used_table_names:
                    table_name = f"{base_name}_{suffix}"
                    suffix += 1
                used_table_names.add(table_name)

                tab = Table(displayName=table_name, ref=f"A1:{last_col_letter}{max_row}")
                tab.tableStyleInfo = TableStyleInfo(
                    name="TableStyleLight1",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                sheet.add_table(tab)

            # Sütun genişliklerini içeriğe göre ayarla (min 15, max 50)
            for col_idx in range(1, max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                adjusted_width = min(max(int(max_lengths[col_idx] * 1.3) + 4, 15), 50)
                sheet.column_dimensions[col_letter].width = adjusted_width

            # İlk satırı dondur (sabit kalsın)
            sheet.freeze_panes = 'A2'

        wb.save(filepath)
    except Exception as e:
        print(f"Excel stili uygulanırken hata oluştu: {e}")

