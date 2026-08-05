import json
from PySide6.QtCore import QObject, Slot, Signal
import json
import logging
import os
import datetime as _dt

# Türkiye kalıcı olarak UTC+3'tür (2016'dan beri yaz saati uygulaması yok), bu yüzden
# sabit offset güvenli ve tzdata bağımlılığı gerektirmez. Bazı zaman sütunları naive
# (tz'siz) ve UTC olarak yazılmış (Python utcnow()), bazıları TIMESTAMPTZ. İkisini de
# doğru Türkiye yerel saatine çevirip gg.aa.yyyy SS:DD formatında döndürür.
_TR_TZ = _dt.timezone(_dt.timedelta(hours=3))

def fmt_tr_datetime(dt, with_time=True):
    """Bir datetime'ı Türkiye yerel saatine çevirip formatlar. None -> ''.
    Naive datetime'lar UTC kabul edilir (repair_records.created_at gibi utcnow() ile yazılanlar)."""
    if not dt:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=_dt.timezone.utc)
    local = dt.astimezone(_TR_TZ)
    return local.strftime("%d.%m.%Y %H:%M" if with_time else "%d.%m.%Y")

def get_cache_dirs():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # We write to a single external folder.
    # Vite middleware will serve from this folder.
    # Production PySide static server should also be able to serve from it.
    d1 = os.path.join(base_dir, 'api_cache')
    os.makedirs(d1, exist_ok=True)
    return [d1]

def write_to_cache(filename, json_data):
    dirs = get_cache_dirs()
    for d in dirs:
        try:
            with open(os.path.join(d, filename), "w", encoding="utf-8") as f:
                f.write(json_data)
        except Exception as e:
            logging.error(f"Failed to write cache {filename} to {d}: {e}")

from config.database import SessionLocal
from config.auth import verify_password
from models.user import User

from sqlalchemy import event
from sqlalchemy.orm import Session

def clear_api_cache(session=None):
    """Veritabanında değişiklik olduğunda cache'i temizler, böylece UI sadece yeni veriyi bekler."""
    dirs = get_cache_dirs()
    for d in dirs:
        for filename in ["parts.json", "stock.json", "critical.json", "price_matrix_items.json", "price_matrix_prices.json"]:
            path = os.path.join(d, filename)
            if os.path.exists(path):
                try: 
                    os.remove(path)
                except Exception as e: 
                    logging.error(f"Failed to clear cache {path}: {e}")

event.listen(Session, 'after_commit', clear_api_cache)

# service_statu kayıtları İngilizce tutulur (language='en'). Servis ekranı gibi kullanıcıya
# dönük yerlerde statü adının Türkçe gösterilmesi için kod -> Türkçe eşlemesi. Haritada
# olmayan bir kod, DB'deki short_name'e (fallback) düşer. Bkz. statu_label_tr.
SERVICE_STATU_TR = {
    100: "Ön bildirim yapıldı",
    101: "Depo kabulü tamamlandı",
    102: "İlk teste aktarıldı",
    103: "İlk test bekleniyor",
    104: "İlk test tamamlandı",
    105: "Üretim planlama onayı bekleniyor",
    106: "Müşteri onayına gönderilecek",
    107: "Müşteri onayı bekleniyor",
    109: "Üretim aşamasında",
    124: "Son teste teslim edilecek",
    125: "Son teste kabul edildi",
    126: "Depoya sevk edilecek",
    127: "Müşteriye sevkiyat bekleniyor",
    128: "Serbest bırakıldı",
    130: "Montaj bekleniyor",
    131: "L1 Montaj yapılacak",
    132: "L2 Montaj yapılacak",
    133: "Montaj tamamlandı",
    134: "RMA incelemesi",
    135: "İade öncesi son teste gönderildi",
    136: "Müşteri onay/red alındı",
    137: "Ara teste teslim edilecek",
    138: "Ara test bekleniyor",
}


def statu_label_tr(code, fallback_name=None):
    """Statü kodunu 'Türkçe ad (kod)' biçiminde döner. Harita yoksa DB short_name'ine
    (fallback_name), o da yoksa kodun kendisine düşer."""
    name = SERVICE_STATU_TR.get(code) or fallback_name or str(code)
    return f"{name} ({code})"


# Otomatik iş akışıyla yönetilen sabit sistem depoları. Bu depolar arasındaki
# manuel transferler (bkz. transfer_stock) SYSTEM_TRANSFER_RULES ile kısıtlanır.
SYSTEM_LOCATION_KINDS = {
    "good_stock": "Good Stock",
    "doa_stock": "DOA Stock",
    "repair_stock": "Repair Stock",
    "scrap_stock": "Scrap Stock",
    "out_stock": "Out Stock",
}

# Depolar arası manuel "Stok Transferi" akışının izin verdiği kaynak->hedef
# eşleşmeleri. Bir kaynak kind burada yoksa (ör. custom raf lokasyonu) kısıtlama
# uygulanmaz.
SYSTEM_TRANSFER_RULES = {
    "good_stock": {"repair_stock"},
    "repair_stock": {"out_stock", "doa_stock"},
    "doa_stock": {"good_stock", "scrap_stock"},
    "out_stock": set(),
    "scrap_stock": set(),
    "wip_stock": set(),
}

# Müşteri Fiyat Matrisi'nde markanın yanındaki "Ürün Tipi" filtresi için kullanılır.
# warehouse.parts'ta bu ayrımı tutan ayrı bir kolon YOK; warehouse.parts.model metnindeki
# anahtar kelimelerden türetilir (iPad/Tab → TABLET, Watch → SMARTWATCH, MacBook → LAPTOP,
# AirPods/Buds/Kulaklık → EARPHONE, diğerleri → SMART PHONE, ki veri setinin büyük
# çoğunluğu zaten telefon parçasıdır). Üretilen değerler KASITLI OLARAK warehouse.
# product_category.code ile birebir aynıdır (SMART PHONE/TABLET/LAPTOP/EARPHONE/
# SMARTWATCH) - böylece get_price_matrix_product_types Türkçe ekran adını (short_name:
# 'Akıllı Telefon', 'Dizüstü Bilgisayar', 'Bluetooth Kulaklık' vb.) o var olan referans
# tablosundan JOIN ile çeker, kendi metnini uydurmaz. get_price_matrix_* fonksiyonları
# arasında tutarlı kalması için tek bir yerde tanımlanır.
PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL = """
    CASE
        WHEN model ILIKE '%iPad%' OR model ILIKE '%Tab %' OR model ILIKE 'Tab %' OR model ILIKE '%Tablet%' THEN 'TABLET'
        WHEN model ILIKE '%Watch%' THEN 'SMARTWATCH'
        WHEN model ILIKE '%MacBook%' OR model ILIKE '%Laptop%' OR model ILIKE '%Notebook%' THEN 'LAPTOP'
        WHEN model ILIKE '%AirPods%' OR model ILIKE '%Buds%' OR model ILIKE '%Kulaklık%' OR model ILIKE '%Earphone%' OR model ILIKE '%Headphone%' THEN 'EARPHONE'
        ELSE 'SMART PHONE'
    END
"""

# work_orders.work_order_type için desteklenen değerler. SERVICE, mevcut/varsayılan
# akıştır (Service Record'a bağlı tamir süreci). PRODUCTION, bir Recipe'ye (ItemBOM,
# bkz. target_part_id) bağlı yarı mamul üretim süreci içindir; Service Record gerektirmez.
WORK_ORDER_TYPE_SERVICE = "SERVICE"
WORK_ORDER_TYPE_PRODUCTION = "PRODUCTION"
WORK_ORDER_TYPES = {WORK_ORDER_TYPE_SERVICE, WORK_ORDER_TYPE_PRODUCTION}

# material_requests.status akışı: WAITING (issued=0) -> PARTIAL (0 < issued < required)
# -> ISSUED (issued >= required). Sadece Production Work Order'lar için kullanılır;
# Service Work Order akışıyla hiçbir ilişkisi yoktur.
MATERIAL_REQUEST_STATUS_WAITING = "WAITING"
MATERIAL_REQUEST_STATUS_PARTIAL = "PARTIAL"
MATERIAL_REQUEST_STATUS_ISSUED = "ISSUED"


def _compute_material_request_status(issued_quantity, required_quantity):
    """issued/required miktarına göre material_requests.status değerini hesaplar."""
    if issued_quantity <= 0:
        return MATERIAL_REQUEST_STATUS_WAITING
    if issued_quantity < required_quantity:
        return MATERIAL_REQUEST_STATUS_PARTIAL
    return MATERIAL_REQUEST_STATUS_ISSUED


# Production Work Order durum akışı: BEKLIYOR -> URETIMDE -> TAMAMLANDI. Service Work
# Order'ın kendi status sözlüğünden (Beklemede/Devam Ediyor/Tamamlandı/...) tamamen
# bağımsızdır; aynı work_orders.status sütununu paylaşırlar ama değer kümeleri farklıdır,
# bu yüzden Service tarafında hiçbir davranış değişikliği olmaz.
PRODUCTION_WO_STATUS_WAITING = "BEKLIYOR"
PRODUCTION_WO_STATUS_IN_PRODUCTION = "URETIMDE"
PRODUCTION_WO_STATUS_COMPLETED = "TAMAMLANDI"

# Müşteriler sayfası toplu (Excel) yükleme modülü için "Flow (İş Akışı)" alanının
# kabul ettiği değer kümesi warehouse.service_request_type.code'dan (bkz.
# WebBridge._get_flow_values / get_flow_values) canlı olarak okunur - burada
# statik bir liste tutulmaz, tek doğruluk kaynağı DB'dir.

# Toplu yüklemede zorunlu olan sütunlar (şablon başlığı -> customers alanı).
CUSTOMER_BULK_REQUIRED_COLUMNS = [
    ("IMEI Numarası", "imei_number"),
    ("Seri Numarası", "serial_number"),
    ("Internal ID", "internal_id"),
    ("Cihaz Modeli", "cihaz_modeli"),
    ("Flow (İş Akışı)", "flow"),
    ("Müşteri Şikayeti", "customer_reported_complaint"),
    ("Giriş Tarihi", "intake_date"),
]


def _get_system_location_id(db, kind):
    """Verilen kind'a ('good_stock' vb.) sahip sistem deposunun id'sini döner."""
    from models.location import Location
    loc = db.query(Location).filter(Location.kind == kind).first()
    return loc.id if loc else None


def _derive_apple_part_codes(model_text):
    """iPhone modelini (örn. 'iPhone 15 Pro Max') parça kataloğundaki kısaltılmış
    kodlara ('iP15PM') çevirir. warehouse.parts.model alanı Apple parçalarında
    'iPhone 15' değil 'iP15' formatını kullanıyor - düz metin (LIKE) eşleştirmesi
    bu yüzden hiçbir zaman tutmuyordu ve Parça Teslim ekranı iPhone cihazlarda hep
    boş kalıyordu. Eşleşme yoksa boş liste döner (çağıran taraf genel LIKE'a düşer)."""
    import re
    if not model_text:
        return []
    t = re.sub(r'\s+', ' ', model_text.strip().lower())

    m = re.match(r'^iphone\s*(xs\s*max|xs|xr|x)\b', t)
    if m:
        key = re.sub(r'\s+', ' ', m.group(1))
        return [{'x': 'iPX', 'xr': 'iPXR', 'xs': 'iPXS', 'xs max': 'iPXSM'}.get(key, '')]

    m = re.match(r'^iphone\s*se\s*(\d)?', t)
    if m:
        return [f"iPSE{m.group(1) or ''}"]

    m = re.match(r'^iphone\s*(\d+)\s*(pro\s*max|pro|plus|mini|e)?', t)
    if m:
        num = m.group(1)
        qualifier = (m.group(2) or '').strip()
        suffix = {'': '', 'plus': 'P', 'pro': 'PR', 'pro max': 'PM', 'mini': 'M', 'e': 'e'}.get(qualifier, '')
        return [f'iP{num}{suffix}']

    return []


def _build_part_display_name(brand, model, color, part_category, name, item_code):
    """Parça için kullanıcıya gösterilecek ismi (marka+model+renk+kategori, yoksa
    ad, o da yoksa item_code) üretir. get_stock_status ile aynı öncelik sırasını kullanır."""
    display_name = " ".join(filter(None, [brand, model, color, part_category])).strip()
    if not display_name:
        display_name = (name or "").strip()
    if not display_name:
        display_name = item_code or "Parça"
    return display_name


class WebBridge(QObject):
    """React (JavaScript) ile Python (PySide6) arasındaki köprü sınıfı."""

    def __init__(self, parent=None):
        super().__init__(parent)
        # Uzak DB'ye her Slot çağrısında yeniden gidilmesin diye, neredeyse hiç değişmeyen
        # referans/liste tabloları (statü listesi, depo durum listesi, flow değerleri, görev
        # grupları, garanti türleri, parça kategorileri) için basit bir TTL bellek önbelleği.
        # bkz. _cached_json / _invalidate_cache.
        self._ref_cache = {}
        self._ensure_performance_indexes()
        self._ensure_department_column()
        self._ensure_status_column()
        self._ensure_repair_records_extra_columns()
        self._ensure_batch_entries_diagnosis_column()
        self._ensure_stock_movement_columns()
        self._ensure_service_records_table()
        self._ensure_work_orders_table()
        self._ensure_work_order_type_columns()
        self._ensure_service_id_columns()
        self._ensure_production_work_order_lifecycle_columns()
        self._ensure_material_requests_table()
        self._ensure_production_tables()
        self._ensure_work_order_parts_table()
        self._ensure_statu_history_table()
        self._ensure_location_kind_column()
        self._ensure_system_locations()
        self._ensure_part_category_columns()
        self._ensure_part_extra_columns()
        self._ensure_user_gorev_column()
        self._ensure_user_fullname_column()
        self._ensure_item_bom_data()
        self._ensure_item_model_lookup()
        self._ensure_batch_entries_table()
        self._ensure_label_templates_table()
        self._ensure_customer_decision_transitions()
        self._schema_cache = None
        # Yazdırılacak etiketin kağıt ölçüsü (mm). Ekran, window.print() öncesi
        # set_label_page_size ile günceller; main_window bunu QPrinter'a uygular.
        self.son_etiket_olcusu = None
        # Kullanıcının seçtiği yazıcı kağıt formunun adı ("30384 PC Postage 2-Part"
        # gibi). Boşsa ölçüye en yakın form seçilir. set_label_form ile güncellenir;
        # main_window._kagit_ayarla bunu QPrinter'a uygular.
        self.etiket_form_adi = ""
        # Son yazdırma işinin sonucu. main_window._yazdir doldurur, ekran
        # get_last_print_result ile okur; böylece sessiz başarısızlıklar görünür olur.
        self.son_yazdirma_sonucu = None
        # Yazıcı seçim penceresinden ÖNCE kendi baskı önizlememiz açılsın mı.
        # Windows'un yazdırma penceresindeki önizleme alanı "Bu uygulama yazdırma
        # önizlemesini desteklemiyor" der: o alanı doldurmak uygulamanın WinRT
        # IPrintDocumentPageSource sözleşmesini uygulamasını gerektirir, Qt uygulamaz.
        # Bu yüzden önizlemeyi kendimiz gösteriyoruz - bkz. main_window._baski_onizleme.
        # Ekran, window.print() öncesi set_print_preview ile günceller (otomatik
        # basımda kapatılır; teknisyen barkodu okutunca pencere çıkmamalı).
        self.baski_onizleme_istendi = True
        # Yazdırma penceresinin teması ve basılacak etiket sayısı; ekran
        # set_print_preview ile bildirir.
        self.baski_temasi = "dark"
        self.baski_etiket_sayisi = 0

        # Giriş ekranındaki veritabanı rozeti için önbelleği arka planda ısıt.
        # İlk çağrı ~950 ms sürüyor (motor kurulumu + bağlantı + köprü gidiş-dönüşü) ve
        # rozet o kadar süre "kontrol ediliyor" kalıyordu. Burada peşinen yapılırsa
        # ekran açıldığında cevap hazır olur, çağrı ~3 ms'ye düşer.
        # Ayrı iş parçacığı: pencere açılışını hiçbir koşulda geciktirmemeli.
        try:
            import threading
            threading.Thread(target=self.get_db_status, daemon=True).start()
        except Exception:
            pass  # ısıtma yapılamazsa rozet sadece ilk seferinde yavaş olur

    @Slot(result=str)
    def get_schema_introspection(self):
        """
        Veritabanındaki 'warehouse' şemasını okuyarak Frontend Schema Mapper
        için gerekli { tables: [], edges: [] } JSON yapısını döndürür.
        Önbellekleme (Caching) kullanarak performansı maksimize eder.
        """
        if self._schema_cache:
            return self._schema_cache
            
        try:
            from sqlalchemy import inspect
            db = SessionLocal()
            engine = db.get_bind()
            inspector = inspect(engine)
            schema_name = 'warehouse'
            schema_names = inspector.get_schema_names()
            
            # Eğer 'warehouse' şeması yoksa varsayılan şemayı kullan (None)
            if schema_name not in schema_names:
                schema_name = None
                
            table_names = inspector.get_table_names(schema=schema_name)
            
            # Eğer hala tablo bulunamadıysa boş dön
            if not table_names:
                return json.dumps({"tables": [], "edges": []})
            
            tables = []
            edges = []
            
            x_pos, y_pos = 50, 50
            
            for t_name in table_names:
                columns = inspector.get_columns(t_name, schema=schema_name)
                pk_constraint = inspector.get_pk_constraint(t_name, schema=schema_name)
                pks = pk_constraint.get('constrained_columns', []) if pk_constraint else []
                fks = inspector.get_foreign_keys(t_name, schema=schema_name)
                
                fields = []
                for col in columns:
                    col_name = col['name']
                    # Tip dönüştürme
                    col_type_str = str(col['type']).lower()
                    fe_type = 'string'
                    if 'int' in col_type_str:
                        fe_type = 'int'
                    elif 'bool' in col_type_str:
                        fe_type = 'boolean'
                    elif 'time' in col_type_str or 'date' in col_type_str:
                        fe_type = 'timestamp'
                    
                    is_pk = col_name in pks
                    is_fk = False
                    fk_ref = None
                    
                    # FK işlemleri
                    for fk in fks:
                        if col_name in fk['constrained_columns']:
                            is_fk = True
                            fe_type = 'relation'
                            ref_table = fk['referred_table']
                            ref_col = fk['referred_columns'][0] if fk['referred_columns'] else 'id'
                            fk_ref = {
                                'tableId': f"tbl_{ref_table}",
                                'fieldId': f"f_{ref_table}_{ref_col}"
                            }
                            
                            # Edge oluştur (İstenilen format: id, source, target)
                            edges.append({
                                'id': f"fk-{t_name}-{ref_table}-{col_name}",
                                'sourceTableId': f"tbl_{t_name}",
                                'sourceFieldId': f"f_{t_name}_{col_name}",
                                'targetTableId': f"tbl_{ref_table}",
                                'targetFieldId': f"f_{ref_table}_{ref_col}",
                                'relationType': 'many-to-one'
                            })
                            break
                            
                    fields.append({
                        'id': f"f_{t_name}_{col_name}",
                        'dbName': col_name,
                        'feName': col_name,
                        'type': fe_type,
                        'isPK': is_pk,
                        'isFK': is_fk,
                        'fkRef': fk_ref
                    })
                
                # Tablo feName oluştur (snake_case -> PascalCase)
                fe_table_name = "".join(word.capitalize() for word in t_name.split('_'))
                
                tables.append({
                    'id': f"tbl_{t_name}",
                    'dbName': t_name,
                    'feName': fe_table_name,
                    'x': x_pos,
                    'y': y_pos,
                    'fields': fields
                })
                
                # Izgara (Grid) dizilimi hesapla
                x_pos += 320
                if x_pos > 1500:
                    x_pos = 50
                    y_pos += 350
                    
            db.close()
            
            result_json = json.dumps({'tables': tables, 'edges': edges})
            self._schema_cache = result_json
            return result_json
            
        except Exception as e:
            logging.error(f"[WebBridge] get_schema_introspection hatası: {e}")
            return json.dumps({'tables': [], 'edges': []})

    @Slot(str, str, str, result=str)
    def drop_schema_table(self, table_name, confirm_name, username):
        """Schema Mapper > Görsel Şema'da seçili tabloyu warehouse şemasından KALICI
        olarak siler (DROP TABLE - tablo ve içindeki TÜM veri kaybolur, geri alınamaz).
        Çok yıkıcı bir işlem olduğundan: (1) sadece admin/developer çalıştırabilir,
        (2) confirm_name, table_name ile birebir eşleşmelidir (frontend'in 'tablo
        adını yazarak onayla' modalıyla eşleşir - burada da tekrar doğrulanır, frontend
        bypass edilse bile korunur), (3) tablo gerçekten warehouse şemasında var mı
        introspection ile doğrulanır (rastgele/keyfi SQL çalıştırılmaz), (4) başka
        tablolardan FK ile referans alınıyorsa Postgres DROP'u reddeder, CASCADE
        kullanılmaz - bu durumda hata olduğu gibi kullanıcıya gösterilir."""
        from sqlalchemy import inspect, MetaData, Table
        db = SessionLocal()
        try:
            table_name = (table_name or "").strip()
            confirm_name = (confirm_name or "").strip()
            if not table_name:
                return json.dumps({"success": False, "message": "Tablo adı boş olamaz."})
            if table_name != confirm_name:
                return json.dumps({"success": False, "message": "Onay için yazılan tablo adı eşleşmiyor."})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                return json.dumps({"success": False, "message": "Bu işlem sadece Admin/Developer tarafından yapılabilir."})

            engine = db.get_bind()
            inspector = inspect(engine)
            real_tables = inspector.get_table_names(schema="warehouse")
            if table_name not in real_tables:
                return json.dumps({"success": False, "message": f"'{table_name}' warehouse şemasında bulunamadı."})

            meta = MetaData(schema="warehouse")
            tbl = Table(table_name, meta, autoload_with=engine)
            tbl.drop(engine)

            self._schema_cache = None
            print(f"[WebBridge] UYARI: '{username}' kullanıcısı warehouse.{table_name} tablosunu SİLDİ (DROP TABLE).")
            return json.dumps({"success": True, "message": f"'{table_name}' tablosu kalıcı olarak silindi."}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": f"Tablo silinemedi: {str(e)}"})
        finally:
            db.close()

    def _find_reference_excel_file(self):
        """Proje kök dizininde MioCreate referans veri dosyasını arar.
        Hem eski isimlendirmeyi ('...dosya...') hem de mevcut 'MioCreate.xlsx' adını destekler."""
        import os
        candidates = [
            f for f in os.listdir('.')
            if f.lower().endswith('.xlsx')
            and not f.startswith('~$')
            and ('dosya' in f.lower() or 'miocreate' in f.lower())
        ]
        return candidates[0] if candidates else None

    def _ensure_item_bom_data(self):
        """ItemBOM tablosunun verilerini Excel dosyasından okuyarak veri tabanına senkronize eder."""
        from sqlalchemy import text
        from models.part import Part
        from models.item_bom import ItemBOM
        import openpyxl

        db = SessionLocal()
        try:
            # Check if table already has data
            count = db.execute(text("SELECT COUNT(*) FROM warehouse.item_bom;")).scalar()
            if count > 0:
                return

            print("[WebBridge] ItemBOM tablosu boş. Excel'den veri içe aktarılıyor...")
            fname = self._find_reference_excel_file()
            if not fname:
                print("[WebBridge] Referans Excel dosyası (MioCreate.xlsx) bulunamadı.")
                return
            wb = openpyxl.load_workbook(fname, data_only=True)
            
            # Read Item sheet for names, types
            ws_item = wb['Item']
            item_rows = list(ws_item.iter_rows(values_only=True))
            h_idx = next(i for i, r in enumerate(item_rows) if r and 'code' in [str(x).lower() for x in r])
            headers_item = item_rows[h_idx]
            code_col = next(i for i, h in enumerate(headers_item) if h == 'code')
            shortname_col = next(i for i, h in enumerate(headers_item) if h == 'shortName')
            category_col = next(i for i, h in enumerate(headers_item) if h == 'itemCategory')
            type_col = next(i for i, h in enumerate(headers_item) if h == 'itemType')

            item_info_map = {}
            for r in item_rows[h_idx+1:]:
                item_code_val = r[code_col]
                s_name = r[shortname_col]
                cat_val = r[category_col]
                type_val = r[type_col]
                if item_code_val:
                    item_info_map[str(item_code_val)] = {
                        "name": str(s_name) if s_name else str(item_code_val),
                        "item_category": str(cat_val) if cat_val else None,
                        "part_type": str(type_val) if type_val else None
                    }
            
            # Read ItemBom sheet
            ws_bom = wb['ItemBom']
            bom_rows = list(ws_bom.iter_rows(values_only=True))
            h_idx_bom = next(i for i, r in enumerate(bom_rows) if r and 'UretilenParcaKodu' in [str(x) for x in r])
            headers_bom = bom_rows[h_idx_bom]
            
            parent_col = next(i for i, h in enumerate(headers_bom) if h == 'UretilenParcaKodu')
            child1_col = next(i for i, h in enumerate(headers_bom) if h == 'Tuketilen Parca_1')
            qty1_col = next(i for i, h in enumerate(headers_bom) if h == 'Tuketilen Parca_1_Miktar')
            child2_col = next(i for i, h in enumerate(headers_bom) if h == 'Tuketilen Parca_2')
            qty2_col = next(i for i, h in enumerate(headers_bom) if h == 'Tuketilen Parca_2_Miktar')
            
            bom_data = []
            unique_codes = set()
            
            for r in bom_rows[h_idx_bom+1:]:
                parent = r[parent_col]
                child1 = r[child1_col]
                qty1 = r[qty1_col]
                child2 = r[child2_col]
                qty2 = r[qty2_col]
                
                if not parent:
                    continue
                
                unique_codes.add(parent)
                children = []
                if child1:
                    unique_codes.add(child1)
                    children.append((child1, int(qty1) if qty1 else 1))
                if child2:
                    unique_codes.add(child2)
                    children.append((child2, int(qty2) if qty2 else 1))
                    
                bom_data.append({
                    'parent': parent,
                    'children': children
                })
            
            wb.close()
            
            # Insert missing parts
            existing_parts = db.query(Part).filter(Part.item_code.in_(list(unique_codes))).all()
            existing_codes = {p.item_code for p in existing_parts}
            missing_codes = unique_codes - existing_codes
            
            for code in missing_codes:
                info = item_info_map.get(code, {"name": code, "item_category": None, "part_type": None})
                new_part = Part(
                    item_code=code,
                    name=info["name"],
                    item_category=info.get("item_category"),
                    part_type=info["part_type"],
                    status="Aktif",
                    stock_tracking_type="Stok Takipli",
                    critical_limit=10
                )
                db.add(new_part)
            
            if missing_codes:
                db.commit()
            
            # Insert BOMs
            for item in bom_data:
                parent = item['parent']
                for child, qty in item['children']:
                    new_bom = ItemBOM(
                        parent_item_id=parent,
                        child_item_id=child,
                        quantity=qty
                    )
                    db.add(new_bom)
            
            db.commit()
            print(f"[WebBridge] ItemBOM Excel verisi başarıyla senkronize edildi. Toplam {len(bom_data)} reçete eklendi.")
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] ItemBOM senkronizasyon hatası: {e}")
        finally:
            db.close()

    def _insert_item_models_batch(self, db, batch):
        """warehouse.item_models tablosuna tek seferde birden çok satır ekler.
        30k satırı tek tek INSERT etmek uzak veritabanına çok fazla round-trip
        açıp bağlantının zaman aşımına uğramasına/kopmasına yol açıyordu."""
        from sqlalchemy import text
        values_sql = ", ".join(f"(:code{i}, :model{i}, :brand{i})" for i in range(len(batch)))
        params = {}
        for i, row in enumerate(batch):
            params[f"code{i}"] = row["code"]
            params[f"model{i}"] = row["model"]
            params[f"brand{i}"] = row["brand"]
        db.execute(text(f"""
            INSERT INTO warehouse.item_models (item_code, model, brand)
            VALUES {values_sql}
            ON CONFLICT (item_code) DO UPDATE SET model = EXCLUDED.model, brand = EXCLUDED.brand;
        """), params)

    def _ensure_item_model_lookup(self):
        """Parça kodu girildiğinde 'Model' alanının otomatik doldurulabilmesi için
        ProductBom (item -> productFamily) ve ProductFamily (productFamily -> shortName)
        sayfalarından warehouse.item_models (item_code -> model) eşleşme tablosunu kurar."""
        from sqlalchemy import text
        import openpyxl

        db = SessionLocal()
        try:
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.item_models (
                    item_code VARCHAR(100) PRIMARY KEY,
                    model TEXT
                );
            """))
            db.execute(text("ALTER TABLE warehouse.item_models ALTER COLUMN model TYPE TEXT;"))
            db.execute(text("ALTER TABLE warehouse.item_models ADD COLUMN IF NOT EXISTS brand TEXT;"))
            db.commit()

            # brand sütunu sonradan eklendiği için, marka verisi henüz işlenmemiş
            # kurulumlarda (model dolu ama brand boş) tabloyu yeniden içe aktarmamız gerekir.
            count = db.execute(text("SELECT COUNT(*) FROM warehouse.item_models WHERE brand IS NOT NULL AND brand <> '';")).scalar()
            if count > 0:
                return

            fname = self._find_reference_excel_file()
            if not fname:
                print("[WebBridge] Referans Excel dosyası (MioCreate.xlsx) bulunamadı.")
                return

            print("[WebBridge] item_models tablosu boş. Excel'den Parça Kodu -> Model eşleşmesi içe aktarılıyor...")
            wb = openpyxl.load_workbook(fname, data_only=True)

            # ProductFamily: kod (örn. 'iP11') -> okunabilir model adı (örn. 'iPhone 11')
            ws_family = wb['ProductFamily']
            family_rows = list(ws_family.iter_rows(values_only=True))
            h_idx_family = next(i for i, r in enumerate(family_rows) if r and 'code' in [str(x).lower() for x in r])
            headers_family = family_rows[h_idx_family]
            fam_code_col = next(i for i, h in enumerate(headers_family) if h == 'code')
            fam_shortname_col = next(i for i, h in enumerate(headers_family) if h == 'shortName')
            fam_brand_col = next((i for i, h in enumerate(headers_family) if h == 'brand'), None)

            family_name_map = {}
            family_brand_map = {}
            for r in family_rows[h_idx_family + 1:]:
                fam_code = r[fam_code_col]
                fam_name = r[fam_shortname_col]
                if fam_code:
                    family_name_map[str(fam_code)] = str(fam_name) if fam_name else str(fam_code)
                    if fam_brand_col is not None and r[fam_brand_col]:
                        family_brand_map[str(fam_code)] = str(r[fam_brand_col])

            # ProductBom: item_code -> productFamily kodları (bir parça birden fazla modelde kullanılabilir)
            ws_bom = wb['ProductBom']
            bom_rows = list(ws_bom.iter_rows(values_only=True))
            h_idx_bom = next(i for i, r in enumerate(bom_rows) if r and 'item' in [str(x).lower() for x in r])
            headers_bom = bom_rows[h_idx_bom]
            item_col = next(i for i, h in enumerate(headers_bom) if h == 'item')
            family_col = next(i for i, h in enumerate(headers_bom) if h == 'productFamily')

            item_families = {}
            for r in bom_rows[h_idx_bom + 1:]:
                item_code = r[item_col]
                fam_code = r[family_col]
                if not item_code or not fam_code:
                    continue
                item_families.setdefault(str(item_code), set()).add(str(fam_code))

            wb.close()

            inserted = 0
            batch = []
            batch_size = 500
            for item_code, fam_codes in item_families.items():
                model_names = sorted({family_name_map.get(fc, fc) for fc in fam_codes})
                model_str = ', '.join(model_names)
                if not model_str:
                    continue
                brand_names = sorted({family_brand_map[fc] for fc in fam_codes if fc in family_brand_map})
                brand_str = ', '.join(brand_names)
                batch.append({"code": item_code, "model": model_str, "brand": brand_str or None})
                if len(batch) >= batch_size:
                    self._insert_item_models_batch(db, batch)
                    inserted += len(batch)
                    batch = []
            if batch:
                self._insert_item_models_batch(db, batch)
                inserted += len(batch)

            db.commit()
            print(f"[WebBridge] item_models eşleşmesi tamamlandı. Toplam {inserted} parça kodu için model belirlendi.")
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] item_models senkronizasyon hatası: {e}")
        finally:
            db.close()

    def _ensure_user_gorev_column(self):
        """warehouse.users tablosuna gorev sütunu yoksa ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.users ADD COLUMN IF NOT EXISTS gorev VARCHAR(100);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] users.gorev kolonu eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_user_fullname_column(self):
        """warehouse.users tablosuna fullname sütunu yoksa ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.users ADD COLUMN IF NOT EXISTS fullname VARCHAR(150);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] users.fullname kolonu eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_production_tables(self):
        """warehouse.production_runs ve production_materials tablolarını yoksa oluşturur."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.production_runs (
                    id SERIAL PRIMARY KEY,
                    target_part_id INTEGER REFERENCES warehouse.parts(id),
                    quantity_produced INTEGER NOT NULL,
                    location_id INTEGER REFERENCES warehouse.locations(id),
                    source_location_id INTEGER REFERENCES warehouse.locations(id),
                    produced_by VARCHAR(150),
                    notes TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """))
            db.execute(text("""
                ALTER TABLE warehouse.production_runs 
                ADD COLUMN IF NOT EXISTS source_location_id INTEGER REFERENCES warehouse.locations(id);
            """))
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.production_materials (
                    id SERIAL PRIMARY KEY,
                    production_run_id INTEGER REFERENCES warehouse.production_runs(id) ON DELETE CASCADE,
                    part_id INTEGER REFERENCES warehouse.parts(id),
                    quantity_consumed INTEGER NOT NULL
                );
            """))
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.produced_units (
                    id SERIAL PRIMARY KEY,
                    production_run_id INTEGER REFERENCES warehouse.production_runs(id) ON DELETE CASCADE,
                    serial_number VARCHAR(100) NOT NULL,
                    is_returned BOOLEAN DEFAULT FALSE,
                    return_reason VARCHAR(500),
                    returned_at TIMESTAMP WITH TIME ZONE,
                    return_location_id INTEGER REFERENCES warehouse.locations(id)
                );
            """))
            db.execute(text("ALTER TABLE warehouse.produced_units DROP CONSTRAINT IF EXISTS produced_units_serial_number_key;"))
            db.execute(text("ALTER TABLE warehouse.produced_units ADD COLUMN IF NOT EXISTS is_returned BOOLEAN DEFAULT FALSE;"))
            db.execute(text("ALTER TABLE warehouse.produced_units ADD COLUMN IF NOT EXISTS return_reason VARCHAR(500);"))
            db.execute(text("ALTER TABLE warehouse.produced_units ADD COLUMN IF NOT EXISTS returned_at TIMESTAMP WITH TIME ZONE;"))
            db.execute(text("ALTER TABLE warehouse.produced_units ADD COLUMN IF NOT EXISTS return_location_id INTEGER REFERENCES warehouse.locations(id);"))
            db.execute(text("ALTER TABLE warehouse.produced_units ADD COLUMN IF NOT EXISTS returned_materials VARCHAR(2000);"))
            db.execute(text("ALTER TABLE warehouse.produced_units ADD COLUMN IF NOT EXISTS replacement_requested_qty INTEGER DEFAULT 0;"))

            # Phonecheck "Parts" ham JSON'u - kritik parca orijinallik kontrolunun kaynagi.
            # Eski satirlar NULL kalir (o kayitlarda parca durumu "Belirtilmemis" gorunur);
            # yeni her Phonecheck sorgusunda dolar.
            db.execute(text("ALTER TABLE warehouse.phonecheck_test_results ADD COLUMN IF NOT EXISTS parts TEXT;"))
            
            # Clean up old records to avoid data inconsistency with the new unique serial number system
            run_count = db.execute(text("SELECT COUNT(*) FROM warehouse.production_runs")).scalar() or 0
            unit_count = db.execute(text("SELECT COUNT(*) FROM warehouse.produced_units")).scalar() or 0
            if run_count > 0 and unit_count == 0:
                db.execute(text("TRUNCATE warehouse.production_runs RESTART IDENTITY CASCADE;"))

            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] production tabloları oluşturulamadı: {e}")
        finally:
            db.close()

    def _ensure_work_orders_table(self):
        """warehouse.work_orders tablosu yoksa oluşturur."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.work_orders (
                    id SERIAL PRIMARY KEY,
                    service_record_id INTEGER REFERENCES warehouse.service_records(id),
                    description TEXT,
                    assigned_technician VARCHAR(150),
                    priority VARCHAR(20) DEFAULT 'Orta',
                    start_date DATE,
                    end_date DATE,
                    parts_used TEXT,
                    status VARCHAR(30) DEFAULT 'Beklemede',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """))
            db.execute(text("ALTER TABLE warehouse.work_orders ADD COLUMN IF NOT EXISTS source_location_id INTEGER REFERENCES warehouse.locations(id);"))
            db.execute(text("ALTER TABLE warehouse.work_orders ADD COLUMN IF NOT EXISTS stock_settled_at TIMESTAMP;"))
            db.execute(text("ALTER TABLE warehouse.work_orders ADD COLUMN IF NOT EXISTS return_reason VARCHAR(500);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] work_orders tablosu oluşturulamadı: {e}")
        finally:
            db.close()

    def _ensure_work_order_type_columns(self):
        """warehouse.work_orders tablosuna work_order_type ve target_part_id sütunlarını ekler.

        work_order_type: DEFAULT 'SERVICE' olduğu için mevcut kayıtlar ve mevcut Service
        Work Order akışı (create_work_order/update_work_order) hiç değişmeden çalışmaya
        devam eder. PRODUCTION tipi için service_record_id NULL kalır; bunun yerine
        target_part_id üzerinden bir Recipe'ye (warehouse.item_bom, parent_item_id =
        hedef parçanın item_code'u) bağlanır.
        """
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                ALTER TABLE warehouse.work_orders
                ADD COLUMN IF NOT EXISTS work_order_type VARCHAR(20) NOT NULL DEFAULT 'SERVICE';
            """))
            db.execute(text("""
                ALTER TABLE warehouse.work_orders
                ADD COLUMN IF NOT EXISTS target_part_id INTEGER REFERENCES warehouse.parts(id);
            """))
            db.execute(text("""
                ALTER TABLE warehouse.work_orders
                ADD COLUMN IF NOT EXISTS planned_quantity INTEGER;
            """))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_work_orders_type ON warehouse.work_orders(work_order_type);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] work_order_type sütunları eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_production_work_order_lifecycle_columns(self):
        """warehouse.work_orders tablosuna Production Work Order'ın üretim yaşam
        döngüsü (BEKLIYOR -> URETIMDE -> TAMAMLANDI) için gereken sütunları ekler.
        Hepsi nullable olduğu için Service Work Order kayıtlarında hep NULL kalır ve
        mevcut Service akışı hiç etkilenmez."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.work_orders ADD COLUMN IF NOT EXISTS started_at TIMESTAMP;"))
            db.execute(text("ALTER TABLE warehouse.work_orders ADD COLUMN IF NOT EXISTS completed_at TIMESTAMP;"))
            db.execute(text("ALTER TABLE warehouse.work_orders ADD COLUMN IF NOT EXISTS produced_quantity INTEGER;"))
            db.execute(text("ALTER TABLE warehouse.work_orders ADD COLUMN IF NOT EXISTS scrap_quantity INTEGER;"))
            db.execute(text("ALTER TABLE warehouse.work_orders ADD COLUMN IF NOT EXISTS production_notes TEXT;"))
            db.execute(text("UPDATE warehouse.work_orders SET status = 'URETIMDE' WHERE work_order_type = 'PRODUCTION' AND status = 'BEKLIYOR';"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] production work order lifecycle sütunları eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_material_requests_table(self):
        """warehouse.material_requests tablosunu yoksa oluşturur. Bir Production Work
        Order'ın Recipe'sindeki (item_bom) her satır için bir Material Request kaydı
        tutulur. remaining_quantity kalıcı sütun değil; okuma sırasında
        (required_quantity + fire_quantity - issued_quantity) olarak hesaplanır.
        fire_quantity: teknisyenden fire olarak DOA Stock'a iade edilip sisteme
        işlenmiş miktardır (bkz. report_material_fire) — depocunun ek teslim
        yapabileceği miktarı, fire bildirilmeden büyütmeden, açar."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.material_requests (
                    id SERIAL PRIMARY KEY,
                    work_order_id INTEGER NOT NULL REFERENCES warehouse.work_orders(id) ON DELETE CASCADE,
                    part_id INTEGER NOT NULL REFERENCES warehouse.parts(id),
                    required_quantity INTEGER NOT NULL,
                    issued_quantity INTEGER NOT NULL DEFAULT 0,
                    status VARCHAR(20) NOT NULL DEFAULT 'WAITING',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """))
            db.execute(text("ALTER TABLE warehouse.material_requests ADD COLUMN IF NOT EXISTS fire_quantity INTEGER NOT NULL DEFAULT 0;"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_material_requests_work_order_id ON warehouse.material_requests(work_order_id);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] material_requests tablosu oluşturulamadı: {e}")
        finally:
            db.close()

    def _ensure_work_order_parts_table(self):
        """warehouse.work_order_parts tablosu yoksa oluşturur (Parça Tedarik Durumu modülü)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.work_order_parts (
                    id SERIAL PRIMARY KEY,
                    work_order_id INTEGER NOT NULL REFERENCES warehouse.work_orders(id) ON DELETE CASCADE,
                    part_id INTEGER NOT NULL REFERENCES warehouse.parts(id),
                    quantity INTEGER NOT NULL DEFAULT 1,
                    status VARCHAR(30) NOT NULL DEFAULT 'Stokta Var',
                    delivered_location_id INTEGER REFERENCES warehouse.locations(id),
                    delivery_movement_id INTEGER REFERENCES warehouse.stock_movements(id),
                    delivered_by VARCHAR(150),
                    delivered_at TIMESTAMP,
                    waiting_notes TEXT,
                    marked_waiting_by VARCHAR(150),
                    marked_waiting_at TIMESTAMP,
                    reversal_movement_id INTEGER REFERENCES warehouse.stock_movements(id),
                    reverted_by VARCHAR(150),
                    reverted_at TIMESTAMP,
                    requested_by VARCHAR(150),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """))
            db.execute(text("""
                ALTER TABLE warehouse.work_order_parts
                ADD COLUMN IF NOT EXISTS delivered_location_id INTEGER REFERENCES warehouse.locations(id);
            """))
            db.execute(text("""
                ALTER TABLE warehouse.work_order_parts
                ADD COLUMN IF NOT EXISTS delivery_movement_id INTEGER REFERENCES warehouse.stock_movements(id);
            """))
            db.execute(text("""
                ALTER TABLE warehouse.work_order_parts
                ADD COLUMN IF NOT EXISTS reversal_movement_id INTEGER REFERENCES warehouse.stock_movements(id);
            """))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_wop_work_order_id ON warehouse.work_order_parts(work_order_id);"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_wop_status ON warehouse.work_order_parts(status);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] work_order_parts tablosu oluşturulamadı: {e}")
        finally:
            db.close()

    def _ensure_statu_history_table(self):
        """warehouse.batch_entry_statu_history tablosunu yoksa oluşturur. Cihazın her
        statü geçişini kalıcı loglar (bkz. models/batch_entry_statu_history.py ve
        get_status_history_by_imei)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.batch_entry_statu_history (
                    id SERIAL PRIMARY KEY,
                    batch_entry_id INTEGER,
                    imei VARCHAR(100),
                    old_statu_code INTEGER,
                    new_statu_code INTEGER NOT NULL,
                    staff_name VARCHAR(100),
                    note TEXT,
                    created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                );
            """))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_beh_batch_entry_id ON warehouse.batch_entry_statu_history(batch_entry_id);"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_beh_imei ON warehouse.batch_entry_statu_history(imei);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] batch_entry_statu_history tablosu oluşturulamadı: {e}")
        finally:
            db.close()

    def _record_statu_change(self, db, entry_id, imei, old_code, new_code, staff=None, note=None):
        """Bir statü geçişini history tablosuna ekler (COMMIT ETMEZ — çağıran commit eder,
        böylece geçişle aynı transaction'da atomik kalır). Log yazımı asla asıl işlemi
        bozmamalı; hata olursa sessizce yutulur."""
        try:
            from models.batch_entry_statu_history import BatchEntryStatuHistory
            db.add(BatchEntryStatuHistory(
                batch_entry_id=int(entry_id) if entry_id is not None else None,
                imei=(imei or None),
                old_statu_code=(int(old_code) if old_code is not None else None),
                new_statu_code=int(new_code),
                staff_name=(staff or None),
                note=(note or None),
            ))
        except Exception as e:
            print(f"[WebBridge] statü geçmişi log yazılamadı: {e}")

    def _ensure_service_records_table(self):
        """warehouse.service_records tablosu yoksa oluşturur."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.service_records (
                    id SERIAL PRIMARY KEY,
                    customer_name VARCHAR(150) NOT NULL,
                    customer_phone VARCHAR(30),
                    customer_email VARCHAR(150),
                    company VARCHAR(150),
                    brand VARCHAR(100),
                    model VARCHAR(100),
                    memory VARCHAR(50),
                    product_code VARCHAR(100),
                    color VARCHAR(50),
                    fault_category VARCHAR(100),
                    fault_type VARCHAR(150),
                    customer_complaint TEXT,
                    preliminary_diagnosis TEXT,
                    status VARCHAR(30) DEFAULT 'Arıza Kabul',
                    technician_note TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """))
            db.execute(text("ALTER TABLE warehouse.service_records ADD COLUMN IF NOT EXISTS memory VARCHAR(50);"))
            db.execute(text("ALTER TABLE warehouse.service_records ADD COLUMN IF NOT EXISTS product_code VARCHAR(100);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] service_records tablosu oluşturulamadı: {e}")
        finally:
            db.close()

    def _ensure_stock_movement_columns(self):
        """warehouse.stock_movements tablosuna technician, description ve movement_kind sütunlarını ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.stock_movements ADD COLUMN IF NOT EXISTS technician VARCHAR(150);"))
            db.execute(text("ALTER TABLE warehouse.stock_movements ADD COLUMN IF NOT EXISTS description TEXT;"))
            db.execute(text("ALTER TABLE warehouse.stock_movements ADD COLUMN IF NOT EXISTS movement_kind VARCHAR(20);"))
            # İşlem sonrası kalan miktar - Raporlar > Transfer Hareketleri ekranındaki
            # "kaynak/hedef kalan" sütunları. Eski satırlarda NULL kalır; o dönemin
            # bakiyesi geriye dönük hesaplanamıyor (bkz. config.database'deki açıklama).
            db.execute(text("ALTER TABLE warehouse.stock_movements ADD COLUMN IF NOT EXISTS source_balance_after INTEGER;"))
            db.execute(text("ALTER TABLE warehouse.stock_movements ADD COLUMN IF NOT EXISTS target_balance_after INTEGER;"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] stock_movements kolonları eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_label_templates_table(self):
        """Etiket şablonlarını tutar. Şablonlar veritabanında durur ki tasarım bir kez
        yapılıp tüm makinelerde aynı çıksın; kod değişikliği gerekmesin."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.label_templates (
                    key         VARCHAR(40) PRIMARY KEY,
                    name        VARCHAR(120),
                    width_mm    NUMERIC(6,2) NOT NULL,
                    height_mm   NUMERIC(6,2) NOT NULL,
                    html        TEXT NOT NULL,
                    updated_by  VARCHAR(100),
                    updated_at  TIMESTAMPTZ DEFAULT NOW()
                );
            """))
            # Tasarımın 90 derece döndürülerek basılıp basılmayacağı. Etiket rulosu
            # fiziksel olarak dikeydir (54 mm geniş); yatay bir tasarım ancak
            # döndürülerek sığar - bkz. EtiketYazdirModal::yazdirmaCss.
            db.execute(text("ALTER TABLE warehouse.label_templates "
                            "ADD COLUMN IF NOT EXISTS rotate BOOLEAN DEFAULT FALSE;"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] label_templates tablosu olusturulamadi: {e}")
        finally:
            db.close()

    def _ensure_customer_decision_transitions(self):
        """136 (Müşteri Onay/Red Geldi) statüsünün çıkış geçişlerini garanti eder.

        Akış şemasına göre 136'dan iki yol çıkar:
          136 -> 109  "İade Edilmeyecek - Müşteri Onayı Geldi"  (onay: cihaz üretime girer)
          136 -> 124  "İade Edilecek - Müşteri Reddetti"        (red: onarılmadan son teste)

        Sahadaki durum: 136_109 tanımlıydı ama KAPALI, 136_124 ise hiç yoktu. Yani müşteri
        onayına giden cihaz 136'ya düştüğü anda kilitleniyordu - 136'dan çıkan aktif hiçbir
        geçiş yoktu. Burası idempotenttir; her açılışta çalışır, varsa dokunmaz.
        """
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                UPDATE warehouse.service_statu_map
                   SET enabled = TRUE
                 WHERE parent_statu = 136 AND child_statu = 109 AND enabled IS NOT TRUE;
            """))
            db.execute(text("""
                INSERT INTO warehouse.service_statu_map
                    (id, code, parent_statu, child_statu, is_positive, is_user_change_statu,
                     to_dest, short_name, full_name, description, enabled, order_number)
                SELECT gen_random_uuid(), '136_124', 136, 124, FALSE, TRUE,
                       'MNG1_AS', 'İade Edilecek - Müşteri Reddetti',
                       'İade Edilecek - Müşteri Reddetti',
                       'Musteri onarimi reddetti; cihaz uretime girmeden son teste teslim edilir.',
                       TRUE, 1
                 WHERE NOT EXISTS (
                    SELECT 1 FROM warehouse.service_statu_map
                     WHERE parent_statu = 136 AND child_statu = 124
                 );
            """))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] 136 musteri karari gecisleri hazirlanamadi: {e}")
        finally:
            db.close()

    @Slot(result=str)
    def get_label_templates(self):
        """Kayıtlı etiket şablonlarını döner. Kayıt yoksa ekran gömülü varsayılanı kullanır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT key, name, width_mm, height_mm, html, rotate, updated_by, updated_at
                FROM warehouse.label_templates
            """)).mappings().all()
            return json.dumps({"success": True, "templates": [{
                "key": r["key"], "name": r["name"] or "",
                "widthMm": float(r["width_mm"]), "heightMm": float(r["height_mm"]),
                "html": r["html"] or "",
                "rotate": bool(r["rotate"]),
                "updatedBy": r["updated_by"] or "",
                "updatedAt": r["updated_at"].strftime("%d.%m.%Y %H:%M") if r["updated_at"] else "",
            } for r in rows]}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, float, float, str, bool, str, result=str)
    def save_label_template(self, key, name, width_mm, height_mm, html, rotate, username):
        """Bir etiket şablonunu kaydeder (varsa günceller)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            k = (key or "").strip()
            if not k:
                return json.dumps({"success": False, "message": "Şablon anahtarı boş olamaz."})
            if not (html or "").strip():
                return json.dumps({"success": False, "message": "Şablon içeriği boş olamaz."})
            if float(width_mm) <= 0 or float(height_mm) <= 0:
                return json.dumps({"success": False, "message": "Ölçüler sıfırdan büyük olmalı."})
            db.execute(text("""
                INSERT INTO warehouse.label_templates (key, name, width_mm, height_mm, html, rotate, updated_by, updated_at)
                VALUES (:k, :n, :w, :h, :html, :r, :u, NOW())
                ON CONFLICT (key) DO UPDATE SET
                    name = EXCLUDED.name, width_mm = EXCLUDED.width_mm,
                    height_mm = EXCLUDED.height_mm, html = EXCLUDED.html,
                    rotate = EXCLUDED.rotate,
                    updated_by = EXCLUDED.updated_by, updated_at = NOW()
            """), {"k": k, "n": (name or "").strip(), "w": float(width_mm),
                   "h": float(height_mm), "html": html, "r": bool(rotate),
                   "u": (username or "").strip()})
            db.commit()
            return json.dumps({"success": True, "message": "Şablon kaydedildi."}, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_label_template(self, key):
        """Şablonu siler - ekran o etiket için gömülü varsayılana geri döner."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("DELETE FROM warehouse.label_templates WHERE key = :k"),
                       {"k": (key or "").strip()})
            db.commit()
            return json.dumps({"success": True, "message": "Varsayılan şablona dönüldü."},
                              ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    def _ensure_location_kind_column(self):
        """warehouse.locations tablosuna kind sütunu yoksa ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.locations ADD COLUMN IF NOT EXISTS kind VARCHAR(20);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] locations.kind kolonu eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_system_locations(self):
        """Good/DOA/Repair/Scrap/Out Stock sistem depolarını yoksa oluşturur."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            for kind, name in SYSTEM_LOCATION_KINDS.items():
                db.execute(text("""
                    INSERT INTO warehouse.locations (name, kind)
                    SELECT :name, :kind
                    WHERE NOT EXISTS (SELECT 1 FROM warehouse.locations WHERE kind = :kind)
                """), {"name": name, "kind": kind})
                db.execute(text("""
                    UPDATE warehouse.locations SET name = :name WHERE kind = :kind
                """), {"name": name, "kind": kind})
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] sistem depoları oluşturulamadı: {e}")
        finally:
            db.close()

    def _ensure_part_category_columns(self):
        """warehouse.part_categories tablosuna Parça Kategorisi modülü sütunlarını ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("CREATE SCHEMA IF NOT EXISTS warehouse;"))
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.part_categories (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(100) NOT NULL
                );
            """))
            db.execute(text("ALTER TABLE warehouse.part_categories ADD COLUMN IF NOT EXISTS departments VARCHAR(255);"))
            db.execute(text("ALTER TABLE warehouse.part_categories ADD COLUMN IF NOT EXISTS stock_tracking_type VARCHAR(20) DEFAULT 'Stok Takipli';"))
            db.execute(text("ALTER TABLE warehouse.part_categories ADD COLUMN IF NOT EXISTS is_active BOOLEAN DEFAULT true;"))
            db.execute(text("ALTER TABLE warehouse.part_categories ADD COLUMN IF NOT EXISTS description TEXT;"))
            db.execute(text("ALTER TABLE warehouse.part_categories ADD COLUMN IF NOT EXISTS part_type VARCHAR(100);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] part_categories kolonları eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_part_extra_columns(self):
        """warehouse.parts tablosuna part_category_id, barcode ve part_type sütunlarını ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.parts ADD COLUMN IF NOT EXISTS part_category_id INTEGER REFERENCES warehouse.part_categories(id);"))
            db.execute(text("ALTER TABLE warehouse.parts ADD COLUMN IF NOT EXISTS barcode VARCHAR(100);"))
            db.execute(text("ALTER TABLE warehouse.parts ADD COLUMN IF NOT EXISTS part_type VARCHAR(100);"))
            db.execute(text("ALTER TABLE warehouse.parts ADD COLUMN IF NOT EXISTS brand VARCHAR(100);"))
            db.execute(text("ALTER TABLE warehouse.parts ADD COLUMN IF NOT EXISTS model VARCHAR(100);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] parts ek kolonları eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_department_column(self):
        """warehouse.parts tablosuna department sütunu yoksa ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.parts ADD COLUMN IF NOT EXISTS department VARCHAR(255);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] department kolonu eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_status_column(self):
        """warehouse.parts tablosuna status sütunu yoksa ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.parts ADD COLUMN IF NOT EXISTS status VARCHAR(20) DEFAULT 'Aktif';"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] status kolonu eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_performance_indexes(self):
        """Bu oturumdaki performans taramasında saptanan, indekssiz oldukları için yavaş
        çalışan sorgulara ait indeksleri ekler: item_models.item_code (get_item_codes'un
        ORDER BY'ı ~30 bin satırı indeksiz sıralıyordu), stock(part_id, location_id)
        (get_stock_for_part gibi parça bazlı stok sorguları için), ve batch_entries -
        warehouse.batch_entries id dışında HİÇ indeksi yoktu (7600+ satıra büyüdü ve
        büyümeye devam ediyor), oysa bu tablo IMEI/seri no/internal id/batch no ile
        LOWER(TRIM(...)) karşılaştırmasıyla bu oturumda eklenen HEMEN HEMEN HER Slot'ta
        (lookup_batch_entry, create_batch_entry'nin aktif-servis kontrolü,
        get_repair_operations_by_imei, sync_customers_to_batch_entries vb.) sorgulanıyor -
        LOWER(TRIM(...)) ifadesi kullanıldığından düz kolon indeksi değil, ifade (expression)
        indeksi gerekiyor."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_item_models_item_code ON warehouse.item_models (item_code);"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_stock_part_location ON warehouse.stock (part_id, location_id);"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_batch_entries_imei_lower ON warehouse.batch_entries (LOWER(TRIM(imei_number)));"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_batch_entries_serial_lower ON warehouse.batch_entries (LOWER(TRIM(serial_number)));"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_batch_entries_internal_lower ON warehouse.batch_entries (LOWER(TRIM(internal_id)));"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_batch_entries_batch_no_lower ON warehouse.batch_entries (LOWER(TRIM(batch_no)));"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_batch_entries_customer_no_lower ON warehouse.batch_entries (LOWER(TRIM(customer_no)));"))
            # repair_records şu an küçük (bu oturumda ~16 satır) ama batch_entries'in
            # aynı oturum içinde 82'den 7600+'e çıktığı gözlemlendiğinden, esas arama
            # sütununa (service_record_id - get_repair_operations_by_imei,
            # submit_dismantle_decision, get_repair_records vb. hemen her yerde
            # kullanılıyor) tablo büyümeden önce indeks eklenir.
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_repair_records_service_record_id ON warehouse.repair_records (service_record_id);"))
            # product_bom_node (33.800+ satır) PK dışında hiç indekse sahip değildi;
            # get_product_boms (sayfalı liste + COUNT(*)) ve get_parts_for_device
            # (Demontaj ekranındaki 'Parça Seçiniz' - LOWER(TRIM(...)) ile eşleşiyor)
            # her ikisi de parent_product_code üzerinden filtreliyor, get_product_boms
            # ayrıca child_item_code'u warehouse.item ile JOIN'liyor.
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_product_bom_node_parent ON warehouse.product_bom_node (parent_product_code);"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_product_bom_node_parent_lower ON warehouse.product_bom_node (LOWER(TRIM(parent_product_code)));"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_product_bom_node_child ON warehouse.product_bom_node (child_item_code);"))
            # warehouse.parts (30 bin+ satır): Müşteri Fiyat Matrisi'ndeki marka/kategori
            # filtreleri UPPER(brand) eşitliği ve item_category eşitliği ile sorguluyor.
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_parts_brand_upper ON warehouse.parts (UPPER(brand));"))
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_parts_item_category ON warehouse.parts (item_category);"))
            # Marka + Model (get_price_matrix_models, get_price_matrix_categories,
            # get_price_matrix_items - modele göre daraltma) birlikte sorgulanıyor.
            db.execute(text("CREATE INDEX IF NOT EXISTS idx_parts_brand_upper_model ON warehouse.parts (UPPER(brand), model);"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] performans indeksleri eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_repair_records_extra_columns(self):
        """warehouse.repair_records tablosuna Demontaj ekranı için part_item_code/item_fault_code,
        Onarım Parçaları ekranı için supply_status_code (warehouse.item_supply_status.code -
        Depo Durum), Depo > Parça Teslim ekranı için supply_requested_by/supply_requested_at
        (Depo Durum'u en son kim/ne zaman değiştirdi) ve Teknisyene Atama ekranı için
        assigned_technician/assigned_by/assigned_at (kaydın hangi teknisyene, kim tarafından,
        ne zaman atandığı - statü 1001 'Teknisyene Atandı' ile birlikte yazılır) sütunlarını
        yoksa ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.repair_records ADD COLUMN IF NOT EXISTS part_item_code VARCHAR(100);"))
            db.execute(text("ALTER TABLE warehouse.repair_records ADD COLUMN IF NOT EXISTS item_fault_code VARCHAR(255);"))
            db.execute(text("ALTER TABLE warehouse.repair_records ADD COLUMN IF NOT EXISTS supply_status_code VARCHAR(255);"))
            db.execute(text("ALTER TABLE warehouse.repair_records ADD COLUMN IF NOT EXISTS supply_requested_by VARCHAR(100);"))
            db.execute(text("ALTER TABLE warehouse.repair_records ADD COLUMN IF NOT EXISTS supply_requested_at TIMESTAMP;"))
            # Teknisyene Atama - work_orders.assigned_technician ile ayni isimlendirme
            db.execute(text("ALTER TABLE warehouse.repair_records ADD COLUMN IF NOT EXISTS assigned_technician VARCHAR(150);"))
            db.execute(text("ALTER TABLE warehouse.repair_records ADD COLUMN IF NOT EXISTS assigned_by VARCHAR(100);"))
            db.execute(text("ALTER TABLE warehouse.repair_records ADD COLUMN IF NOT EXISTS assigned_at TIMESTAMP;"))
            # Geriye dönük düzeltme: Onarım Havuzu ekranı eskiden teknisyen atamasını
            # yanlışlıkla supply_requested_by'a yazıyordu (bu sütun aslında Parça Teslim'de
            # depo durumunu kimin değiştirdiğini tutar). Bu yüzden havuzdan atanan teknisyen
            # Üretim Kaydını Görüntüle ekranında (assigned_technician okur) görünmüyor, hatta
            # depocu parça durumunu değiştirince üzerine yazılıp kayboluyordu. Atama artık
            # kanonik assigned_technician sütununa yazılıyor; henüz taşınmamış 1001 kayıtları
            # (assigned_technician boş ama supply_requested_by geçerli bir kullanıcı) buraya kopyalanır.
            db.execute(text("""
                UPDATE warehouse.repair_records rr
                SET assigned_technician = rr.supply_requested_by
                WHERE rr.repair_result_type_code = 1001
                  AND (rr.assigned_technician IS NULL OR TRIM(rr.assigned_technician) = '')
                  AND rr.supply_requested_by IS NOT NULL
                  AND EXISTS (SELECT 1 FROM warehouse.users u WHERE u.username = rr.supply_requested_by)
            """))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] repair_records ek kolonları eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_batch_entries_diagnosis_column(self):
        """warehouse.batch_entries tablosuna customer_diagnosis sütunu yoksa ekler.
        Bağlı bir service_records/work_order zinciri olmayan cihazlarda (device_ref=IMEI
        durumu) 'Müşteri Arıza Tespiti' bunun yerine buraya, IMEI'ye bağlı olarak yazılır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.batch_entries ADD COLUMN IF NOT EXISTS customer_diagnosis TEXT;"))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] batch_entries customer_diagnosis kolonu eklenemedi: {e}")
        finally:
            db.close()

    def _ensure_service_id_columns(self):
        """warehouse.batch_entries/service_records/work_orders tablolarına service_id (UUID)
        sütunu yoksa ekler. service_id, bir cihazın sisteme girdiği andan müşteriye sevkine
        (statü 128) kadar süren TEK bir servis döngüsünü temsil eden benzersiz koddur - aynı
        cihaz (IMEI/seri no) tekrar girdiğinde, önceki döngü kapanmadıysa (statü 128 değilse)
        yeni giriş engellenir (bkz. _find_active_service_for_device); kapandıysa yeni bir
        service_id üretilir. Mevcut satırlar için her satır kendi service_id'sini alır -
        geçmiş kayıtlar arasında hangi girişlerin "aynı döngü" olduğu tahmin edilmeye çalışılmaz."""
        import uuid
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("ALTER TABLE warehouse.batch_entries ADD COLUMN IF NOT EXISTS service_id UUID;"))
            db.execute(text("ALTER TABLE warehouse.service_records ADD COLUMN IF NOT EXISTS service_id UUID;"))
            db.execute(text("ALTER TABLE warehouse.work_orders ADD COLUMN IF NOT EXISTS service_id UUID;"))
            db.commit()

            missing_ids = db.execute(text(
                "SELECT id FROM warehouse.batch_entries WHERE service_id IS NULL"
            )).fetchall()
            for (entry_id,) in missing_ids:
                db.execute(text("UPDATE warehouse.batch_entries SET service_id = :sid WHERE id = :id"),
                           {"sid": str(uuid.uuid4()), "id": entry_id})
            if missing_ids:
                db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] service_id kolonları eklenemedi: {e}")
        finally:
            db.close()

    @Slot(str, str, result=str)
    def login(self, username, password):
        """React üzerinden gelen giriş isteğini işler."""
        print(f"[WebBridge] Login request received for username: {username}")
        db = SessionLocal()
        try:
            user = db.query(User).filter(User.username == username).first()
            if not user:
                return json.dumps({"success": False, "message": "Kullanıcı bulunamadı"})
            
            if not verify_password(password, user.password_hash):
                return json.dumps({"success": False, "message": "Hatalı şifre"})

            # Başarılı giriş
            user_data = {
                "id": user.id,
                "username": user.username,
                "tc_no": user.tc_no or "",
                "fullname": user.fullname or "",
                "role": user.role,
                "gorev": user.gorev or "",
                "account_enabled": user.account_enabled if user.account_enabled is not None else True,
                "team_leader": user.team_leader or "",
                "operation_manager": user.operation_manager or "",
                "administrative_manager": user.administrative_manager or ""
            }
            return json.dumps({"success": True, "user": user_data})
        except Exception as e:
            return json.dumps({"success": False, "message": f"Veritabanı hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(result=str)
    def get_users(self):
        """Tüm kullanıcıları getirir. Diğer birçok ekranda (İrsaliye, Servis Onarımları vb.)
        teknisyen/kullanıcı seçimi için de çağrıldığından, 60 saniye önbelleklenir; kullanıcı
        create/update/delete edildiğinde önbellek anında geçersiz kılınır."""
        def _compute():
            db = SessionLocal()
            try:
                users = db.query(User).all()
                users_list = []
                for u in users:
                    users_list.append({
                        "id": u.id,
                        "username": u.username,
                        "tc_no": u.tc_no or "",
                        "fullname": u.fullname or "",
                        "role": u.role,
                        "gorev": u.gorev or "",
                        "account_enabled": u.account_enabled if u.account_enabled is not None else True,
                        "team_leader": u.team_leader or "",
                        "operation_manager": u.operation_manager or "",
                        "administrative_manager": u.administrative_manager or ""
                    })
                return json.dumps({"success": True, "users": users_list})
            except Exception as e:
                return json.dumps({"success": False, "message": f"Kullanıcılar getirilemedi: {str(e)}"})
            finally:
                db.close()
        return self._cached_json("users", 60, _compute)

    @Slot(str, str, str, str, str, str, bool, str, str, str, result=str)
    def create_user(self, username, tc_no, password, role, gorev, fullname, account_enabled, team_leader, operation_manager, administrative_manager):
        """Yeni bir kullanıcı oluşturur."""
        from config.auth import get_password_hash
        db = SessionLocal()
        try:
            # Var olanı kontrol et
            if db.query(User).filter(User.username == username).first():
                return json.dumps({"success": False, "message": "Bu kullanıcı adı zaten alınmış"})
            if db.query(User).filter(User.tc_no == tc_no).first():
                return json.dumps({"success": False, "message": "Bu TC kimlik numarası zaten kullanımda"})
            
            hashed_pwd = get_password_hash(password)
            new_user = User(
                username=username,
                tc_no=tc_no,
                password_hash=hashed_pwd,
                role=role,
                gorev=gorev or None,
                fullname=fullname or None,
                account_enabled=account_enabled,
                team_leader=team_leader or None,
                operation_manager=operation_manager or None,
                administrative_manager=administrative_manager or None
            )
            db.add(new_user)
            db.commit()
            self._invalidate_cache("users")
            return json.dumps({"success": True, "message": "Kullanıcı başarıyla oluşturuldu"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Kullanıcı oluşturulamadı: {str(e)}"})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, bool, str, str, str, result=str)
    def update_user(self, user_id_str, username, tc_no, password, role, gorev, fullname, account_enabled, team_leader, operation_manager, administrative_manager):
        """Var olan bir kullanıcıyı günceller."""
        import sys
        print(f"[WebBridge] update_user called with ID: '{user_id_str}', username: '{username}', tc_no: '{tc_no}', gorev: '{gorev}', fullname: '{fullname}', account_enabled: {account_enabled}")
        sys.stdout.flush()
        from config.auth import get_password_hash
        db = SessionLocal()
        try:
            user_id = int(user_id_str)
            user = db.query(User).filter(User.id == user_id).first()
            if not user:
                print("[WebBridge] User not found.")
                sys.stdout.flush()
                return json.dumps({"success": False, "message": "Kullanıcı bulunamadı"})
            
            # Başka bir kullanıcının aynı kullanıcı adını kullanıp kullanmadığını kontrol et
            if username != user.username and db.query(User).filter(User.username == username).first():
                print("[WebBridge] Username already taken.")
                sys.stdout.flush()
                return json.dumps({"success": False, "message": "Bu kullanıcı adı zaten alınmış"})
                
            if tc_no != user.tc_no and db.query(User).filter(User.tc_no == tc_no).first():
                print("[WebBridge] TC No already taken.")
                sys.stdout.flush()
                return json.dumps({"success": False, "message": "Bu TC kimlik numarası zaten kullanımda"})
            
            user.username = username
            user.tc_no = tc_no
            user.role = role
            user.gorev = gorev or None
            user.fullname = fullname or None
            user.account_enabled = account_enabled
            user.team_leader = team_leader or None
            user.operation_manager = operation_manager or None
            user.administrative_manager = administrative_manager or None
            
            # Şifre gönderilmişse güncelle
            if password and len(password.strip()) > 0:
                print("[WebBridge] Updating password.")
                sys.stdout.flush()
                user.password_hash = get_password_hash(password)
                
            db.commit()
            self._invalidate_cache("users")
            print("[WebBridge] User updated successfully. Role is now:", user.role)
            sys.stdout.flush()
            return json.dumps({"success": True, "message": "Kullanıcı başarıyla güncellendi"})
        except Exception as e:
            print(f"[WebBridge] Update error: {str(e)}")
            sys.stdout.flush()
            db.rollback()
            return json.dumps({"success": False, "message": f"Güncelleme hatası: {str(e)}"})
        finally:
            db.close()

    # ==========================
    # PARTS (PARÇALAR) MODÜLÜ
    # ==========================

    @Slot(int, int, str, str, str, str, result=str)
    def get_parts_paginated(self, page, limit, search_term, filter_category, sort_key, sort_dir):
        from sqlalchemy import text
        from core.mapper import map_item_to_part
        db = SessionLocal()
        try:
            offset = (page - 1) * limit
            
            # WHERE clause build
            where_clauses = ["1=1"]
            params = {}
            if search_term:
                where_clauses.append("(i.code ILIKE :search OR i.short_name ILIKE :search)")
                params['search'] = f"%{search_term}%"
            if filter_category:
                where_clauses.append("i.item_category = :cat")
                params['cat'] = filter_category
                
            where_sql = " AND ".join(where_clauses)
            
            # Sorting logic
            valid_sort_keys = {
                "name": "i.short_name",
                "item_code": "i.code",
                "barcode": "i.imei",
                "item_category": "i.item_category",
                "status": "i.enabled",
                "brand": "p.brand",
                "model": "p.model"
            }
            order_by_sql = "ORDER BY i.short_name ASC"
            if sort_key in valid_sort_keys:
                direction = "DESC" if sort_dir.upper() == "DESC" else "ASC"
                order_by_sql = f"ORDER BY {valid_sort_keys[sort_key]} {direction}"

            # Total count query
            count_query = text(f"SELECT COUNT(*) FROM warehouse.item i WHERE {where_sql}")
            total_count = db.execute(count_query, params).scalar()

            # warehouse.parts, item ile ayni satirlari item_code=code eslemesiyle tasiyan
            # ayri bir tablo; brand/model buradan geliyor (product_model.brand her zaman
            # NULL oldugundan eski pm join'i hicbir zaman gercek deger uretmiyordu).
            # department icin correlated subquery kasitli olarak korunuyor: WHERE+ORDER BY+
            # LIMIT once uygulandigindan (get_part_categories'in aksine burada LIMIT var),
            # subquery sadece o sayfadaki (ör. 100) satir icin calisiyor - bunu JOIN+GROUP BY'a
            # cevirmek denendi ama GROUP BY, LIMIT'ten once TUM eslenen item tablosunun
            # gruplanmasini zorunlu kildigindan (317ms -> 2400ms+) performansi ciddi
            # kotulestirdi, bu yuzden orijinal haline geri donuldu.
            query = text(f"""
                SELECT
                    i.id, i.code, i.short_name, i.color, i.item_type, i.item_category, i.enabled,
                    p.brand, p.model,
                    (SELECT string_agg(icm.mission, ', ') FROM warehouse.item_category_mission icm WHERE icm.item_category = i.item_category OR icm.item_category = i.code) AS department
                FROM warehouse.item i
                LEFT JOIN warehouse.parts p ON p.item_code = i.code
                WHERE {where_sql}
                {order_by_sql}
                LIMIT :limit OFFSET :offset
            """)
            params['limit'] = limit
            params['offset'] = offset
            
            rows = db.execute(query, params).mappings().all()
            
            parts_list = [map_item_to_part(row) for row in rows]
            
            return json.dumps({
                "success": True, 
                "parts": parts_list,
                "total_count": total_count,
                "current_page": page,
                "total_pages": (total_count + limit - 1) // limit if limit > 0 else 1
            })
        except Exception as e:
            print(f"[WebBridge] get_parts_paginated error: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()
            
    @Slot(result=str)
    def get_parts(self):
        # Backward compatibility or fallback
        return self.get_parts_paginated(1, 100, "", "", "", "")

    @Slot(result=str)
    def get_item_boms(self):
        """Tüm ItemBOM (Recipe) kayıtlarını, parent ve child parça bilgileriyle birlikte
        getirir. item_bom küçük bir tablo olduğu için (get_parts/get_products/get_stock'un
        aksine) dosya cache'i kullanılmaz, doğrudan sorgulanır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            bom_result = db.execute(text("""
                SELECT b.id, b.parent_item_id, b.child_item_id, b.quantity,
                       p_parent.name AS parent_name, p_parent.id AS parent_part_id,
                       p_child.name AS child_name, p_child.id AS child_part_id
                FROM warehouse.item_bom b
                LEFT JOIN warehouse.parts p_parent ON p_parent.item_code = b.parent_item_id
                LEFT JOIN warehouse.parts p_child ON p_child.item_code = b.child_item_id
                ORDER BY b.parent_item_id, b.child_item_id;
            """)).mappings().all()

            bom_map = {}
            for row in bom_result:
                parent_code = row["parent_item_id"]
                if parent_code not in bom_map:
                    bom_map[parent_code] = {
                        "parent_item_id": parent_code,
                        "parent_part_id": str(row["parent_part_id"]) if row["parent_part_id"] else "",
                        "parent_name": row["parent_name"] or parent_code,
                        "materials": []
                    }
                bom_map[parent_code]["materials"].append({
                    "child_item_id": row["child_item_id"],
                    "child_part_id": str(row["child_part_id"]) if row["child_part_id"] else "",
                    "child_name": row["child_name"] or row["child_item_id"],
                    "quantity": int(row["quantity"])
                })

            # Reçetesi (BOM) olmayan ama 'Mamül' veya 'Yarı Mamül' olan parçaları da listeye ekle
            products = db.execute(text("""
                SELECT id, item_code, name 
                FROM warehouse.parts 
                WHERE part_type ILIKE '%Mamül%' OR part_type ILIKE '%Mamul%'
                   OR part_category ILIKE '%Mamül%' OR part_category ILIKE '%Mamul%'
                   OR item_category ILIKE '%Mamül%' OR item_category ILIKE '%Mamul%'
            """)).mappings().all()

            for p in products:
                parent_code = p["item_code"]
                if parent_code and parent_code not in bom_map:
                    bom_map[parent_code] = {
                        "parent_item_id": parent_code,
                        "parent_part_id": str(p["id"]),
                        "parent_name": p["name"] or parent_code,
                        "materials": []
                    }

            return json.dumps({"success": True, "item_boms": list(bom_map.values())}, ensure_ascii=False)
        except Exception as e:
            print(f"[WebBridge] get_item_boms hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, result=str)
    def get_product_boms(self, page="1", page_size="50", search_term="", model_filter="", status_filter=""):
        """Sayfalanmis (LIMIT/OFFSET) ve filtrelemeli sekilde doner."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            page = max(1, int(page or 1))
            page_size = min(1000, max(1, int(page_size or 50)))
            offset = (page - 1) * page_size

            where_clauses = []
            params = {"limit": page_size, "offset": offset}

            if search_term and str(search_term).strip():
                where_clauses.append("(b.parent_product_code ILIKE :search OR b.child_item_code ILIKE :search OR i.short_name ILIKE :search)")
                params["search"] = f"%{str(search_term).strip()}%"

            if model_filter and str(model_filter).strip():
                where_clauses.append("b.parent_product_code = :model_filter")
                params["model_filter"] = str(model_filter).strip()

            if status_filter and str(status_filter).strip() and str(status_filter).strip().lower() != "tümü":
                if status_filter.lower() == "aktif":
                    where_clauses.append("b.enabled = true")
                else:
                    where_clauses.append("b.enabled = false")

            where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

            count_sql = f"""
                SELECT COUNT(*)
                FROM warehouse.product_bom_node b
                LEFT JOIN warehouse.item i ON i.code = b.child_item_code
                {where_sql};
            """
            total = db.execute(text(count_sql), params).scalar()

            data_sql = f"""
                SELECT b.id, b.parent_product_code, b.child_item_code, b.quantity, b.enabled,
                       i.short_name AS child_name, i.id AS child_part_id
                FROM warehouse.product_bom_node b
                LEFT JOIN warehouse.item i ON i.code = b.child_item_code
                {where_sql}
                ORDER BY b.parent_product_code, b.child_item_code
                LIMIT :limit OFFSET :offset
            """
            rows = db.execute(text(data_sql), params).fetchall()

            boms = []
            for row in rows:
                boms.append({
                    "id": str(row.id),
                    "product_model": row.parent_product_code,
                    "child_item_code": row.child_item_code,
                    "child_name": row.child_name,
                    "quantity": row.quantity,
                    "status": "Aktif" if row.enabled else "Pasif",
                    "created_at": "",
                    "updated_at": "",
                    "child_part_id": str(row.child_part_id) if row.child_part_id else None
                })

            return json.dumps({
                "success": True,
                "boms": boms,
                "total": total,
                "page": page,
                "page_size": page_size
            }, ensure_ascii=False)
        except Exception as e:
            print(f"[WebBridge] get_product_boms hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def create_product_bom(self, product_model, child_item_code, quantity):
        """warehouse.product_bom_node'a yeni bir reçete satırı ekler - Demontaj ekranındaki
        'Parça Seçiniz' (bkz. get_parts_for_device) bu tabloyu okur, bu yüzden burada
        eklenen her satır o ekranda da görünür."""
        from models.product_bom_node import ProductBomNode
        db = SessionLocal()
        try:
            new_bom = ProductBomNode(
                parent_product_code=product_model,
                child_item_code=child_item_code,
                quantity=int(quantity) if quantity else 1
            )
            db.add(new_bom)
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def bulk_import_product_bom(self, rows_json):
        """Toplu (Excel) reçete (BOM) satırı içe aktarma. create_product_bom'un tek satırlık
        eşdeğeriyle aynı kuralları (yalnızca zorunlu alan kontrolü; parent/child eşleşmesi
        veya mükerrer kontrolü tek satırlık akışta da yoktu) kullanır, ama N ayrı çağrı/commit
        yerine tüm satırları tek transaction'da yazar. Herhangi bir satırda zorunlu alan
        eksikse hiçbir satır kaydedilmez."""
        from models.product_bom_node import ProductBomNode
        db = SessionLocal()
        try:
            try:
                rows = json.loads(rows_json or "[]")
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Geçersiz dosya verisi.", "errors": []})

            if not rows:
                return json.dumps({"success": False, "message": "Dosyada içe aktarılacak satır bulunamadı.", "errors": []})

            errors = []
            valid_rows = []

            for idx, row in enumerate(rows):
                row_num = idx + 2
                row = row or {}

                def get_val(key):
                    v = row.get(key)
                    return str(v).strip() if v is not None else ""

                product_model = get_val("product_model")
                child_item_code = get_val("child_item_code")
                quantity_raw = get_val("quantity")

                if not product_model:
                    errors.append({"row": row_num, "field": "Cihaz Modeli", "message": "Cihaz Modeli (product_model) boş olamaz."})
                if not child_item_code:
                    errors.append({"row": row_num, "field": "Alt Parça Kodu", "message": "Alt Parça Kodu (child_item_code) boş olamaz."})

                quantity = 1
                if quantity_raw:
                    try:
                        quantity = int(float(quantity_raw))
                    except ValueError:
                        errors.append({"row": row_num, "field": "Miktar", "message": f"\"{quantity_raw}\" geçerli bir sayı değil."})

                valid_rows.append({
                    "parent_product_code": product_model,
                    "child_item_code": child_item_code,
                    "quantity": quantity,
                })

            if errors:
                return json.dumps({"success": False, "message": f"{len(errors)} hata bulundu, hiçbir satır içe aktarılmadı.", "errors": errors})

            for r in valid_rows:
                db.add(ProductBomNode(
                    parent_product_code=r["parent_product_code"],
                    child_item_code=r["child_item_code"],
                    quantity=r["quantity"],
                ))

            db.commit()
            return json.dumps({"success": True, "message": f"{len(valid_rows)} reçete satırı başarıyla içe aktarıldı.", "imported": len(valid_rows)})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"İçe aktarma hatası: {str(e)}", "errors": []})
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def update_product_bom(self, bom_id, product_model, child_item_code, quantity):
        from models.product_bom_node import ProductBomNode
        db = SessionLocal()
        try:
            bom = db.query(ProductBomNode).filter(ProductBomNode.id == bom_id).first()
            if not bom:
                return json.dumps({"success": False, "message": "BOM bulunamadı."})

            bom.parent_product_code = product_model
            bom.child_item_code = child_item_code
            bom.quantity = int(quantity) if quantity else 1
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_product_bom(self, bom_id):
        from models.product_bom_node import ProductBomNode
        db = SessionLocal()
        try:
            bom = db.query(ProductBomNode).filter(ProductBomNode.id == bom_id).first()
            if bom:
                db.delete(bom)
                db.commit()
                return json.dumps({"success": True})
            return json.dumps({"success": False, "message": "BOM bulunamadı"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def toggle_product_bom_status(self, bom_id):
        from models.product_bom_node import ProductBomNode
        db = SessionLocal()
        try:
            bom = db.query(ProductBomNode).filter(ProductBomNode.id == bom_id).first()
            if not bom:
                return json.dumps({"success": False, "message": "BOM kaydı bulunamadı."})

            bom.enabled = not bom.enabled
            db.commit()
            new_status = "Aktif" if bom.enabled else "Pasif"
            return json.dumps({"success": True, "message": f"Durum '{new_status}' olarak güncellendi."})
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] toggle_product_bom_status hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_item_model(self, item_code):
        """Parça Kodu girildiğinde Model alanını otomatik doldurmak için warehouse.item_models
        (ProductBom/ProductFamily'den türetilmiş) ve mevcut parts kayıtlarını sorgular."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            code = (item_code or "").strip()
            if not code:
                return json.dumps({"success": False, "model": "", "brand": ""})

            row = db.execute(
                text("SELECT model, brand FROM warehouse.item_models WHERE item_code = :code"),
                {"code": code}
            ).first()
            if row and (row[0] or row[1]):
                return json.dumps({"success": True, "model": row[0] or "", "brand": row[1] or ""})

            row2 = db.execute(
                text("SELECT model, brand FROM warehouse.parts WHERE item_code = :code AND ((model IS NOT NULL AND model <> '') OR (brand IS NOT NULL AND brand <> '')) LIMIT 1"),
                {"code": code}
            ).first()
            if row2 and (row2[0] or row2[1]):
                return json.dumps({"success": True, "model": row2[0] or "", "brand": row2[1] or ""})

            return json.dumps({"success": False, "model": "", "brand": ""})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_item_codes_by_model(self, model_name):
        """Ürün Ağacı (BOM) ekranında Model seçilince, Ana Parça/Alt Parça alanlarını
        sadece o modele ait parça kodlarıyla sınırlamak için warehouse.item_models'i
        (virgülle ayrılmış model listesi) sorgular."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            name = (model_name or "").strip()
            if not name:
                return json.dumps({"success": False, "item_codes": []})

            rows = db.execute(text("""
                SELECT item_code FROM warehouse.item_models
                WHERE :name = ANY(string_to_array(model, ', '))
                ORDER BY item_code
            """), {"name": name}).all()
            codes = [r[0] for r in rows]
            return json.dumps({"success": True, "item_codes": codes})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_item_codes(self):
        """MioCreate.xlsx ProductBom sayfasından (warehouse.item_models üzerinden) bilinen
        tüm parça kodlarını döner. 'Yeni Stok Kartı Ekle' formundaki Parça Kodu alanı
        bu listeyi otomatik tamamlama (datalist) için kullanır."""
        filename = "item_codes.json"
        path = os.path.join(get_cache_dirs()[0], filename)
        fetch_url = f"http://localhost:5173/api_cache/{filename}" if os.getenv("DEV_MODE", "1") == "1" else f"/api_cache/{filename}"
        if os.path.exists(path):
            return json.dumps({"success": True, "fetch_url": fetch_url})

        from sqlalchemy import text
        db = SessionLocal()
        try:
            result = db.execute(text(
                "SELECT item_code FROM warehouse.item_models ORDER BY item_code"
            )).all()
            codes = [row[0] for row in result if row[0]]
            json_data = json.dumps({"success": True, "item_codes": codes})
            write_to_cache(filename, json_data)
            return json.dumps({"success": True, "fetch_url": fetch_url})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, str, str, str, str, str, result=str)
    def create_part(self, name, item_code, barcode, brand, model, item_category, part_category, part_category_id, stock_tracking_type, department, status, critical_limit, memory, part_type):
        """Yeni parça ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            code = item_code.strip()
            if not code:
                return json.dumps({"success": False, "message": "Parça Kodu zorunludur"})

            part_name = name.strip()
            if not part_name:
                part_name = f"{brand.strip()} {model.strip()}".strip() or code

            # Check if item_code already exists
            existing = db.execute(text("SELECT id FROM warehouse.parts WHERE item_code = :code"), {"code": code}).scalar()
            if existing:
                return json.dumps({
                    "success": False, 
                    "message": f"'{code}' kodlu parça zaten sistemde kayıtlı. Eklemek yerine arama çubuğundan bulup düzenleyebilirsiniz."
                })

            sql = """
                INSERT INTO warehouse.parts (name, item_code, barcode, brand, model, item_category, part_category, part_category_id, stock_tracking_type, department, status, critical_limit, memory, part_type)
                VALUES (:name, :code, :barcode, :brand, :model, :icat, :pcat, :pcat_id, :stt, :dept, :status, :critical_limit, :memory, :part_type)
            """
            if part_type in ["Labour", "Service", "Cost", "SparePartLabour", "Labour (İşçilik)", "Stoksuz Parça / Hizmet"]:
                stock_tracking_type = "Stok Takipsiz"
            
            db.execute(text(sql), {
                "name": part_name, "code": code, "barcode": barcode or None,
                "brand": brand or None, "model": model or None,
                "icat": item_category or None, "pcat": part_category or None,
                "pcat_id": int(part_category_id) if part_category_id.strip() else None,
                "stt": stock_tracking_type or "Stok Takipli",
                "dept": department or None,
                "status": status or "Aktif",
                "critical_limit": int(critical_limit) if critical_limit.strip() else 50,
                "memory": memory or None,
                "part_type": part_type or None
            })
            db.commit()
            return json.dumps({"success": True, "message": "Parça eklendi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Kayıt hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, result=str)
    def bulk_import_parts(self, rows_json):
        """Toplu (Excel) parça içe aktarma. Tüm satırları önce doğrular; herhangi bir
        satırda zorunlu alan eksikse veya item_code zaten kayıtlıysa/dosyada tekrarlıysa
        HİÇBİR satır kaydedilmez, tüm hatalar satır numarasıyla birlikte tek seferde döner.
        create_part'ın tek satırlık eşdeğeriyle aynı doğrulama/varsayılan kurallarını kullanır,
        ancak N ayrı QWebChannel çağrısı + N ayrı DB commit yerine tek transaction'da çalışır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            try:
                rows = json.loads(rows_json or "[]")
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Geçersiz dosya verisi.", "errors": []})

            if not rows:
                return json.dumps({"success": False, "message": "Dosyada içe aktarılacak satır bulunamadı.", "errors": []})

            existing_codes = {r[0] for r in db.execute(text("SELECT item_code FROM warehouse.parts")).all()}

            errors = []
            seen_codes_in_file = {}
            valid_rows = []

            for idx, row in enumerate(rows):
                row_num = idx + 2  # 1. satır başlık; ilk veri satırı Excel'de 2. satır
                row = row or {}

                def get_val(key):
                    v = row.get(key)
                    return str(v).strip() if v is not None else ""

                code = get_val("item_code")
                name = get_val("name")
                barcode = get_val("barcode")
                item_category = get_val("item_category")
                part_category = get_val("part_category")
                status = get_val("status")
                part_type = get_val("part_type")

                if not code:
                    errors.append({"row": row_num, "field": "Parça Kodu", "message": "Parça Kodu (item_code) boş olamaz."})
                elif code in existing_codes:
                    errors.append({"row": row_num, "field": "Parça Kodu", "message": f"\"{code}\" kodlu parça zaten sistemde kayıtlı."})
                elif code in seen_codes_in_file:
                    errors.append({"row": row_num, "field": "Parça Kodu", "message": f"\"{code}\" dosyada birden fazla satırda tekrarlanıyor (satır {seen_codes_in_file[code]})."})
                else:
                    seen_codes_in_file[code] = row_num

                stock_tracking_type = "Stok Takipsiz" if part_type in [
                    "Labour", "Service", "Cost", "SparePartLabour", "Labour (İşçilik)", "Stoksuz Parça / Hizmet"
                ] else "Stok Takipli"

                valid_rows.append({
                    "name": name or (code or None),
                    "item_code": code or None,
                    "barcode": barcode or None,
                    "item_category": item_category or None,
                    "part_category": part_category or None,
                    "stock_tracking_type": stock_tracking_type,
                    "status": status or "Aktif",
                    "part_type": part_type or None,
                })

            if errors:
                return json.dumps({"success": False, "message": f"{len(errors)} hata bulundu, hiçbir satır içe aktarılmadı.", "errors": errors})

            for r in valid_rows:
                db.execute(text("""
                    INSERT INTO warehouse.parts (
                        name, item_code, barcode, item_category, part_category,
                        stock_tracking_type, status, critical_limit, part_type
                    ) VALUES (
                        :name, :item_code, :barcode, :item_category, :part_category,
                        :stock_tracking_type, :status, 50, :part_type
                    )
                """), r)

            db.commit()
            return json.dumps({"success": True, "message": f"{len(valid_rows)} parça başarıyla içe aktarıldı.", "imported": len(valid_rows)})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"İçe aktarma hatası: {str(e)}", "errors": []})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, result=str)
    def update_part(self, part_id_str, name, item_code, barcode, brand, model, item_category, part_category, part_category_id, stock_tracking_type, department, status, critical_limit, memory, part_type):
        """Var olan bir parçayı günceller."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            part_id = int(part_id_str)
            code = item_code.strip()
            if not code:
                return json.dumps({"success": False, "message": "Parça Kodu zorunludur"})

            part_name = name.strip()
            if not part_name:
                part_name = f"{brand.strip()} {model.strip()}".strip() or code

            sql = """
                UPDATE warehouse.parts
                SET name = :name, item_code = :code, barcode = :barcode, brand = :brand,
                    model = :model, item_category = :icat, part_category = :pcat,
                    part_category_id = :pcat_id, stock_tracking_type = :stt,
                    department = :dept, status = :status, critical_limit = :critical_limit,
                    memory = :memory, part_type = :part_type
                WHERE id = :id
            """
            if part_type in ["Labour", "Service", "Cost", "SparePartLabour", "Labour (İşçilik)", "Stoksuz Parça / Hizmet"]:
                stock_tracking_type = "Stok Takipsiz"

            db.execute(text(sql), {
                "name": part_name, "code": code, "barcode": barcode or None,
                "brand": brand or None, "model": model or None,
                "icat": item_category or None, "pcat": part_category or None,
                "pcat_id": int(part_category_id) if part_category_id.strip() else None,
                "stt": stock_tracking_type or "Stok Takipli",
                "dept": department or None,
                "status": status or "Aktif",
                "critical_limit": int(critical_limit) if critical_limit.strip() else 50,
                "memory": memory or None,
                "part_type": part_type or None,
                "id": part_id
            })
            db.commit()
            return json.dumps({"success": True, "message": "Parça güncellendi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Güncelleme hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_part(self, part_id_str):
        """Belirtilen id'ye sahip parçayı siler (Stok miktarı 0 ise parçayı ve ilişkili tüm kayıtlarını temizler)."""
        from sqlalchemy import text, func
        from models.stock import Stock
        from models.location import Location
        db = SessionLocal()
        try:
            part_id = int(part_id_str)

            # Stok Miktarı Kontrolü — sadece fiziksel/depoda bulunan lokasyonlar sayılır
            # (Out Stock / Scrap Stock lokasyonlarındaki miktar, ürünün depodan çıktığının
            # kaydıdır; parçanın silinmesini engellememeli).
            total_stock_qty = db.query(func.coalesce(func.sum(Stock.quantity), 0)) \
                .join(Location, Stock.location_id == Location.id) \
                .filter(Stock.part_id == part_id, Location.kind.in_(("good_stock", "doa_stock", "repair_stock"))) \
                .scalar() or 0
            if total_stock_qty > 0:
                return json.dumps({"success": False, "message": f"Bu parçanın stokta {total_stock_qty} adet ürünü var. Silmeden önce stok miktarını sıfırlayınız."})

            # İrsaliye geçmişinde gösterilmeye devam etsin diye, parça silinmeden önce
            # görünen adının anlık görüntüsünü alıyoruz.
            part_row = db.execute(text("""
                SELECT item_code, brand, model, color, part_category, name
                FROM warehouse.parts WHERE id = :id
            """), {"id": part_id}).mappings().first()
            snapshot_name = _build_part_display_name(
                part_row.get("brand") if part_row else None,
                part_row.get("model") if part_row else None,
                part_row.get("color") if part_row else None,
                part_row.get("part_category") if part_row else None,
                part_row.get("name") if part_row else None,
                part_row.get("item_code") if part_row else None,
            )

            queries = [
                "DELETE FROM warehouse.stock WHERE part_id = :id",
                # İrsaliye geçmişi korunsun diye hareket kayıtları silinmiyor, sadece
                # silinen parçaya olan referans temizleniyor; ekranda isim anlık
                # görüntüsü + "(silindi)" ibaresiyle gösterilir.
                "UPDATE warehouse.stock_movements SET part_id = NULL, part_name_snapshot = :snapshot_name WHERE part_id = :id",
                "DELETE FROM warehouse.inbound_entries WHERE part_id = :id",
                "DELETE FROM warehouse.outbound_entries WHERE part_id = :id",
                "DELETE FROM warehouse.work_order_parts WHERE part_id = :id",
                "DELETE FROM warehouse.production_materials WHERE part_id = :id",
                "DELETE FROM warehouse.bom_items WHERE part_id = :id OR parent_item_id = :id",
                "DELETE FROM warehouse.item_bom WHERE part_id = :id OR parent_item_id = :id",
                "DELETE FROM warehouse.part_supplier_prices WHERE part_id = :id",
                "DELETE FROM warehouse.part_suppliers WHERE part_id = :id",
                "UPDATE warehouse.production_runs SET target_part_id = NULL WHERE target_part_id = :id",
                "UPDATE warehouse.work_orders SET target_part_id = NULL WHERE target_part_id = :id",
                "DELETE FROM warehouse.parts WHERE id = :id"
            ]

            for q in queries:
                try:
                    with db.begin_nested():
                        db.execute(text(q), {"id": part_id, "snapshot_name": snapshot_name})
                except Exception as ex:
                    logging.warning(f"delete_part subquery bypass: {ex}")

            db.commit()
            clear_api_cache()
            return json.dumps({"success": True, "message": "Parça başarıyla silindi."})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Silme hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_parts_bulk(self, part_ids_csv):
        """Birden fazla parçayı toplu olarak siler."""
        from sqlalchemy import text, func
        from models.stock import Stock
        from models.location import Location
        db = SessionLocal()
        try:
            ids = [int(x.strip()) for x in part_ids_csv.split(",") if x.strip()]
            if not ids:
                return json.dumps({"success": False, "message": "Silinecek parça seçilmedi."})

            safe_ids = []
            skipped_count = 0
            for pid in ids:
                # Sadece fiziksel/depoda bulunan lokasyonlar sayılır (bkz. delete_part)
                total_stock_qty = db.query(func.coalesce(func.sum(Stock.quantity), 0)) \
                    .join(Location, Stock.location_id == Location.id) \
                    .filter(Stock.part_id == pid, Location.kind.in_(("good_stock", "doa_stock", "repair_stock"))) \
                    .scalar() or 0
                if total_stock_qty == 0:
                    safe_ids.append(pid)
                else:
                    skipped_count += 1
                    
            if not safe_ids:
                return json.dumps({"success": False, "message": "Seçilen parçaların tamamının stokta ürünü var. Önce stok miktarlarını sıfırlayınız."})
                
            queries = [
                "DELETE FROM warehouse.stock WHERE part_id = :id",
                # İrsaliye geçmişi korunsun diye hareket kayıtları silinmiyor, sadece
                # silinen parçaya olan referans temizleniyor; ekranda isim anlık
                # görüntüsü + "(silindi)" ibaresiyle gösterilir.
                "UPDATE warehouse.stock_movements SET part_id = NULL, part_name_snapshot = :snapshot_name WHERE part_id = :id",
                "DELETE FROM warehouse.inbound_entries WHERE part_id = :id",
                "DELETE FROM warehouse.outbound_entries WHERE part_id = :id",
                "DELETE FROM warehouse.work_order_parts WHERE part_id = :id",
                "DELETE FROM warehouse.production_materials WHERE part_id = :id",
                "DELETE FROM warehouse.bom_items WHERE part_id = :id OR parent_item_id = :id",
                "DELETE FROM warehouse.item_bom WHERE part_id = :id OR parent_item_id = :id",
                "DELETE FROM warehouse.part_supplier_prices WHERE part_id = :id",
                "DELETE FROM warehouse.part_suppliers WHERE part_id = :id",
                "UPDATE warehouse.production_runs SET target_part_id = NULL WHERE target_part_id = :id",
                "UPDATE warehouse.work_orders SET target_part_id = NULL WHERE target_part_id = :id",
                "DELETE FROM warehouse.parts WHERE id = :id"
            ]

            for pid in safe_ids:
                part_row = db.execute(text("""
                    SELECT item_code, brand, model, color, part_category, name
                    FROM warehouse.parts WHERE id = :id
                """), {"id": pid}).mappings().first()
                snapshot_name = _build_part_display_name(
                    part_row.get("brand") if part_row else None,
                    part_row.get("model") if part_row else None,
                    part_row.get("color") if part_row else None,
                    part_row.get("part_category") if part_row else None,
                    part_row.get("name") if part_row else None,
                    part_row.get("item_code") if part_row else None,
                )
                for q in queries:
                    try:
                        with db.begin_nested():
                            db.execute(text(q), {"id": pid, "snapshot_name": snapshot_name})
                    except Exception as ex:
                        logging.warning(f"delete_parts_bulk subquery bypass: {ex}")

            db.commit()
            clear_api_cache()
            
            msg = f"{len(safe_ids)} parça başarıyla silindi."
            if skipped_count > 0:
                msg += f" {skipped_count} adet parça stokta ürünü olduğu için silinemedi."
            return json.dumps({"success": True, "message": msg})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Silme hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_user(self, user_id_str):
        """Belirtilen id'ye sahip kullanıcıyı siler."""
        import sys
        print(f"[WebBridge] delete_user called with ID: {user_id_str}")
        sys.stdout.flush()
        from models.user import User
        db = SessionLocal()
        try:
            user_id = int(user_id_str)
            user = db.query(User).filter(User.id == user_id).first()
            if not user:
                return json.dumps({"success": False, "message": "Kullanıcı bulunamadı"})
            
            if user.role == "Admin":
                admin_count = db.query(User).filter(User.role == "Admin").count()
                if admin_count <= 1:
                    return json.dumps({"success": False, "message": "Sistemdeki son Admin kullanıcısı silinemez!"})
            
            db.delete(user)
            db.commit()
            self._invalidate_cache("users")
            return json.dumps({"success": True, "message": "Kullanıcı silindi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Silme hatası: {str(e)}"})
        finally:
            db.close()

    # --- YENİ EKLENEN LOKASYON FONKSİYONLARI ---
    @Slot(result=str)
    def get_locations(self):
        """Çalışma zamanında nadiren değiştiğinden 5 dakika önbelleklenir (bkz. _cached_json)."""
        def _compute():
            from models.location import Location
            db = SessionLocal()
            try:
                locs = db.query(Location).all()
                return json.dumps({"success": True, "locations": [{"id": l.id, "name": l.name, "kind": l.kind, "description": l.description} for l in locs]})
            except Exception as e:
                return json.dumps({"success": False, "message": str(e)})
            finally:
                db.close()
        return self._cached_json("locations", 300, _compute)

    @Slot(result=str)
    def get_system_locations(self):
        """Good/DOA/Repair/Scrap/Out Stock sistem depolarını döner. Çalışma zamanında
        nadiren değiştiğinden 5 dakika önbelleklenir."""
        def _compute():
            from models.location import Location
            db = SessionLocal()
            try:
                locs = db.query(Location).filter(Location.kind.isnot(None)).all()
                return json.dumps({"success": True, "locations": [{"id": l.id, "name": l.name, "kind": l.kind} for l in locs]})
            except Exception as e:
                return json.dumps({"success": False, "message": str(e)})
            finally:
                db.close()
        return self._cached_json("system_locations", 300, _compute)

    @Slot(str, str, result=str)
    def create_location(self, name, description):
        from models.location import Location
        db = SessionLocal()
        try:
            if db.query(Location).filter(Location.name == name).first():
                return json.dumps({"success": False, "message": "Bu lokasyon zaten var"})
            loc = Location(name=name, description=description)
            db.add(loc)
            db.commit()
            self._invalidate_cache("locations", "system_locations")
            return json.dumps({"success": True, "message": "Lokasyon eklendi", "id": loc.id})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_location(self, id_str):
        from models.location import Location
        from models.stock import Stock
        from models.stock_movement import StockMovement
        from sqlalchemy import or_
        db = SessionLocal()
        try:
            loc_id = int(id_str)
            loc = db.query(Location).filter(Location.id == loc_id).first()
            if loc:
                if loc.kind:
                    return json.dumps({"success": False, "message": "Bu depo sistem tarafından otomatik yönetiliyor, silinemez."})
                
                stock_count = db.query(Stock).filter(Stock.location_id == loc_id, Stock.quantity > 0).count()
                if stock_count > 0:
                    return json.dumps({"success": False, "message": "Bu depoda stoklu ürünler var, silinemez."})
                
                movement_count = db.query(StockMovement).filter(or_(StockMovement.source_location_id == loc_id, StockMovement.target_location_id == loc_id)).count()
                if movement_count > 0:
                    return json.dumps({"success": False, "message": "Bu deponun geçmiş stok hareketi var, silinemez."})

                db.delete(loc)
                db.commit()
                self._invalidate_cache("locations", "system_locations")
                return json.dumps({"success": True})
            return json.dumps({"success": False, "message": "Bulunamadı"})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()


    @Slot(result=str)
    def get_product_families(self):
        """Aktif ürün ailesi (cihaz modeli) adlarını getirir. MioCreate.xlsx -> ProductFamily'den seed edilmiştir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT id, name
                FROM warehouse.product_families
                WHERE is_active IS TRUE
                ORDER BY name ASC
            """)).mappings().all()
            families = [{"id": r["id"], "name": r["name"]} for r in rows]
            return json.dumps({"success": True, "product_families": families})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_mission_groups(self):
        """Görev gruplarını getirir. MioCreate.xlsx -> MissionGroup'tan seed edilmiştir (organization.mission_groups).
        Sadece üretim/onarım ile ilgili gruplar döner (department='Üretim'). Çalışma
        zamanında pratikte hiç değişmediğinden 5 dakika önbelleklenir."""
        def _compute():
            from sqlalchemy import text
            db = SessionLocal()
            try:
                rows = db.execute(text("""
                    SELECT code, short_name, order_number
                    FROM organization.mission_groups
                    WHERE department = 'Üretim'
                    ORDER BY order_number NULLS LAST, short_name ASC
                """)).mappings().all()
                groups = [{"code": r["code"], "short_name": r["short_name"], "order_number": r["order_number"]} for r in rows]
                return json.dumps({"success": True, "mission_groups": groups}, ensure_ascii=False)
            except Exception as e:
                return json.dumps({"success": False, "message": str(e)})
            finally:
                db.close()
        return self._cached_json("mission_groups", 300, _compute)

    @Slot(str, result=str)
    def get_mission_for_item_category(self, item_category):
        """warehouse.item_category_mission'dan, verilen parça kategorisi için önerilen
        Onarım Takımı'nı (organization.mission_groups.code) döner. item_category_mission.mission
        'TEC_' önekli (TEC_CASE, TEC_BATTERY...) iken mission_groups.code önek taşımaz (CASE,
        BATTERY...) - TEC_ öneki kırpılarak eşleştirilir. Birden fazla etkin (enabled) aday
        varsa, TEC_L1REPAIR/L2REPAIR/L3REPAIR gibi genel montaj kodları yerine o kategoriye
        özel uzman ekip (ör. TEC_CAMERA, TEC_BATTERY) tercih edilir. Demontaj ekranında parça
        seçilince 'Onarım Takımı' alanının otomatik önerilmesi için kullanılır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            category = (item_category or "").strip()
            if not category:
                return json.dumps({"success": True, "mission_code": None})

            rows = db.execute(text("""
                SELECT mission FROM warehouse.item_category_mission
                WHERE LOWER(TRIM(item_category)) = LOWER(:cat) AND enabled = TRUE
            """), {"cat": category}).fetchall()

            candidates = [r[0].strip() for r in rows if r[0]]
            if not candidates:
                return json.dumps({"success": True, "mission_code": None})

            generic = {"TEC_L1REPAIR", "TEC_L2REPAIR", "TEC_L3REPAIR"}
            specialist = [c for c in candidates if c.upper() not in generic]
            chosen = specialist[0] if specialist else candidates[0]

            bare_code = chosen[4:] if chosen.upper().startswith("TEC_") else chosen

            exists = db.execute(text(
                "SELECT code FROM organization.mission_groups WHERE code = :c"
            ), {"c": bare_code}).first()

            return json.dumps({"success": True, "mission_code": bare_code if exists else None})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_missions_for_item_category(self, item_category):
        """warehouse.item_category_mission'da verilen parça kategorisi için tanımlı TÜM etkin
        (enabled) departmanları (organization.mission_groups.code) döner - TEC_ öneki kırpılarak
        eşleştirilir (bkz. get_mission_for_item_category). Demontaj ekranındaki 'Onarım Takımı'
        dropdown'unu, sadece o kategori için gerçekten tanımlı departmanlarla sınırlamak için
        kullanılır. Eşleşme yoksa boş liste döner (çağıran taraf tüm departmanlara geri düşebilir)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            category = (item_category or "").strip()
            if not category:
                return json.dumps({"success": True, "mission_codes": []})

            rows = db.execute(text("""
                SELECT DISTINCT mission FROM warehouse.item_category_mission
                WHERE LOWER(TRIM(item_category)) = LOWER(:cat) AND enabled = TRUE
            """), {"cat": category}).fetchall()

            bare_codes = set()
            for (mission,) in rows:
                if not mission:
                    continue
                m = mission.strip()
                bare_codes.add(m[4:] if m.upper().startswith("TEC_") else m)

            if not bare_codes:
                return json.dumps({"success": True, "mission_codes": []})

            existing = db.execute(text(
                "SELECT code FROM organization.mission_groups WHERE code = ANY(:codes)"
            ), {"codes": list(bare_codes)}).fetchall()
            valid_codes = [r[0] for r in existing]

            return json.dumps({"success": True, "mission_codes": valid_codes}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_missions(self, department_filter=""):
        """Görevleri/rolleri getirir. MioCreate.xlsx -> Mission'dan seed edilmiştir (organization.missions).
        department_filter doluysa (ör. 'Üretim') sadece o departmandaki görevler döner, boşsa hepsi döner."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            sql = """
                SELECT m.id, m.code, m.short_name, m.full_name, m.description, m.cost_center,
                       m.department, m.order_number,
                       mg.code AS mission_group_code, mg.short_name AS mission_group_name,
                       mw.code AS mission_workgroup_code, mw.short_name AS mission_workgroup_name,
                       m.team_leader_mission_code, tl.short_name AS team_leader_name,
                       m.operation_manager_mission_code, om.short_name AS operation_manager_name,
                       m.administrative_manager_mission_code, am.short_name AS administrative_manager_name
                FROM organization.missions m
                LEFT JOIN organization.mission_groups mg ON mg.id = m.mission_group_id
                LEFT JOIN organization.mission_workgroups mw ON mw.id = m.mission_workgroup_id
                LEFT JOIN organization.missions tl ON tl.code = m.team_leader_mission_code
                LEFT JOIN organization.missions om ON om.code = m.operation_manager_mission_code
                LEFT JOIN organization.missions am ON am.code = m.administrative_manager_mission_code
            """
            params = {}
            if department_filter and department_filter.strip():
                sql += " WHERE m.department = :dept"
                params["dept"] = department_filter.strip()
            sql += " ORDER BY m.order_number NULLS LAST, m.short_name ASC"

            rows = db.execute(text(sql), params).mappings().all()
            missions = [{
                "id": str(r["id"]),
                "code": r["code"],
                "short_name": r["short_name"],
                "full_name": r["full_name"] or "",
                "description": r["description"] or "",
                "cost_center": r["cost_center"] or "",
                "department": r["department"] or "",
                "order_number": r["order_number"],
                "mission_group_code": r["mission_group_code"] or "",
                "mission_group_name": r["mission_group_name"] or "",
                "mission_workgroup_code": r["mission_workgroup_code"] or "",
                "mission_workgroup_name": r["mission_workgroup_name"] or "",
                "team_leader_mission_code": r["team_leader_mission_code"] or "",
                "team_leader_name": r["team_leader_name"] or "",
                "operation_manager_mission_code": r["operation_manager_mission_code"] or "",
                "operation_manager_name": r["operation_manager_name"] or "",
                "administrative_manager_mission_code": r["administrative_manager_mission_code"] or "",
                "administrative_manager_name": r["administrative_manager_name"] or "",
            } for r in rows]
            return json.dumps({"success": True, "missions": missions}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_mission_workgroups(self):
        """Atölye/masa (mission workgroup) listesini getirir. MioCreate.xlsx -> MissionWorkgroup'tan seed edilmiştir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT id, code, short_name
                FROM organization.mission_workgroups
                ORDER BY short_name ASC
            """)).mappings().all()
            workgroups = [{"id": str(r["id"]), "code": r["code"], "short_name": r["short_name"]} for r in rows]
            return json.dumps({"success": True, "mission_workgroups": workgroups}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_item_faults(self):
        """Arıza tespiti referans listesini getirir. MioCreate.xlsx -> ItemFault'tan seed edilmiştir.
        Demontaj ekranındaki 'Arıza Tespiti' dropdown'u için kullanılır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT code, short_name, item_category
                FROM warehouse.item_fault
                ORDER BY short_name ASC
            """)).mappings().all()
            faults = [{"code": r["code"], "short_name": r["short_name"] or "", "item_category": r["item_category"] or ""} for r in rows]
            return json.dumps({"success": True, "item_faults": faults}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_item_faults_by_category(self, item_category):
        """warehouse.item_fault'tan, verilen item_category'ye (Parça Seçiniz'de seçilen
        parçanın kategorisi) uygun arıza tespiti seçeneklerini getirir. Demontaj ekranındaki
        'Arıza Tespiti' dropdown'unun kaynağıdır - seçilen parçanın kategorisine göre filtrelenir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            category = (item_category or "").strip()
            if not category:
                return json.dumps({"success": True, "item_faults": []})

            rows = db.execute(text("""
                SELECT code, short_name FROM warehouse.item_fault
                WHERE LOWER(TRIM(item_category)) = LOWER(:cat)
                ORDER BY short_name ASC
            """), {"cat": category}).mappings().all()
            faults = [{"code": r["code"], "short_name": r["short_name"] or "N/A"} for r in rows]
            return json.dumps({"success": True, "item_faults": faults}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    def _cached_json(self, key, ttl_seconds, compute_fn):
        """Uzak DB'ye her seferinde gitmek yerine, nadiren değişen referans/liste
        Slot'larının (statü listesi, depo durum, flow değerleri, görev grupları, garanti
        türleri, parça kategorileri vb.) sonucunu ttl_seconds boyunca bellekte tutar.
        compute_fn, JSON string döner (Slot'un normal dönüş değeriyle aynı format).
        Yalnızca success=true sonuçlar önbelleklenir - geçici bir DB/ağ hatası
        (ör. bu oturumda gözlemlenen uzak DB gecikme/kopma sorunları) dakikalarca
        önbelleklenip tekrar tekrar aynı hatayı döndürmesin."""
        import time, json as _json
        now = time.time()
        entry = self._ref_cache.get(key)
        if entry and (now - entry[0]) < ttl_seconds:
            return entry[1]
        value = compute_fn()
        try:
            ok = bool(_json.loads(value).get("success"))
        except Exception:
            ok = False
        if ok:
            self._ref_cache[key] = (now, value)
        else:
            self._ref_cache.pop(key, None)
        return value

    def _invalidate_cache(self, *keys):
        """Bir referans tablosu yazıldığında (create/update/delete) ilgili _cached_json
        anahtarını önbellekten düşürür ki bir sonraki okuma güncel veriyi görsün."""
        for k in keys:
            self._ref_cache.pop(k, None)

    def _get_flow_values(self, db):
        """warehouse.service_request_type.code'dan geçerli Flow (Akış Durumu) değerlerini
        döner - veritabanında sadece bu değerler kullanılabilir olmalıdır (boş/NULL kod
        satırları hariç tutulur). get_flow_values Slot'u ve dahili doğrulama/şablon
        kodu (generate_customer_bulk_template, bulk_import_customers) bunu paylaşır."""
        from sqlalchemy import text
        rows = db.execute(text("""
            SELECT DISTINCT TRIM(code) AS code FROM warehouse.service_request_type
            WHERE code IS NOT NULL AND TRIM(code) <> ''
            ORDER BY 1
        """)).fetchall()
        return [r[0] for r in rows]

    @Slot(result=str)
    def get_flow_values(self):
        """Batch Girişi, Parça Kategorileri ve Tedarikçiler sayfalarındaki Flow seçim
        listelerinin tek kaynağıdır (bkz. _get_flow_values). warehouse.service_request_type
        çalışma zamanında pratikte hiç değişmediğinden 5 dakika önbelleklenir."""
        def _compute():
            db = SessionLocal()
            try:
                flows = self._get_flow_values(db)
                return json.dumps({"success": True, "flows": flows}, ensure_ascii=False)
            except Exception as e:
                return json.dumps({"success": False, "message": str(e)})
            finally:
                db.close()
        return self._cached_json("flow_values", 300, _compute)

    # Qt slotu DEĞİLDİR - flow (Akış Durumu) değerini kanonik KODA çevirir.
    def _kanonik_flow(self, db, ham_flow):
        """batch_entries.flow bazen KODU ('To refurbish') bazen KISA ADI ('Refurbish')
        tutuyor. Canlı veride 7644 cihaz kısa ad, 33 cihaz kod - yani neredeyse tamamı
        kısa ad. Kural metinleri ('to refurbish', 'to rma') ve
        service_request_item_category.service_request_type ise KOD ile yazılmış.

        Ham değerle karşılaştırma yapılırsa 'Refurbish' akışındaki cihazlar hiçbir kurala
        uymuyor ve hepsi 'Müşteri Onayı Alınacak' tarafına düşüyordu. Burada değer
        service_request_type üzerinden (code VEYA short_name eşleşmesiyle) koda çevrilir.
        Eşleşme bulunamazsa ham değer olduğu gibi döner.
        """
        from sqlalchemy import text
        f = (ham_flow or "").strip()
        if not f:
            return ""
        row = db.execute(text("""
            SELECT code FROM warehouse.service_request_type
            WHERE LOWER(TRIM(code)) = LOWER(:f) OR LOWER(TRIM(short_name)) = LOWER(:f)
            LIMIT 1
        """), {"f": f}).first()
        return (row[0] or "").strip() if row and row[0] else f

    # Müşteri onayı hiç aranmayan akışlar (kanonik KOD ile yazılır).
    ONAY_GEREKTIRMEYEN_FLOWLAR = {"to rma", "to refurbish"}

    @Slot(str, result=str)
    def get_approved_categories_for_flow(self, flow):
        """warehouse.service_request_item_category'den, verilen Flow (Akış Durumu) için
        önceden onaylanmış (is_customer_approved=TRUE) parça kategorilerini döner. Demontaj
        ekranındaki 'Üretime Aktar'/'Müşteri Onayı Alınacak' butonunun canlı önizlemesi
        (submit_dismantle_decision ile aynı mantık) burada da kullanılır.

        Ayrıca noApprovalNeeded döner: bu akışta müşteri onayı hiç aranmıyorsa True.
        Bu bilgi eskiden ekranda ayrıca kodlanmıştı; iki yerde ayrı yazılınca kural
        zamanla ayrışıyordu, artık tek kaynak burası."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            ham = (flow or "").strip()
            if not ham:
                return json.dumps({"success": True, "categories": [], "noApprovalNeeded": False,
                                   "canonicalFlow": ""}, ensure_ascii=False)
            kanonik = self._kanonik_flow(db, ham)
            rows = db.execute(text("""
                SELECT DISTINCT item_category FROM warehouse.service_request_item_category
                WHERE LOWER(TRIM(service_request_type)) = LOWER(:flow) AND is_customer_approved = TRUE
            """), {"flow": kanonik}).fetchall()
            categories = [r[0] for r in rows if r[0]]
            return json.dumps({
                "success": True,
                "categories": categories,
                "canonicalFlow": kanonik,
                "noApprovalNeeded": kanonik.lower() in self.ONAY_GEREKTIRMEYEN_FLOWLAR,
            }, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_service_request_types_by_category(self, item_category):
        """warehouse.service_request_item_category'den, verilen item_category'ye (Parça
        Seçiniz'de seçilen parçanın kategorisi) uygun serviceRequestType değerlerini getirir.
        Demontaj ekranındaki 'Arıza Tespiti' dropdown'unun kaynağıdır - seçilen parçanın
        kategorisine göre filtrelenir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            category = (item_category or "").strip()
            if not category:
                return json.dumps({"success": True, "service_request_types": []})

            rows = db.execute(text("""
                SELECT DISTINCT service_request_type, is_customer_approved
                FROM warehouse.service_request_item_category
                WHERE LOWER(TRIM(item_category)) = LOWER(:cat)
                ORDER BY service_request_type ASC
            """), {"cat": category}).mappings().all()
            types = [{
                "code": (r["service_request_type"] or "").strip(),
                "short_name": (r["service_request_type"] or "").strip(),
                "is_customer_approved": bool(r["is_customer_approved"]),
            } for r in rows]
            return json.dumps({"success": True, "service_request_types": types}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_parts_for_device(self, device_model_text):
        """Verilen cihazın modelini (ör. batch_entries.model - serbest metin) önce
        product_family'ye eşleştirip, o cihazın reçetesini (BOM) warehouse.product_bom_node'dan
        okur (parent_product_code = product_family.code) ve reçetedeki her child_item_code'u
        warehouse.parts'tan çözerek döner. Demontaj ekranındaki 'Parça Seçiniz' kutusunun
        kaynağıdır - Product Bom sayfasında (create/update/delete/toggle_product_bom_status,
        aynı product_bom_node tablosu) tanımlanan reçeteyle birebir aynıdır. Model çözülemezse
        veya o modele ait reçete hiç girilmemişse boş liste + açıklayıcı bir 'warning' döner."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            model_text = (device_model_text or "").strip()
            if not model_text:
                return json.dumps({"success": True, "parts": [], "resolved_model": None, "warning": None})

            fam = db.execute(text("""
                SELECT code, short_name FROM warehouse.product_family
                WHERE LOWER(code) = LOWER(:m) OR LOWER(short_name) = LOWER(:m)
                ORDER BY (LOWER(code) = LOWER(:m)) DESC
                LIMIT 1
            """), {"m": model_text}).mappings().first()

            if not fam or not fam["code"]:
                return json.dumps({
                    "success": True, "parts": [], "resolved_model": None,
                    "warning": f"\"{model_text}\" bir ürün ailesiyle eşleştirilemedi."
                }, ensure_ascii=False)

            # Product Bom sayfasındaki 'Cihaz Modeli' seçici ayrı, eski bir referans tablosundan
            # (warehouse.product_families - sadece okunabilir isim, code yok) besleniyor, bu yüzden
            # kullanıcılar parent_product_code'a code (iP12PR) yerine okunabilir isim (iPhone 12 Pro)
            # girmiş olabilir - her iki ihtimal de büyük/küçük harf duyarsız kontrol edilir.
            # Ayrıca sadece Parça Kategorileri'nde "Ön Fiyat Verebilir" (item_category.is_pre_approved)
            # işaretli kategorilerdeki parçalar gösterilir.
            rows = db.execute(text("""
                SELECT p.id, p.item_code, p.name, p.brand, p.model, p.color,
                       p.item_category, p.part_category, p.part_type, b.quantity
                FROM warehouse.product_bom_node b
                JOIN warehouse.parts p ON p.item_code = b.child_item_code
                WHERE LOWER(TRIM(b.parent_product_code)) IN (LOWER(:code), LOWER(:short_name))
                  AND b.enabled = TRUE
                  AND EXISTS (
                      SELECT 1 FROM warehouse.item_category ic
                      WHERE LOWER(TRIM(ic.short_name)) = LOWER(TRIM(p.item_category))
                        AND ic.is_pre_approved = TRUE
                  )
                ORDER BY p.item_category, p.item_code
            """), {"code": fam["code"], "short_name": fam["short_name"]}).mappings().all()

            parts = [{
                "id": str(r["id"]),
                "item_code": r["item_code"] or "",
                "name": r["name"] or "",
                "brand": r["brand"] or "",
                "model": r["model"] or "",
                "color": r["color"] or "",
                "item_category": r["item_category"] or "",
                "part_category": r["part_category"] or "",
                "part_type": r["part_type"] or "",
                "quantity": r["quantity"] or 1,
            } for r in rows]

            warning = None
            if not parts:
                bom_exists = db.execute(text("""
                    SELECT 1 FROM warehouse.product_bom_node b
                    WHERE LOWER(TRIM(b.parent_product_code)) IN (LOWER(:code), LOWER(:short_name))
                      AND b.enabled = TRUE
                    LIMIT 1
                """), {"code": fam["code"], "short_name": fam["short_name"]}).first()
                if bom_exists:
                    warning = f"{fam['short_name']} için reçete (BOM) tanımlı ama hiçbir parça kategorisi 'Ön Fiyat Verebilir' olarak işaretli değil. Parça Kategorileri sayfasından güncelleyebilirsiniz."
                else:
                    warning = f"{fam['short_name']} için henüz reçete (BOM) tanımlanmamış. Product Bom sayfasından ekleyebilirsiniz."

            return json.dumps({"success": True, "parts": parts, "resolved_model": fam["short_name"], "warning": warning}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_repair_item_operation_types(self):
        """İşlem tipi referans listesini getirir (Onar / Parça Değişim).
        MioCreate.xlsx -> RepairItemOperationType'tan seed edilmiştir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT code, short_name
                FROM warehouse.repair_item_operation_type
                ORDER BY order_number ASC NULLS LAST
            """)).mappings().all()
            types = [{"code": r["code"], "short_name": r["short_name"] or ""} for r in rows]
            return json.dumps({"success": True, "operation_types": types}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_repair_item_warranties(self):
        """Ücret tipi (Ücretli/Ücretsiz Onarım) referans listesini getirir.
        MioCreate.xlsx -> RepairItemWarranty'den seed edilmiştir (IW=Ücretsiz, OOW=Ücretli).
        Çalışma zamanında pratikte hiç değişmediğinden 5 dakika önbelleklenir."""
        def _compute():
            from sqlalchemy import text
            db = SessionLocal()
            try:
                rows = db.execute(text("""
                    SELECT code, short_name, is_paid_for
                    FROM warehouse.repair_item_warranty
                    ORDER BY order_number ASC NULLS LAST
                """)).mappings().all()
                warranties = [{"code": r["code"], "short_name": r["short_name"] or "", "is_paid_for": bool(r["is_paid_for"])} for r in rows]
                return json.dumps({"success": True, "warranties": warranties}, ensure_ascii=False)
            except Exception as e:
                return json.dumps({"success": False, "message": str(e)})
            finally:
                db.close()
        return self._cached_json("repair_item_warranties", 300, _compute)

    @Slot(str, result=str)
    def get_test_detected_parts(self, device_ref):
        """Test aşamasında (QAC) tespit edilen, onarımda beklenen parçaları getirir.
        Bu tabloyu şu an hiçbir ekran yazmıyor (QAC test ekranı henüz yok) — bu yüzden
        gerçek veri gelene kadar hep boş liste döner."""
        from models.test_detected_part import TestDetectedPart
        db = SessionLocal()
        try:
            if not device_ref or not str(device_ref).strip():
                return json.dumps({"success": True, "parts": []})
            rows = db.query(TestDetectedPart).filter(TestDetectedPart.device_ref == str(device_ref).strip()).all()
            parts = [{
                "id": str(r.id),
                "symptomCode": r.symptom_code or "",
                "partCategory": r.part_category or "",
                "partItemCode": r.part_item_code or "",
            } for r in rows]
            return json.dumps({"success": True, "parts": parts}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    def _resolve_mission_fk_fields(self, db, mission_group_code, mission_workgroup_code):
        """mission_group_code/mission_workgroup_code metinlerini ilgili tabloların UUID id'lerine çevirir.
        Kod verilmiş ama bulunamamışsa (None, None, hata_mesaji) döner."""
        from models.mission_group import MissionGroup
        from models.mission_workgroup import MissionWorkgroup

        mission_group_id = None
        if mission_group_code and mission_group_code.strip():
            mg = db.query(MissionGroup).filter(MissionGroup.code == mission_group_code.strip()).first()
            if not mg:
                return None, None, f"Görev grubu kodu bulunamadı: {mission_group_code}"
            mission_group_id = mg.id

        mission_workgroup_id = None
        if mission_workgroup_code and mission_workgroup_code.strip():
            mw = db.query(MissionWorkgroup).filter(MissionWorkgroup.code == mission_workgroup_code.strip()).first()
            if not mw:
                return None, None, f"Atölye/masa kodu bulunamadı: {mission_workgroup_code}"
            mission_workgroup_id = mw.id

        return mission_group_id, mission_workgroup_id, None

    @Slot(str, str, str, str, str, str, str, str, str, str, str, str, result=str)
    def create_mission(self, code, short_name, full_name, description, cost_center, department,
                        order_number, mission_group_code, mission_workgroup_code,
                        team_leader_mission_code, operation_manager_mission_code, administrative_manager_mission_code):
        """Yeni bir görev/rol (organization.missions) ekler."""
        import uuid
        from models.mission import Mission
        db = SessionLocal()
        try:
            code = (code or "").strip()
            short_name = (short_name or "").strip()
            if not code or not short_name:
                return json.dumps({"success": False, "message": "Kod ve kısa ad zorunludur."})

            mission_group_id, mission_workgroup_id, err = self._resolve_mission_fk_fields(db, mission_group_code, mission_workgroup_code)
            if err:
                return json.dumps({"success": False, "message": err})

            rec = Mission(
                id=uuid.uuid4(),
                code=code,
                short_name=short_name,
                full_name=full_name.strip() if full_name else None,
                description=description.strip() if description else None,
                cost_center=cost_center.strip() if cost_center else None,
                department=department.strip() if department else None,
                order_number=float(order_number) if order_number and order_number.strip() else None,
                mission_group_id=mission_group_id,
                mission_workgroup_id=mission_workgroup_id,
                team_leader_mission_code=team_leader_mission_code.strip() if team_leader_mission_code else None,
                operation_manager_mission_code=operation_manager_mission_code.strip() if operation_manager_mission_code else None,
                administrative_manager_mission_code=administrative_manager_mission_code.strip() if administrative_manager_mission_code else None,
            )
            db.add(rec)
            db.commit()
            return json.dumps({"success": True, "id": str(rec.id)})
        except Exception as e:
            db.rollback()
            if "unique" in str(e).lower() or "duplicate" in str(e).lower():
                return json.dumps({"success": False, "message": f"Bu kod zaten kullanılıyor: {code}"})
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, str, str, str, str, result=str)
    def update_mission(self, mission_id, code, short_name, full_name, description, cost_center, department,
                        order_number, mission_group_code, mission_workgroup_code,
                        team_leader_mission_code, operation_manager_mission_code, administrative_manager_mission_code):
        """Var olan bir görevi/rolü (organization.missions) günceller."""
        from models.mission import Mission
        db = SessionLocal()
        try:
            rec = db.query(Mission).filter(Mission.id == mission_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Görev bulunamadı."})

            code = (code or "").strip()
            short_name = (short_name or "").strip()
            if not code or not short_name:
                return json.dumps({"success": False, "message": "Kod ve kısa ad zorunludur."})

            mission_group_id, mission_workgroup_id, err = self._resolve_mission_fk_fields(db, mission_group_code, mission_workgroup_code)
            if err:
                return json.dumps({"success": False, "message": err})

            rec.code = code
            rec.short_name = short_name
            rec.full_name = full_name.strip() if full_name else None
            rec.description = description.strip() if description else None
            rec.cost_center = cost_center.strip() if cost_center else None
            rec.department = department.strip() if department else None
            rec.order_number = float(order_number) if order_number and order_number.strip() else None
            rec.mission_group_id = mission_group_id
            rec.mission_workgroup_id = mission_workgroup_id
            rec.team_leader_mission_code = team_leader_mission_code.strip() if team_leader_mission_code else None
            rec.operation_manager_mission_code = operation_manager_mission_code.strip() if operation_manager_mission_code else None
            rec.administrative_manager_mission_code = administrative_manager_mission_code.strip() if administrative_manager_mission_code else None
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            if "unique" in str(e).lower() or "duplicate" in str(e).lower():
                return json.dumps({"success": False, "message": f"Bu kod zaten kullanılıyor: {code}"})
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_mission(self, mission_id):
        """Bir görevi/rolü (organization.missions) siler. Başka görevlerin amir zincirinde
        bu görevin koduna referans varsa engellemez, sadece bilgi notu döner."""
        from sqlalchemy import text
        from models.mission import Mission
        db = SessionLocal()
        try:
            rec = db.query(Mission).filter(Mission.id == mission_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Görev bulunamadı."})

            ref_count = db.execute(text("""
                SELECT COUNT(*) FROM organization.missions
                WHERE team_leader_mission_code = :code
                   OR operation_manager_mission_code = :code
                   OR administrative_manager_mission_code = :code
            """), {"code": rec.code}).scalar()

            db.delete(rec)
            db.commit()
            note = f"{ref_count} görevin amir zincirinde bu koda referans vardı, otomatik güncellenmedi." if ref_count else ""
            return json.dumps({"success": True, "note": note})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # --- FLOW -> DGD İŞÇİLİK KODU EŞLEŞMESİ ---

    @Slot(result=str)
    def get_flow_dgd_mappings(self):
        """warehouse.flow_dgd_mapping'deki tüm Flow -> DGD işçilik kodu eşleşmelerini döner."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT id, flow_code, dgd_item_code, enabled
                FROM warehouse.flow_dgd_mapping ORDER BY flow_code
            """)).mappings().all()
            items = [{"id": str(r["id"]), "flow_code": r["flow_code"], "dgd_item_code": r["dgd_item_code"], "enabled": bool(r["enabled"])} for r in rows]
            return json.dumps({"success": True, "mappings": items}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def create_flow_dgd_mapping(self, flow_code, dgd_item_code):
        """Yeni bir Flow -> DGD işçilik kodu eşleşmesi (warehouse.flow_dgd_mapping) ekler."""
        import uuid
        from models.flow_dgd_mapping import FlowDgdMapping
        db = SessionLocal()
        try:
            flow_code = (flow_code or "").strip()
            dgd_item_code = (dgd_item_code or "").strip()
            if not flow_code or not dgd_item_code:
                return json.dumps({"success": False, "message": "Flow ve DGD kodu zorunludur."})

            rec = FlowDgdMapping(id=uuid.uuid4(), flow_code=flow_code, dgd_item_code=dgd_item_code, enabled=True)
            db.add(rec)
            db.commit()
            return json.dumps({"success": True, "id": str(rec.id)})
        except Exception as e:
            db.rollback()
            if "unique" in str(e).lower() or "duplicate" in str(e).lower():
                return json.dumps({"success": False, "message": f"Bu Flow için zaten bir DGD kodu tanımlı: {flow_code}"})
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def update_flow_dgd_mapping(self, mapping_id, flow_code, dgd_item_code):
        """Var olan bir Flow -> DGD işçilik kodu eşleşmesini günceller."""
        from models.flow_dgd_mapping import FlowDgdMapping
        db = SessionLocal()
        try:
            rec = db.query(FlowDgdMapping).filter(FlowDgdMapping.id == mapping_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Eşleşme bulunamadı."})

            flow_code = (flow_code or "").strip()
            dgd_item_code = (dgd_item_code or "").strip()
            if not flow_code or not dgd_item_code:
                return json.dumps({"success": False, "message": "Flow ve DGD kodu zorunludur."})

            rec.flow_code = flow_code
            rec.dgd_item_code = dgd_item_code
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            if "unique" in str(e).lower() or "duplicate" in str(e).lower():
                return json.dumps({"success": False, "message": f"Bu Flow için zaten bir DGD kodu tanımlı: {flow_code}"})
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_flow_dgd_mapping(self, mapping_id):
        """Bir Flow -> DGD işçilik kodu eşleşmesini siler."""
        from models.flow_dgd_mapping import FlowDgdMapping
        db = SessionLocal()
        try:
            rec = db.query(FlowDgdMapping).filter(FlowDgdMapping.id == mapping_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Eşleşme bulunamadı."})
            db.delete(rec)
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # --- PARÇA KATEGORİSİ MODÜLÜ ---
    @Slot(result=str)
    def get_part_categories(self):
        """Tüm Parça Kategorilerini getirir. (Modül 2: ItemCategory ve ItemCategoryMission üzerinden beslenir)
        Departman listesi eskiden her satır için ayrı bir correlated subquery ile
        hesaplanıyordu (196 kategori x 301 mission satırı taranıyordu, tek bir LEFT
        JOIN + GROUP BY yerine) - uzak DB'de saniyeler süren en yavaş Slot'tu, tek geçişli
        JOIN'e çevrildi. Kategori listesi create/update/delete ile değişebildiğinden kısa
        (60sn) bir TTL ile önbelleklenir; yazma Slot'ları önbelleği geçersiz kılar."""
        def _compute():
            from sqlalchemy import text
            import json
            db = SessionLocal()
            try:
                rows = db.execute(text("""
                    SELECT
                        ic.id,
                        ic.short_name AS name,
                        '' AS part_type,
                        '' AS flow,
                        string_agg(DISTINCT icm.mission, ', ') AS departments,
                        'Stok Takipli' AS stock_tracking_type,
                        NULL AS default_location_id,
                        '' AS default_location_name,
                        ic.enabled AS is_active,
                        '' AS description,
                        ic.item_labour AS item_labour,
                        ic.is_pre_approved AS is_pre_approved
                    FROM warehouse.item_category ic
                    LEFT JOIN warehouse.item_category_mission icm
                        ON icm.item_category = ic.short_name OR icm.item_category = ic.code
                    GROUP BY ic.id, ic.short_name, ic.enabled, ic.item_labour, ic.is_pre_approved
                    ORDER BY ic.short_name ASC
                """)).mappings().all()

                categories = []
                for r in rows:
                    categories.append({
                        "id": str(r["id"]),
                        "name": r["name"] or "",
                        "part_type": r["part_type"] or "",
                        "flow": r["flow"] or "",
                        "departments": r["departments"] or "",
                        "stock_tracking_type": r["stock_tracking_type"] or "Stok Takipli",
                        "default_location_id": str(r["default_location_id"]) if r["default_location_id"] else "",
                        "default_location_name": r["default_location_name"] or "",
                        "is_active": r["is_active"] if r["is_active"] is not None else True,
                        "description": r["description"] or "",
                        "labour_level": r["item_labour"] or "",
                        "can_pre_price": bool(r["is_pre_approved"])
                    })

                return json.dumps({"success": True, "categories": categories})
            except Exception as e:
                print(f"[WebBridge] get_part_categories error: {e}")
                return json.dumps({"success": False, "message": str(e)})
            finally:
                db.close()
        return self._cached_json("part_categories", 60, _compute)

    @Slot(str, str, str, str, str, str, str, result=str)
    def create_part_category(self, name, part_type, flow, departments, stock_tracking_type, default_location_id, description):
        """Yeni Parça Kategorisi ekler."""
        from models.part_category import PartCategory
        db = SessionLocal()
        try:
            name = (name or "").strip()
            if not name:
                return json.dumps({"success": False, "message": "Kategori adı zorunludur"})
            if db.query(PartCategory).filter(PartCategory.name == name).first():
                return json.dumps({"success": False, "message": "Bu kategori zaten var"})
            cat = PartCategory(
                name=name,
                part_type=part_type or None,
                flow=flow or None,
                departments=departments or None,
                stock_tracking_type=stock_tracking_type or "Stok Takipli",
                is_active=True,
                description=description or None
            )
            db.add(cat)
            db.commit()
            self._invalidate_cache("part_categories")
            return json.dumps({"success": True, "message": "Kategori eklendi", "id": cat.id})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, result=str)
    def update_part_category(self, id_str, name, part_type, flow, departments, stock_tracking_type, default_location_id, is_active, description):
        """Var olan bir Parça Kategorisini günceller."""
        from models.part_category import PartCategory
        db = SessionLocal()
        try:
            cat_id = int(id_str)
            name = (name or "").strip()
            if not name:
                return json.dumps({"success": False, "message": "Kategori adı zorunludur"})
            cat = db.query(PartCategory).filter(PartCategory.id == cat_id).first()
            if not cat:
                return json.dumps({"success": False, "message": "Kategori bulunamadı"})
            if db.query(PartCategory).filter(PartCategory.name == name, PartCategory.id != cat_id).first():
                return json.dumps({"success": False, "message": "Bu isimde başka bir kategori zaten var"})
            cat.name = name
            cat.part_type = part_type or None
            cat.flow = flow or None
            cat.departments = departments or None
            cat.stock_tracking_type = stock_tracking_type or "Stok Takipli"
            cat.is_active = (is_active == "true" or is_active == "1" or is_active == "True")
            cat.description = description or None
            db.commit()
            self._invalidate_cache("part_categories")
            return json.dumps({"success": True, "message": "Kategori güncellendi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_part_category(self, id_str):
        from sqlalchemy import text
        from models.part_category import PartCategory
        db = SessionLocal()
        try:
            cat_id = int(id_str)
            cat = db.query(PartCategory).filter(PartCategory.id == cat_id).first()
            if not cat:
                return json.dumps({"success": False, "message": "Bulunamadı"})
            linked = db.execute(text("SELECT COUNT(*) FROM warehouse.parts WHERE part_category_id = :id"), {"id": cat_id}).scalar()
            if linked:
                return json.dumps({"success": False, "message": f"Bu kategoriye bağlı {linked} parça var, önce onları başka kategoriye taşıyın."})
            db.delete(cat)
            db.commit()
            self._invalidate_cache("part_categories")
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # ==========================
    # SERVİS KAYITLARI MODÜLÜ
    # ==========================

    @Slot(str, result=str)
    def get_repair_details_by_imei(self, imei_number):
        """IMEI numarasına göre cihaz bilgilerini ve onarım kaydını getirir."""
        from models.batch_entry import BatchEntry
        from models.service_repair import ServiceRepair
        db = SessionLocal()
        try:
            # 1. Cihaz bilgilerini BatchEntries'ten al
            batch_entry = db.query(BatchEntry).filter(BatchEntry.imei_number == imei_number).first()
            if not batch_entry:
                return json.dumps({"success": False, "message": f"Bu IMEI numarasına ({imei_number}) ait bir cihaz (BatchEntry) bulunamadı."})
                
            # 2. Servis onarım kaydını al
            repair_record = db.query(ServiceRepair).filter(ServiceRepair.imei_number == imei_number).first()
            
            # 3. Verileri birleştir
            data = {
                # Cihaz Bilgileri (BatchEntry - Salt Okunur)
                "imei_number": batch_entry.imei_number,
                "customer_name": batch_entry.customer_name or "",
                "customer_phone": "", # BatchEntry'de yok, eklenebilir veya boş kalabilir
                "brand": batch_entry.model or "", # BatchEntry'deki model genelde marka+model veya sadece model tutuyor
                "model": batch_entry.model or "",
                "memory": batch_entry.gb or "",
                "color": batch_entry.color or "",
                "customer_complaint": batch_entry.defects or "",
                "flow": batch_entry.flow or "",
                "created_at": batch_entry.created_at.strftime('%d.%m.%Y %H:%M') if batch_entry.created_at else "",
                
                # Onarım Bilgileri (ServiceRepair - Düzenlenebilir)
                "repair_id": str(repair_record.id) if repair_record else None,
                "technician_name": repair_record.technician_name if repair_record else "",
                "fault_description": repair_record.fault_description if repair_record else (batch_entry.defects or ""),
                "repair_notes": repair_record.repair_notes if repair_record else "",
                "status": repair_record.status if repair_record else "Arıza Kabul",
                "warranty_status": repair_record.warranty_status if repair_record else ""
            }
            
            return json.dumps({"success": True, "data": data})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def save_service_repair(self, data_json):
        """IMEI'ye göre servis onarım kaydını günceller veya oluşturur."""
        from models.service_repair import ServiceRepair
        db = SessionLocal()
        try:
            data = json.loads(data_json)
            imei_number = data.get("imei_number")
            if not imei_number:
                return json.dumps({"success": False, "message": "IMEI numarası eksik."})
                
            record = db.query(ServiceRepair).filter(ServiceRepair.imei_number == imei_number).first()
            
            if record:
                # Güncelle
                record.technician_name = data.get("technician_name", record.technician_name)
                record.fault_description = data.get("fault_description", record.fault_description)
                record.repair_notes = data.get("repair_notes", record.repair_notes)
                record.status = data.get("status", record.status)
                record.warranty_status = data.get("warranty_status", record.warranty_status)
            else:
                # Yeni oluştur
                new_record = ServiceRepair(
                    imei_number=imei_number,
                    technician_name=data.get("technician_name", ""),
                    fault_description=data.get("fault_description", ""),
                    repair_notes=data.get("repair_notes", ""),
                    status=data.get("status", "Arıza Kabul"),
                    warranty_status=data.get("warranty_status", "")
                )
                db.add(new_record)
                
            db.commit()
            return json.dumps({"success": True, "message": "Onarım kaydı başarıyla kaydedildi."})
        except Exception as e:
            db.rollback()
            import traceback
            traceback.print_exc()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()


    @Slot(result=str)
    def get_service_records(self):
        """Tüm servis kayıtlarını getirir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT id, customer_name, customer_phone, customer_email, company,
                       brand, model, memory, product_code, imei_number, color, fault_category, fault_type,
                       customer_complaint, preliminary_diagnosis, status, technician_note, created_at
                FROM warehouse.service_records
                ORDER BY id DESC
                LIMIT 200
            """)).mappings().all()
            records = []
            for row in rows:
                records.append({
                    "id": str(row["id"]),
                    "customer_name": row["customer_name"] or "",
                    "customer_phone": row["customer_phone"] or "",
                    "customer_email": row["customer_email"] or "",
                    "company": row["company"] or "",
                    "brand": row["brand"] or "",
                    "model": row["model"] or "",
                    "memory": row["memory"] or "",
                    "product_code": row["product_code"] or "",
                    "imei_number": row["imei_number"] or "",
                    "color": row["color"] or "",
                    "fault_category": row["fault_category"] or "",
                    "fault_type": row["fault_type"] or "",
                    "customer_complaint": row["customer_complaint"] or "",
                    "preliminary_diagnosis": row["preliminary_diagnosis"] or "",
                    "status": row["status"] or "Arıza Kabul",
                    "technician_note": row["technician_note"] or "",
                    "created_at": row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else ""
                })
            return json.dumps({"success": True, "records": records})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, result=str)
    def create_service_record(self, customer_name, customer_phone, customer_email, company,
                               brand, model, memory, product_code, imei_number, color, fault_category, fault_type,
                               customer_complaint, preliminary_diagnosis, status, technician_note):
        """Yeni servis kaydı ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            name = customer_name.strip()
            if not name:
                return json.dumps({"success": False, "message": "Müşteri adı zorunludur"})

            db.execute(text("""
                INSERT INTO warehouse.service_records (
                    customer_name, customer_phone, customer_email, company,
                    brand, model, memory, product_code, imei_number, color, fault_category, fault_type,
                    customer_complaint, preliminary_diagnosis, status, technician_note
                ) VALUES (
                    :name, :phone, :email, :company,
                    :brand, :model, :memory, :code, :imei, :color, :fcat, :ftype,
                    :complaint, :diagnosis, :status, :note
                )
            """), {
                "name": name, "phone": customer_phone or None, "email": customer_email or None,
                "company": company or None, "brand": brand or None, "model": model or None,
                "memory": memory or None, "code": product_code or None, "imei": imei_number or None, "color": color or None,
                "fcat": fault_category or None, "ftype": fault_type or None,
                "complaint": customer_complaint or None, "diagnosis": preliminary_diagnosis or None,
                "status": status or "Arıza Kabul", "note": technician_note or None
            })
            db.commit()
            return json.dumps({"success": True, "message": "Servis kaydı eklendi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Kayıt hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, result=str)
    def update_service_record(self, record_id_str, customer_name, customer_phone, customer_email, company,
                               brand, model, memory, product_code, imei_number, color, fault_category, fault_type,
                               customer_complaint, preliminary_diagnosis, status, technician_note):
        """Var olan bir servis kaydını günceller."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            record_id = int(record_id_str)
            name = customer_name.strip()
            if not name:
                return json.dumps({"success": False, "message": "Müşteri adı zorunludur"})

            db.execute(text("""
                UPDATE warehouse.service_records
                SET customer_name = :name, customer_phone = :phone, customer_email = :email, company = :company,
                    brand = :brand, model = :model, memory = :memory, product_code = :code, imei_number = :imei, color = :color,
                    fault_category = :fcat, fault_type = :ftype,
                    customer_complaint = :complaint, preliminary_diagnosis = :diagnosis,
                    status = :status, technician_note = :note
                WHERE id = :id
            """), {
                "name": name, "phone": customer_phone or None, "email": customer_email or None,
                "company": company or None, "brand": brand or None, "model": model or None,
                "memory": memory or None, "code": product_code or None, "imei": imei_number or None, "color": color or None,
                "fcat": fault_category or None, "ftype": fault_type or None,
                "complaint": customer_complaint or None, "diagnosis": preliminary_diagnosis or None,
                "status": status or "Arıza Kabul", "note": technician_note or None,
                "id": record_id
            })
            db.commit()
            return json.dumps({"success": True, "message": "Servis kaydı güncellendi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Güncelleme hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_service_record(self, record_id_str):
        """Belirtilen id'ye sahip servis kaydını siler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            record_id = int(record_id_str)
            db.execute(text("DELETE FROM warehouse.service_records WHERE id = :id"), {"id": record_id})
            db.commit()
            return json.dumps({"success": True, "message": "Servis kaydı silindi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Silme hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(result=str)
    def generate_customer_bulk_template(self):
        """Müşteriler sayfası toplu (Excel) yükleme şablonunu üretir. Cihaz Modeli ve Flow
        (İş Akışı) sütunlarına Excel Data Validation ile açılır liste eklenir; zorunlu
        sütun başlıkları kırmızı ile işaretlenir. export_table_to_excel ile aynı
        konvansiyonu kullanır: Downloads klasörüne kaydeder ve dosyayı otomatik açar."""
        import os
        from pathlib import Path
        from sqlalchemy import text
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill, Alignment
        from core.excel_utils import style_excel_file

        db = SessionLocal()
        try:
            model_rows = db.execute(text("""
                SELECT DISTINCT brand, model FROM warehouse.products
                WHERE brand IS NOT NULL AND model IS NOT NULL AND brand <> '' AND model <> ''
                ORDER BY brand, model
            """)).mappings().all()
            device_models = [f"{r['brand']} {r['model']}".strip() for r in model_rows] or ["Tanımlı ürün bulunamadı"]

            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Toplu Cihaz Girişi"

            headers = [c[0] for c in CUSTOMER_BULK_REQUIRED_COLUMNS] + [
                "Müşteri Adı", "Müşteri Telefon", "Müşteri E-posta", "Firma"
            ]
            required_col_count = len(CUSTOMER_BULK_REQUIRED_COLUMNS)
            sheet.append(headers)

            flow_values = self._get_flow_values(db) or ["To refurbish"]

            # Örnek satır, kullanıcıya beklenen formatı gösterir.
            sheet.append([
                "353XXXXXXXXXXXX", "SN-000123", "INT-000123",
                device_models[0], flow_values[0],
                "Ekran kırık, dokunmatik çalışmıyor", "2026-01-15",
                "Ahmet Yılmaz", "05XXXXXXXXX", "", ""
            ])

            max_data_row = 500

            # Gizli "Listeler" sayfası: dropdown kaynakları buradan referans alınır
            # (Cihaz Modeli listesi 255 karakter inline sınırını aşabileceği için).
            list_sheet = wb.create_sheet("Listeler")
            list_sheet["A1"] = "Cihaz Modelleri"
            for i, dm in enumerate(device_models, start=2):
                list_sheet.cell(row=i, column=1, value=dm)
            list_sheet.sheet_state = "hidden"

            model_range = f"Listeler!$A$2:$A${len(device_models) + 1}"
            model_dv = DataValidation(type="list", formula1=f"={model_range}", allow_blank=True, showErrorMessage=True)
            model_dv.error = "Lütfen listeden geçerli bir Cihaz Modeli seçin."
            model_dv.errorTitle = "Geçersiz Cihaz Modeli"
            sheet.add_data_validation(model_dv)

            flow_list = ",".join(flow_values)
            flow_dv = DataValidation(type="list", formula1=f'"{flow_list}"', allow_blank=True, showErrorMessage=True)
            flow_dv.error = "Lütfen listeden geçerli bir Flow (İş Akışı) değeri seçin."
            flow_dv.errorTitle = "Geçersiz Flow"
            sheet.add_data_validation(flow_dv)

            model_col_letter = openpyxl.utils.get_column_letter(headers.index("Cihaz Modeli") + 1)
            flow_col_letter = openpyxl.utils.get_column_letter(headers.index("Flow (İş Akışı)") + 1)
            model_dv.add(f"{model_col_letter}2:{model_col_letter}{max_data_row}")
            flow_dv.add(f"{flow_col_letter}2:{flow_col_letter}{max_data_row}")

            # Giriş Tarihi sütununu metin olarak biçimlendir (kullanıcı YYYY-AA-GG girer);
            # Excel'in kendi tarih otomatik-biçimlendirmesiyle karışmasın diye.
            intake_col_letter = openpyxl.utils.get_column_letter(headers.index("Giriş Tarihi") + 1)
            for row_idx in range(2, max_data_row + 1):
                sheet[f"{intake_col_letter}{row_idx}"].number_format = "@"

            downloads_path = str(Path.home() / "Downloads")
            filename = "musteriler_toplu_yukleme_sablonu.xlsx"
            file_path = os.path.join(downloads_path, filename)
            counter = 1
            base_name, ext = os.path.splitext(filename)
            while os.path.exists(file_path):
                file_path = os.path.join(downloads_path, f"{base_name}_{counter}{ext}")
                counter += 1

            wb.save(file_path)

            try:
                style_excel_file(file_path)
            except Exception:
                pass

            # Zorunlu sütun başlıklarını kırmızıyla vurgula (style_excel_file'dan SONRA,
            # üzerine yazılmasın diye tekrar açıp kaydediyoruz).
            wb2 = openpyxl.load_workbook(file_path)
            sheet2 = wb2["Toplu Cihaz Girişi"]
            # style_excel_file artık zebra/kenarlık yerine Excel'in yerleşik "Table" özelliğini
            # kullanıyor; Table'ın XML'de ayrıca sakladığı sütun adları, aşağıda hücre değerini
            # değiştirdiğimizde otomatik güncellenmez - elle senkronize etmezsek Excel dosyayı
            # bozuk/onarım-gerekli olarak işaretler.
            table_cols_by_idx = {}
            if sheet2.tables:
                tbl = next(iter(sheet2.tables.values()))
                for i, col in enumerate(tbl.tableColumns, start=1):
                    table_cols_by_idx[i] = col
            required_fill = PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid")
            for col_idx in range(1, required_col_count + 1):
                cell = sheet2.cell(row=1, column=col_idx)
                cell.value = f"{cell.value} *"
                cell.fill = required_fill
                cell.font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if col_idx in table_cols_by_idx:
                    table_cols_by_idx[col_idx].name = str(cell.value)
            sheet2["A3"] = "(*) işaretli sütunlar zorunludur. Örnek satırı (2. satır) silip kendi verilerinizi girin."
            wb2.save(file_path)

            os.startfile(file_path)
            return json.dumps({"success": True})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def bulk_import_customers(self, rows_json):
        """Toplu (Excel) müşteri/cihaz kabul içe aktarma. Tüm satırları önce doğrular;
        herhangi bir satırda herhangi bir zorunlu alan eksikse veya geçersizse HİÇBİR
        satır kaydedilmez, tüm hatalar satır numarasıyla birlikte tek seferde döner."""
        from sqlalchemy import text
        from datetime import datetime
        db = SessionLocal()
        try:
            try:
                rows = json.loads(rows_json or "[]")
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Geçersiz dosya verisi.", "errors": []})

            if not rows:
                return json.dumps({"success": False, "message": "Dosyada içe aktarılacak satır bulunamadı.", "errors": []})

            # Cihaz Modeli -> (brand, model, product_code) eşlemesi için ürün listesini çek.
            product_rows = db.execute(text("""
                SELECT brand, model, item_code FROM warehouse.products
                WHERE brand IS NOT NULL AND model IS NOT NULL
            """)).mappings().all()
            model_lookup = {f"{r['brand']} {r['model']}".strip().lower(): r for r in product_rows}

            existing_imeis = {r[0] for r in db.execute(text(
                "SELECT imei_number FROM warehouse.customers WHERE imei_number IS NOT NULL"
            )).all()}
            existing_serials = {r[0] for r in db.execute(text(
                "SELECT serial_number FROM warehouse.customers WHERE serial_number IS NOT NULL"
            )).all()}

            valid_flow_values = self._get_flow_values(db)

            errors = []
            seen_imeis_in_file = {}
            seen_serials_in_file = {}
            valid_rows = []

            for idx, row in enumerate(rows):
                row_num = idx + 2  # 1. satır başlık; ilk veri satırı Excel'de 2. satır
                row = row or {}

                def get_val(key):
                    v = row.get(key)
                    return str(v).strip() if v is not None else ""

                imei = get_val("imei_number")
                serial = get_val("serial_number")
                internal_id = get_val("internal_id")
                cihaz_modeli = get_val("cihaz_modeli")
                flow = get_val("flow")
                complaint = get_val("customer_reported_complaint")
                intake_date = get_val("intake_date")

                for label, value in [
                    ("IMEI Numarası", imei), ("Seri Numarası", serial), ("Internal ID", internal_id),
                    ("Cihaz Modeli", cihaz_modeli), ("Flow (İş Akışı)", flow),
                    ("Müşteri Şikayeti", complaint), ("Giriş Tarihi", intake_date)
                ]:
                    if not value:
                        errors.append({"row": row_num, "field": label, "message": f"{label} boş olamaz."})

                if flow and flow not in valid_flow_values:
                    errors.append({"row": row_num, "field": "Flow (İş Akışı)", "message": f"Geçersiz değer: \"{flow}\". Geçerli değerler: {', '.join(valid_flow_values)}"})

                product = None
                if cihaz_modeli:
                    product = model_lookup.get(cihaz_modeli.strip().lower())
                    if not product:
                        errors.append({"row": row_num, "field": "Cihaz Modeli", "message": f"\"{cihaz_modeli}\" sistemde tanımlı bir ürün değil."})

                if intake_date:
                    try:
                        datetime.strptime(intake_date[:10], "%Y-%m-%d")
                    except ValueError:
                        errors.append({"row": row_num, "field": "Giriş Tarihi", "message": f"\"{intake_date}\" geçerli bir tarih değil (YYYY-AA-GG bekleniyor)."})

                if imei:
                    if imei in existing_imeis:
                        errors.append({"row": row_num, "field": "IMEI Numarası", "message": f"\"{imei}\" zaten sistemde kayıtlı."})
                    elif imei in seen_imeis_in_file:
                        errors.append({"row": row_num, "field": "IMEI Numarası", "message": f"\"{imei}\" dosyada birden fazla satırda tekrarlanıyor (satır {seen_imeis_in_file[imei]})."})
                    else:
                        seen_imeis_in_file[imei] = row_num

                if serial:
                    if serial in existing_serials:
                        errors.append({"row": row_num, "field": "Seri Numarası", "message": f"\"{serial}\" zaten sistemde kayıtlı."})
                    elif serial in seen_serials_in_file:
                        errors.append({"row": row_num, "field": "Seri Numarası", "message": f"\"{serial}\" dosyada birden fazla satırda tekrarlanıyor (satır {seen_serials_in_file[serial]})."})
                    else:
                        seen_serials_in_file[serial] = row_num

                valid_rows.append({
                    "imei_number": imei or None, "serial_number": serial or None, "internal_id": internal_id or None,
                    "flow": flow or None, "customer_reported_complaint": complaint or None,
                    "intake_date": intake_date[:10] if intake_date else None,
                    "brand": product["brand"] if product else None,
                    "model": product["model"] if product else None,
                    "product_code": product["item_code"] if product else None,
                    "customer_name": get_val("customer_name") or None,
                    "customer_phone": get_val("customer_phone") or None,
                    "customer_email": get_val("customer_email") or None,
                    "company": get_val("company") or None,
                })

            if errors:
                return json.dumps({"success": False, "message": f"{len(errors)} hata bulundu, hiçbir satır içe aktarılmadı.", "errors": errors})

            for r in valid_rows:
                db.execute(text("""
                    INSERT INTO warehouse.customers (
                        imei_number, serial_number, internal_id, flow, customer_reported_complaint,
                        intake_date, brand, model, product_code,
                        customer_name, customer_phone, customer_email, company
                    ) VALUES (
                        :imei_number, :serial_number, :internal_id, :flow, :customer_reported_complaint,
                        :intake_date, :brand, :model, :product_code,
                        :customer_name, :customer_phone, :customer_email, :company
                    )
                """), r)

            db.commit()
            return json.dumps({"success": True, "message": f"{len(valid_rows)} müşteri kaydı başarıyla içe aktarıldı.", "imported": len(valid_rows)})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"İçe aktarma hatası: {str(e)}", "errors": []})
        finally:
            db.close()

    # ==========================
    # İŞ EMİRLERİ MODÜLÜ
    # ==========================

    @Slot(result=str)
    def get_work_orders(self):
        """Tüm iş emirlerini, bağlı olduğu servis kaydı bilgileriyle birlikte getirir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT w.id, w.service_record_id, w.description, w.assigned_technician, w.priority,
                       w.start_date, w.end_date, w.parts_used, w.status, w.created_at,
                       w.source_location_id, w.stock_settled_at,
                       w.work_order_type, w.target_part_id, w.planned_quantity,
                       w.started_at, w.completed_at, w.produced_quantity, w.scrap_quantity, w.production_notes, w.department,
                       s.customer_name, s.brand, s.model, s.fault_category, s.fault_type,
                       tp.item_code AS target_part_code, tp.name AS target_part_name
                FROM warehouse.work_orders w
                LEFT JOIN warehouse.service_records s ON s.id = w.service_record_id
                LEFT JOIN warehouse.parts tp ON tp.id = w.target_part_id
                ORDER BY w.id DESC
                LIMIT 200
            """)).mappings().all()
            orders = []
            for row in rows:
                orders.append({
                    "id": str(row["id"]),
                    "work_order_type": row["work_order_type"] or WORK_ORDER_TYPE_SERVICE,
                    "service_record_id": str(row["service_record_id"]) if row["service_record_id"] else "",
                    "customer_name": row["customer_name"] or "",
                    "brand": row["brand"] or "",
                    "model": row["model"] or "",
                    "fault_category": row["fault_category"] or "",
                    "fault_type": row["fault_type"] or "",
                    "target_part_id": str(row["target_part_id"]) if row["target_part_id"] else "",
                    "target_part_code": row["target_part_code"] or "",
                    "target_part_name": row["target_part_name"] or "",
                    "planned_quantity": row["planned_quantity"] if row["planned_quantity"] is not None else "",
                    "started_at": row["started_at"].strftime("%Y-%m-%d %H:%M") if row["started_at"] else "",
                    "completed_at": row["completed_at"].strftime("%Y-%m-%d %H:%M") if row["completed_at"] else "",
                    "produced_quantity": row["produced_quantity"] if row["produced_quantity"] is not None else "",
                    "scrap_quantity": row["scrap_quantity"] if row["scrap_quantity"] is not None else "",
                    "production_notes": row["production_notes"] or "",
                    "department": row["department"] or "",
                    "description": row["description"] or "",
                    "assigned_technician": row["assigned_technician"] or "",
                    "priority": row["priority"] or "Orta",
                    "start_date": row["start_date"].strftime("%Y-%m-%d") if row["start_date"] else "",
                    "end_date": row["end_date"].strftime("%Y-%m-%d") if row["end_date"] else "",
                    "parts_used": row["parts_used"] or "[]",
                    "status": row["status"] or "Beklemede",
                    "source_location_id": str(row["source_location_id"]) if row["source_location_id"] else "",
                    "stock_settled_at": row["stock_settled_at"].strftime("%Y-%m-%d %H:%M") if row["stock_settled_at"] else "",
                    "created_at": row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else ""
                })
            return json.dumps({"success": True, "work_orders": orders})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, result=str)
    def create_work_order(self, service_record_id, description, assigned_technician, priority,
                           start_date, end_date, parts_used, status, source_location_id):
        """Yeni iş emri ekler. parts_used doluysa kaynak depodan (Good/DOA Stock) Repair Stock'a
        otomatik transfer yapar; stok yetersizse iş emri hiç oluşturulmaz."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        db = SessionLocal()
        try:
            lines = []
            if parts_used and parts_used.strip():
                try:
                    lines = [l for l in json.loads(parts_used) if l.get("part_id") and int(l.get("quantity") or 0) > 0]
                except (ValueError, TypeError):
                    lines = []

            src_loc_id = int(source_location_id) if source_location_id and source_location_id.strip() else None

            if lines:
                if not src_loc_id:
                    return json.dumps({"success": False, "message": "Parça kullanılan bir iş emri için kaynak depo seçmelisiniz."})
                repair_stock_id = _get_system_location_id(db, "repair_stock")
                if not repair_stock_id:
                    return json.dumps({"success": False, "message": "Repair Stock deposu bulunamadı."})

                agg = {}
                for line in lines:
                    pid = int(line["part_id"])
                    agg[pid] = agg.get(pid, 0) + int(line["quantity"])

                for part_id, qty in agg.items():
                    stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == src_loc_id).first()
                    if not stock or stock.quantity < qty:
                        return json.dumps({"success": False, "message": f"Parça #{part_id} için seçilen depoda yeterli stok yok."})

                for part_id, qty in agg.items():
                    stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == src_loc_id).first()
                    stock.quantity -= qty
                    target_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == repair_stock_id).first()
                    if target_stock:
                        target_stock.quantity += qty
                    else:
                        db.add(Stock(part_id=part_id, location_id=repair_stock_id, quantity=qty))
                    db.add(StockMovement(
                        type="İş Emri: Tamire Alındı",
                        movement_kind="Transfer",
                        quantity=qty,
                        part_id=part_id,
                        source_location_id=src_loc_id,
                        target_location_id=repair_stock_id,
                        created_by=assigned_technician or "system"
                    ))

            new_id = db.execute(text("""
                INSERT INTO warehouse.work_orders (
                    service_record_id, description, assigned_technician, priority,
                    start_date, end_date, parts_used, status, source_location_id
                ) VALUES (
                    :sr_id, :desc, :tech, :priority, :start, :end, :parts, :status, :src_loc
                ) RETURNING id
            """), {
                "sr_id": int(service_record_id) if service_record_id.strip() else None,
                "desc": description or None,
                "tech": assigned_technician or None,
                "priority": priority or "Orta",
                "start": start_date or None,
                "end": end_date or None,
                "parts": parts_used or None,
                "status": status or "Beklemede",
                "src_loc": src_loc_id
            }).scalar()
            db.commit()
            return json.dumps({"success": True, "message": "İş emri eklendi", "id": str(new_id)})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Kayıt hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, result=str)
    def update_work_order(self, order_id_str, service_record_id, description, assigned_technician, priority,
                           start_date, end_date, parts_used, status):
        """Var olan bir iş emrini günceller. Durum Tamamlandı/Başarısız/İptal'e geçerse Repair
        Stock'taki parçaları otomatik olarak ilgili depoya taşır (yalnızca bir kez, stock_settled_at guard'ı)."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        db = SessionLocal()
        try:
            order_id = int(order_id_str)
            current = db.execute(text("""
                SELECT status, parts_used, source_location_id, stock_settled_at
                FROM warehouse.work_orders WHERE id = :id
            """), {"id": order_id}).mappings().first()
            if not current:
                return json.dumps({"success": False, "message": "İş emri bulunamadı."})

            new_status = status or "Beklemede"
            settle_now = False

            if (new_status != current["status"] and current["stock_settled_at"] is None
                    and new_status in ("Tamamlandı", "Başarısız", "İptal")):
                lines = []
                if current["parts_used"]:
                    try:
                        lines = [l for l in json.loads(current["parts_used"]) if l.get("part_id") and int(l.get("quantity") or 0) > 0]
                    except (ValueError, TypeError):
                        lines = []
                src_loc_id = current["source_location_id"]

                if lines and src_loc_id:
                    repair_stock_id = _get_system_location_id(db, "repair_stock")
                    if new_status == "Tamamlandı":
                        target_id = _get_system_location_id(db, "out_stock")
                        movement_kind = "Outbound"
                        mov_type = "Servis Tamamlandı: Out Stock'a Alındı"
                    elif new_status == "Başarısız":
                        target_id = _get_system_location_id(db, "scrap_stock")
                        movement_kind = "Scrap"
                        mov_type = "Tamir Başarısız: Scrap Stock'a Alındı"
                    else:  # İptal
                        target_id = src_loc_id
                        movement_kind = "Transfer"
                        mov_type = "İş Emri İptal: Depoya İade"

                    if not repair_stock_id or not target_id:
                        return json.dumps({"success": False, "message": "Sistem depoları bulunamadı."})

                    agg = {}
                    for line in lines:
                        pid = int(line["part_id"])
                        agg[pid] = agg.get(pid, 0) + int(line["quantity"])

                    for part_id, qty in agg.items():
                        repair_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == repair_stock_id).first()
                        if not repair_stock or repair_stock.quantity < qty:
                            return json.dumps({"success": False, "message": f"Repair Stock'ta parça #{part_id} için yeterli miktar yok."})

                    for part_id, qty in agg.items():
                        repair_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == repair_stock_id).first()
                        repair_stock.quantity -= qty
                        target_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == target_id).first()
                        if target_stock:
                            target_stock.quantity += qty
                        else:
                            db.add(Stock(part_id=part_id, location_id=target_id, quantity=qty))
                        db.add(StockMovement(
                            type=mov_type,
                            movement_kind=movement_kind,
                            quantity=qty,
                            part_id=part_id,
                            source_location_id=repair_stock_id,
                            target_location_id=target_id,
                            created_by=assigned_technician or "system"
                        ))
                settle_now = True

            settle_clause = ", stock_settled_at = NOW()" if settle_now else ""
            db.execute(text(f"""
                UPDATE warehouse.work_orders
                SET service_record_id = :sr_id, description = :desc, assigned_technician = :tech,
                    priority = :priority, start_date = :start, end_date = :end,
                    parts_used = :parts, status = :status{settle_clause}
                WHERE id = :id
            """), {
                "sr_id": int(service_record_id) if service_record_id.strip() else None,
                "desc": description or None,
                "tech": assigned_technician or None,
                "priority": priority or "Orta",
                "start": start_date or None,
                "end": end_date or None,
                "parts": parts_used or None,
                "status": new_status,
                "id": order_id
            })
            db.commit()
            return json.dumps({"success": True, "message": "İş emri güncellendi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Güncelleme hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_work_order(self, order_id_str):
        """Belirtilen id'ye sahip iş emrini siler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            order_id = int(order_id_str)
            delivered_count = db.execute(text("""
                SELECT COUNT(*) FROM warehouse.work_order_parts
                WHERE work_order_id = :id AND status = 'Teslim Edildi'
            """), {"id": order_id}).scalar()
            if delivered_count:
                return json.dumps({"success": False, "message": "Teslim edilmiş parçaları olan bir iş emri silinemez. Önce parça teslimatlarını geri alın."})
            db.execute(text("DELETE FROM warehouse.work_orders WHERE id = :id"), {"id": order_id})
            db.commit()
            return json.dumps({"success": True, "message": "İş emri silindi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Silme hatası: {str(e)}"})
        finally:
            db.close()

    # ==========================
    # PRODUCTION WORK ORDER (Yarı Mamul Üretim İş Emri)
    # Service Work Order akışıyla aynı work_orders tablosunu paylaşır; work_order_type
    # sütunu üzerinden ayrışır. Service Record'a bağlı değildir; bunun yerine
    # target_part_id ile bir Recipe'ye (warehouse.item_bom) bağlanır. Bu aşamada
    # malzeme talebi (Material Request) veya stok hareketi oluşturulmaz.
    # ==========================

    @Slot(str, str, str, str, str, str, result=str)
    def create_production_work_order(self, target_part_id, description, priority, planned_quantity, assigned_technician, department):
        """PRODUCTION tipinde yeni bir iş emri oluşturur. target_part_id, üretilecek yarı
        mamulün parça id'sidir; bu parçanın item_code'una karşılık gelen bir Recipe
        (warehouse.item_bom kaydı) bulunmalıdır. Service Record gerekmez. Recipe'deki her
        BOM satırı için bir Material Request kaydı (WAITING durumunda) oluşturulur. Durum
        BEKLIYOR ile başlar (bkz. start_production_work_order/complete_production_work_order).
        Bu aşamada stok düşme, depo transferi veya üretim tamamlama yapılmaz."""
        from sqlalchemy import text
        from models.part import Part
        db = SessionLocal()
        try:
            if not target_part_id or not target_part_id.strip():
                return json.dumps({"success": False, "message": "Üretilecek parça (Recipe) seçmelisiniz."})

            part_id = int(target_part_id)
            part = db.query(Part).filter(Part.id == part_id).first()
            if not part:
                return json.dumps({"success": False, "message": "Parça bulunamadı."})

            bom_rows = db.execute(text("""
                SELECT b.child_item_id, b.quantity, cp.id AS child_part_id
                FROM warehouse.item_bom b
                LEFT JOIN warehouse.parts cp ON cp.item_code = b.child_item_id
                WHERE b.parent_item_id = :code
            """), {"code": part.item_code}).mappings().all()
            if not bom_rows:
                return json.dumps({"success": False, "message": "Bu parça için tanımlı bir Recipe (ItemBOM) bulunamadı."})

            if not planned_quantity or not planned_quantity.strip():
                return json.dumps({"success": False, "message": "Planlanan Üretim Adedi zorunludur."})
            try:
                qty = int(planned_quantity)
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Planlanan Üretim Adedi geçerli bir sayı olmalıdır."})
            if qty <= 0:
                return json.dumps({"success": False, "message": "Planlanan Üretim Adedi sıfırdan büyük olmalıdır."})
            multiplier = qty

            new_id = db.execute(text("""
                INSERT INTO warehouse.work_orders (
                    work_order_type, target_part_id, description, priority, planned_quantity,
                    assigned_technician, department, status
                ) VALUES (
                    :wtype, :target, :desc, :priority, :qty, :tech, :dept, :status
                ) RETURNING id
            """), {
                "wtype": WORK_ORDER_TYPE_PRODUCTION,
                "target": part_id,
                "desc": description or None,
                "priority": priority or "Orta",
                "qty": qty,
                "tech": assigned_technician or None,
                "dept": department or None,
                "status": PRODUCTION_WO_STATUS_WAITING
            }).scalar()

            for bom_row in bom_rows:
                if not bom_row["child_part_id"]:
                    print(f"[WebBridge] Material Request atlandı, parça bulunamadı: {bom_row['child_item_id']}")
                    continue
                required_qty = int(bom_row["quantity"]) * multiplier
                db.execute(text("""
                    INSERT INTO warehouse.material_requests (work_order_id, part_id, required_quantity, issued_quantity, status)
                    VALUES (:wid, :pid, :req, 0, :status)
                """), {
                    "wid": new_id,
                    "pid": bom_row["child_part_id"],
                    "req": required_qty,
                    "status": MATERIAL_REQUEST_STATUS_WAITING
                })

            db.commit()
            return json.dumps({"success": True, "message": "Üretim iş emri eklendi", "id": str(new_id)})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Kayıt hatası: {str(e)}"})
        finally:
            db.close()

    # ==========================
    # PRODUCTION WORK ORDER YAŞAM DÖNGÜSÜ (BEKLIYOR -> URETIMDE -> TAMAMLANDI)
    # Sadece PRODUCTION tipi work order'lar için çalışır; Service Work Order'ın kendi
    # status akışını (create_work_order/update_work_order) hiç etkilemez. Bir iş emri
    # BEKLIYOR durumunda oluşturulur; kullanıcı "Başlat" ile URETIMDE'ye alır
    # (start_production_work_order). Malzeme teslimi (issue_material_request) yalnızca
    # URETIMDE durumunda yapılabilir; bu nedenle BEKLIYOR aşamasında malzeme kilitlidir.
    # ==========================

    @Slot(str, str, result=str)
    def start_production_work_order(self, work_order_id_str, username):
        """PRODUCTION tipi bir iş emrini BEKLIYOR durumundan URETIMDE durumuna geçirir.
        (Geçmişten kalan BEKLIYOR durumundaki iş emirlerinin başlatılabilmesi için eklendi)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            work_order_id = int(work_order_id_str)
            
            row = db.execute(
                text("SELECT status, work_order_type FROM warehouse.work_orders WHERE id = :id FOR UPDATE"),
                {"id": work_order_id}
            ).mappings().first()
            
            if not row:
                return json.dumps({"success": False, "message": "İş emri bulunamadı."})
                
            if row["work_order_type"] != WORK_ORDER_TYPE_PRODUCTION:
                return json.dumps({"success": False, "message": "Sadece üretim iş emirleri başlatılabilir."})
                
            if row["status"] != PRODUCTION_WO_STATUS_WAITING:
                return json.dumps({"success": False, "message": f"Bu iş emri {PRODUCTION_WO_STATUS_WAITING} durumunda değil (şu an: {row['status']})."})
                
            db.execute(text("""
                UPDATE warehouse.work_orders 
                SET status = :status, started_at = CURRENT_TIMESTAMP
                WHERE id = :id
            """), {"status": PRODUCTION_WO_STATUS_IN_PRODUCTION, "id": work_order_id})
            
            db.commit()
            return json.dumps({"success": True, "message": "İş emri üretime alındı."})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Başlatma hatası: {str(e)}"})
        finally:
            db.close()


    @Slot(str, str, str, str, str, result=str)
    def complete_production_work_order(self, work_order_id_str, produced_quantity_str, scrap_quantity_str, production_notes, username):
        """PRODUCTION tipi bir iş emrini URETIMDE durumundan TAMAMLANDI durumuna geçirir.
        Üretilen Adet + Fire Adedi, Planlanan Üretim'e eşit olmak zorundadır; değilse
        işlem reddedilir ve hiçbir kayıt değişmez. Üretilen Adet kadar hedef parça Good
        Stock'a eklenir; Fire Adedi kadar hedef parça Scrap Stock'a eklenir (bu, hammadde
        fire'ından farklıdır -- burada bahsedilen, sonuçta kullanılamaz çıkan bitmiş
        ürün miktarıdır). Üretilen Adet > 0 ise, "Hızlı Üretim" (create_production_run)
        ile aynı production_runs/produced_units/production_materials kayıtları açılır ki
        Üretim Raporu'nda görünsün ve aynı iade/değişim akışıyla (delete_production_run)
        yönetilebilsin. Tüketilen hammadde miktarı, her Material Request'in
        (issued_quantity - fire_quantity) değerinden hesaplanır."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        db = SessionLocal()
        try:
            work_order_id = int(work_order_id_str)
            try:
                produced_quantity = int(produced_quantity_str)
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Üretilen Adet geçerli bir sayı olmalıdır."})
            try:
                scrap_quantity = int(scrap_quantity_str)
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Fire Adedi geçerli bir sayı olmalıdır."})

            if produced_quantity < 0 or scrap_quantity < 0:
                return json.dumps({"success": False, "message": "Üretilen Adet ve Fire Adedi negatif olamaz."})

            row = db.execute(
                text("""SELECT id, work_order_type, status, planned_quantity, target_part_id
                        FROM warehouse.work_orders WHERE id = :id FOR UPDATE"""),
                {"id": work_order_id}
            ).mappings().first()
            if not row:
                return json.dumps({"success": False, "message": "İş emri bulunamadı."})
            if row["work_order_type"] != WORK_ORDER_TYPE_PRODUCTION:
                return json.dumps({"success": False, "message": "Bu işlem sadece Production Work Order'lar için geçerlidir."})
            if row["status"] != PRODUCTION_WO_STATUS_IN_PRODUCTION:
                return json.dumps({"success": False, "message": f"Sadece {PRODUCTION_WO_STATUS_IN_PRODUCTION} durumundaki iş emirleri tamamlanabilir."})
            if row["planned_quantity"] is None:
                return json.dumps({"success": False, "message": "Bu iş emrinde Planlanan Üretim Adedi tanımlı değil, tamamlanamaz."})

            planned_quantity = row["planned_quantity"]
            target_part_id = row["target_part_id"]
            if produced_quantity + scrap_quantity != planned_quantity:
                return json.dumps({
                    "success": False,
                    "message": f"Üretilen Adet ({produced_quantity}) + Fire Adedi ({scrap_quantity}) = {produced_quantity + scrap_quantity}, "
                                f"Planlanan Üretim'e ({planned_quantity}) eşit olmalıdır."
                })

            good_stock_id = _get_system_location_id(db, "good_stock")
            scrap_stock_id = _get_system_location_id(db, "scrap_stock")
            wip_stock_id = _get_system_location_id(db, "repair_stock")  # Redirect WIP to repair_stock
            if not good_stock_id:
                return json.dumps({"success": False, "message": "Good Stock deposu bulunamadı."})
            if scrap_quantity > 0 and not scrap_stock_id:
                return json.dumps({"success": False, "message": "Scrap Stock deposu bulunamadı."})
            if not wip_stock_id:
                return json.dumps({"success": False, "message": "Repair Stock deposu bulunamadı."})

            # Tüketilen hammaddeler: her malzeme talebinin fiilen üretime giden kısmı
            # (issued - fire). Fire olarak DOA'ya iade edilenler zaten oradan çıkarılmıştı.
            material_rows = db.execute(text("""
                SELECT part_id, issued_quantity, fire_quantity, required_quantity
                FROM warehouse.material_requests
                WHERE work_order_id = :wid
            """), {"wid": work_order_id}).mappings().all()

            # KONTROL: Teknisyene verilen net malzeme (issued - fire), üretimi tamamlamak için yeterli mi?
            for mr in material_rows:
                net_issued = mr["issued_quantity"] - mr["fire_quantity"]
                required = mr["required_quantity"]
                if net_issued < required:
                    part_name = db.execute(text("SELECT name FROM warehouse.parts WHERE id = :pid"), {"pid": mr["part_id"]}).scalar()
                    return json.dumps({
                        "success": False, 
                        "message": f"Teknisyene verilen malzeme yetersiz! {part_name} için en az {required} adet teslim edilmeli (Şu anki net teslim: {net_issued})."
                    })

            net_materials = [(m["part_id"], m["issued_quantity"] - m["fire_quantity"]) for m in material_rows if (m["issued_quantity"] - m["fire_quantity"]) > 0]

            if produced_quantity > 0:
                existing = db.execute(text("""
                    SELECT id FROM warehouse.stock WHERE part_id = :pid AND location_id = :lid
                """), {"pid": target_part_id, "lid": good_stock_id}).first()
                if existing:
                    db.execute(text("UPDATE warehouse.stock SET quantity = quantity + :qty WHERE id = :id"),
                               {"qty": produced_quantity, "id": existing[0]})
                else:
                    db.execute(text("""
                        INSERT INTO warehouse.stock (part_id, location_id, quantity) VALUES (:pid, :lid, :qty)
                    """), {"pid": target_part_id, "lid": good_stock_id, "qty": produced_quantity})

                run_id = db.execute(text("""
                    INSERT INTO warehouse.production_runs (target_part_id, quantity_produced, source_location_id, location_id, produced_by, notes)
                    VALUES (:tgt, :qty, :slid, :tlid, :by, :notes) RETURNING id
                """), {
                    "tgt": target_part_id, "qty": produced_quantity, "slid": good_stock_id, "tlid": good_stock_id,
                    "by": username or None, "notes": f"Üretim İş Emri #{work_order_id} tamamlandı (Fire: {scrap_quantity} adet)" + (f" - {production_notes}" if production_notes else "")
                }).scalar()

                next_id = db.execute(text("SELECT nextval(pg_get_serial_sequence('warehouse.produced_units', 'id'))")).scalar()
                serial_num = f"{next_id:015d}"
                db.execute(text("""
                    INSERT INTO warehouse.produced_units (id, production_run_id, serial_number)
                    VALUES (:id, :run_id, :serial)
                """), {"id": next_id, "run_id": run_id, "serial": serial_num})

                for part_id, qty_consumed in net_materials:
                    db.execute(text("""
                        INSERT INTO warehouse.production_materials (production_run_id, part_id, quantity_consumed)
                        VALUES (:run_id, :pid, :qty)
                    """), {"run_id": run_id, "pid": part_id, "qty": qty_consumed})
                    
                    # Deduct from Repair Stock
                    wip_stock_entry = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == wip_stock_id).first()
                    if wip_stock_entry:
                        wip_stock_entry.quantity -= qty_consumed
                        
                    db.add(StockMovement(
                        type="Üretim İçin Malzeme Tüketimi",
                        movement_kind="Outbound",
                        quantity=qty_consumed,
                        part_id=part_id,
                        source_location_id=wip_stock_id,
                        created_by=username or None,
                        description=f"İş Emri {work_order_id:015d} tamamlandı, malzemeler tüketildi"
                    ))

                db.add(StockMovement(
                    type="Üretim",
                    movement_kind="Inbound",
                    quantity=produced_quantity,
                    part_id=target_part_id,
                    target_location_id=good_stock_id,
                    created_by=username or None,
                    description=f"Üretim İş Emri #{work_order_id} ({serial_num}) tamamlandı"
                ))

            if scrap_quantity > 0:
                existing_scrap = db.execute(text("""
                    SELECT id FROM warehouse.stock WHERE part_id = :pid AND location_id = :lid
                """), {"pid": target_part_id, "lid": scrap_stock_id}).first()
                if existing_scrap:
                    db.execute(text("UPDATE warehouse.stock SET quantity = quantity + :qty WHERE id = :id"),
                               {"qty": scrap_quantity, "id": existing_scrap[0]})
                else:
                    db.execute(text("""
                        INSERT INTO warehouse.stock (part_id, location_id, quantity) VALUES (:pid, :lid, :qty)
                    """), {"pid": target_part_id, "lid": scrap_stock_id, "qty": scrap_quantity})

                db.add(StockMovement(
                    type="Üretim Fire (Hurda)",
                    movement_kind="Inbound",
                    quantity=scrap_quantity,
                    part_id=target_part_id,
                    target_location_id=scrap_stock_id,
                    created_by=username or None,
                    description=f"Üretim İş Emri #{work_order_id} tamamlanırken fire çıkan {scrap_quantity} adet hurdaya ayrıldı"
                ))

            db.execute(text("""
                UPDATE warehouse.work_orders
                SET status = :status, completed_at = CURRENT_TIMESTAMP,
                    produced_quantity = :produced, scrap_quantity = :scrap, production_notes = :notes
                WHERE id = :id
            """), {
                "status": PRODUCTION_WO_STATUS_COMPLETED,
                "produced": produced_quantity,
                "scrap": scrap_quantity,
                "notes": production_notes or None,
                "id": work_order_id
            })
            db.commit()
            return json.dumps({"success": True, "message": "Üretim tamamlandı", "status": PRODUCTION_WO_STATUS_COMPLETED})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"İşlem hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_material_requests(self, work_order_id_str):
        """Bir Production Work Order'a bağlı Material Request kayıtlarını, parça
        bilgileriyle birlikte getirir. Salt okunurdur; stok düşme/depo transferi bu
        aşamada yapılmaz. remaining_quantity, (required_quantity + fire_quantity -
        issued_quantity) olarak canlı hesaplanır -- fire_quantity, bildirilmiş fire
        kadar ek teslim hakkı açar (bkz. report_material_fire). unit_quantity (reçetedeki
        birim başına miktar), required_quantity // planned_quantity olarak hesaplanır --
        bu, iş emri oluşturulduğu andaki değeri yansıtır; reçete sonradan değişmiş olsa
        bile bu iş emri için kullanılan orijinal değeri gösterir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            work_order_id = int(work_order_id_str)
            rows = db.execute(text("""
                SELECT mr.id, mr.work_order_id, mr.part_id,
                       mr.required_quantity, mr.issued_quantity, mr.fire_quantity,
                       (mr.required_quantity + mr.fire_quantity - mr.issued_quantity) AS remaining_quantity,
                       mr.status, mr.created_at,
                       p.item_code, p.name AS part_name_raw, p.brand, p.model, p.color, p.part_category,
                       wo.planned_quantity
                FROM warehouse.material_requests mr
                LEFT JOIN warehouse.parts p ON p.id = mr.part_id
                LEFT JOIN warehouse.work_orders wo ON wo.id = mr.work_order_id
                WHERE mr.work_order_id = :wid
                ORDER BY mr.id ASC
            """), {"wid": work_order_id}).mappings().all()

            requests = []
            for row in rows:
                part_name = " ".join(filter(None, [row["brand"], row["model"], row["color"], row["part_category"]])) or (row["part_name_raw"] or "")
                planned_qty = row["planned_quantity"]
                unit_quantity = (row["required_quantity"] // planned_qty) if planned_qty else None
                requests.append({
                    "id": str(row["id"]),
                    "work_order_id": str(row["work_order_id"]),
                    "part_id": str(row["part_id"]),
                    "part_name": part_name,
                    "item_code": row["item_code"] or "",
                    "unit_quantity": unit_quantity,
                    "required_quantity": row["required_quantity"],
                    "issued_quantity": row["issued_quantity"],
                    "fire_quantity": row["fire_quantity"],
                    "remaining_quantity": row["remaining_quantity"],
                    "status": row["status"] or MATERIAL_REQUEST_STATUS_WAITING,
                    "created_at": row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else ""
                })
            return json.dumps({"success": True, "material_requests": requests})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def issue_material_request(self, mr_id_str, quantity_str, username):
        """Bir Material Request satırının bir kısmını veya tamamını Good Stock'tan teslim
        eder (Malzeme Teslim / Material Issue). Kısmi teslimi destekler: WAITING ->
        PARTIAL -> ISSUED, issued_quantity/(required_quantity + fire_quantity) oranına
        göre otomatik hesaplanır. fire_quantity, bildirilmiş fire kadar ek teslim hakkı
        açar (bkz. report_material_fire) -- fire bildirilmeden limit büyümez. Stok
        yetersizse işlem iptal edilir, hiçbir kayıt değişmez. Başarılı teslimde bir
        StockMovement kaydı açılır. Sadece PRODUCTION tipi Work Order'lara aittir;
        Service Work Order akışını hiçbir şekilde etkilemez. Üretim tamamlama, yarı
        mamul oluşturma bu aşamada yapılmaz."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        db = SessionLocal()
        try:
            mr_id = int(mr_id_str)
            try:
                quantity = int(quantity_str)
            except (ValueError, TypeError):
                quantity = 0
            if quantity <= 0:
                return json.dumps({"success": False, "message": "Teslim miktarı 0'dan büyük olmalıdır."})

            row = db.execute(
                text("""SELECT id, work_order_id, part_id, required_quantity, issued_quantity, fire_quantity, status
                        FROM warehouse.material_requests WHERE id = :id FOR UPDATE"""),
                {"id": mr_id}
            ).mappings().first()
            if not row:
                return json.dumps({"success": False, "message": "Malzeme talebi bulunamadı."})

            wo_row = db.execute(
                text("SELECT work_order_type, status FROM warehouse.work_orders WHERE id = :id"),
                {"id": row["work_order_id"]}
            ).mappings().first()
            wo_type = wo_row["work_order_type"] if wo_row else None
            if wo_type != WORK_ORDER_TYPE_PRODUCTION:
                return json.dumps({"success": False, "message": "Malzeme teslimi sadece Production Work Order'lar için yapılabilir."})
            if wo_row["status"] != PRODUCTION_WO_STATUS_IN_PRODUCTION:
                return json.dumps({"success": False, "message": f"Malzeme teslimi sadece {PRODUCTION_WO_STATUS_IN_PRODUCTION} durumundaki iş emirleri için yapılabilir (şu an: {wo_row['status']})."})

            effective_limit = row["required_quantity"] + row["fire_quantity"]
            remaining = effective_limit - row["issued_quantity"]
            if quantity > remaining:
                return json.dumps({"success": False, "message": f"Kalan miktardan ({remaining}) fazla teslim edilemez."})

            good_stock_id = _get_system_location_id(db, "good_stock")
            wip_stock_id = _get_system_location_id(db, "repair_stock")  # Redirect WIP to repair_stock
            if not good_stock_id:
                return json.dumps({"success": False, "message": "Good Stock deposu bulunamadı."})
            if not wip_stock_id:
                return json.dumps({"success": False, "message": "Repair Stock deposu bulunamadı."})

            stock = db.query(Stock).filter(Stock.part_id == row["part_id"], Stock.location_id == good_stock_id).first()
            available = stock.quantity if stock else 0
            if available < quantity:
                return json.dumps({"success": False, "message": f"Good Stock'ta yeterli stok yok. Mevcut: {available}, İstenen: {quantity}."})

            stock.quantity -= quantity
            
            wip_stock_entry = db.query(Stock).filter(Stock.part_id == row["part_id"], Stock.location_id == wip_stock_id).first()
            if wip_stock_entry:
                wip_stock_entry.quantity += quantity
            else:
                db.add(Stock(part_id=row["part_id"], location_id=wip_stock_id, quantity=quantity))

            db.add(StockMovement(
                type="Üretim İçin Malzeme Teslimi",
                movement_kind="Transfer",
                quantity=quantity,
                part_id=row["part_id"],
                source_location_id=good_stock_id,
                target_location_id=wip_stock_id,
                created_by=username or None,
                technician=username or None,
                description=f"Hedef: İş Emri {row['work_order_id']:015d} - Material Request #{mr_id} teslimi"
            ))

            new_issued = row["issued_quantity"] + quantity
            new_status = _compute_material_request_status(new_issued, effective_limit)

            db.execute(text("""
                UPDATE warehouse.material_requests
                SET issued_quantity = :issued, status = :status
                WHERE id = :id
            """), {"issued": new_issued, "status": new_status, "id": mr_id})

            db.commit()
            return json.dumps({
                "success": True,
                "message": "Malzeme teslim edildi",
                "issued_quantity": new_issued,
                "remaining_quantity": effective_limit - new_issued,
                "status": new_status
            })
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Teslim hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def report_material_fire(self, mr_id_str, fire_qty_str, username):
        """Bir Material Request'e ait, teknisyenden fire (kullanılamayan/bozuk) olarak
        geri gelen malzemeyi DOA Stock'a iade eder ve fire_quantity'yi artırır."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        db = SessionLocal()
        try:
            mr_id = int(mr_id_str)
            try:
                fire_qty = int(fire_qty_str)
            except (ValueError, TypeError):
                fire_qty = 0
            if fire_qty <= 0:
                return json.dumps({"success": False, "message": "Fire miktarı 0'dan büyük olmalıdır."})

            row = db.execute(
                text("""SELECT id, work_order_id, part_id, issued_quantity, fire_quantity
                        FROM warehouse.material_requests WHERE id = :id FOR UPDATE"""),
                {"id": mr_id}
            ).mappings().first()
            if not row:
                return json.dumps({"success": False, "message": "Malzeme talebi bulunamadı."})

            unaccounted = row["issued_quantity"] - row["fire_quantity"]
            if fire_qty > unaccounted:
                return json.dumps({"success": False, "message": f"En fazla {unaccounted} adet fire bildirebilirsiniz (teslim edilmiş, henüz fire işlenmemiş miktarı aşamaz)."})

            doa_stock_id = _get_system_location_id(db, "doa_stock")
            repair_stock_id = _get_system_location_id(db, "repair_stock")
            if not doa_stock_id:
                return json.dumps({"success": False, "message": "DOA Stock deposu bulunamadı."})
            if not repair_stock_id:
                return json.dumps({"success": False, "message": "Repair Stock deposu bulunamadı."})

            repair_stock_entry = db.query(Stock).filter(Stock.part_id == row["part_id"], Stock.location_id == repair_stock_id).first()
            if repair_stock_entry:
                repair_stock_entry.quantity -= fire_qty

            doa_stock = db.query(Stock).filter(Stock.part_id == row["part_id"], Stock.location_id == doa_stock_id).first()
            if doa_stock:
                doa_stock.quantity += fire_qty
            else:
                db.add(Stock(part_id=row["part_id"], location_id=doa_stock_id, quantity=fire_qty))

            db.add(StockMovement(
                type="Fire İadesi",
                movement_kind="Transfer",
                quantity=fire_qty,
                part_id=row["part_id"],
                source_location_id=repair_stock_id,
                target_location_id=doa_stock_id,
                created_by=username or None,
                technician=username or None,
                description=f"Kaynak: İş Emri {row['work_order_id']:015d} - Material Request #{mr_id} fire iadesi"
            ))

            new_fire_total = row["fire_quantity"] + fire_qty
            db.execute(text("""
                UPDATE warehouse.material_requests
                SET fire_quantity = :fire
                WHERE id = :id
            """), {"fire": new_fire_total, "id": mr_id})

            db.commit()
            return json.dumps({
                "success": True,
                "message": "Fire bildirildi ve DOA stoğa iade edildi",
                "fire_quantity": new_fire_total
            })
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Fire bildirme hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def return_bom_part_to_doa(self, part_id_str, return_qty_str, source_location_id_str, username):
        """Hızlı Tekrar Üretim (BOM) reçetesindeki bir parçayı seçili lokasyondan DOA Stock'a iade eder."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        db = SessionLocal()
        try:
            part_id = int(part_id_str)
            try:
                return_qty = int(return_qty_str)
            except (ValueError, TypeError):
                return_qty = 0
            if return_qty <= 0:
                return json.dumps({"success": False, "message": "İade miktarı 0'dan büyük olmalıdır."})
            
            source_location_id = int(source_location_id_str) if source_location_id_str else None
            if not source_location_id:
                return json.dumps({"success": False, "message": "Geçerli bir kaynak lokasyon seçmelisiniz."})

            doa_stock_id = _get_system_location_id(db, "doa_stock")
            if not doa_stock_id:
                return json.dumps({"success": False, "message": "DOA Stock deposu bulunamadı."})

            source_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == source_location_id).first()
            if source_stock and source_stock.quantity >= return_qty:
                source_stock.quantity -= return_qty
            else:
                return json.dumps({"success": False, "message": "Kaynak depoda yeterli stok yok."})

            doa_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == doa_stock_id).first()
            if doa_stock:
                doa_stock.quantity += return_qty
            else:
                db.add(Stock(part_id=part_id, location_id=doa_stock_id, quantity=return_qty))

            db.add(StockMovement(
                type="DOA İade",
                movement_kind="Transfer",
                quantity=return_qty,
                part_id=part_id,
                source_location_id=source_location_id,
                target_location_id=doa_stock_id,
                created_by=username or None,
                technician=username or None,
                description=f"Hızlı Tekrar Üretim reçetesi parçası ({return_qty} adet) DOA stoğa geri alındı"
            ))

            db.commit()
            return json.dumps({"success": True, "message": "Parça DOA depoya iade edildi"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"DOA İade hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def issue_extra_bom_materials(self, part_id_str, extra_qty_str, source_location_id_str, username):
        """Reçete/İş Emri için seçilen depodan ekstra malzeme/parça çıkışı yapar."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        from models.location import Location
        db = SessionLocal()
        try:
            part_id = int(part_id_str)
            try:
                extra_qty = int(extra_qty_str)
            except (ValueError, TypeError):
                extra_qty = 0
            if extra_qty <= 0:
                return json.dumps({"success": False, "message": "Ekstra miktar 0'dan büyük olmalıdır."})

            source_loc_id = int(source_location_id_str) if (source_location_id_str and str(source_location_id_str).isdigit()) else 0
            source_loc = db.query(Location).filter(Location.id == source_loc_id).first()
            if not source_loc:
                source_loc = db.query(Location).filter(Location.kind == "good_stock").first()
                if source_loc:
                    source_loc_id = source_loc.id

            if not source_loc_id:
                return json.dumps({"success": False, "message": "Kaynak depo bulunamadı."})

            stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == source_loc_id).first()
            if not stock or stock.quantity < extra_qty:
                available = stock.quantity if stock else 0
                loc_name = source_loc.name if source_loc else "Seçilen depo"
                return json.dumps({"success": False, "message": f"{loc_name}'da yeterli stok yok. Mevcut: {available}, İstenen: {extra_qty}."})

            stock.quantity -= extra_qty
            db.add(StockMovement(
                type="Ekstra Malzeme Çıkışı",
                movement_kind="Outbound",
                quantity=extra_qty,
                part_id=part_id,
                source_location_id=source_loc_id,
                created_by=username or None,
                technician=username or None,
                description=f"İş Emri / Reçete için ekstra parça çıkışı ({extra_qty} adet) - Depo: {source_loc.name}"
            ))

            db.commit()
            clear_api_cache()
            return json.dumps({"success": True, "message": "Ekstra parça çıkışı yapıldı"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Ekstra parça çıkışı hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def receive_extra_bom_materials(self, part_id_str, extra_qty_str, target_location_id_str, technician):
        """Hızlı Tekrar Üretim reçetesi için seçilen depoya ekstra malzeme/parça girişi yapar."""
        from models.stock import Stock
        from models.stock_movement import StockMovement
        db = SessionLocal()
        try:
            part_id = int(part_id_str)
            target_location_id = int(target_location_id_str) if target_location_id_str else None
            try:
                extra_qty = int(extra_qty_str)
            except (ValueError, TypeError):
                extra_qty = 0
            if extra_qty <= 0:
                return json.dumps({"success": False, "message": "Ekstra miktar 0'dan büyük olmalıdır."})
            
            if not target_location_id:
                return json.dumps({"success": False, "message": "Geçerli bir lokasyon seçmelisiniz."})

            target_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == target_location_id).first()
            if target_stock:
                target_stock.quantity += extra_qty
            else:
                target_stock = Stock(part_id=part_id, location_id=target_location_id, quantity=extra_qty)
                db.add(target_stock)

            db.add(StockMovement(
                type="Ekstra Malzeme Girişi",
                movement_kind="Inbound",
                quantity=extra_qty,
                part_id=part_id,
                target_location_id=target_location_id,
                created_by=technician or None,
                technician=technician or None,
                description=f"Hızlı Tekrar Üretim için ekstra parça girişi ({extra_qty} adet)"
            ))

            db.commit()
            return json.dumps({"success": True, "message": "Ekstra parça girişi yapıldı"})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Ekstra parça girişi hatası: {str(e)}"})
        finally:
            db.close()

    # ==========================
    # PARÇA TEDARİK DURUMU (İş Emri Parça Satırları / Stok Teslim-Bekleme-Geri Alma)
    # ==========================

    @Slot(str, result=str)
    def get_work_order_parts_by_imei(self, imei_number):
        """Bir IMEI numarasına ait parça satırlarını (ve teknisyen bilgilerini) getirir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            if not imei_number or not imei_number.strip():
                return json.dumps({"success": False, "message": "IMEI numarası boş olamaz"})
            
            imei_clean = imei_number.strip()
            
            wo_id = None
            pu_run_id = None
            pu_row = db.execute(text("""
                SELECT pr.id AS run_id, pr.work_order_id, pr.target_part_id, pr.produced_by AS assigned_technician
                FROM warehouse.produced_units pu
                JOIN warehouse.production_runs pr ON pr.id = pu.production_run_id
                WHERE pu.serial_number = :serial
            """), {"serial": imei_clean}).mappings().first()
            
            if pu_row:
                pu_run_id = pu_row['run_id']
                wo_id = pu_row['work_order_id']

            if not wo_id and not pu_run_id and imei_clean.isdigit():
                possible_wo_id = None
                if len(imei_clean) == 15 and imei_clean.startswith('1'):
                    possible_wo_id = int(imei_clean[1:])
                else:
                    possible_wo_id = int(imei_clean)
                    
                wo_check = db.execute(text("SELECT id FROM warehouse.work_orders WHERE id = :id AND work_order_type = 'PRODUCTION'"), {"id": possible_wo_id}).scalar()
                if wo_check:
                    wo_id = wo_check

            rows = []
            
            if pu_run_id:
                rows = db.execute(text("""
                    SELECT 'pm_' || pm.id::text AS id, pm.production_run_id AS work_order_id, pm.part_id, pm.quantity_consumed AS quantity, 
                           'Üretimde Kullanıldı' AS status,
                           NULL AS delivered_location_id, NULL AS delivery_movement_id, NULL AS delivered_by, NULL AS delivered_at,
                           '' AS waiting_notes, NULL AS marked_waiting_by, NULL AS marked_waiting_at,
                           NULL AS reversal_movement_id, NULL AS reverted_by, NULL AS reverted_at,
                           NULL AS requested_by, NULL AS created_at,
                           p.brand, p.model, p.color, COALESCE(p.part_category, p.item_category) AS part_category, p.item_code, p.name AS part_name_raw,
                           '' AS delivered_location_name,
                           :tech AS assigned_technician,
                           tp.brand AS sr_brand, tp.model AS sr_model, '' AS sr_memory, 'Üretim (Tamamlanmış)' AS customer_name
                    FROM warehouse.production_materials pm
                    LEFT JOIN warehouse.parts p ON p.id = pm.part_id
                    LEFT JOIN warehouse.parts tp ON tp.id = :target_part_id
                    WHERE pm.production_run_id = :run_id
                    ORDER BY pm.id DESC
                """), {
                    "run_id": pu_run_id, 
                    "tech": pu_row["assigned_technician"], 
                    "target_part_id": pu_row["target_part_id"]
                }).mappings().all()

            elif wo_id:
                rows = db.execute(text("""
                    SELECT 'mr_' || mr.id::text AS id, mr.work_order_id, mr.part_id, 
                           mr.required_quantity AS required_qty,
                           mr.issued_quantity AS issued_qty,
                           (mr.required_quantity - mr.issued_quantity) AS remaining_qty,
                           CASE 
                               WHEN mr.status = 'CANCELLED' THEN mr.required_quantity
                               WHEN mr.issued_quantity >= mr.required_quantity THEN mr.issued_quantity
                               ELSE (mr.required_quantity - mr.issued_quantity)
                           END AS quantity, 
                           CASE 
                               WHEN mr.status = 'CANCELLED' THEN 'İptal Edildi'
                               WHEN mr.issued_quantity >= mr.required_quantity THEN 'Teslim Edildi' 
                               WHEN mr.issued_quantity > 0 THEN 'Kısmi Teslim'
                               ELSE 'Tedarik Bekleniyor' 
                           END AS status,
                           NULL AS delivered_location_id, '' AS delivered_location_name,
                           NULL AS delivery_movement_id, NULL AS delivered_by, NULL AS delivered_at,
                           '' AS waiting_notes, NULL AS marked_waiting_by, NULL AS marked_waiting_at,
                           NULL AS reversal_movement_id, NULL AS reverted_by, NULL AS reverted_at,
                           NULL AS requested_by, NULL AS created_at,
                           p.brand, p.model, p.color, COALESCE(p.part_category, p.item_category) AS part_category, p.item_code, p.name AS part_name_raw,
                           wo.assigned_technician,
                           tp.brand AS sr_brand, tp.model AS sr_model, '' AS sr_memory, 'Üretim İş Emri' AS customer_name
                    FROM warehouse.material_requests mr
                    JOIN warehouse.work_orders wo ON wo.id = mr.work_order_id
                    LEFT JOIN warehouse.parts p ON p.id = mr.part_id
                    LEFT JOIN warehouse.parts tp ON tp.id = wo.target_part_id
                    WHERE wo.id = :id
                    ORDER BY mr.id DESC
                """), {"id": wo_id}).mappings().all()

                if not rows:
                    pm_rows = db.execute(text("""
                        SELECT 'pm_' || pm.id::text AS id, pr.work_order_id, pm.part_id, pm.quantity_consumed AS quantity, 
                               'Üretimde Kullanıldı' AS status,
                               NULL AS delivered_location_id, '' AS delivered_location_name,
                               NULL AS delivery_movement_id, NULL AS delivered_by, NULL AS delivered_at,
                               '' AS waiting_notes, NULL AS marked_waiting_by, NULL AS marked_waiting_at,
                               NULL AS reversal_movement_id, NULL AS reverted_by, NULL AS reverted_at,
                               NULL AS requested_by, NULL AS created_at,
                               p.brand, p.model, p.color, COALESCE(p.part_category, p.item_category) AS part_category, p.item_code, p.name AS part_name_raw,
                               pr.produced_by AS assigned_technician,
                               tp.brand AS sr_brand, tp.model AS sr_model, '' AS sr_memory, 'Üretim (Tamamlanmış)' AS customer_name
                        FROM warehouse.production_materials pm
                        JOIN warehouse.production_runs pr ON pr.id = pm.production_run_id
                        LEFT JOIN warehouse.parts p ON p.id = pm.part_id
                        LEFT JOIN warehouse.parts tp ON tp.id = pr.target_part_id
                        WHERE pr.work_order_id = :id
                        ORDER BY pm.id DESC
                    """), {"id": wo_id}).mappings().all()
                    
                    if pm_rows:
                        rows = pm_rows


            if not rows:
                rows = db.execute(text("""
                    SELECT wop.id, wop.work_order_id, wop.part_id, wop.quantity, wop.status,
                           wop.quantity AS required_qty,
                           CASE WHEN wop.status = 'Teslim Edildi' THEN wop.quantity ELSE 0 END AS issued_qty,
                           CASE WHEN wop.status = 'Teslim Edildi' THEN 0 ELSE wop.quantity END AS remaining_qty,
                           wop.delivered_location_id, wop.delivery_movement_id, wop.delivered_by, wop.delivered_at,
                           wop.waiting_notes, wop.marked_waiting_by, wop.marked_waiting_at,
                           wop.reversal_movement_id, wop.reverted_by, wop.reverted_at,
                           wop.requested_by, wop.created_at,
                           p.brand, p.model, p.color, COALESCE(p.part_category, p.item_category) AS part_category, p.item_code, p.name AS part_name_raw,
                           dl.name AS delivered_location_name,
                           wo.assigned_technician,
                           sr.brand AS sr_brand, sr.model AS sr_model, sr.memory AS sr_memory, sr.customer_name
                    FROM warehouse.work_order_parts wop
                    JOIN warehouse.work_orders wo ON wo.id = wop.work_order_id
                    JOIN warehouse.service_records sr ON sr.id = wo.service_record_id
                    LEFT JOIN warehouse.parts p ON p.id = wop.part_id
                    LEFT JOIN warehouse.locations dl ON dl.id = wop.delivered_location_id
                    WHERE sr.imei_number = :imei
                    ORDER BY wop.id DESC
                """), {"imei": imei_clean}).mappings().all()

            parts = []
            for row in rows:
                part_name = " ".join(filter(None, [row["brand"], row["model"], row["color"], row["part_category"]])) or (row["part_name_raw"] or "")
                parts.append({
                    "id": str(row["id"]),
                    "work_order_id": str(row["work_order_id"]),
                    "part_id": str(row["part_id"]),
                    "part_name": part_name,
                    "part_category": row["part_category"] or "",
                    "item_code": row["item_code"] or "",
                    "assigned_technician": row["assigned_technician"] or "",
                    "quantity": row["quantity"],
                    "required_qty": int(row["required_qty"]) if row.get("required_qty") is not None else row["quantity"],
                    "issued_qty": int(row["issued_qty"]) if row.get("issued_qty") is not None else (row["quantity"] if row["status"] == "Teslim Edildi" else 0),
                    "remaining_qty": int(row["remaining_qty"]) if row.get("remaining_qty") is not None else (0 if row["status"] == "Teslim Edildi" else row["quantity"]),
                    "status": row["status"] or "Stokta Var",
                    "delivered_location_id": str(row["delivered_location_id"]) if row["delivered_location_id"] else "",
                    "delivered_location_name": row["delivered_location_name"] or "",
                    "delivery_movement_id": str(row["delivery_movement_id"]) if row["delivery_movement_id"] else "",
                    "delivered_by": row["delivered_by"] or "",
                    "delivered_at": row["delivered_at"].strftime("%Y-%m-%d %H:%M") if row["delivered_at"] else "",
                    "waiting_notes": row["waiting_notes"] or "",
                    "marked_waiting_by": row["marked_waiting_by"] or "",
                    "marked_waiting_at": row["marked_waiting_at"].strftime("%Y-%m-%d %H:%M") if row["marked_waiting_at"] else "",
                    "reverted_by": row["reverted_by"] or "",
                    "reverted_at": row["reverted_at"].strftime("%Y-%m-%d %H:%M") if row["reverted_at"] else "",
                    "requested_by": row["requested_by"] or "",
                    "created_at": row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else ""
                })
            if rows:
                first_row = rows[0]
                device_info = f"{first_row['sr_brand'] or ''} {first_row['sr_model'] or ''} {first_row['sr_memory'] or ''}".strip()
                batch_info = first_row["customer_name"] or ""
            else:
                device_info = ""
                batch_info = ""

            # Identify work_order_type
            wo_type = None
            recipe_materials = []
            if wo_id:
                wo_row = db.execute(text("SELECT work_order_type, target_part_id FROM warehouse.work_orders WHERE id = :id"), {"id": wo_id}).mappings().first()
                if wo_row:
                    wo_type = wo_row["work_order_type"]
                    if wo_type == 'PRODUCTION' and wo_row["target_part_id"]:
                        target_part = db.execute(text("SELECT item_code FROM warehouse.parts WHERE id = :id"), {"id": wo_row["target_part_id"]}).mappings().first()
                        if target_part and target_part["item_code"]:
                            materials = db.execute(text("""
                                SELECT p_child.id AS child_part_id, p_child.name AS child_name, p_child.item_code AS child_item_code
                                FROM warehouse.item_bom b
                                JOIN warehouse.parts p_child ON p_child.item_code = b.child_item_id
                                WHERE b.parent_item_id = :parent_code
                            """), {"parent_code": target_part["item_code"]}).mappings().all()
                            recipe_materials = [{"id": str(m["child_part_id"]), "name": m["child_name"], "item_code": m["child_item_code"]} for m in materials]

            # Query stock movements related to this work order
            movements = []
            if wo_id:
                from models.stock_movement import StockMovement
                from models.part import Part
                from models.location import Location
                from sqlalchemy.orm import aliased
                
                MovSourceLoc = aliased(Location)
                MovTargetLoc = aliased(Location)
                
                wo_id_str_std = f"#{wo_id}"
                wo_id_str_pad = f"{wo_id:015d}"
                
                mov_rows = db.query(StockMovement, Part, MovSourceLoc, MovTargetLoc)                     .outerjoin(Part, StockMovement.part_id == Part.id)                     .outerjoin(MovSourceLoc, StockMovement.source_location_id == MovSourceLoc.id)                     .outerjoin(MovTargetLoc, StockMovement.target_location_id == MovTargetLoc.id)                     .filter(
                        (StockMovement.description.like(f"%{wo_id_str_std}%")) | 
                        (StockMovement.description.like(f"%{wo_id_str_pad}%"))
                    ).order_by(StockMovement.created_at.desc()).all()
                
                for mov, p, sloc, tloc in mov_rows:
                    source_name = sloc.name if sloc else "-"
                    target_name = tloc.name if tloc else "-"
                    
                    # Fix fallback names for cleaner UI
                    if not sloc:
                        if "İade" in mov.type and "İptal" not in mov.type:
                            source_name = "Good Stock"
                        elif "İptali" in mov.type:
                            source_name = "Good Stock"
                        elif mov.type == "Giriş":
                            source_name = "Dış Kaynak"
                            
                    if not tloc:
                        if "Çıkış" in mov.type or "Tüketimi" in mov.type or ("İptal" in mov.type and "İptali" not in mov.type) or mov.type == "Servis Kullanımı":
                            target_name = "Kullanım/Tüketim"
                        elif mov.type == "Çıkış":
                            target_name = "Dış Kaynak"
                            
                    movements.append({
                        "id": mov.id,
                        "type": mov.type,
                        "quantity": mov.quantity,
                        "part_name": p.name if p else (mov.part_name_snapshot or "Bilinmeyen Parça"),
                        "source_location": source_name,
                        "target_location": target_name,
                        "created_by": mov.created_by or "-",
                        "created_at": mov.created_at.strftime("%Y-%m-%d %H:%M") if mov.created_at else "",
                        "description": mov.description or ""
                    })

            return json.dumps({
                "success": True, 
                "parts": parts, 
                "movements": movements,
                "device_info": device_info, 
                "batch_info": batch_info,
                "work_order_id": str(wo_id) if wo_id else None,
                "work_order_type": wo_type,
                "recipe_materials": recipe_materials
            })
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_work_order_parts(self, work_order_id_str):
        """Bir iş emrine ait parça satırlarını, parça/lokasyon bilgileriyle birlikte getirir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            work_order_id = int(work_order_id_str)
            rows = db.execute(text("""
                SELECT wop.id, wop.work_order_id, wop.part_id, wop.quantity, wop.status,
                       wop.delivered_location_id, wop.delivery_movement_id, wop.delivered_by, wop.delivered_at,
                       wop.waiting_notes, wop.marked_waiting_by, wop.marked_waiting_at,
                       wop.reversal_movement_id, wop.reverted_by, wop.reverted_at,
                       wop.requested_by, wop.created_at,
                       p.brand, p.model, p.color, COALESCE(p.part_category, p.item_category) AS part_category, p.item_code, p.name AS part_name_raw,
                       dl.name AS delivered_location_name
                FROM warehouse.work_order_parts wop
                LEFT JOIN warehouse.parts p ON p.id = wop.part_id
                LEFT JOIN warehouse.locations dl ON dl.id = wop.delivered_location_id
                WHERE wop.work_order_id = :wid
                ORDER BY wop.id ASC
            """), {"wid": work_order_id}).mappings().all()

            parts = []
            for row in rows:
                part_name = " ".join(filter(None, [row["brand"], row["model"], row["color"], row["part_category"]])) or (row["part_name_raw"] or "")
                parts.append({
                    "id": str(row["id"]),
                    "work_order_id": str(row["work_order_id"]),
                    "part_id": str(row["part_id"]),
                    "part_name": part_name,
                    "item_code": row["item_code"] or "",
                    "quantity": row["quantity"],
                    "required_qty": int(row["required_qty"]) if row.get("required_qty") is not None else row["quantity"],
                    "issued_qty": int(row["issued_qty"]) if row.get("issued_qty") is not None else (row["quantity"] if row["status"] == "Teslim Edildi" else 0),
                    "remaining_qty": int(row["remaining_qty"]) if row.get("remaining_qty") is not None else (0 if row["status"] == "Teslim Edildi" else row["quantity"]),
                    "status": row["status"] or "Stokta Var",
                    "delivered_location_id": str(row["delivered_location_id"]) if row["delivered_location_id"] else "",
                    "delivered_location_name": row["delivered_location_name"] or "",
                    "delivery_movement_id": str(row["delivery_movement_id"]) if row["delivery_movement_id"] else "",
                    "delivered_by": row["delivered_by"] or "",
                    "delivered_at": row["delivered_at"].strftime("%Y-%m-%d %H:%M") if row["delivered_at"] else "",
                    "waiting_notes": row["waiting_notes"] or "",
                    "marked_waiting_by": row["marked_waiting_by"] or "",
                    "marked_waiting_at": row["marked_waiting_at"].strftime("%Y-%m-%d %H:%M") if row["marked_waiting_at"] else "",
                    "reverted_by": row["reverted_by"] or "",
                    "reverted_at": row["reverted_at"].strftime("%Y-%m-%d %H:%M") if row["reverted_at"] else "",
                    "requested_by": row["requested_by"] or "",
                    "created_at": row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else ""
                })
            return json.dumps({"success": True, "parts": parts})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def add_work_order_parts_bulk(self, work_order_id_str, rows_json, username):
        """Yeni oluşturulan bir iş emri için taslak parça satırlarını toplu olarak kaydeder."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            work_order_id = int(work_order_id_str)
            try:
                rows = json.loads(rows_json or "[]")
            except (ValueError, TypeError):
                rows = []

            # AYNI PARÇA İKİ KEZ EKLENEMEZ - hem gelen listenin kendi içinde hem de iş emrinde
            # zaten bulunan satırlara karşı. İptal edilmiş satırlar sayılmaz (bkz.
            # remove_work_order_part: satırı silmez, status='İptal Edildi' yapar).
            var_olanlar = {
                r[0] for r in db.execute(text("""
                    SELECT part_id FROM warehouse.work_order_parts
                     WHERE work_order_id = :wid AND COALESCE(status, '') <> 'İptal Edildi'
                """), {"wid": work_order_id}).fetchall()
            }

            inserted = 0
            atlanan = 0
            for row in rows:
                part_id = row.get("part_id")
                try:
                    qty = int(row.get("quantity") or 0)
                except (ValueError, TypeError):
                    qty = 0
                if not part_id or qty < 1:
                    continue
                pid = int(part_id)
                if pid in var_olanlar:
                    atlanan += 1
                    continue
                db.execute(text("""
                    INSERT INTO warehouse.work_order_parts (work_order_id, part_id, quantity, status, requested_by)
                    VALUES (:wid, :pid, :qty, 'Stokta Var', :req)
                """), {"wid": work_order_id, "pid": pid, "qty": qty, "req": username or None})
                var_olanlar.add(pid)
                inserted += 1
            db.commit()
            # Atlananlar SESSİZCE yutulmaz: çağıran ekran kaç satırın tekrar olduğunu bilsin.
            return json.dumps({"success": True, "inserted": inserted, "skipped_duplicates": atlanan})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()


    @Slot(str, str, str, str, result=str)
    def add_material_request(self, work_order_id_str, part_id_str, quantity_str, username):
        """Uretim is emrine manuel olarak ekstra malzeme talebi ekler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            wo_id = int(work_order_id_str)
            part_id = int(part_id_str)
            qty = int(quantity_str)
            
            db.execute(text("""
                INSERT INTO warehouse.material_requests (work_order_id, part_id, required_quantity, issued_quantity, fire_quantity)
                VALUES (:wid, :pid, :qty, 0, 0)
            """), {"wid": wo_id, "pid": part_id, "qty": qty})
            
            row = db.execute(text("""
                SELECT 'mr_' || mr.id::text AS id, mr.work_order_id, mr.part_id, mr.required_quantity AS quantity, 
                       'Tedarik Bekleniyor' AS status, p.brand, p.model, p.color, 
                       COALESCE(p.part_category, p.item_category) AS part_category, p.item_code, p.name AS part_name_raw
                FROM warehouse.material_requests mr
                LEFT JOIN warehouse.parts p ON p.id = mr.part_id
                WHERE mr.work_order_id = :wid AND mr.part_id = :pid
                ORDER BY mr.id DESC LIMIT 1
            """), {"wid": wo_id, "pid": part_id}).mappings().first()
            
            db.commit()
            
            part_name = " ".join(filter(None, [row['brand'], row['model'], row['color'], row['part_category']])) or (row['part_name_raw'] or "")
            part_obj = {
                "id": str(row["id"]),
                "work_order_id": str(row["work_order_id"]),
                "part_id": str(row["part_id"]),
                "quantity": row["quantity"],
                "status": row["status"],
                "part_name": part_name,
                "part_category": row["part_category"] or "",
                "item_code": row["item_code"] or ""
            }
            return json.dumps({"success": True, "part": part_obj})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def add_work_order_part(self, work_order_id_str, part_id_str, quantity_str, username):
        """Kayıtlı bir iş emrine tek bir parça satırı ekler ve eklenen satırı döner."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            work_order_id = int(work_order_id_str)
            part_id = int(part_id_str)
            qty = int(quantity_str) if quantity_str and int(quantity_str) > 0 else 1

            # AYNI PARÇA İKİ KEZ EKLENEMEZ. İptal edilmiş satırlar sayılmaz: remove_work_order_part
            # satırı silmez, status='İptal Edildi' yapar - iptal edilen parça yeniden eklenebilmeli.
            mevcut = db.execute(text("""
                SELECT id FROM warehouse.work_order_parts
                 WHERE work_order_id = :wid AND part_id = :pid
                   AND COALESCE(status, '') <> 'İptal Edildi'
                 LIMIT 1
            """), {"wid": work_order_id, "pid": part_id}).first()
            if mevcut:
                return json.dumps({
                    "success": False,
                    "message": "Bu parça iş emrine zaten eklenmiş. Aynı parça ikinci kez eklenemez; "
                               "adet değiştirmek için mevcut satırı düzenleyin."
                }, ensure_ascii=False)

            # Check available Good Stock quantity
            good_stock_loc = _get_system_location_id(db, "good_stock")
            if not good_stock_loc:
                return json.dumps({"success": False, "message": "Good Stock deposu bulunamadı."})
            
            from models.stock import Stock
            stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == good_stock_loc).first()
            available_qty = stock.quantity if stock else 0
            if qty > available_qty:
                return json.dumps({"success": False, "message": f"Yetersiz stok! Bu parçadan Good Stock deposunda en fazla {available_qty} adet mevcuttur."})

            new_id = db.execute(text("""
                INSERT INTO warehouse.work_order_parts (work_order_id, part_id, quantity, status, requested_by)
                VALUES (:wid, :pid, :qty, 'Stokta Var', :req)
                RETURNING id
            """), {"wid": work_order_id, "pid": part_id, "qty": qty, "req": username or None}).scalar()
            db.commit()

            row = db.execute(text("""
                SELECT wop.id, wop.work_order_id, wop.part_id, wop.quantity, wop.status, wop.created_at,
                       p.brand, p.model, p.color, p.part_category, p.item_code, p.name AS part_name_raw
                FROM warehouse.work_order_parts wop
                LEFT JOIN warehouse.parts p ON p.id = wop.part_id
                WHERE wop.id = :id
            """), {"id": new_id}).mappings().first()

            part_name = " ".join(filter(None, [row["brand"], row["model"], row["color"], row["part_category"]])) or (row["part_name_raw"] or "")
            part = {
                "id": str(row["id"]),
                "work_order_id": str(row["work_order_id"]),
                "part_id": str(row["part_id"]),
                "part_name": part_name,
                "item_code": row["item_code"] or "",
                "quantity": row["quantity"],
                "status": row["status"],
                "delivered_location_id": "", "delivered_location_name": "", "delivery_movement_id": "",
                "delivered_by": "", "delivered_at": "",
                "waiting_notes": "", "marked_waiting_by": "", "marked_waiting_at": "",
                "reverted_by": "", "reverted_at": "",
                "requested_by": username or "",
                "created_at": row["created_at"].strftime("%Y-%m-%d %H:%M") if row["created_at"] else ""
            }
            return json.dumps({"success": True, "part": part})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def deliver_work_order_part(self, wop_id_str, location_id_str, username):
        """'Depodan Teslim Al': stoktan düşer, StockMovement kaydı açar, satırı 'Teslim Edildi' yapar."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        db = SessionLocal()
        try:
            wop_id = int(wop_id_str)
            location_id = int(location_id_str)

            row = db.execute(
                text("SELECT id, work_order_id, part_id, quantity, status FROM warehouse.work_order_parts WHERE id = :id FOR UPDATE"),
                {"id": wop_id}
            ).mappings().first()
            if not row:
                return json.dumps({"success": False, "message": "Parça satırı bulunamadı."})
            if row["status"] == "Teslim Edildi":
                return json.dumps({"success": False, "message": "Bu parça zaten teslim edilmiş."})

            qty = row["quantity"]
            stock = db.query(Stock).filter(Stock.part_id == row["part_id"], Stock.location_id == location_id).first()
            if not stock or stock.quantity < qty:
                return json.dumps({"success": False, "message": "Seçilen lokasyonda yeterli stok yok."})

            from models.location import Location
            repair_stock_loc = db.query(Location).filter(Location.kind == "repair_stock").first()
            if not repair_stock_loc:
                return json.dumps({"success": False, "message": "Repair Stock deposu bulunamadı."})

            # Transfer: decrease from Good Stock, increase in Repair Stock
            stock.quantity -= qty
            
            repair_stock_entry = db.query(Stock).filter(Stock.part_id == row["part_id"], Stock.location_id == repair_stock_loc.id).first()
            if repair_stock_entry:
                repair_stock_entry.quantity += qty
            else:
                db.add(Stock(part_id=row["part_id"], location_id=repair_stock_loc.id, quantity=qty))

            movement = StockMovement(
                type="Stok Çıkışı (Teknisyene)",
                movement_kind="Transfer",
                quantity=qty,
                part_id=row["part_id"],
                source_location_id=location_id,
                target_location_id=repair_stock_loc.id,
                created_by=username or None,
                technician=username or None,
                description=f"İş Emri #{row['work_order_id']} için teknisyene teslim edildi (Good -> Repair)"
            )
            db.add(movement)
            db.flush()

            db.execute(text("""
                UPDATE warehouse.work_order_parts
                SET status = 'Teslim Edildi', delivered_location_id = :loc, delivery_movement_id = :mov,
                    delivered_by = :user, delivered_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
                WHERE id = :id
            """), {"loc": location_id, "mov": movement.id, "user": username or None, "id": wop_id})
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def mark_work_order_part_waiting(self, wop_id_str, notes, username):
        """Sağ tık aksiyonu: parçayı 'Tedarik Bekleniyor' olarak işaretler (stok hareketi yok)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            wop_id = int(wop_id_str)
            result = db.execute(text("""
                UPDATE warehouse.work_order_parts
                SET status = 'Stokta Var', waiting_notes = :notes,
                    marked_waiting_by = :user, marked_waiting_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
                WHERE id = :id AND status != 'Teslim Edildi'
            """), {"notes": notes or None, "user": username or None, "id": wop_id})
            if result.rowcount == 0:
                db.rollback()
                return json.dumps({"success": False, "message": "Zaten teslim edilmiş bir parça bekliyor olarak işaretlenemez."})
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def revert_work_order_part_status(self, wop_id_str, username, return_qty_str=None):
        """Durumu geri alır: Tedarik Bekleniyor -> Stokta Var, veya Teslim Edildi -> Stokta Var (stok iadeli)."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        db = SessionLocal()
        try:
            if str(wop_id_str).startswith('pm_'):
                return json.dumps({"success": False, "message": "Üretimi tamamlanmış cihazların parçaları tekil olarak iade edilemez. Lütfen cihazın tamamını iade edin."})

            if str(wop_id_str).startswith('mr_'):
                # Production Work Order material request
                mr_id = int(str(wop_id_str).replace('mr_', ''))
                mr = db.execute(text("SELECT id, part_id, issued_quantity FROM warehouse.material_requests WHERE id = :id FOR UPDATE"), {"id": mr_id}).mappings().first()
                if not mr:
                    return json.dumps({"success": False, "message": "Material Request bulunamadı."})
                qty_to_return = int(return_qty_str) if return_qty_str and int(return_qty_str) > 0 else mr["issued_quantity"]
                if qty_to_return <= 0:
                    return json.dumps({"success": False, "message": "İade edilecek teslim edilmiş miktar yok."})
                if qty_to_return > mr["issued_quantity"]:
                    return json.dumps({"success": False, "message": f"En fazla {mr['issued_quantity']} adet iade edebilirsiniz."})

                good_stock_loc = _get_system_location_id(db, "good_stock")
                if not good_stock_loc:
                    return json.dumps({"success": False, "message": "Good Stock lokasyonu bulunamadı."})

                # Add to Good Stock
                existing_stock = db.query(Stock).filter(Stock.location_id == good_stock_loc, Stock.part_id == mr["part_id"]).first()
                if existing_stock:
                    existing_stock.quantity += qty_to_return
                else:
                    db.add(Stock(location_id=good_stock_loc, part_id=mr["part_id"], quantity=qty_to_return))

                db.add(StockMovement(
                    type="Stock Return",
                    movement_kind="Inbound",
                    part_id=mr["part_id"],
                    quantity=qty_to_return,
                    target_location_id=good_stock_loc,
                    created_by=username,
                    description=f"Üretim siparişinden Stoğa Geri Alındı (MR #{mr['id']})"
                ))

                db.execute(text("UPDATE warehouse.material_requests SET issued_quantity = issued_quantity - :qty WHERE id = :id"), {"qty": qty_to_return, "id": mr_id})
                db.commit()
                return json.dumps({"success": True, "message": "Parçalar stoğa geri alındı."})
            
            wop_id = int(wop_id_str)
            row = db.execute(
                text("""SELECT id, work_order_id, part_id, quantity, status, delivered_location_id, delivery_movement_id
                        FROM warehouse.work_order_parts WHERE id = :id FOR UPDATE"""),
                {"id": wop_id}
            ).mappings().first()
            if not row:
                return json.dumps({"success": False, "message": "Parça satırı bulunamadı."})

            if row["status"] == "Stokta Var":
                return json.dumps({"success": False, "message": "Bu parça zaten başlangıç durumunda."})

            if row["status"] in ("Tedarik Bekleniyor", "İptal Edildi"):
                # waiting_notes/marked_waiting_by/marked_waiting_at kasıtlı olarak silinmiyor:
                # Tedarik Talepleri geçmişinde bu satırın bir talep olduğu bilgisi korunur (Onaylandı olarak görünür).
                db.execute(text("""
                    UPDATE warehouse.work_order_parts
                    SET status = 'Stokta Var',
                        reverted_by = :user, reverted_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
                    WHERE id = :id
                """), {"user": username or None, "id": wop_id})
                db.commit()
                return json.dumps({"success": True})

            # status == 'Teslim Edildi' -> stok iadesi + telafi hareketi
            qty = int(return_qty_str) if return_qty_str and int(return_qty_str) > 0 else row["quantity"]
            if qty > row["quantity"]:
                return json.dumps({"success": False, "message": f"En fazla {row['quantity']} adet iade edebilirsiniz."})
                
            location_id = row["delivered_location_id"]
            stock = db.query(Stock).filter(Stock.part_id == row["part_id"], Stock.location_id == location_id).first()
            if stock:
                stock.quantity += qty
            else:
                stock = Stock(part_id=row["part_id"], location_id=location_id, quantity=qty)
                db.add(stock)

            reversal = StockMovement(
                type="Teslimat İptali",
                quantity=qty,
                part_id=row["part_id"],
                target_location_id=location_id,
                created_by=username or None,
                description=f"İş Emri #{row['work_order_id']} teslimatı geri alındı (orijinal hareket #{row['delivery_movement_id']})"
            )
            db.add(reversal)
            db.flush()

            if qty == row["quantity"]:
                db.execute(text("""
                    UPDATE warehouse.work_order_parts
                    SET status = 'Stoğa Geri Alındı', delivered_location_id = NULL, delivery_movement_id = NULL,
                        delivered_by = NULL, delivered_at = NULL,
                        reversal_movement_id = :rev, reverted_by = :user, reverted_at = CURRENT_TIMESTAMP,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = :id
                """), {"rev": reversal.id, "user": username or None, "id": wop_id})
            else:
                # Partial return to good stock: split row
                db.execute(text("""
                    UPDATE warehouse.work_order_parts
                    SET quantity = quantity - :rqty, updated_at = CURRENT_TIMESTAMP
                    WHERE id = :id
                """), {"rqty": qty, "id": wop_id})
                
                db.execute(text("""
                    INSERT INTO warehouse.work_order_parts (work_order_id, part_id, quantity, status, delivered_location_id, created_at, updated_at, reverted_by, reverted_at, reversal_movement_id)
                    VALUES (:wo_id, :part_id, :qty, 'Stoğa Geri Alındı', :delivered_loc, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP, :user, CURRENT_TIMESTAMP, :rev)
                """), {
                    "wo_id": row["work_order_id"],
                    "part_id": row["part_id"],
                    "qty": qty,
                    "delivered_loc": row["delivered_location_id"],
                    "user": username or None,
                    "rev": reversal.id
                })
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def return_part_to_doa(self, wop_id_str, return_qty_str, username):
        """'DOA Stoğa Geri Al': Teslim edilmiş bir parçanın belirtilen miktarını DOA Stock'a taşır."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        from models.location import Location
        
        if str(wop_id_str).startswith('pm_'):
            return json.dumps({"success": False, "message": "Üretimi tamamlanmış cihazların parçaları tekil olarak DOA'ya alınamaz. Lütfen 'Üretim İş Emirleri' kısmından fire/iade işlemi yapın."})

        if str(wop_id_str).startswith('mr_'):
            # Production Work Order material request
            mr_id = str(wop_id_str).replace('mr_', '')
            # DOA Stoğa al demek fire bildirmekle aynı mantık (DOA stoğa atar, issued_quantity'den eksiltip fire_quantity'ye ekler)
            return self.report_material_fire(mr_id, return_qty_str, username)

        db = SessionLocal()
        try:
            wop_id = int(wop_id_str)
            return_qty = int(return_qty_str) if return_qty_str and int(return_qty_str) > 0 else 0
            if return_qty <= 0:
                return json.dumps({"success": False, "message": "Geçerli bir miktar giriniz."})

            row = db.execute(
                text("SELECT id, work_order_id, part_id, quantity, status, delivered_location_id FROM warehouse.work_order_parts WHERE id = :id FOR UPDATE"),
                {"id": wop_id}
            ).mappings().first()
            if not row:
                return json.dumps({"success": False, "message": "Parça satırı bulunamadı."})
            if row["status"] != "Teslim Edildi":
                return json.dumps({"success": False, "message": "Sadece teslim edilmiş parçalar DOA stoğa geri alınabilir."})

            if return_qty > row["quantity"]:
                return json.dumps({"success": False, "message": f"En fazla {row['quantity']} adet geri alabilirsiniz."})

            doa_loc = db.query(Location).filter(Location.kind == "doa_stock").first()
            if not doa_loc:
                return json.dumps({"success": False, "message": "DOA Stock lokasyonu bulunamadı."})

            src_loc_id = row["delivered_location_id"]
            if not src_loc_id:
                repair_loc = db.query(Location).filter(Location.kind == "repair_stock").first()
                src_loc_id = repair_loc.id if repair_loc else None

            target_stock = db.query(Stock).filter(Stock.part_id == row["part_id"], Stock.location_id == doa_loc.id).first()
            if target_stock:
                target_stock.quantity += return_qty
            else:
                db.add(Stock(part_id=row["part_id"], location_id=doa_loc.id, quantity=return_qty))

            movement = StockMovement(
                type="DOA İade",
                movement_kind="Transfer",
                quantity=return_qty,
                part_id=row["part_id"],
                source_location_id=src_loc_id,
                target_location_id=doa_loc.id,
                created_by=username or None,
                technician=username or None,
                description=f"İş Emri #{row['work_order_id']} parçası ({return_qty} adet) DOA stoğa geri alındı"
            )
            db.add(movement)

            if return_qty == row["quantity"]:
                db.execute(text("""
                    UPDATE warehouse.work_order_parts
                    SET status = 'DOA İade', updated_at = CURRENT_TIMESTAMP
                    WHERE id = :id
                """), {"id": wop_id})
            else:
                # Partial return: split the row
                # 1. Deduct quantity from original "Teslim Edildi" row
                db.execute(text("""
                    UPDATE warehouse.work_order_parts
                    SET quantity = quantity - :rqty, updated_at = CURRENT_TIMESTAMP
                    WHERE id = :id
                """), {"rqty": return_qty, "id": wop_id})
                
                # 2. Insert new row representing the returned quantity with status "Kısmi İade Edildi"
                db.execute(text("""
                    INSERT INTO warehouse.work_order_parts (work_order_id, part_id, quantity, status, delivered_location_id, created_at, updated_at)
                    VALUES (:wo_id, :part_id, :qty, 'Kısmi İade Edildi', :delivered_loc, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """), {
                    "wo_id": row["work_order_id"],
                    "part_id": row["part_id"],
                    "qty": return_qty,
                    "delivered_loc": row["delivered_location_id"]
                })

            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def remove_work_order_part(self, payload_str):
        """Bir parça satırını iptal edildi olarak günceller ve sebebini not düşer. (Payload JSON formatında)"""
        from sqlalchemy import text
        import json
        db = SessionLocal()
        try:
            try:
                payload = json.loads(payload_str)
                wop_id_str = str(payload.get('id', ''))
                reason_str = str(payload.get('reason', ''))
                username = str(payload.get('username', 'System'))
            except Exception:
                # Geriye dönük uyumluluk (sadece ID gelirse)
                wop_id_str = str(payload_str)
                reason_str = ''
                username = 'System'

            if wop_id_str.startswith('pm_'):
                return json.dumps({"success": False, "message": "Üretimi tamamlanmış cihazların parçaları silinemez/iptal edilemez."})

            if wop_id_str.startswith('mr_'):
                mr_id = int(wop_id_str.replace('mr_', ''))
                mr = db.execute(text("SELECT id, part_id, issued_quantity FROM warehouse.material_requests WHERE id = :id FOR UPDATE"), {"id": mr_id}).mappings().first()
                if not mr:
                    return json.dumps({"success": False, "message": "Material Request bulunamadı."})
                
                qty_to_return = mr["issued_quantity"]
                if qty_to_return > 0:
                    # Auto return to Good Stock
                    from models.stock import Stock
                    from models.stock_movement import StockMovement
                    good_stock_loc = _get_system_location_id(db, "good_stock")
                    repair_stock_id = _get_system_location_id(db, "repair_stock")
                    
                    # 1. Decrement from Repair Stock (technician stock)
                    if repair_stock_id:
                        rep_stock = db.query(Stock).filter(Stock.location_id == repair_stock_id, Stock.part_id == mr["part_id"]).first()
                        if rep_stock:
                            rep_stock.quantity = max(0, rep_stock.quantity - qty_to_return)
                            
                    # 2. Increment in Good Stock (warehouse stock)
                    if good_stock_loc:
                        existing_stock = db.query(Stock).filter(Stock.location_id == good_stock_loc, Stock.part_id == mr["part_id"]).first()
                        if existing_stock:
                            existing_stock.quantity += qty_to_return
                        else:
                            db.add(Stock(location_id=good_stock_loc, part_id=mr["part_id"], quantity=qty_to_return))

                        db.add(StockMovement(
                            type="Stock Return",
                            movement_kind="Inbound",
                            part_id=mr["part_id"],
                            quantity=qty_to_return,
                            target_location_id=good_stock_loc,
                            created_by=username or "System",
                            description=f"İptal nedeniyle Stoğa Geri Alındı (MR #{mr['id']})"
                        ))
                
                # Material Request tablosunda durumu CANCELLED yapıyoruz
                db.execute(text("UPDATE warehouse.material_requests SET status = 'CANCELLED', issued_quantity = 0 WHERE id = :id"), {"id": mr_id})
                db.commit()
                return json.dumps({"success": True})

            wop_id = int(wop_id_str)
            # Service parçaları için
            row = db.execute(
                text("SELECT id, work_order_id, part_id, quantity, status, delivered_location_id, delivery_movement_id FROM warehouse.work_order_parts WHERE id = :id FOR UPDATE"),
                {"id": wop_id}
            ).mappings().first()
            
            if not row:
                return json.dumps({"success": False, "message": "İptal edilecek satır bulunamadı."})

            if row["status"] in ("Teslim Edildi", "Kısmi İade Edildi"):
                # Auto return to source location (Good Stock)
                from models.stock import Stock
                from models.stock_movement import StockMovement
                qty = row["quantity"]
                location_id = row["delivered_location_id"]
                repair_stock_id = _get_system_location_id(db, "repair_stock")
                
                # 1. Decrement from Repair Stock (technician stock)
                if repair_stock_id:
                    rep_stock = db.query(Stock).filter(Stock.location_id == repair_stock_id, Stock.part_id == row["part_id"]).first()
                    if rep_stock:
                        rep_stock.quantity = max(0, rep_stock.quantity - qty)
                
                # 2. Increment in Good Stock
                if location_id:
                    stock = db.query(Stock).filter(Stock.part_id == row["part_id"], Stock.location_id == location_id).first()
                    if stock:
                        stock.quantity += qty
                    else:
                        db.add(Stock(part_id=row["part_id"], location_id=location_id, quantity=qty))

                    reversal = StockMovement(
                        type="Teslimat İptali",
                        quantity=qty,
                        part_id=row["part_id"],
                        target_location_id=location_id,
                        created_by=username or "System",
                        description=f"İptal nedeniyle teslimat geri alındı (İş Emri #{row['work_order_id']}, orijinal hareket #{row['delivery_movement_id']})"
                    )
                    db.add(reversal)

            db.execute(text("""
                UPDATE warehouse.work_order_parts 
                SET status = 'İptal Edildi', waiting_notes = :reason, delivered_location_id = NULL, delivery_movement_id = NULL
                WHERE id = :id
            """), {"id": wop_id, "reason": reason_str})
            
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_service_repair_details(self, work_order_id_str):
        """Servis iş emrine ait arıza tespit, tanı kartları, onarım aşamaları bilgilerini getirir."""
        from sqlalchemy import text
        import json
        db = SessionLocal()
        try:
            wo_id = int(work_order_id_str)
            wo = db.execute(text("""
                SELECT id, service_record_id, assigned_technician
                FROM warehouse.work_orders WHERE id = :id
            """), {"id": wo_id}).mappings().first()
            if not wo:
                return json.dumps({"success": False, "message": "İş emri bulunamadı."})
            
            sr_id = wo["service_record_id"]
            if not sr_id:
                return json.dumps({"success": False, "message": "Bu iş emri bir servis kaydına bağlı değil."})
                
            sr = db.execute(text("""
                SELECT id, customer_name, brand, model, color, memory, imei_number, imei_serial,
                       customer_complaint, preliminary_diagnosis, technician_note
                FROM warehouse.service_records WHERE id = :id
            """), {"id": sr_id}).mappings().first()
            
            if not sr:
                return json.dumps({"success": False, "message": "Servis kaydı bulunamadı."})

            diag_data = None
            raw_note = sr["technician_note"] or ""
            if raw_note.strip().startswith('{') and raw_note.strip().endswith('}'):
                try:
                    diag_data = json.loads(raw_note)
                except Exception:
                    pass
            
            if not diag_data or not isinstance(diag_data, dict):
                diag_data = {
                    "diagnostics": {
                        "lcd": "OK",
                        "mp_camera": "OK",
                        "b_camera": "OK",
                        "battery_cycle": "0",
                        "battery_health": "100"
                    },
                    "stages": [
                        {"group_name": "Kasa Onarımı", "staff_name": wo["assigned_technician"] or "", "count": 1, "status": "Beklemede", "start_time": "", "finish_time": ""},
                        {"group_name": "Kamera Onarımı", "staff_name": "", "count": 1, "status": "Beklemede", "start_time": "", "finish_time": ""},
                        {"group_name": "Ekran Onarımı", "staff_name": "", "count": 1, "status": "Beklemede", "start_time": "", "finish_time": ""},
                        {"group_name": "L1 Onarımı", "staff_name": "", "count": 1, "status": "Beklemede", "start_time": "", "finish_time": ""}
                    ],
                    "price": "0.00"
                }

            res = {
                "success": True,
                "service_record_id": sr["id"],
                "customer_name": sr["customer_name"] or "",
                "brand": sr["brand"] or "",
                "model": sr["model"] or "",
                "color": sr["color"] or "",
                "memory": sr["memory"] or "",
                "imei_number": sr["imei_number"] or sr["imei_serial"] or "",
                "customer_complaint": sr["customer_complaint"] or "",
                "preliminary_diagnosis": sr["preliminary_diagnosis"] or "",
                "diagnostics": diag_data.get("diagnostics", {}),
                "stages": diag_data.get("stages", []),
                "price": diag_data.get("price", "0.00")
            }
            return json.dumps(res)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def save_service_repair_details(self, work_order_id_str, details_json):
        """Servis iş emrine ait arıza tespit, tanı kartları, onarım aşamaları bilgilerini kaydeder."""
        from sqlalchemy import text
        import json
        db = SessionLocal()
        try:
            wo_id = int(work_order_id_str)
            wo = db.execute(text("SELECT service_record_id FROM warehouse.work_orders WHERE id = :id"), {"id": wo_id}).mappings().first()
            if not wo or not wo["service_record_id"]:
                return json.dumps({"success": False, "message": "İş emri veya bağlı servis kaydı bulunamadı."})
                
            db.execute(text("""
                UPDATE warehouse.service_records
                SET technician_note = :note
                WHERE id = :id
            """), {"note": details_json, "id": wo["service_record_id"]})
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def cancel_supply_request(self, wop_id_str, username):
        """Bir tedarik talebini iptal eder (satır silinmez, durum 'İptal Edildi' olur; geçmişte görünmeye devam eder)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            wop_id = int(wop_id_str)
            result = db.execute(text("""
                UPDATE warehouse.work_order_parts
                SET status = 'İptal Edildi', reverted_by = :user, reverted_at = CURRENT_TIMESTAMP, updated_at = CURRENT_TIMESTAMP
                WHERE id = :id AND status != 'Teslim Edildi'
            """), {"user": username or None, "id": wop_id})
            if result.rowcount == 0:
                db.rollback()
                return json.dumps({"success": False, "message": "Teslim edilmiş bir talep iptal edilemez, önce geri alın."})
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, result=str)
    def create_supply_request(self, work_order_id_str, part_id_str, quantity_str, notes, username):
        """Teknisyenin doğrudan tedarik talebi oluşturması: satır doğrudan 'Tedarik Bekleniyor' olarak eklenir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            work_order_id = int(work_order_id_str)
            part_id = int(part_id_str)
            qty = int(quantity_str) if quantity_str and int(quantity_str) > 0 else 1
            db.execute(text("""
                INSERT INTO warehouse.work_order_parts
                    (work_order_id, part_id, quantity, status, waiting_notes, marked_waiting_by, marked_waiting_at, requested_by)
                VALUES
                    (:wid, :pid, :qty, 'Stokta Var', :notes, :user, CURRENT_TIMESTAMP, :user)
            """), {"wid": work_order_id, "pid": part_id, "qty": qty, "notes": notes or None, "user": username or None})
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_supply_requests(self):
        """Tüm iş emirlerinde 'Tedarik Bekleniyor' durumundaki parça satırlarını getirir (Tedarik İstekleri sayfası)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT wop.id, wop.work_order_id, wop.quantity, wop.waiting_notes, wop.marked_waiting_by, wop.marked_waiting_at,
                       p.id AS part_id, p.brand, p.model, p.color, COALESCE(p.part_category, p.item_category) AS part_category, p.item_code, p.name AS part_name_raw,
                       w.assigned_technician, w.priority, w.status AS work_order_status,
                       s.customer_name, s.brand AS device_brand, s.model AS device_model
                FROM warehouse.work_order_parts wop
                JOIN warehouse.work_orders w ON w.id = wop.work_order_id
                LEFT JOIN warehouse.service_records s ON s.id = w.service_record_id
                LEFT JOIN warehouse.parts p ON p.id = wop.part_id
                WHERE wop.status IN ('Tedarik Bekleniyor', 'Stokta Var')
                ORDER BY wop.marked_waiting_at ASC
            """)).mappings().all()

            requests = []
            for row in rows:
                part_name = " ".join(filter(None, [row["brand"], row["model"], row["color"], row["part_category"]])) or (row["part_name_raw"] or "")
                requests.append({
                    "id": str(row["id"]),
                    "work_order_id": str(row["work_order_id"]),
                    "part_id": str(row["part_id"]) if row["part_id"] else "",
                    "part_name": part_name,
                    "item_code": row["item_code"] or "",
                    "quantity": row["quantity"],
                    "customer_name": row["customer_name"] or "",
                    "device_brand": row["device_brand"] or "",
                    "device_model": row["device_model"] or "",
                    "assigned_technician": row["assigned_technician"] or "",
                    "priority": row["priority"] or "Orta",
                    "work_order_status": row["work_order_status"] or "",
                    "waiting_notes": row["waiting_notes"] or "",
                    "marked_waiting_by": row["marked_waiting_by"] or "",
                    "marked_waiting_at": row["marked_waiting_at"].strftime("%Y-%m-%d %H:%M") if row["marked_waiting_at"] else ""
                })
            return json.dumps({"success": True, "requests": requests})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_supply_request_history(self, username):
        """Oturumdaki kullanıcının kendi oluşturduğu tedarik taleplerini getirir (Tedarik Talepleri sayfası).
        Başka kullanıcıların talepleri dahil edilmez; depocunun tüm talepleri gördüğü kuyruk için
        bkz. get_supply_requests (ayrı, kullanıcıya göre filtrelenmeyen bir Slot)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT wop.id, wop.work_order_id, wop.quantity, wop.status, wop.waiting_notes,
                       wop.marked_waiting_by, wop.marked_waiting_at,
                       p.id AS part_id, p.brand, p.model, p.color, COALESCE(p.part_category, p.item_category) AS part_category, p.item_code, p.name AS part_name_raw,
                       w.assigned_technician, w.priority, w.status AS work_order_status,
                       s.customer_name, s.brand AS device_brand, s.model AS device_model
                FROM warehouse.work_order_parts wop
                JOIN warehouse.work_orders w ON w.id = wop.work_order_id
                LEFT JOIN warehouse.service_records s ON s.id = w.service_record_id
                LEFT JOIN warehouse.parts p ON p.id = wop.part_id
                WHERE wop.marked_waiting_at IS NOT NULL AND wop.requested_by = :username
                ORDER BY wop.marked_waiting_at DESC
            """), {"username": username or None}).mappings().all()

            requests = []
            for row in rows:
                part_name = " ".join(filter(None, [row["brand"], row["model"], row["color"], row["part_category"]])) or (row["part_name_raw"] or "")
                requests.append({
                    "id": str(row["id"]),
                    "work_order_id": str(row["work_order_id"]),
                    "part_id": str(row["part_id"]) if row["part_id"] else "",
                    "part_name": part_name,
                    "item_code": row["item_code"] or "",
                    "quantity": row["quantity"],
                    "status": row["status"] or "Tedarik Bekleniyor",
                    "customer_name": row["customer_name"] or "",
                    "device_brand": row["device_brand"] or "",
                    "device_model": row["device_model"] or "",
                    "assigned_technician": row["assigned_technician"] or "",
                    "priority": row["priority"] or "Orta",
                    "work_order_status": row["work_order_status"] or "",
                    "waiting_notes": row["waiting_notes"] or "",
                    "marked_waiting_by": row["marked_waiting_by"] or "",
                    "marked_waiting_at": row["marked_waiting_at"].strftime("%Y-%m-%d %H:%M") if row["marked_waiting_at"] else ""
                })
            return json.dumps({"success": True, "requests": requests})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # ==========================
    # ÜRETİM MODÜLÜ (Yarı Mamul Üretimi / Malzeme Tüketimi / Üretim Geçmişi)
    # ==========================

    @Slot(result=str)
    def get_production_runs(self):
        """Tüm üretilen cihaz kayıtlarını, benzersiz seri numaraları (Cihaz Kimlik ID) ve tükettikleri malzemelerle birlikte getirir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            units = db.execute(text("""
                SELECT pu.id AS unit_id, pu.serial_number, pu.is_returned, pu.return_reason, pu.returned_at, pu.return_location_id, pu.returned_materials, pu.replacement_requested_qty,
                       pr.id AS run_id, pr.target_part_id, pr.quantity_produced, pr.location_id, pr.source_location_id,
                       pr.produced_by, pr.notes, pr.created_at, pr.department, pr.scrap_quantity, pr.work_order_id,
                       p.brand AS target_brand, p.model AS target_model,
                       p.item_code AS target_code, p.name AS target_name,
                       l.name AS location_name,
                       sl.name AS source_location_name,
                       rl.name AS return_location_name
                FROM warehouse.produced_units pu
                JOIN warehouse.production_runs pr ON pr.id = pu.production_run_id
                LEFT JOIN warehouse.parts p ON p.id = pr.target_part_id
                LEFT JOIN warehouse.locations l ON l.id = pr.location_id
                LEFT JOIN warehouse.locations sl ON sl.id = pr.source_location_id
                LEFT JOIN warehouse.locations rl ON rl.id = pu.return_location_id
                ORDER BY pu.id DESC
            """)).mappings().all()

            materials = db.execute(text("""
                SELECT pm.production_run_id, pm.part_id, pm.quantity_consumed,
                       p.brand, p.model, p.item_code, p.name
                FROM warehouse.production_materials pm
                LEFT JOIN warehouse.parts p ON p.id = pm.part_id
            """)).mappings().all()

            materials_by_run = {}
            for m in materials:
                part_label = f'{m["brand"] or ""} {m["model"] or ""}'.strip() or (m["name"] or "")
                materials_by_run.setdefault(m["production_run_id"], []).append({
                    "part_id": str(m["part_id"]) if m["part_id"] else "",
                    "part_name": part_label,
                    "item_code": m["item_code"] or "",
                    "quantity_consumed": m["quantity_consumed"]
                })

            result = []
            for u in units:
                target_label = f'{u["target_brand"] or ""} {u["target_model"] or ""}'.strip() or (u["target_name"] or "")
                run_qty = u["quantity_produced"]
                
                # Tüm parti için toplam malzeme tüketimini ekle
                unit_materials = []
                for m in materials_by_run.get(u["run_id"], []):
                    qty = float(m["quantity_consumed"])
                    if qty.is_integer():
                        qty = int(qty)
                    else:
                        qty = round(qty, 2)
                        
                    unit_materials.append({
                        "part_id": m["part_id"],
                        "part_name": m["part_name"],
                        "item_code": m["item_code"],
                        "quantity_consumed": qty
                    })

                result.append({
                    "id": str(u["run_id"]),  # Geri alma işlemleri için run_id
                    "unit_id": str(u["unit_id"]),
                    "serial_number": u["serial_number"],
                    "is_returned": bool(u["is_returned"]),
                    "return_reason": u["return_reason"] or "",
                    "returned_at": u["returned_at"].strftime("%Y-%m-%d %H:%M") if u["returned_at"] else "",
                    "return_location_id": str(u["return_location_id"]) if u["return_location_id"] else "",
                    "return_location_name": u["return_location_name"] or "",
                    "returned_materials": json.loads(u["returned_materials"]) if u["returned_materials"] else [],
                    "replacement_requested_qty": int(u["replacement_requested_qty"]) if u["replacement_requested_qty"] else 0,
                    "target_part_id": str(u["target_part_id"]) if u["target_part_id"] else "",
                    "target_part_name": target_label,
                    "target_item_code": u["target_code"] or "",
                    "quantity_produced": u["quantity_produced"],
                    "location_id": str(u["location_id"]) if u["location_id"] else "",
                    "location_name": u["location_name"] or "",
                    "source_location_id": str(u["source_location_id"]) if u["source_location_id"] else "",
                    "source_location_name": u["source_location_name"] or "",
                    "produced_by": u["produced_by"] or "",
                    "department": u["department"] or "",
                    "scrap_quantity": u["scrap_quantity"] or 0,
                    "work_order_id": str(u["work_order_id"]) if u["work_order_id"] else "",
                    "notes": u["notes"] or "",
                    "created_at": u["created_at"].strftime("%Y-%m-%d %H:%M") if u["created_at"] else "",
                    "materials": unit_materials
                })
            return json.dumps({"success": True, "production_runs": result})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, result=str)
    def create_production_run(self, target_part_id, quantity_produced, source_location_id, target_location_id, produced_by, notes, materials_json, department, scrap_quantity_str):
        """Hammadde tüketip yarı mamul/ürün stoku oluşturan bir üretim kaydı ekler."""
        from sqlalchemy import text
        from models.stock_movement import StockMovement
        import json as json_module
        db = SessionLocal()
        try:
            qty = int(quantity_produced)
            src_loc_id = int(source_location_id)
            tgt_loc_id = int(target_location_id)
            tgt_id = int(target_part_id)
            materials = json_module.loads(materials_json or "[]")
            scrap_qty = int(scrap_quantity_str) if scrap_quantity_str else 0
            dept = department or None

            if qty <= 0:
                return json.dumps({"success": False, "message": "Üretilecek miktar sıfırdan büyük olmalıdır."})

            # Stok yeterliliğini kontrol et (hangi lokasyonda olursa olsun toplam stok esas alınır)
            for m in materials:
                part_id = int(m["part_id"])
                needed = int(m["quantity_consumed"])
                total_available = db.execute(text("""
                    SELECT COALESCE(SUM(quantity), 0) FROM warehouse.stock WHERE part_id = :pid
                """), {"pid": part_id}).scalar()
                if total_available < needed:
                    return json.dumps({"success": False, "message": f"Yetersiz stok (parça id {part_id}): mevcut {total_available}, gerekli {needed}"})

            # Hammaddeleri, stoğu nerede varsa oradan düş (birden fazla lokasyona/satıra yayılmış olabilir).
            # Her düşülen (parça, lokasyon, miktar) üçlüsünü, hareket geçmişine daha sonra
            # StockMovement kaydı açabilmek için ayrıca not ediyoruz.
            consumption_records = []
            for m in materials:
                part_id = int(m["part_id"])
                remaining = int(m["quantity_consumed"])
                rows = db.execute(text("""
                    SELECT id, location_id, quantity FROM warehouse.stock
                    WHERE part_id = :pid AND quantity > 0
                    ORDER BY id
                    FOR UPDATE
                """), {"pid": part_id}).all()
                for stock_id, stock_location_id, stock_qty in rows:
                    if remaining <= 0:
                        break
                    take = min(stock_qty, remaining)
                    db.execute(text("""
                        UPDATE warehouse.stock SET quantity = quantity - :take WHERE id = :id
                    """), {"take": take, "id": stock_id})
                    remaining -= take
                    consumption_records.append((part_id, stock_location_id, take))

            # Üretilen parçanın stokunu artır (yoksa oluştur)
            existing = db.execute(text("""
                SELECT id FROM warehouse.stock WHERE part_id = :pid AND location_id = :lid
            """), {"pid": tgt_id, "lid": tgt_loc_id}).first()
            if existing:
                db.execute(text("UPDATE warehouse.stock SET quantity = quantity + :qty WHERE id = :id"),
                           {"qty": qty, "id": existing[0]})
            else:
                db.execute(text("""
                    INSERT INTO warehouse.stock (part_id, location_id, quantity) VALUES (:pid, :lid, :qty)
                """), {"pid": tgt_id, "lid": tgt_loc_id, "qty": qty})

            # Hızlı Üretim için de bir İş Emri (Work Order) oluştur ve tamamlandı işaretle
            wo_id = db.execute(text("""
                INSERT INTO warehouse.work_orders (
                    work_order_type, target_part_id, description, priority, planned_quantity,
                    assigned_technician, department, status, completed_at, produced_quantity, scrap_quantity, production_notes
                ) VALUES (
                    :wtype, :tgt, :desc, 'Orta', :qty, :tech, :dept, :status, CURRENT_TIMESTAMP, :qty, :scrap, :notes
                ) RETURNING id
            """), {
                "wtype": WORK_ORDER_TYPE_PRODUCTION,
                "tgt": tgt_id, "desc": "Hızlı Üretim (Otomatik İş Emri)",
                "qty": qty, "tech": produced_by or None, "dept": dept,
                "status": PRODUCTION_WO_STATUS_COMPLETED,
                "scrap": scrap_qty, "notes": notes or None
            }).scalar()

            # Üretim kaydını oluştur
            run_id = db.execute(text("""
                INSERT INTO warehouse.production_runs (target_part_id, quantity_produced, source_location_id, location_id, produced_by, notes, department, scrap_quantity, work_order_id)
                VALUES (:tgt, :qty, :slid, :tlid, :by, :notes, :dept, :scrap, :wo_id) RETURNING id
            """), {
                "tgt": tgt_id, "qty": qty, "slid": src_loc_id, "tlid": tgt_loc_id,
                "by": produced_by or None, "notes": notes or None, "dept": dept, "scrap": scrap_qty, "wo_id": wo_id
            }).scalar()

            # Tek bir ortak serial number (Cihaz Kimlik ID) oluştur ve tek satır olarak ekle
            next_id = db.execute(text("SELECT nextval(pg_get_serial_sequence('warehouse.produced_units', 'id'))")).scalar()
            serial_num = f"{next_id:015d}"

            db.execute(text("""
                INSERT INTO warehouse.produced_units (id, production_run_id, serial_number)
                VALUES (:id, :run_id, :serial)
            """), {"id": next_id, "run_id": run_id, "serial": serial_num})

            for m in materials:
                db.execute(text("""
                    INSERT INTO warehouse.production_materials (production_run_id, part_id, quantity_consumed)
                    VALUES (:run_id, :pid, :qty)
                """), {"run_id": run_id, "pid": int(m["part_id"]), "qty": int(m["quantity_consumed"])})

            # Hareket geçmişi (audit trail): tüketilen her hammadde için bir çıkış hareketi,
            # üretilen yarı mamul için bir giriş hareketi. Böylece Depo sayfasındaki "Son
            # Hareket Tarihi" ve Stok Hareketleri raporu üretim faaliyetini de yansıtır.
            for part_id, stock_location_id, take in consumption_records:
                db.add(StockMovement(
                    type="Üretim İçin Malzeme Tüketimi",
                    movement_kind="Outbound",
                    quantity=take,
                    part_id=part_id,
                    source_location_id=stock_location_id,
                    created_by=produced_by or None,
                    description=f"Üretim Kaydı #{run_id} ({serial_num}) için tüketildi"
                ))
            db.add(StockMovement(
                type="Üretim",
                movement_kind="Inbound",
                quantity=qty,
                part_id=tgt_id,
                target_location_id=tgt_loc_id,
                created_by=produced_by or None,
                description=f"Üretim Kaydı #{run_id} ({serial_num}) ile üretildi"
            ))

            db.commit()
            return json.dumps({"success": True, "message": "Üretim kaydı oluşturuldu", "serial_number": serial_num, "run_id": run_id})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"Üretim kaydı hatası: {str(e)}"})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_production_run(self, params_json):
        """Belirtilen üretilmiş cihaz birimini (produced_unit) iade eder.
        Sorunlu parçalar seçilen iade deposuna, sorunsuzlar Good Stock'a (id=26) aktarılır.
        params_json: JSON string {unit_id, return_location_id, return_reason, defective_parts: [{part_id, defective}]}
        """
        # Dış try: PySide6'nın slot'tan "" döndürmesini engeller
        try:
            return self._do_delete_production_run(params_json)
        except Exception as e:
            return json.dumps({"success": False, "message": f"Slot hatası: {str(e)}"})

    def _do_delete_production_run(self, params_json):
        from sqlalchemy import text
        from datetime import datetime
        from models.stock_movement import StockMovement
        db = SessionLocal()
        result_str = json.dumps({"success": False, "message": "Bilinmeyen hata"})
        try:
            params = json.loads(params_json)
            unit_id = int(params["unit_id"])
            return_location_id = int(params["return_location_id"])
            return_reason = params.get("return_reason") or "Belirtilmedi"
            # Sorunlu parça miktarlarını çöz: {part_id: defective_qty}
            defective_qtys = {}
            for entry in (params.get("defective_parts") or []):
                try:
                    p_id = int(entry.get("part_id"))
                    def_qty = int(entry.get("defective_qty", 0))
                    if def_qty > 0:
                        defective_qtys[p_id] = def_qty
                except (ValueError, TypeError):
                    pass

            # Değişim istenecek parça miktarlarını çöz: {part_id: replacement_qty}
            replacement_qtys = {}
            for entry in (params.get("replacement_parts") or []):
                try:
                    p_id = int(entry.get("part_id"))
                    rep_qty = int(entry.get("replacement_qty", 0))
                    if rep_qty > 0:
                        replacement_qtys[p_id] = rep_qty
                except (ValueError, TypeError):
                    pass

            replacement_qty = max(replacement_qtys.values()) if replacement_qtys else 0
            GOOD_STOCK_ID = _get_system_location_id(db, "good_stock")
            if not GOOD_STOCK_ID:
                result_str = json.dumps({"success": False, "message": "Good Stock deposu bulunamadı."})
                return result_str
            
            # 1. Üretilmiş cihaz kaydını ve bağlı olduğu üretim koşusunu çek
            unit = db.execute(text("""
                SELECT pu.id, pu.serial_number, pu.is_returned,
                       pr.id AS run_id, pr.target_part_id, pr.quantity_produced, pr.location_id, pr.source_location_id
                FROM warehouse.produced_units pu
                JOIN warehouse.production_runs pr ON pr.id = pu.production_run_id
                WHERE pu.id = :uid
            """), {"uid": unit_id}).mappings().first()
            
            if not unit:
                result_str = json.dumps({"success": False, "message": "Üretilen cihaz kaydı bulunamadı."})
                return result_str
                
            if unit["is_returned"]:
                result_str = json.dumps({"success": False, "message": "Bu cihaz zaten iade edilmiş."})
                return result_str
                
            run_id = unit["run_id"]
            target_part_id = unit["target_part_id"]
            quantity_produced = unit["quantity_produced"]
            location_id = unit["location_id"]
            
            # 2. Üretilen parçanın stoğunu kontrol et
            target_stock = db.execute(text("""
                SELECT id, quantity FROM warehouse.stock
                WHERE part_id = :pid AND location_id = :lid
                FOR UPDATE
            """), {"pid": target_part_id, "lid": location_id}).first()
            
            if not target_stock or target_stock[1] < quantity_produced:
                current_qty = target_stock[1] if target_stock else 0
                result_str = json.dumps({
                    "success": False, 
                    "message": f"Üretilen parçanın stoğu yetersiz ({current_qty} adet var, {quantity_produced} adet gerekli). İade gerçekleştirilemez."
                })
                return result_str
                
            # 2.5. Değişim (replacement) talep edildiyse, iadeyi mutasyona başlamadan önce
            # fizibilite kontrolü yap: reçete var mı, gereken hammadde stokta yeterli mi?
            # Yetersizse tüm iade işlemi iptal edilir (hiçbir şey değişmez) — kısmi başarı
            # istenmiyor, ya iade+değişim birlikte gerçekleşir ya da hiçbiri.


            # 2.5. Değişim (replacement) talep edildiyse, hammadde stok kontrolü yap
            replacement_materials = []
            if replacement_qty > 0:
                for part_id, rep_qty in replacement_qtys.items():
                    replacement_materials.append((part_id, rep_qty))

                for part_id, needed in replacement_materials:
                    total_available = db.execute(text("""
                        SELECT COALESCE(SUM(quantity), 0) FROM warehouse.stock WHERE part_id = :pid
                    """), {"pid": part_id}).scalar()
                    if total_available < needed:
                        raise Exception(
                            f"Değişim üretimi için yetersiz stok (parça id {part_id}): mevcut {total_available}, "
                            f"gerekli {needed}. İade işlemi iptal edildi."
                        )

            # 3. Tüketilen malzemeleri çek
            materials = db.execute(text("""
                SELECT part_id, quantity_consumed
                FROM warehouse.production_materials
                WHERE production_run_id = :run_id
            """), {"run_id": run_id}).all()
            
            # 4. Üretilen parçanın stoğunu batch miktarı kadar düş
            db.execute(text("""
                UPDATE warehouse.stock
                SET quantity = quantity - :qty
                WHERE id = :id
            """), {"qty": quantity_produced, "id": target_stock[0]})

            db.add(StockMovement(
                type="Üretim İadesi/İptal",
                movement_kind="Outbound",
                quantity=quantity_produced,
                part_id=target_part_id,
                source_location_id=location_id,
                created_by=None,
                description=f"Üretilen cihaz {unit['serial_number']} iade edildi ({return_reason})"
            ))

            # 5. Her malzemeyi sorunlu/sorunsuz durumuna göre farklı depoya ekle
            returned_mats = []
            for m in materials:
                m_part_id = m[0]
                total_qty = m[1] # Tüm batch için tüketilen miktar
                
                # Sorunlu miktarını al ve sınırla (en fazla total_qty kadar olabilir)
                def_qty = min(total_qty, defective_qtys.get(m_part_id, 0))
                good_qty = total_qty - def_qty
                
                part_row = db.execute(text("SELECT brand, model, name, item_code FROM warehouse.parts WHERE id = :pid"), {"pid": m_part_id}).first()
                part_label = ""
                item_code = ""
                if part_row:
                    part_label = f"{part_row[0] or ''} {part_row[1] or ''}".strip() or (part_row[2] or "")
                    item_code = part_row[3] or ""

                returned_mats.append({
                    "part_id": str(m_part_id),
                    "part_name": part_label,
                    "item_code": item_code,
                    "defective_qty": def_qty,
                    "good_qty": good_qty,
                    "total_qty": total_qty
                })
                
                # 5a. Sorunlu olanları seçilen iade lokasyonuna aktar
                if def_qty > 0:
                    existing_m = db.execute(text("""
                        SELECT id FROM warehouse.stock
                        WHERE part_id = :pid AND location_id = :lid
                        FOR UPDATE
                    """), {"pid": m_part_id, "lid": return_location_id}).first()
                    
                    if existing_m:
                        db.execute(text("""
                            UPDATE warehouse.stock
                            SET quantity = quantity + :qty
                            WHERE id = :id
                        """), {"qty": def_qty, "id": existing_m[0]})
                    else:
                        db.execute(text("""
                            INSERT INTO warehouse.stock (part_id, location_id, quantity)
                            VALUES (:pid, :lid, :qty)
                        """), {"pid": m_part_id, "lid": return_location_id, "qty": def_qty})

                    db.add(StockMovement(
                        type="Üretim İadesi - Sorunlu Malzeme",
                        movement_kind="Inbound",
                        quantity=def_qty,
                        part_id=m_part_id,
                        target_location_id=return_location_id,
                        created_by=None,
                        description=f"Üretilen cihaz {unit['serial_number']} iadesinden sorunlu malzeme"
                    ))

                # 5b. Sorunsuz olanları doğrudan Good Stock'a aktar
                if good_qty > 0:
                    existing_m = db.execute(text("""
                        SELECT id FROM warehouse.stock
                        WHERE part_id = :pid AND location_id = :lid
                        FOR UPDATE
                    """), {"pid": m_part_id, "lid": GOOD_STOCK_ID}).first()
                    
                    if existing_m:
                        db.execute(text("""
                            UPDATE warehouse.stock
                            SET quantity = quantity + :qty
                            WHERE id = :id
                        """), {"qty": good_qty, "id": existing_m[0]})
                    else:
                        db.execute(text("""
                            INSERT INTO warehouse.stock (part_id, location_id, quantity)
                            VALUES (:pid, :lid, :qty)
                        """), {"pid": m_part_id, "lid": GOOD_STOCK_ID, "qty": good_qty})

                    db.add(StockMovement(
                        type="Üretim İadesi - Sorunsuz Malzeme",
                        movement_kind="Inbound",
                        quantity=good_qty,
                        part_id=m_part_id,
                        target_location_id=GOOD_STOCK_ID,
                        created_by=None,
                        description=f"Üretilen cihaz {unit['serial_number']} iadesinden sorunsuz malzeme"
                    ))

            # 5.5. Değişim üretimi: fizibilitesi 2.5'te doğrulanmış replacement_materials'ı
            # tüket, replacement_qty kadar yeni bir üretim partisi (yeni seri no) oluştur.
            # Bu, orijinal iade edilen partiden bağımsız yeni bir production_runs kaydıdır;
            # notes alanı üzerinden hangi iadenin karşılığı olduğu izlenebilir.
            if replacement_qty > 0:
                replacement_consumption_records = []
                for part_id, needed in replacement_materials:
                    remaining = needed
                    rows = db.execute(text("""
                        SELECT id, location_id, quantity FROM warehouse.stock
                        WHERE part_id = :pid AND quantity > 0
                        ORDER BY id
                        FOR UPDATE
                    """), {"pid": part_id}).all()
                    for stock_id, stock_location_id, stock_qty in rows:
                        if remaining <= 0:
                            break
                        take = min(stock_qty, remaining)
                        db.execute(text("UPDATE warehouse.stock SET quantity = quantity - :take WHERE id = :id"),
                                   {"take": take, "id": stock_id})
                        remaining -= take
                        replacement_consumption_records.append((part_id, stock_location_id, take))

                existing_repl = db.execute(text("""
                    SELECT id FROM warehouse.stock WHERE part_id = :pid AND location_id = :lid
                """), {"pid": target_part_id, "lid": GOOD_STOCK_ID}).first()
                if existing_repl:
                    db.execute(text("UPDATE warehouse.stock SET quantity = quantity + :qty WHERE id = :id"),
                               {"qty": replacement_qty, "id": existing_repl[0]})
                else:
                    db.execute(text("""
                        INSERT INTO warehouse.stock (part_id, location_id, quantity) VALUES (:pid, :lid, :qty)
                    """), {"pid": target_part_id, "lid": GOOD_STOCK_ID, "qty": replacement_qty})

                replacement_run_id = db.execute(text("""
                    INSERT INTO warehouse.production_runs (target_part_id, quantity_produced, source_location_id, location_id, produced_by, notes)
                    VALUES (:tgt, :qty, :slid, :tlid, :by, :notes) RETURNING id
                """), {
                    "tgt": target_part_id, "qty": replacement_qty, "slid": GOOD_STOCK_ID, "tlid": GOOD_STOCK_ID,
                    "by": None, "notes": f"'{unit['serial_number']}' iadesi icin otomatik degisim uretimi (neden: {return_reason})"
                }).scalar()

                replacement_next_id = db.execute(text("SELECT nextval(pg_get_serial_sequence('warehouse.produced_units', 'id'))")).scalar()
                replacement_serial = f"{replacement_next_id:015d}"
                db.execute(text("""
                    INSERT INTO warehouse.produced_units (id, production_run_id, serial_number)
                    VALUES (:id, :run_id, :serial)
                """), {"id": replacement_next_id, "run_id": replacement_run_id, "serial": replacement_serial})

                for part_id, needed in replacement_materials:
                    db.execute(text("""
                        INSERT INTO warehouse.production_materials (production_run_id, part_id, quantity_consumed)
                        VALUES (:run_id, :pid, :qty)
                    """), {"run_id": replacement_run_id, "pid": part_id, "qty": needed})

                for part_id, stock_location_id, take in replacement_consumption_records:
                    db.add(StockMovement(
                        type="Değişim Üretimi İçin Malzeme Tüketimi",
                        movement_kind="Outbound",
                        quantity=take,
                        part_id=part_id,
                        source_location_id=stock_location_id,
                        created_by=None,
                        description=f"Değişim Üretimi #{replacement_run_id} ({replacement_serial}) için tüketildi — iade edilen {unit['serial_number']} yerine"
                    ))
                db.add(StockMovement(
                    type="Değişim Üretimi",
                    movement_kind="Inbound",
                    quantity=replacement_qty,
                    part_id=target_part_id,
                    target_location_id=GOOD_STOCK_ID,
                    created_by=None,
                    description=f"Değişim Üretimi #{replacement_run_id} ({replacement_serial}) — iade edilen {unit['serial_number']} yerine üretildi"
                ))

            # 6. Cihaz kaydını iade edildi olarak işaretle ve nedenini kaydet
            db.execute(text("""
                UPDATE warehouse.produced_units
                SET is_returned = TRUE,
                    return_reason = :reason,
                    returned_at = :now,
                    return_location_id = :ret_loc_id,
                    returned_materials = :returned_mats,
                    replacement_requested_qty = :replacement_qty
                WHERE id = :uid
            """), {
                "reason": return_reason,
                "now": datetime.utcnow(),
                "ret_loc_id": return_location_id,
                "uid": unit_id,
                "returned_mats": json.dumps(returned_mats),
                "replacement_qty": replacement_qty
            })
            
            db.commit()
            result_str = json.dumps({"success": True, "message": "Üretim iade/değişim işlemi başarıyla tamamlandı."})
            return result_str
        except Exception as e:
            try:
                db.rollback()
            except Exception:
                pass
            result_str = json.dumps({"success": False, "message": f"İade hatası: {str(e)}"})
            return result_str
        finally:
            try:
                db.close()
            except Exception:
                pass

    # --- YENİ EKLENEN ÜRÜN (TELEFON) VE TEDARİKÇİ FONKSİYONLARI ---
    # Products verileri artik kendi 'products' tablosundan cekiliyor (parts'tan bagimsiz).

    @Slot(str, str, str, str, str, str, result=str)
    def get_products(self, page="1", page_size="50", search_term="", category_filter="", sort_key="", sort_dir=""):
        import re
        from sqlalchemy import text
        from core.product_code_generator import normalize_brand
        db = SessionLocal()
        try:
            page = max(1, int(page or 1))
            limit = min(1000, max(1, int(page_size or 50)))

            where_clauses = ["pm.enabled = TRUE"]
            params = {}

            if search_term and str(search_term).strip():
                where_clauses.append("(pmf.code ILIKE :search OR pm.short_name ILIKE :search OR b.short_name ILIKE :search)")
                params["search"] = f"%{str(search_term).strip()}%"

            # ILIKE Smart Phone check (Esnek)
            if category_filter and str(category_filter).strip():
                where_clauses.append("(pmf.short_name ILIKE :category OR :category ILIKE '%phone%' OR :category ILIKE '%telefon%' AND pmf.short_name ILIKE '%phone%')")
                params["category"] = f"%{str(category_filter).strip()}%"

            where_sql = "WHERE " + " AND ".join(where_clauses)

            # product_model.product_family serbest metin bir kolon: pf.code (kısaltma,
            # örn. iP12PM) veya pf.short_name (tam ad) ile eşleşebilir - LOWER() ile
            # her iki ihtimali de büyük/küçük harf duyarsız kontrol eder. Aynı isme
            # sahip birden çok aile satırı olabileceğinden LATERAL + LIMIT 1 ile tekilleştirilir,
            # code eşleşmesi short_name eşleşmesine tercih edilir.
            family_join = """
                LEFT JOIN LATERAL (
                    SELECT pf.code, pf.short_name, pf.brand
                    FROM warehouse.product_family pf
                    WHERE LOWER(pf.code) = LOWER(pm.product_family)
                       OR LOWER(pf.short_name) = LOWER(pm.product_family)
                    ORDER BY (LOWER(pf.code) = LOWER(pm.product_family)) DESC
                    LIMIT 1
                ) pmf ON TRUE
            """

            # warehouse.brand: marka basina tek, temiz gorunum adi (Samsung, Xiaomi, ...).
            # product_model.brand her zaman NULL oldugundan gercek marka bilgisini
            # product_family.brand'dan (serbest metin, ayni markanin 20 farkli
            # buyuk/kucuk harf varyasyonu var) normalize_brand ile okuyup bu tabloya eslistiriyoruz.
            brand_rows = db.execute(text("SELECT code, short_name FROM warehouse.brand")).fetchall()
            brand_display = {}
            for br in brand_rows:
                key = normalize_brand(br.short_name or br.code)
                brand_display[key] = br.short_name or br.code

            # pm.short_name = "{model adi} {hafiza}GB/TB" seklinde. Once TUM eslesen
            # satirlari cekip Python'da model (hafizasiz) bazinda grupluyoruz, boylece
            # ayni telefonun farkli hafiza secenekleri tek satirda, Hafiza sutununda
            # yan yana gosterilebiliyor. Veri boyutu (~1000 satir) SQL tarafinda
            # aggregate etmeye gerek birakmayacak kadar kucuk.
            data_sql = f"""
                SELECT pm.id, pm.short_name,
                       COALESCE(b.short_name, '') as legacy_brand_name,
                       pmf.brand as family_brand,
                       COALESCE(pmf.short_name, 'Bilinmiyor') as family_name,
                       pmf.code as family_code
                FROM warehouse.product_model pm
                LEFT JOIN warehouse.brand b ON b.id::text = pm.brand OR b.short_name = pm.brand
                {family_join}
                {where_sql}
                ORDER BY pm.short_name ASC;
            """
            rows = db.execute(text(data_sql), params).fetchall()

            mem_re = re.compile(r"\s*(\d+\s*(?:GB|TB))\s*$", re.IGNORECASE)

            def split_memory(short_name):
                m = mem_re.search(short_name or "")
                if not m:
                    return (short_name or "").strip(), None
                mem = m.group(1).replace(" ", "").upper()
                model_no_mem = short_name[: m.start()].strip()
                return model_no_mem, mem

            def memory_sort_key(mem):
                mm = re.match(r"(\d+)(GB|TB)", mem)
                if not mm:
                    return 0
                n, unit = mm.groups()
                return int(n) * (1024 if unit == "TB" else 1)

            groups = {}
            order = []
            for row in rows:
                model_no_mem, mem = split_memory(row.short_name)
                display_model = row.family_name if row.family_name != "Bilinmiyor" else (model_no_mem or row.short_name)
                brand_name = brand_display.get(normalize_brand(row.family_brand)) or row.legacy_brand_name or "Bilinmiyor"
                group_key = (brand_name, row.family_code or display_model)
                if group_key not in groups:
                    groups[group_key] = {
                        "id": str(row.id),
                        "item_code": row.family_code or "",
                        "brand": brand_name,
                        "model": display_model,
                        "category": row.family_name,
                        "memories": [],
                        "color": ""
                    }
                    order.append(group_key)
                if mem and mem not in groups[group_key]["memories"]:
                    groups[group_key]["memories"].append(mem)

            grouped = []
            for key in order:
                g = groups[key]
                g["memories"].sort(key=memory_sort_key)
                g["memory"] = ", ".join(g["memories"])
                del g["memories"]
                grouped.append(g)

            total = len(grouped)
            offset = (page - 1) * limit
            res = grouped[offset: offset + limit]

            # Return format requested by user { data: [...], total: N }
            return json.dumps({
                "success": True,
                "data": res,
                "total": total,
                "page": page,
                "limit": limit
            }, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, result=str)
    def create_product(self, item_code, brand, model, memory, color, name):
        from models.product import Product
        db = SessionLocal()
        try:
            product = Product(
                item_code=item_code or None,
                brand=brand,
                model=model,
                memory=memory,
                color=color
            )
            db.add(product)
            db.commit()
            return json.dumps({"success": True, "id": product.id})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def bulk_import_products(self, rows_json):
        """Toplu (Excel) ürün içe aktarma. create_product'ın tek satırlık eşdeğeriyle aynı
        şekilde diğer alanları zorunlu tutmaz, ama create_product'ın atladığı item_code
        unique kısıtını (tek satırlık akışta hata sessizce yutulup kullanıcıya hiç
        gösterilmiyordu) burada satır numarasıyla birlikte raporlar. N ayrı çağrı/commit
        yerine tüm satırları tek transaction'da yazar."""
        from sqlalchemy import text
        from models.product import Product
        db = SessionLocal()
        try:
            try:
                rows = json.loads(rows_json or "[]")
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Geçersiz dosya verisi.", "errors": []})

            if not rows:
                return json.dumps({"success": False, "message": "Dosyada içe aktarılacak satır bulunamadı.", "errors": []})

            existing_codes = {r[0] for r in db.execute(text(
                "SELECT item_code FROM warehouse.products WHERE item_code IS NOT NULL"
            )).all()}

            errors = []
            seen_codes_in_file = {}
            valid_rows = []

            for idx, row in enumerate(rows):
                row_num = idx + 2
                row = row or {}

                def get_val(key):
                    v = row.get(key)
                    return str(v).strip() if v is not None else ""

                item_code = get_val("item_code")
                if item_code:
                    if item_code in existing_codes:
                        errors.append({"row": row_num, "field": "Ürün Kodu", "message": f"\"{item_code}\" kodlu ürün zaten sistemde kayıtlı."})
                    elif item_code in seen_codes_in_file:
                        errors.append({"row": row_num, "field": "Ürün Kodu", "message": f"\"{item_code}\" dosyada birden fazla satırda tekrarlanıyor (satır {seen_codes_in_file[item_code]})."})
                    else:
                        seen_codes_in_file[item_code] = row_num

                valid_rows.append({
                    "item_code": item_code or None,
                    "brand": get_val("brand"),
                    "model": get_val("model"),
                    "memory": get_val("memory"),
                    "color": get_val("color"),
                })

            if errors:
                return json.dumps({"success": False, "message": f"{len(errors)} hata bulundu, hiçbir satır içe aktarılmadı.", "errors": errors})

            for r in valid_rows:
                db.add(Product(**r))

            db.commit()
            return json.dumps({"success": True, "message": f"{len(valid_rows)} ürün başarıyla içe aktarıldı.", "imported": len(valid_rows)})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"İçe aktarma hatası: {str(e)}", "errors": []})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, result=str)
    def update_product(self, product_id_str, item_code, brand, model, memory, color, name):
        from models.product import Product
        db = SessionLocal()
        try:
            product_id = int(product_id_str)
            product = db.query(Product).filter(Product.id == product_id).first()
            if not product:
                return json.dumps({"success": False, "message": "Ürün bulunamadı"})

            product.item_code = item_code or None
            product.brand = brand
            product.model = model
            product.memory = memory
            product.color = color

            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_product(self, product_id_str):
        from models.product import Product
        db = SessionLocal()
        try:
            product_id = int(product_id_str)
            product = db.query(Product).filter(Product.id == product_id).first()
            if not product:
                return json.dumps({"success": False, "message": "Ürün bulunamadı"})
            db.delete(product)
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()


    @Slot(result=str)
    def get_customers(self):
        """Müşteriler sayfası için: warehouse.customers tablosundaki tüm kayıtlar
        (parts tablosundan tamamen bağımsız, gerçek bir müşteri/cihaz kabul tablosu)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT id, customer_name, customer_phone, customer_email, company,
                       imei_number, serial_number, internal_id, brand, model, product_code,
                       flow, customer_reported_complaint, intake_date, created_at,
                       code, short_name, currency, customer_language, use_mio
                FROM warehouse.customers
                ORDER BY id DESC
                LIMIT 500
            """)).mappings().all()
            customers = []
            for r in rows:
                customers.append({
                    "id": str(r["id"]),
                    "customer_name": r["customer_name"] or "",
                    "customer_phone": r["customer_phone"] or "",
                    "customer_email": r["customer_email"] or "",
                    "company": r["company"] or "",
                    "imei_number": r["imei_number"] or "",
                    "serial_number": r["serial_number"] or "",
                    "internal_id": r["internal_id"] or "",
                    "brand": r["brand"] or "",
                    "model": r["model"] or "",
                    "product_code": r["product_code"] or "",
                    "flow": r["flow"] or "",
                    "customer_reported_complaint": r["customer_reported_complaint"] or "",
                    "intake_date": r["intake_date"].strftime("%Y-%m-%d") if r["intake_date"] else "",
                    "created_at": r["created_at"].strftime("%Y-%m-%d %H:%M") if r["created_at"] else "",
                    "code": r["code"] or "",
                    "short_name": r["short_name"] or "",
                    "currency": r["currency"] or "",
                    "customer_language": r["customer_language"] or "",
                    "use_mio": bool(r["use_mio"])
                })
            return json.dumps({"success": True, "customers": customers})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, result=str)
    def create_customer(self, customer_name, customer_phone, customer_email, company,
                         imei_number, serial_number, internal_id, cihaz_modeli, flow,
                         customer_reported_complaint, intake_date,
                         code, short_name, currency, customer_language, use_mio):
        """Yeni bir müşteri/cihaz kabul kaydı ekler (manuel tek-kayıt formu)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            name = (customer_name or "").strip()
            if not name:
                return json.dumps({"success": False, "message": "Müşteri adı zorunludur."})

            product = None
            if cihaz_modeli and cihaz_modeli.strip():
                product = db.execute(text("""
                    SELECT brand, model, item_code FROM warehouse.products
                    WHERE LOWER(TRIM(brand || ' ' || model)) = LOWER(:cm) LIMIT 1
                """), {"cm": cihaz_modeli.strip()}).mappings().first()

            db.execute(text("""
                INSERT INTO warehouse.customers (
                    customer_name, customer_phone, customer_email, company,
                    imei_number, serial_number, internal_id, brand, model, product_code,
                    flow, customer_reported_complaint, intake_date,
                    code, short_name, currency, customer_language, use_mio
                ) VALUES (
                    :name, :phone, :email, :company,
                    :imei, :serial, :internal_id, :brand, :model, :product_code,
                    :flow, :complaint, :intake_date,
                    :code, :short_name, :currency, :customer_language, :use_mio
                )
            """), {
                "name": name, "phone": customer_phone or None, "email": customer_email or None,
                "company": company or None,
                "imei": imei_number or None, "serial": serial_number or None,
                "internal_id": internal_id or None,
                "brand": product["brand"] if product else None,
                "model": product["model"] if product else None,
                "product_code": product["item_code"] if product else None,
                "flow": flow or None, "complaint": customer_reported_complaint or None,
                "intake_date": intake_date or None,
                "code": code or None, "short_name": short_name or None,
                "currency": currency or None, "customer_language": customer_language or None,
                "use_mio": use_mio == "true"
            })
            db.commit()
            return json.dumps({"success": True, "message": "Müşteri kaydı eklendi."})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, str, result=str)
    def update_customer(self, customer_id_str, customer_name, customer_phone, customer_email, company,
                         imei_number, serial_number, internal_id, cihaz_modeli, flow,
                         customer_reported_complaint, intake_date,
                         code, short_name, currency, customer_language, use_mio):
        """Var olan bir müşteri/cihaz kabul kaydını günceller."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            customer_id = int(customer_id_str)
            name = (customer_name or "").strip()
            if not name:
                return json.dumps({"success": False, "message": "Müşteri adı zorunludur."})

            product = None
            if cihaz_modeli and cihaz_modeli.strip():
                product = db.execute(text("""
                    SELECT brand, model, item_code FROM warehouse.products
                    WHERE LOWER(TRIM(brand || ' ' || model)) = LOWER(:cm) LIMIT 1
                """), {"cm": cihaz_modeli.strip()}).mappings().first()

            db.execute(text("""
                UPDATE warehouse.customers
                SET customer_name = :name, customer_phone = :phone, customer_email = :email, company = :company,
                    imei_number = :imei, serial_number = :serial, internal_id = :internal_id,
                    brand = :brand, model = :model, product_code = :product_code,
                    flow = :flow, customer_reported_complaint = :complaint, intake_date = :intake_date,
                    code = :code, short_name = :short_name, currency = :currency,
                    customer_language = :customer_language, use_mio = :use_mio
                WHERE id = :id
            """), {
                "name": name, "phone": customer_phone or None, "email": customer_email or None,
                "company": company or None,
                "imei": imei_number or None, "serial": serial_number or None,
                "internal_id": internal_id or None,
                "brand": product["brand"] if product else None,
                "model": product["model"] if product else None,
                "product_code": product["item_code"] if product else None,
                "flow": flow or None, "complaint": customer_reported_complaint or None,
                "intake_date": intake_date or None,
                "code": code or None, "short_name": short_name or None,
                "currency": currency or None, "customer_language": customer_language or None,
                "use_mio": use_mio == "true",
                "id": customer_id
            })
            db.commit()
            return json.dumps({"success": True, "message": "Müşteri kaydı güncellendi."})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_customer(self, customer_id_str):
        """Belirtilen id'ye sahip müşteri kaydını siler."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            customer_id = int(customer_id_str)
            db.execute(text("DELETE FROM warehouse.customers WHERE id = :id"), {"id": customer_id})
            db.commit()
            return json.dumps({"success": True, "message": "Müşteri kaydı silindi."})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # ==========================
    # STOK & DEPO YÖNETİMİ
    # ==========================

    @Slot(result=str)
    def get_stock_status(self):
        filename = "stock.json"
        path = os.path.join(get_cache_dirs()[0], filename)
        fetch_url = f"/api_cache/{filename}"
        if os.path.exists(path):
            return json.dumps({"success": True, "fetch_url": fetch_url})

        from sqlalchemy import text
        db = SessionLocal()
        try:
            # warehouse.stock ~30 bin satır olabildiğinden, her satır için ayrı ayrı
            # warehouse.stock_movements'a bakan correlated subquery yerine, movements
            # tek seferde (part_id, location_id) bazında önceden MAX(created_at)'e
            # indirgenip tek bir LEFT JOIN ile eşleştirilir - stock_movements büyüdükçe
            # correlated subquery'nin aksine ölçeklenir.
            stocks = db.execute(text("""
                WITH movement_touch AS (
                    SELECT part_id, source_location_id AS location_id, created_at
                    FROM warehouse.stock_movements WHERE source_location_id IS NOT NULL
                    UNION ALL
                    SELECT part_id, target_location_id AS location_id, created_at
                    FROM warehouse.stock_movements WHERE target_location_id IS NOT NULL
                ),
                last_movement AS (
                    SELECT part_id, location_id, MAX(created_at) AS last_movement_at
                    FROM movement_touch
                    GROUP BY part_id, location_id
                )
                SELECT s.id, p.id as part_id, p.brand, p.model, p.color, p.part_category, p.name as pname, p.item_code,
                       l.id as location_id, l.name as location_name, l.kind as location_kind,
                       s.quantity, p.critical_limit,
                       lm.last_movement_at
                FROM warehouse.stock s
                JOIN warehouse.parts p ON s.part_id = p.id
                JOIN warehouse.locations l ON s.location_id = l.id
                LEFT JOIN last_movement lm ON lm.part_id = s.part_id AND lm.location_id = s.location_id
                ORDER BY s.id DESC
            """)).mappings().all()
            res = []
            for row in stocks:
                lm_at = row.get("last_movement_at")
                date_str = lm_at.strftime("%d.%m.%Y %H:%M") if lm_at else "-"
                part_name = " ".join(filter(None, [row.get("brand"), row.get("model"), row.get("color"), row.get("part_category")]))
                if not part_name:
                    part_name = (row.get("pname") or "").strip()
                if not part_name:
                    part_name = row.get("item_code") or "İsimsiz Parça"
                    
                res.append({
                    "id": row["id"],
                    "part_id": row["part_id"],
                    "item_code": row["item_code"] or "-",
                    "brand": row.get("brand") or "",
                    "model": row.get("model") or "",
                    "part_name": part_name,
                    "location_id": row["location_id"],
                    "location_name": row["location_name"],
                    "location_kind": row["location_kind"],
                    "quantity": row["quantity"],
                    "critical_limit": row["critical_limit"] or 50,
                    "updated_at": date_str,
                    "date": date_str
                })
            json_data = json.dumps({"success": True, "stock": res})
            write_to_cache("stock.json", json_data)
            return json.dumps({"success": True, "fetch_url": fetch_url})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_stock_for_part(self, part_id):
        """Tek bir parçanın tüm lokasyonlardaki stok miktarlarını döner - warehouse.stock
        ~30 bin satır olduğundan, İrsaliye ekranı gibi sadece TEK bir parçanın stok
        durumuna ihtiyaç duyan yerlerin get_stock_status() ile TÜM tabloyu indirmesi
        gerekmesin diye eklendi (bkz. idx_stock_part_location)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            pid = int(part_id)
            rows = db.execute(text("""
                SELECT s.location_id, l.name AS location_name, l.kind AS location_kind, s.quantity
                FROM warehouse.stock s
                JOIN warehouse.locations l ON l.id = s.location_id
                WHERE s.part_id = :pid
            """), {"pid": pid}).mappings().all()
            stock = [{"part_id": pid, "location_id": r["location_id"], "location_name": r["location_name"], "location_kind": r["location_kind"], "quantity": r["quantity"]} for r in rows]
            return json.dumps({"success": True, "stock": stock}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_stock_by_item_code(self, item_code):
        """Bir warehouse.item/parts kodunun (repair_records.part_item_code) Good Stock
        depodaki toplam miktarını döner. Onarım Parçaları ekranındaki Depo Durum/Depo
        Parça sütunlarının 'depoda stok var mı' kontrolünün kaynağıdır - idx_stock_part_location
        ve idx_batch_entries benzeri, tek parçaya odaklı, tüm stok tablosunu indirmeyen bir sorgu."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            item_code = (item_code or "").strip()
            if not item_code:
                return json.dumps({"success": True, "item_code": item_code, "quantity": 0})
            row = db.execute(text("""
                SELECT COALESCE(SUM(s.quantity), 0) AS qty
                FROM warehouse.parts p
                JOIN warehouse.stock s ON s.part_id = p.id
                JOIN warehouse.locations l ON l.id = s.location_id
                WHERE p.item_code = :item_code AND l.kind = 'good_stock'
            """), {"item_code": item_code}).mappings().first()
            qty = int(row["qty"]) if row else 0
            return json.dumps({"success": True, "item_code": item_code, "quantity": qty}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def get_stock_status_paged(self, search, page, page_size):
        """Depo sayfası için sunucu taraflı arama + sayfalama. Sadece Good Stock
        deposundaki kayıtları döndürür; büyük stok tablosunun tamamını istemciye
        indirmeden sadece görünen sayfayı çeker."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            page = max(1, int(page) if str(page).isdigit() else 1)
            page_size = min(200, max(1, int(page_size) if str(page_size).isdigit() else 30))
            offset = (page - 1) * page_size
            search = (search or "").strip()

            params = {"limit": page_size, "offset": offset}
            search_clause = ""
            if search:
                search_clause = """
                    AND (
                        p.item_code ILIKE :q OR p.name ILIKE :q OR p.brand ILIKE :q OR
                        p.model ILIKE :q OR p.color ILIKE :q OR p.part_category ILIKE :q OR
                        l.name ILIKE :q OR CAST(s.id AS TEXT) ILIKE :q
                    )
                """
                params["q"] = f"%{search}%"

            rows = db.execute(text(f"""
                SELECT s.id, p.id as part_id, p.brand, p.model, p.color, p.part_category, p.name as pname, p.item_code,
                       l.name as location_name, s.quantity, p.critical_limit,
                       (
                         SELECT MAX(sm.created_at)
                         FROM warehouse.stock_movements sm
                         WHERE sm.part_id = s.part_id AND (sm.source_location_id = s.location_id OR sm.target_location_id = s.location_id)
                       ) as last_movement_at,
                       COUNT(*) OVER() as total_count,
                       COALESCE(SUM(s.quantity) OVER(), 0) as total_qty
                FROM warehouse.stock s
                JOIN warehouse.parts p ON s.part_id = p.id
                JOIN warehouse.locations l ON s.location_id = l.id
                WHERE l.kind = 'good_stock'
                {search_clause}
                ORDER BY s.id DESC
                LIMIT :limit OFFSET :offset
            """), params).mappings().all()

            res = []
            total_count = 0
            total_qty = 0
            for row in rows:
                total_count = row["total_count"]
                total_qty = row["total_qty"]
                lm_at = row.get("last_movement_at")
                date_str = lm_at.strftime("%d.%m.%Y %H:%M") if lm_at else "-"
                part_name = " ".join(filter(None, [row.get("brand"), row.get("model"), row.get("color"), row.get("part_category")]))
                if not part_name:
                    part_name = (row.get("pname") or "").strip()
                if not part_name:
                    part_name = row.get("item_code") or "İsimsiz Parça"

                res.append({
                    "id": row["id"],
                    "part_id": row["part_id"],
                    "item_code": row["item_code"] or "-",
                    "part_name": part_name,
                    "location_name": row["location_name"],
                    "quantity": row["quantity"],
                    "critical_limit": row["critical_limit"] or 50,
                    "updated_at": date_str
                })

            return json.dumps({
                "success": True,
                "stock": res,
                "total": total_count,
                "total_quantity": int(total_qty or 0),
                "page": page,
                "page_size": page_size
            })
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, result=str)
    def transfer_stock(self, part_id, from_loc_id, to_loc_id, qty, username):
        from models.stock import Stock
        from models.stock_movement import StockMovement
        from models.location import Location
        db = SessionLocal()
        try:
            # GİRDİ DOĞRULAMASI - burası köprüye bağlanan herkese açıktır ve
            # @Slot metodlarında rol kontrolü yoktur; ekranın doğrulamasına
            # güvenilemez (bkz. aşağıdaki iki hata).
            try:
                qty = int(qty)
                from_id, to_id = int(from_loc_id), int(to_loc_id)
                part_id = int(part_id)
            except (TypeError, ValueError):
                return json.dumps({"success": False, "message": "Geçersiz parça, depo veya miktar bilgisi."})

            # 1) Miktar pozitif olmalı. Kontrolsüz negatif miktar YOKTAN STOK YARATIYORDU:
            #    "source_stock.quantity < qty" negatif qty'de daima False döndüğü için
            #    yetersiz stok kontrolü geçiliyor, ardından kaynaktan çıkarma işlemi
            #    (quantity -= -5) kaynağı ARTIRIYOR, hedefe ekleme (quantity += -5)
            #    hedefi eksiye düşürüyordu.
            if qty <= 0:
                return json.dumps({"success": False, "message": "Transfer miktarı sıfırdan büyük olmalıdır."})

            # 2) Kaynak ve hedef aynı olamaz. Stok bozulmuyordu (aynı satır önce eksilip
            #    sonra artıyor) ama stock_movements denetim defterine gerçekte hiçbir şey
            #    taşımayan bir "İç Transfer" satırı yazılıyordu.
            if from_id == to_id:
                return json.dumps({"success": False, "message": "Kaynak ve hedef depo aynı olamaz."})

            locs = db.query(Location).filter(Location.id.in_([from_id, to_id])).all()
            loc_by_id = {l.id: l for l in locs}

            # 3) Her iki depo da gerçekten var olmalı. Kaynak sistem deposu değilse
            #    kural denetimi çalışmadığından, olmayan bir hedefe transfer sessizce
            #    öksüz bir stok satırı oluşturuyordu.
            eksik = [str(i) for i in (from_id, to_id) if i not in loc_by_id]
            if eksik:
                return json.dumps({"success": False,
                                   "message": f"Lokasyon bulunamadı: {', '.join(eksik)}"})

            from_kind = loc_by_id[from_id].kind
            to_kind = loc_by_id[to_id].kind

            if from_kind in SYSTEM_TRANSFER_RULES:
                allowed_targets = SYSTEM_TRANSFER_RULES[from_kind]
                if to_kind not in allowed_targets:
                    from_label = SYSTEM_LOCATION_KINDS.get(from_kind, from_kind)
                    if allowed_targets:
                        allowed_labels = " veya ".join(SYSTEM_LOCATION_KINDS.get(k, k) for k in allowed_targets)
                        message = f"{from_label}'tan sadece {allowed_labels} deposuna transfer yapılabilir."
                    else:
                        message = f"{from_label} sadece çıkış deposudur, buradan başka bir depoya transfer yapılamaz."
                    return json.dumps({"success": False, "message": message})

            source_stock = db.query(Stock).with_for_update().filter(Stock.part_id == part_id, Stock.location_id == from_id).first()
            if not source_stock or source_stock.quantity < qty:
                return json.dumps({"success": False, "message": "Yetersiz stok veya lokasyon bulunamadı."})

            source_stock.quantity -= qty

            target_stock = db.query(Stock).with_for_update().filter(Stock.part_id == part_id, Stock.location_id == to_id).first()
            if target_stock:
                target_stock.quantity += qty
            else:
                target_stock = Stock(part_id=part_id, location_id=to_id, quantity=qty)
                db.add(target_stock)

            from_name = loc_by_id[from_id].name
            to_name = loc_by_id[to_id].name

            movement = StockMovement(
                type="İç Transfer",
                movement_kind="Transfer",
                quantity=qty,
                part_id=part_id,
                source_location_id=from_id,
                target_location_id=to_id,
                created_by=username,
                description=f"Stok Transferi: {from_name} -> {to_name}"
            )
            db.add(movement)
            db.commit()
            clear_api_cache()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    def _query_stock_movements(self, db, mov_type, limit):
        from models.stock_movement import StockMovement
        from models.part import Part
        from models.location import Location
        from sqlalchemy.orm import aliased
        SourceLoc = aliased(Location)
        TargetLoc = aliased(Location)
        query = db.query(StockMovement, Part, SourceLoc, TargetLoc) \
            .outerjoin(Part, StockMovement.part_id == Part.id) \
            .outerjoin(SourceLoc, StockMovement.source_location_id == SourceLoc.id) \
            .outerjoin(TargetLoc, StockMovement.target_location_id == TargetLoc.id)

        if mov_type == 'in':
            query = query.filter(StockMovement.type.in_(["Giriş", "İç Transfer", "Yeni Alım", "Inbound", "Transfer"]))
        elif mov_type == 'out':
            query = query.filter(StockMovement.type.in_(["Çıkış", "İç Transfer", "Müşteri Satışı", "Tedarikçiye İade", "Outbound", "Transfer"]))

        results = query.order_by(StockMovement.created_at.desc()).limit(limit).all()

        res = []
        for mov, p, sloc, tloc in results:
            source_name = sloc.name if sloc else None
            target_name = tloc.name if tloc else None

            if not source_name:
                if "İade" in mov.type and "İptal" not in mov.type:
                    source_name = "Good Stock"
                elif "İptali" in mov.type:
                    source_name = "Good Stock"
                elif mov.type == "Giriş":
                    source_name = "Dış Kaynak"
                else:
                    source_name = "Bilinmiyor"

            if not target_name:
                if "Çıkış" in mov.type or "Tüketimi" in mov.type or ("İptal" in mov.type and "İptali" not in mov.type):
                    target_name = "Kullanım/Tüketim"
                elif mov.type == "Çıkış":
                    target_name = "Dış Kaynak"
                else:
                    target_name = "Bilinmiyor"
            res.append({
                "id": mov.id,
                "type": mov.type,
                "quantity": mov.quantity,
                "part_id": mov.part_id,
                "part_name": p.name if p else (f"{mov.part_name_snapshot} (silindi)" if mov.part_name_snapshot else "Silinmiş Parça"),
                "source_location_id": mov.source_location_id,
                "source_location": source_name,
                "target_location_id": mov.target_location_id,
                "target_location": target_name,
                "created_by": mov.created_by,
                "technician": mov.technician or "-",
                "description": mov.description or "-",
                "unit_price": float(mov.unit_price) if mov.unit_price else None,
                "created_at": mov.created_at.strftime("%Y-%m-%d %H:%M") if mov.created_at else ""
            })
        return res

    @Slot(str, result=str)
    def get_stock_movements(self, mov_type):
        # mov_type can be 'in' or 'out' or 'all'
        db = SessionLocal()
        try:
            res = self._query_stock_movements(db, mov_type, 200)
            return json.dumps({"success": True, "movements": res})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(int, result=str)
    def get_recent_stock_movements(self, limit):
        """Dashboard'daki 'Son Hareketler' widget'ı için - Dashboard eskiden sadece
        ilk 5'ini göstermek üzere get_stock_movements('all') ile 200 satırlık tam
        JOIN'i çekip 200 satırın source/target ismini çözüyordu, sonra JS'te 5'e
        kırpıyordu. Bu, ihtiyaç duyulan satır sayısıyla sınırlı ayrı bir Slot."""
        db = SessionLocal()
        try:
            limit = max(1, min(50, int(limit or 5)))
            res = self._query_stock_movements(db, 'all', limit)
            return json.dumps({"success": True, "movements": res})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, result=str)
    def add_inbound_entry(self, part_id, location_id, qty, unit_price, type_str, username):
        from models.stock import Stock
        from models.stock_movement import StockMovement
        from models.location import Location
        db = SessionLocal()
        try:
            part_id = int(part_id)
            qty = int(qty)
            price = float(unit_price) if unit_price else 0.0

            # Stok Girişleri HER ZAMAN Good Stock deposuna yapılır
            target_loc = db.query(Location).filter(Location.kind == "good_stock").first()
            if not target_loc:
                return json.dumps({"success": False, "message": "Good Stock deposu bulunamadı."})
            
            location_id = target_loc.id

            stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == location_id).first()
            if stock:
                stock.quantity += qty
            else:
                stock = Stock(part_id=part_id, location_id=location_id, quantity=qty)
                db.add(stock)

            mov = StockMovement(
                type=type_str or "Giriş",
                movement_kind="Inbound",
                quantity=qty,
                part_id=part_id,
                target_location_id=location_id,
                unit_price=price,
                total_cost=qty * price,
                created_by=username
            )
            db.add(mov)
            db.commit()
            clear_api_cache()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def bulk_import_inbound_entries(self, rows_json, username):
        """Toplu (Excel) İrsaliye Giriş içe aktarma. add_inbound_entry'nin tek satırlık
        eşdeğeriyle aynı kuralları (barkod/miktar/kim zorunlu, barkoda göre parça eşleşmesi,
        giriş her zaman Good Stock deposuna) kullanır; ancak tüm satırları önce doğrulayıp
        HERHANGİ bir hata varsa hiçbir satırı kaydetmeyen all-or-nothing akışa çevrilmiştir
        (eskiden hatalı satırlar sessizce atlanıp devam ediliyordu, kullanıcı neyin neden
        atlandığını göremiyordu). Aynı barkod birden fazla satırda geçiyorsa miktarlar
        Good Stock'a tek seferde toplanarak eklenir, ama her satır için ayrı bir hareket
        (StockMovement) kaydı korunur."""
        from sqlalchemy import text
        from models.stock import Stock
        from models.stock_movement import StockMovement
        from models.location import Location
        db = SessionLocal()
        try:
            try:
                rows = json.loads(rows_json or "[]")
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Geçersiz dosya verisi.", "errors": []})

            if not rows:
                return json.dumps({"success": False, "message": "Dosyada içe aktarılacak satır bulunamadı.", "errors": []})

            target_loc = db.query(Location).filter(Location.kind == "good_stock").first()
            if not target_loc:
                return json.dumps({"success": False, "message": "Good Stock deposu bulunamadı.", "errors": []})

            parts_by_barcode = {
                str(r[0]): r[1] for r in db.execute(text(
                    "SELECT barcode, id FROM warehouse.parts WHERE barcode IS NOT NULL"
                )).all()
            }

            errors = []
            valid_rows = []

            for idx, row in enumerate(rows):
                row_num = idx + 2
                row = row or {}

                def get_val(key):
                    v = row.get(key)
                    return str(v).strip() if v is not None else ""

                barcode = get_val("barcode")
                qty_raw = get_val("qty")
                who = get_val("who")

                if not barcode:
                    errors.append({"row": row_num, "field": "Barkod", "message": "Barkod boş olamaz."})
                if not qty_raw:
                    errors.append({"row": row_num, "field": "Miktar", "message": "Miktar boş olamaz."})
                if not who:
                    errors.append({"row": row_num, "field": "Kim", "message": "Kim (işlemi yapan) boş olamaz."})

                part_id = parts_by_barcode.get(barcode) if barcode else None
                if barcode and not part_id:
                    errors.append({"row": row_num, "field": "Barkod", "message": f"Barkodu \"{barcode}\" olan parça bulunamadı."})

                qty = None
                if qty_raw:
                    try:
                        qty = int(float(qty_raw))
                        if qty <= 0:
                            errors.append({"row": row_num, "field": "Miktar", "message": "Miktar 0'dan büyük olmalıdır."})
                    except ValueError:
                        errors.append({"row": row_num, "field": "Miktar", "message": f"\"{qty_raw}\" geçerli bir sayı değil."})

                price_raw = get_val("price")
                try:
                    price = float(price_raw) if price_raw else 0.0
                except ValueError:
                    price = 0.0

                if part_id and qty:
                    valid_rows.append({
                        "part_id": part_id,
                        "qty": qty,
                        "price": price,
                        "type": get_val("type") or "Yeni Alım",
                        "who": who,
                    })

            if errors:
                return json.dumps({"success": False, "message": f"{len(errors)} hata bulundu, hiçbir satır içe aktarılmadı.", "errors": errors})

            qty_by_part = {}
            for r in valid_rows:
                qty_by_part[r["part_id"]] = qty_by_part.get(r["part_id"], 0) + r["qty"]

            for part_id, total_qty in qty_by_part.items():
                stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == target_loc.id).first()
                if stock:
                    stock.quantity += total_qty
                else:
                    db.add(Stock(part_id=part_id, location_id=target_loc.id, quantity=total_qty))

            for r in valid_rows:
                db.add(StockMovement(
                    type=r["type"],
                    movement_kind="Inbound",
                    quantity=r["qty"],
                    part_id=r["part_id"],
                    target_location_id=target_loc.id,
                    unit_price=r["price"],
                    total_cost=r["qty"] * r["price"],
                    created_by=r["who"] or username,
                ))

            db.commit()
            clear_api_cache()
            return json.dumps({"success": True, "message": f"{len(valid_rows)} giriş kaydı başarıyla içe aktarıldı.", "imported": len(valid_rows)})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"İçe aktarma hatası: {str(e)}", "errors": []})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, result=str)
    def add_outbound_entry(self, part_id, location_id, qty, type_str, username, technician, description):
        """İrsaliye Çıkış: bir depodan dışarı (Out Stock/Scrap Stock) çıkış kaydı açar.
        source_loc.kind sistem depolarından biriyse (good_stock, doa_stock, repair_stock,
        vb.), hedef SYSTEM_TRANSFER_RULES'a göre doğrulanır -- örn. Good Stock sadece
        Repair Stock'a çıkabilir, doğrudan Out/Scrap Stock'a çıkamaz (bkz. transfer_stock,
        aynı kural kontrolü). Kural izin vermiyorsa işlem reddedilir, stok değişmez."""
        from models.stock import Stock
        from models.stock_movement import StockMovement
        from models.location import Location
        db = SessionLocal()
        try:
            part_id = int(part_id)
            location_id = int(location_id)
            qty = int(qty)

            source_loc = db.query(Location).filter(Location.id == int(location_id)).first()
            target_location_id = None
            movement_kind = None
            if source_loc and source_loc.kind in ("good_stock", "doa_stock"):
                target_kind = "scrap_stock" if type_str == "Fire" else "out_stock"

                if source_loc.kind in SYSTEM_TRANSFER_RULES and target_kind not in SYSTEM_TRANSFER_RULES[source_loc.kind]:
                    from_label = SYSTEM_LOCATION_KINDS.get(source_loc.kind, source_loc.kind)
                    allowed_targets = SYSTEM_TRANSFER_RULES[source_loc.kind]
                    if allowed_targets:
                        allowed_labels = " veya ".join(SYSTEM_LOCATION_KINDS.get(k, k) for k in allowed_targets)
                        message = f"{from_label}'tan doğrudan çıkış yapılamaz; sadece {allowed_labels} deposuna transfer yapılabilir."
                    else:
                        message = f"{from_label} sadece çıkış deposudur, buradan başka bir depoya transfer yapılamaz."
                    return json.dumps({"success": False, "message": message})

                target_location_id = _get_system_location_id(db, target_kind)
                movement_kind = "Scrap" if target_kind == "scrap_stock" else "Outbound"

            stock = db.query(Stock).with_for_update().filter(Stock.part_id == part_id, Stock.location_id == location_id).first()
            if not stock or stock.quantity < qty:
                return json.dumps({"success": False, "message": "Yetersiz stok."})

            stock.quantity -= qty

            if target_location_id:
                target_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == target_location_id).first()
                if target_stock:
                    target_stock.quantity += qty
                else:
                    db.add(Stock(part_id=part_id, location_id=target_location_id, quantity=qty))

            mov = StockMovement(
                type=type_str or "Çıkış",
                movement_kind=movement_kind,
                quantity=qty,
                part_id=part_id,
                source_location_id=location_id,
                target_location_id=target_location_id,
                created_by=username,
                technician=technician or None,
                description=description or None
            )
            db.add(mov)
            db.commit()
            clear_api_cache()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def get_reports(self, start_date, end_date):
        from models.stock_movement import StockMovement
        from models.part import Part
        from models.location import Location
        from sqlalchemy.orm import aliased
        from datetime import datetime
        db = SessionLocal()
        try:
            TargetLoc = aliased(Location)
            SourceLoc = aliased(Location)
            
            query = db.query(StockMovement, Part, SourceLoc, TargetLoc) \
                .outerjoin(Part, StockMovement.part_id == Part.id) \
                .outerjoin(SourceLoc, StockMovement.source_location_id == SourceLoc.id) \
                .outerjoin(TargetLoc, StockMovement.target_location_id == TargetLoc.id)
                
            if start_date:
                try:
                    if 'T' in start_date:
                        dt = datetime.fromisoformat(start_date)
                    else:
                        dt = datetime.strptime(start_date, "%Y-%m-%d")
                    query = query.filter(StockMovement.created_at >= dt)
                except Exception as e:
                    print(f"Error parsing start_date '{start_date}': {e}")
            if end_date:
                try:
                    if 'T' in end_date:
                        dt = datetime.fromisoformat(end_date)
                        query = query.filter(StockMovement.created_at <= dt)
                    else:
                        dt = datetime.strptime(end_date, "%Y-%m-%d")
                        import datetime as dt_module
                        dt = dt + dt_module.timedelta(days=1)
                        query = query.filter(StockMovement.created_at < dt)
                except Exception as e:
                    print(f"Error parsing end_date '{end_date}': {e}")
                    
            query = query.order_by(StockMovement.created_at.desc()).limit(10000)
            results = query.all()
            
            res = []
            for mov, p, sloc, tloc in results:
                # determine generic loc
                loc_name = tloc.name if tloc else (sloc.name if sloc else "-")
                if mov.type == "İç Transfer" and sloc and tloc:
                    loc_name = f"{sloc.name} -> {tloc.name}"
                    
                source_name = sloc.name if sloc else None
                target_name = tloc.name if tloc else None
                
                if not source_name:
                    if "İade" in mov.type and "İptal" not in mov.type:
                        source_name = "Good Stock"
                    elif "İptali" in mov.type:
                        source_name = "Good Stock"
                    elif mov.type == "Giriş":
                        source_name = "Dış Kaynak"
                    else:
                        source_name = "Bilinmiyor"
                        
                if not target_name:
                    if "Çıkış" in mov.type or "Tüketimi" in mov.type or ("İptal" in mov.type and "İptali" not in mov.type):
                        target_name = "Kullanım/Tüketim"
                    elif mov.type == "Çıkış":
                        target_name = "Dış Kaynak"
                    else:
                        target_name = "Bilinmiyor"
                res.append({
                    "id": mov.id,
                    "date": mov.created_at.strftime("%Y-%m-%d %H:%M") if mov.created_at else "",
                    "type": mov.type,
                    "part_name": p.name if p else (f"{mov.part_name_snapshot} (silindi)" if mov.part_name_snapshot else "-"),
                    "item_code": p.item_code if p else "-",
                    "location": loc_name,
                    "source_location": source_name,
                    "target_location": target_name,
                    "quantity": mov.quantity,
                    # İşlem sonrası kalan miktar. Bu kolonlar eklenmeden ÖNCE yazılmış
                    # hareketlerde None kalır; ekran o satırlarda "—" gösterir.
                    "source_balance_after": mov.source_balance_after,
                    "target_balance_after": mov.target_balance_after,
                    "user": mov.created_by,
                    "description": mov.description if mov.description else ""
                })
            return json.dumps({"success": True, "reports": res})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # ==========================
    # DEV MODE (Sadece Admin)
    # ==========================

    @Slot(result=str)
    def get_dev_mode(self):
        import os
        return json.dumps({"success": True, "dev_mode": os.getenv("DEV_MODE", "1") == "1"})

    @Slot(bool, result=str)
    def set_dev_mode(self, enabled):
        import dotenv
        import os
        env_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
        try:
            dotenv.set_key(env_file, "DEV_MODE", "1" if enabled else "0")
            dotenv.load_dotenv(env_file, override=True)
            return json.dumps({"success": True})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})

    @Slot(str, str, str, str, str, result=str)
    def update_db_settings(self, host, port, db_name, user, password):
        import dotenv
        import os
        env_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
        try:
            dotenv.set_key(env_file, "PG_HOST", host)
            dotenv.set_key(env_file, "PG_PORT", port)
            dotenv.set_key(env_file, "PG_DATABASE", db_name)
            dotenv.set_key(env_file, "PG_USER", user)
            dotenv.set_key(env_file, "PG_PASSWORD", password)
            dotenv.load_dotenv(env_file, override=True)
            return json.dumps({"success": True})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})


    # ==========================
    # DASHBOARD & EXCEL (NEW)
    # ==========================

    @Slot(result=str)
    def get_dashboard_stats(self):
        """Dashboard, giriş sonrası ilk görülen sayfa olduğundan burası özellikle
        önemli - eskiden 5 ayrı round trip (parça sayısı, kritik stok, bugünkü giriş,
        bugünkü çıkış, lokasyon sayısı) yapıyordu, tek bir sorguda birleştirildi."""
        from sqlalchemy import text
        from datetime import date, datetime, time
        db = SessionLocal()
        try:
            today_start = datetime.combine(date.today(), time.min)
            inbound_types = ["Giriş", "İç Transfer", "Yeni Alım", "Inbound", "Transfer", "Yeni Alım (Tedarikçiden)", "İade Girişi", "Diğer"]
            outbound_types = ["Çıkış", "İç Transfer", "Müşteri Satışı", "Tedarikçiye İade", "Outbound", "Transfer", "Teknik Servis", "Fire", "Fire / Bozuk", "Servis Kullanımı"]

            row = db.execute(text("""
                SELECT
                    (SELECT COUNT(*) FROM warehouse.parts) AS total_parts,
                    (SELECT COUNT(*) FROM warehouse.stock s
                     JOIN warehouse.parts p ON s.part_id = p.id
                     WHERE s.location_id = (SELECT id FROM warehouse.locations WHERE kind = 'good_stock' LIMIT 1)
                       AND s.quantity < COALESCE(p.critical_limit, 50)) AS critical_count,
                    (SELECT COALESCE(SUM(quantity), 0) FROM warehouse.stock_movements
                     WHERE (movement_kind = 'Inbound' OR type = ANY(:inbound_types)) AND created_at >= :today_start) AS todays_inbound,
                    (SELECT COALESCE(SUM(quantity), 0) FROM warehouse.stock_movements
                     WHERE (movement_kind = ANY(:outbound_kinds) OR type = ANY(:outbound_types)) AND created_at >= :today_start) AS todays_outbound,
                    (SELECT COUNT(*) FROM warehouse.locations) AS active_locations
            """), {
                "inbound_types": inbound_types,
                "outbound_types": outbound_types,
                "outbound_kinds": ["Outbound", "Scrap"],
                "today_start": today_start,
            }).mappings().first()

            import json
            return json.dumps({
                "success": True,
                "stats": {
                    "totalParts": str(row["total_parts"] or 0),
                    "criticalStock": str(row["critical_count"] or 0),
                    "todaysInbound": str(int(row["todays_inbound"] or 0)),
                    "todaysOutbound": str(int(row["todays_outbound"] or 0)),
                    "activeLocations": str(row["active_locations"] or 0)
                }
            })
        except Exception as e:
            import json
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_critical_stock(self):
        filename = "critical.json"
        path = os.path.join(get_cache_dirs()[0], filename)
        fetch_url = f"/api_cache/{filename}"
        if os.path.exists(path):
            return json.dumps({"success": True, "fetch_url": fetch_url})

        from models.stock import Stock
        from models.part import Part
        from models.location import Location
        db = SessionLocal()
        try:
            from sqlalchemy import func
            stocks = db.query(Stock, Part, Location).join(Part, Stock.part_id == Part.id).join(Location, Stock.location_id == Location.id).filter(
                Location.kind == "good_stock",
                Stock.quantity < func.coalesce(Part.critical_limit, 50)
            ).all()

            res = []
            for s, p, l in stocks:
                limit = p.critical_limit or 50
                res.append({
                    "id": s.id,
                    "part_name": p.name or "-",
                    "item_code": p.item_code or "-",
                    "location_name": l.name,
                    "quantity": s.quantity,
                    "critical_limit": limit,
                    "status": "Kritik" if s.quantity > 0 else "Tükendi"
                })
            json_data = json.dumps({"success": True, "critical_stock": res}, ensure_ascii=False)
            write_to_cache("critical.json", json_data)
            return json.dumps({"success": True, "fetch_url": fetch_url})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def export_table_to_excel(self, data_json_str, filename):
        """
        Genel amaçlı excel export - Direkt İndirilenler Klasörüne Kaydeder
        """
        from core.excel_utils import style_excel_file
        import json
        import pandas as pd
        import os
        from pathlib import Path
        try:
            data = json.loads(data_json_str)
            if not data:
                return json.dumps({"success": False, "message": "Dışa aktarılacak veri yok."})
                
            downloads_path = str(Path.home() / "Downloads")
            file_path = os.path.join(downloads_path, filename)
            
            # Eğer dosya varsa ismini değiştir
            counter = 1
            base_name, ext = os.path.splitext(filename)
            while os.path.exists(file_path):
                file_path = os.path.join(downloads_path, f"{base_name}_{counter}{ext}")
                counter += 1
                
            df = pd.DataFrame(data)
            df.to_excel(file_path, index=False)
            try:
                style_excel_file(file_path)
            except:
                pass
                
            # Dosyayı otomatik aç (Windows)
            os.startfile(file_path)
            
            return json.dumps({"success": True})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})

    @Slot(str, result=str)
    def export_all_tables_to_excel(self, filename):
        """
        Tüm veritabanı tablolarını tek bir excel dosyasında farklı sheet'lerde dışa aktarır.
        JS üzerinden veri transferini atlayarak (size limitleri aşmamak için) direkt db'den çeker.
        """
        from core.excel_utils import style_excel_file
        import json
        import pandas as pd
        import os
        import re
        from pathlib import Path
        from sqlalchemy import text
        from config.database import get_db
        try:
            downloads_path = str(Path.home() / "Downloads")
            file_path = os.path.join(downloads_path, filename)
            
            # Eğer dosya varsa ismini değiştir
            counter = 1
            base_name, ext = os.path.splitext(filename)
            while os.path.exists(file_path):
                file_path = os.path.join(downloads_path, f"{base_name}_{counter}{ext}")
                counter += 1
                
            # Sheet adlarını temizleme fonksiyonu
            def clean_sheet_name(name):
                # Excel kısıtlamaları: max 31 karakter, : \ / ? * [ ] yasak
                name = re.sub(r'[:\\/?*\[\]]', '_', name)
                return name[:31]
                
            with get_db() as db:
                tables_query = text('''
                    SELECT table_schema, table_name 
                    FROM information_schema.tables 
                    WHERE table_schema IN ('public', 'warehouse', 'auth') 
                      AND table_type = 'BASE TABLE'
                ''')
                tables_result = db.execute(tables_query).fetchall()

                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    used_sheet_names = set()
                    
                    for schema, t_name in tables_result:
                        query = text(f'SELECT * FROM "{schema}"."{t_name}"')
                        result = db.execute(query).fetchall()
                        keys = result[0]._mapping.keys() if result else []
                        data = [dict(zip(keys, row)) for row in result]
                        
                        for row in data:
                            for k, v in row.items():
                                if hasattr(v, 'isoformat'):
                                    row[k] = v.isoformat()
                                    
                        cleaned_name = clean_sheet_name(t_name)
                        
                        # Aynı isim çakışmasını önle
                        original_clean = cleaned_name
                        suffix = 1
                        while cleaned_name in used_sheet_names:
                            suffix_str = f"_{suffix}"
                            cleaned_name = f"{original_clean[:31-len(suffix_str)]}{suffix_str}"
                            suffix += 1
                            
                        used_sheet_names.add(cleaned_name)
                        
                        df = pd.DataFrame(data)
                        df.to_excel(writer, sheet_name=cleaned_name, index=False)
                    
            try:
                style_excel_file(file_path)
            except:
                pass
                
            # Dosyayı otomatik aç (Windows)
            os.startfile(file_path)
            
            return json.dumps({"success": True})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})

    # ==========================================
    # LOCAL DB & DATA FOLDERS (AYARLAR SEKME)
    # ==========================================
    def _get_settings_file(self):
        import os
        from pathlib import Path
        settings_dir = os.path.join(str(Path.home()), ".remalab")
        os.makedirs(settings_dir, exist_ok=True)
        return os.path.join(settings_dir, "settings.json")

    def _read_settings(self):
        import json, os
        settings_file = self._get_settings_file()
        if not os.path.exists(settings_file):
            return {"local_files": [], "data_folders": []}
        try:
            with open(settings_file, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {"local_files": [], "data_folders": []}

    def _write_settings(self, data):
        import json
        settings_file = self._get_settings_file()
        with open(settings_file, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

    @Slot(result=str)
    def get_local_files(self):
        import json, os, datetime
        settings = self._read_settings()
        files = settings.get("local_files", [])
        valid_files = []
        for f in files:
            path = f.get("path")
            if path and os.path.exists(path):
                size_bytes = os.path.getsize(path)
                size_mb = size_bytes / (1024 * 1024)
                mod_time = os.path.getmtime(path)
                mod_date = datetime.datetime.fromtimestamp(mod_time).strftime('%d.%m.%Y %H:%M')
                
                f["size"] = f"{size_mb:.2f} MB"
                f["modified"] = mod_date
                
                # Mock tables/records count if sqlite
                if f.get("type") == "sqlite":
                    f["tables"] = 0
                    f["records"] = 0
                    try:
                        import sqlite3
                        conn = sqlite3.connect(path)
                        cursor = conn.cursor()
                        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                        tables = cursor.fetchall()
                        f["tables"] = len(tables)
                        conn.close()
                    except:
                        pass
                
                valid_files.append(f)
            else:
                f["size"] = "Kayıp"
                f["modified"] = "Bilinmiyor"
                valid_files.append(f)
        
        return json.dumps({"success": True, "local_files": valid_files})

    @Slot(result=str)
    def add_local_file(self):
        from PySide6.QtWidgets import QFileDialog, QApplication
        import json, os, uuid
        
        main_win = QApplication.instance().main_window
        file_path, _ = QFileDialog.getOpenFileName(
            main_win, "Var Olan Veritabanı veya Betiği Seç", "", "Database Files (*.db *.sqlite *.sql);;All Files (*)"
        )
        if not file_path:
            return json.dumps({"success": False, "message": "Seçim iptal edildi"})
            
        settings = self._read_settings()
        
        # Check if already exists
        for f in settings.get("local_files", []):
            if f.get("path") == file_path:
                return json.dumps({"success": False, "message": "Bu dosya zaten listede."})
                
        file_type = "sqlite" if file_path.endswith((".db", ".sqlite")) else "sql"
        new_file = {
            "id": str(uuid.uuid4()),
            "name": os.path.basename(file_path),
            "path": file_path,
            "type": file_type
        }
        
        settings.setdefault("local_files", []).append(new_file)
        self._write_settings(settings)
        return json.dumps({"success": True, "file": new_file})

    @Slot(result=str)
    def create_local_file(self):
        from PySide6.QtWidgets import QFileDialog, QApplication
        import json, os, uuid, sqlite3
        
        main_win = QApplication.instance().main_window
        file_path, _ = QFileDialog.getSaveFileName(
            main_win, "Yeni SQLite Veritabanı Oluştur", "yeni_veritabani.db", "SQLite Database (*.db)"
        )
        if not file_path:
            return json.dumps({"success": False, "message": "İşlem iptal edildi"})
            
        try:
            # Create empty sqlite
            conn = sqlite3.connect(file_path)
            conn.close()
            
            settings = self._read_settings()
            new_file = {
                "id": str(uuid.uuid4()),
                "name": os.path.basename(file_path),
                "path": file_path,
                "type": "sqlite"
            }
            settings.setdefault("local_files", []).append(new_file)
            self._write_settings(settings)
            return json.dumps({"success": True, "file": new_file})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})

    @Slot(str, result=str)
    def delete_local_file(self, file_id):
        import json
        settings = self._read_settings()
        files = settings.get("local_files", [])
        settings["local_files"] = [f for f in files if f.get("id") != file_id]
        self._write_settings(settings)
        return json.dumps({"success": True})

    @Slot(result=str)
    def open_project_guide(self):
        """Proje rehberi PDF'ini (docs/RemaLab_WMS_Birlesik_Dokuman.pdf) sistemin
        varsayilan PDF goruntuleyicisiyle acar. Bu dosya elle hazirlanmis/aciklamali
        bir surum oldugu icin generate_project_guide_pdf.py tarafindan UZERINE YAZILMAZ
        (otomatik uretim ayri dosyayi - RemaLab_WMS_Proje_Rehberi.pdf - hedefler)."""
        import json, os, sys
        import subprocess
        try:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            pdf_path = os.path.join(base_dir, "docs", "RemaLab_WMS_Birlesik_Dokuman.pdf")
            if not os.path.exists(pdf_path):
                return json.dumps({"success": False, "message": "Proje rehberi PDF'i bulunamadı."})
            if os.name == 'nt':
                os.startfile(pdf_path)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', pdf_path])
            else:
                subprocess.Popen(['xdg-open', pdf_path])
            return json.dumps({"success": True})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})

    @Slot(str, result=str)
    def open_local_folder(self, file_path):
        import json, os, sys
        import subprocess
        try:
            folder = os.path.dirname(file_path)
            if not os.path.exists(folder):
                return json.dumps({"success": False, "message": "Klasör bulunamadı."})
            if os.name == 'nt':
                os.startfile(folder)
            elif sys.platform == 'darwin':
                subprocess.Popen(['open', folder])
            else:
                subprocess.Popen(['xdg-open', folder])
            return json.dumps({"success": True})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})

    @Slot(result=str)
    def get_data_folders(self):
        import json
        settings = self._read_settings()
        return json.dumps({"success": True, "data_folders": settings.get("data_folders", [])})

    @Slot(result=str)
    def add_data_folder(self):
        from PySide6.QtWidgets import QFileDialog, QApplication
        import json, os, uuid
        
        main_win = QApplication.instance().main_window
        folder_path = QFileDialog.getExistingDirectory(
            main_win, "Klasör Seç"
        )
        if not folder_path:
            return json.dumps({"success": False, "message": "Seçim iptal edildi"})
            
        settings = self._read_settings()
        
        # Check if already exists
        for f in settings.get("data_folders", []):
            if f.get("path") == folder_path:
                return json.dumps({"success": False, "message": "Bu klasör zaten listede."})
                
        # Determine type based on name heuristically or default to data
        folder_type = "backup" if "backup" in folder_path.lower() or "yedek" in folder_path.lower() else "data"
        new_folder = {
            "id": str(uuid.uuid4()),
            "name": os.path.basename(folder_path) or folder_path,
            "path": folder_path,
            "type": folder_type
        }
        
        settings.setdefault("data_folders", []).append(new_folder)
        self._write_settings(settings)
        return json.dumps({"success": True, "folder": new_folder})

    @Slot(str, result=str)
    def delete_data_folder(self, folder_id):
        import json
        settings = self._read_settings()
        folders = settings.get("data_folders", [])
        settings["data_folders"] = [f for f in folders if f.get("id") != folder_id]
        self._write_settings(settings)
        return json.dumps({"success": True})

    # ==========================
    # DYNAMIC TABLE MANAGEMENT
    # ==========================
    @Slot(result=str)
    def get_all_tables_schema(self):
        """Fetch all tables and their columns from specific schemas."""
        import json
        from sqlalchemy import text
        from config.database import get_db
        try:
            with get_db() as db:
                # First fetch tables
                tables_query = text('''
                    SELECT table_schema, table_name 
                    FROM information_schema.tables 
                    WHERE table_schema IN ('public', 'warehouse', 'auth') 
                      AND table_type = 'BASE TABLE'
                ''')
                tables_result = db.execute(tables_query).fetchall()
                
                tables_data = []
                for schema, t_name in tables_result:
                    cols_query = text('''
                        SELECT column_name 
                        FROM information_schema.columns 
                        WHERE table_schema = :schema AND table_name = :t_name
                    ''')
                    cols_result = db.execute(cols_query, {"schema": schema, "t_name": t_name}).fetchall()
                    columns = [row[0] for row in cols_result]
                    
                    tables_data.append({
                        "id": f"{schema}.{t_name}",
                        "name": f"{t_name} ({schema})",
                        "schema": schema,
                        "table_name": t_name,
                        "columns": columns
                    })
                return json.dumps({"success": True, "tables": tables_data})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})

    @Slot(str, str, result=str)
    def get_table_data(self, schema, table_name):
        """Fetch all rows from a dynamically specified table."""
        import json
        from sqlalchemy import text
        from config.database import get_db
        
        # Security: whitelist check or ensure schema is valid
        if schema not in ['public', 'warehouse', 'auth']:
            return json.dumps({"success": False, "message": "Invalid schema"})
            
        try:
            with get_db() as db:
                query = text(f'SELECT * FROM "{schema}"."{table_name}"')
                result = db.execute(query).fetchall()
                keys = result[0]._mapping.keys() if result else []
                data = [dict(zip(keys, row)) for row in result]
                # convert datetime objects to string
                for row in data:
                    for k, v in row.items():
                        if hasattr(v, 'isoformat'):
                            row[k] = v.isoformat()
                return json.dumps({"success": True, "data": data})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})

    @Slot(str, str, str, result=str)
    def insert_table_data(self, schema, table_name, data_json):
        """Insert row into a specified table."""
        import json
        from sqlalchemy import text
        from config.database import get_db
        
        if schema not in ['public', 'warehouse', 'auth']:
            return json.dumps({"success": False, "message": "Invalid schema"})
            
        try:
            data = json.loads(data_json)
            if not isinstance(data, dict):
                return json.dumps({"success": False, "message": "Data must be a dictionary"})
                
            with get_db() as db:
                # Apply SYSTEM_TRANSFER_RULES if inserting into stock_movements (e.g. via Excel)
                if table_name == 'stock_movements':
                    type_ = data.get('type')
                    movement_kind = data.get('movement_kind')
                    if type_ == "İç Transfer" or movement_kind == "Transfer":
                        from_loc_id = data.get('source_location_id')
                        to_loc_id = data.get('target_location_id')
                        if from_loc_id and to_loc_id:
                            sloc = db.execute(text("SELECT kind, name FROM warehouse.locations WHERE id = :id"), {'id': from_loc_id}).fetchone()
                            tloc = db.execute(text("SELECT kind, name FROM warehouse.locations WHERE id = :id"), {'id': to_loc_id}).fetchone()
                            if sloc and tloc:
                                from_kind = sloc[0]
                                to_kind = tloc[0]
                                if from_kind in SYSTEM_TRANSFER_RULES:
                                    allowed_targets = SYSTEM_TRANSFER_RULES[from_kind]
                                    if to_kind not in allowed_targets:
                                        allowed_labels = ", ".join(allowed_targets)
                                        if not allowed_targets:
                                            msg = f"Excel Hata: {from_kind} sadece çıkış deposudur, buradan başka depoya transfer yapılamaz."
                                        else:
                                            msg = f"Excel Hata: {from_kind}'tan sadece {allowed_labels} deposuna transfer yapılabilir."
                                        return json.dumps({"success": False, "message": msg})

                columns = list(data.keys())
                values = list(data.values())
                placeholders = ', '.join([f':{col}' for col in columns])
                col_names = ', '.join([f'"{col}"' for col in columns])
            
                query = text(f'INSERT INTO "{schema}"."{table_name}" ({col_names}) VALUES ({placeholders})')
                db.execute(query, data)
                db.commit()
                return json.dumps({"success": True})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})

    @Slot(str, str, str, result=str)
    def bulk_insert_table_data(self, schema, table_name, rows_json):
        """insert_table_data'nın tek satırlık eşdeğeriyle aynı davranışı (şema allowlist,
        stock_movements transfer kuralı kontrolü) korur, ama Veri Yönetimi (DataManagement)
        ekranındaki jenerik "herhangi bir tabloya Excel'den toplu satır ekle" akışında N ayrı
        çağrı/commit yerine tüm satırları TEK connection'da işler. Bu ekran rastgele
        tablolara (farklı unique/NOT NULL kısıtları) yazdığından, tek satırlık akışta olduğu
        gibi satırlar birbirinden BAĞIMSIZ değerlendirilir (bir satırın hatası diğerlerini
        engellemez) - SAVEPOINT ile satır başına izole edilip sonda tek commit yapılır."""
        from sqlalchemy import text
        from config.database import get_db

        if schema not in ['public', 'warehouse', 'auth']:
            return json.dumps({"success": False, "message": "Invalid schema", "errors": []})

        try:
            rows = json.loads(rows_json or "[]")
        except (ValueError, TypeError):
            return json.dumps({"success": False, "message": "Geçersiz dosya verisi.", "errors": []})

        if not rows or not isinstance(rows, list):
            return json.dumps({"success": False, "message": "Dosyada içe aktarılacak satır bulunamadı.", "errors": []})

        try:
            with get_db() as db:
                errors = []
                for idx, data in enumerate(rows):
                    row_num = idx + 2
                    if not isinstance(data, dict) or not data:
                        errors.append({"row": row_num, "field": "-", "message": "Satır verisi geçersiz veya boş."})
                        continue

                    if table_name == 'stock_movements':
                        type_ = data.get('type')
                        movement_kind = data.get('movement_kind')
                        if type_ == "İç Transfer" or movement_kind == "Transfer":
                            from_loc_id = data.get('source_location_id')
                            to_loc_id = data.get('target_location_id')
                            blocked_msg = None
                            if from_loc_id and to_loc_id:
                                sloc = db.execute(text("SELECT kind, name FROM warehouse.locations WHERE id = :id"), {'id': from_loc_id}).fetchone()
                                tloc = db.execute(text("SELECT kind, name FROM warehouse.locations WHERE id = :id"), {'id': to_loc_id}).fetchone()
                                if sloc and tloc:
                                    from_kind = sloc[0]
                                    to_kind = tloc[0]
                                    if from_kind in SYSTEM_TRANSFER_RULES:
                                        allowed_targets = SYSTEM_TRANSFER_RULES[from_kind]
                                        if to_kind not in allowed_targets:
                                            allowed_labels = ", ".join(allowed_targets)
                                            if not allowed_targets:
                                                blocked_msg = f"{from_kind} sadece çıkış deposudur, buradan başka depoya transfer yapılamaz."
                                            else:
                                                blocked_msg = f"{from_kind}'tan sadece {allowed_labels} deposuna transfer yapılabilir."
                            if blocked_msg:
                                errors.append({"row": row_num, "field": "type", "message": blocked_msg})
                                continue

                    columns = list(data.keys())
                    placeholders = ', '.join([f':{col}' for col in columns])
                    col_names = ', '.join([f'"{col}"' for col in columns])
                    query = text(f'INSERT INTO "{schema}"."{table_name}" ({col_names}) VALUES ({placeholders})')
                    try:
                        with db.begin_nested():
                            db.execute(query, data)
                    except Exception as row_ex:
                        errors.append({"row": row_num, "field": "-", "message": str(row_ex)})

                valid_count = len(rows) - len(errors)
                if valid_count == 0:
                    db.rollback()
                    return json.dumps({"success": False, "message": f"{len(errors)} hata bulundu, hiçbir satır içe aktarılmadı.", "errors": errors})

                db.commit()
                if errors:
                    return json.dumps({"success": True, "message": f"{valid_count} satır içe aktarıldı, {len(errors)} satır hata verdi.", "imported": valid_count, "errors": errors})
                return json.dumps({"success": True, "message": f"{valid_count} satır başarıyla içe aktarıldı.", "imported": valid_count, "errors": []})
        except Exception as e:
            return json.dumps({"success": False, "message": f"İçe aktarma hatası: {str(e)}", "errors": []})

    def _ensure_batch_entries_table(self):
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("""
                CREATE TABLE IF NOT EXISTS warehouse.batch_entries (
                    id SERIAL PRIMARY KEY,
                    customer_no VARCHAR(100),
                    customer_name VARCHAR(255),
                    imei_number VARCHAR(100),
                    serial_number VARCHAR(100),
                    internal_id VARCHAR(100),
                    batch_no VARCHAR(100),
                    model VARCHAR(255),
                    gb VARCHAR(50),
                    color VARCHAR(50),
                    unit_price NUMERIC(12, 2) DEFAULT 0.00,
                    currency VARCHAR(10) DEFAULT 'EUR',
                    defects TEXT,
                    screen_test VARCHAR(100),
                    power_test VARCHAR(100),
                    flow VARCHAR(100) DEFAULT 'To refurbish',
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                );
            """))
            # Mevcut tabloya currency, is_success, created_by kolonları yoksa ekle
            db.execute(text("""
                DO $$
                BEGIN
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_schema = 'warehouse' AND table_name = 'batch_entries' AND column_name = 'currency'
                    ) THEN
                        ALTER TABLE warehouse.batch_entries ADD COLUMN currency VARCHAR(10) DEFAULT 'EUR';
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_schema = 'warehouse' AND table_name = 'batch_entries' AND column_name = 'is_success'
                    ) THEN
                        ALTER TABLE warehouse.batch_entries ADD COLUMN is_success BOOLEAN DEFAULT false;
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_schema = 'warehouse' AND table_name = 'batch_entries' AND column_name = 'created_by'
                    ) THEN
                        ALTER TABLE warehouse.batch_entries ADD COLUMN created_by VARCHAR(100) DEFAULT 'io';
                    END IF;
                    IF NOT EXISTS (
                        SELECT 1 FROM information_schema.columns
                        WHERE table_schema = 'warehouse' AND table_name = 'batch_entries' AND column_name = 'statu_code'
                    ) THEN
                        ALTER TABLE warehouse.batch_entries ADD COLUMN statu_code INTEGER DEFAULT 100;
                        UPDATE warehouse.batch_entries SET statu_code = 100 WHERE statu_code IS NULL;
                    END IF;
                END $$;
            """))
            # Müşteri para birimlerini batch_entries tablosuna senkronize et
            db.execute(text("""
                UPDATE warehouse.batch_entries b
                SET currency = c.currency
                FROM warehouse.customers c
                WHERE (LOWER(b.customer_name) = LOWER(c.customer_name) OR b.customer_no = c.code)
                  AND c.currency IS NOT NULL AND c.currency != '';
            """))
            db.commit()
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] _ensure_batch_entries_table hatası: {e}")
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def get_batch_entries(self, page="1", page_size="50", search_term="", flow_filter=""):
        from sqlalchemy import text
        db = SessionLocal()
        try:
            page = max(1, int(page or 1))
            page_size = min(1000, max(1, int(page_size or 50)))
            offset = (page - 1) * page_size

            where_clauses = []
            params = {"limit": page_size, "offset": offset}

            if search_term and str(search_term).strip():
                term_raw = str(search_term).strip()
                term_like = f"%{term_raw}%"

                # Önce tam eşleşen Batch No, Internal ID, Seri No, IMEI, Müşteri No veya Müşteri Adı var mı kontrol et
                exact_count = db.execute(text("""
                    SELECT COUNT(*) FROM warehouse.batch_entries
                    WHERE LOWER(TRIM(COALESCE(batch_no, ''))) = LOWER(:t)
                       OR LOWER(TRIM(COALESCE(internal_id, ''))) = LOWER(:t)
                       OR LOWER(TRIM(COALESCE(serial_number, ''))) = LOWER(:t)
                       OR LOWER(TRIM(COALESCE(imei_number, ''))) = LOWER(:t)
                       OR LOWER(TRIM(COALESCE(customer_no, ''))) = LOWER(:t)
                       OR LOWER(TRIM(COALESCE(customer_name, ''))) = LOWER(:t)
                """), {"t": term_raw}).scalar()

                if exact_count > 0:
                    where_clauses.append("""(
                        LOWER(TRIM(COALESCE(batch_no, ''))) = LOWER(:exact_term) OR 
                        LOWER(TRIM(COALESCE(internal_id, ''))) = LOWER(:exact_term) OR 
                        LOWER(TRIM(COALESCE(serial_number, ''))) = LOWER(:exact_term) OR 
                        LOWER(TRIM(COALESCE(imei_number, ''))) = LOWER(:exact_term) OR 
                        LOWER(TRIM(COALESCE(customer_no, ''))) = LOWER(:exact_term) OR
                        LOWER(TRIM(COALESCE(customer_name, ''))) = LOWER(:exact_term)
                    )""")
                    params["exact_term"] = term_raw
                else:
                    where_clauses.append("""(
                        customer_no ILIKE :search OR 
                        customer_name ILIKE :search OR 
                        imei_number ILIKE :search OR 
                        serial_number ILIKE :search OR 
                        internal_id ILIKE :search OR 
                        batch_no ILIKE :search OR 
                        model ILIKE :search OR 
                        defects ILIKE :search
                    )""")
                    params["search"] = term_like

            if flow_filter and str(flow_filter).strip() and str(flow_filter).strip().lower() not in ("tümü", "hepsi"):
                where_clauses.append("flow = :flow_filter")
                params["flow_filter"] = str(flow_filter).strip()

            where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

            count_sql = f"SELECT COUNT(*) FROM warehouse.batch_entries {where_sql};"
            total = db.execute(text(count_sql), params).scalar()

            data_sql = f"""
                SELECT id, customer_no, customer_name, imei_number, serial_number, internal_id, batch_no,
                       model, gb, color, unit_price, currency, defects, screen_test, power_test, flow, created_at, updated_at
                FROM warehouse.batch_entries
                {where_sql}
                ORDER BY id DESC
                LIMIT :limit OFFSET :offset;
            """
            rows = db.execute(text(data_sql), params).mappings().all()

            records = [{
                "id": r["id"],
                "document_date": r["created_at"].strftime("%d.%m.%Y") if r["created_at"] else "-",
                "document_number": r["batch_no"] or r["internal_id"] or r["serial_number"] or r["imei_number"] or "-",
                "customer_no": r["customer_no"] or "",
                "customer_name": r["customer_name"] or "",
                "imei_number": r["imei_number"] or "",
                "serial_number": r["serial_number"] or "",
                "internal_id": r["internal_id"] or "",
                "batch_no": r["batch_no"] or "",
                "model": r["model"] or "",
                "gb": r["gb"] or "",
                "color": r["color"] or "",
                "unit_price": float(r["unit_price"]) if r["unit_price"] is not None else 0.0,
                "currency": r.get("currency", None) or "EUR",
                "defects": r["defects"] or "",
                "screen_test": r["screen_test"] or "",
                "power_test": r["power_test"] or "",
                "flow": r["flow"] or "To refurbish",
                "created_at": r["created_at"].strftime("%d.%m.%Y %H:%M") if r["created_at"] else "-",
                "updated_at": r["updated_at"].strftime("%d.%m.%Y %H:%M") if r["updated_at"] else "-"
            } for r in rows]

            return json.dumps({
                "success": True,
                "records": records,
                "total": total,
                "page": page,
                "page_size": page_size
            }, ensure_ascii=False)
        except Exception as e:
            print(f"[WebBridge] get_batch_entries hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    def _validate_product_model(self, db, model_name):
        """Girilmiş cihaz modelinin sistemde tanımlı (product_model, parts, product_family vb.)
        geçerli bir model olup olmadığını esnek ve akıllı kelime eşleştirmesi ile denetler.
        (Örn. 'iPad Air 5' -> 'iPad Air (5th Gen)' ile eşleşir, ancak 'iPhone 19' engellenir)."""
        import re
        from sqlalchemy import text
        m_clean = (model_name or "").strip()
        if not m_clean:
            return False, "Cihaz modeli boş olamaz. Lütfen tanımlı bir model giriniz."
        
        words = [w for w in re.split(r"[\s\-\(\)\,\.\_\/]+", m_clean.lower()) if len(w) > 0]
        if not words:
            return False, "Geçersiz cihaz modeli."

        params = {f"w{i}": f"%{words[i]}%" for i in range(len(words))}
        
        # 1. warehouse.product_model
        conds_pm = " AND ".join([f"LOWER(short_name) LIKE :w{i}" for i in range(len(words))])
        q_pm = text(f"SELECT 1 FROM warehouse.product_model WHERE {conds_pm} LIMIT 1")
        if db.execute(q_pm, params).first():
            return True, ""

        # 2. warehouse.parts
        conds_parts = " AND ".join([f"LOWER(model) LIKE :w{i}" for i in range(len(words))])
        q_parts = text(f"SELECT 1 FROM warehouse.parts WHERE {conds_parts} LIMIT 1")
        if db.execute(q_parts, params).first():
            return True, ""

        # 3. warehouse.product_family
        conds_fam = " AND ".join([f"LOWER(short_name) LIKE :w{i}" for i in range(len(words))])
        q_fam = text(f"SELECT 1 FROM warehouse.product_family WHERE {conds_fam} LIMIT 1")
        if db.execute(q_fam, params).first():
            return True, ""

        return False, f"Sistemde tanımlı olmayan geçersiz cihaz modeli: '{m_clean}'. Lütfen sistemde tanımlı geçerli bir model giriniz."

    def _validate_new_batch_entry(self, db, d):
        """Yeni batch girişi kurallarını kontrol eder. Hata varsa açıklama metnini,
        yoksa None döner. Hem create_batch_entry (gerçek kayıt) hem de
        validate_batch_entry (içe aktarma öncesi dry-run) tarafından kullanılır."""
        from models.batch_entry import BatchEntry

        # 0) Cihaz Modeli Doğrulaması (Sistemde tanımlı olmayan modelleri engeller)
        model_val = (d.get("model") or "").strip()
        is_valid_m, m_err_msg = self._validate_product_model(db, model_val)
        if not is_valid_m:
            return m_err_msg

        # 0b) Kapasite (GB/TB) mükerrer girişi engelleme.
        #     'model' alanı zaten bir kapasite içeriyorsa (örn. "iPhone 13 mini 128GB")
        #     ve ayrıca 'gb' alanı da doldurulmuşsa, ekranda/kayıtta "iPhone 13 mini
        #     128GB 128GB" gibi ürün ailesinde bulunmayan geçersiz bir kombinasyon oluşur.
        #     Kelime bazlı fuzzy eşleşme aynı kelimenin tekrarını yakalamadığından burada
        #     açıkça reddediyoruz. Kapasite YA modelde YA da GB alanında olmalı, ikisinde
        #     birden değil.
        import re as _re
        gb_val = (d.get("gb") or "").strip()
        model_caps = _re.findall(r"\d+\s*(?:gb|tb)\b", model_val.lower())
        if gb_val and model_caps:
            return (f"Kapasite modelde zaten belirtilmiş ('{model_caps[0].upper().replace(' ', '')}'). "
                    f"GB alanına ikinci kez kapasite girildiğinde '{model_val} {gb_val}' gibi "
                    f"ürün ailesinde tanımlı olmayan geçersiz bir kayıt oluşur. "
                    f"Lütfen ya modeldeki kapasiteyi kaldırın ya da GB seçimini boş bırakın.")

        # 0c) Power Test her zaman zorunludur (fiyatlandırma kuralı Power Test sonucuna
        #     dayanıyor ve Screen Test'ten daha üstün/öncelikli; bu yüzden asla boş
        #     bırakılamaz - Screen Test ise opsiyonel kalabilir).
        power_test_val = (d.get("power_test") or "").strip()
        if not power_test_val:
            return "Power Test alanı zorunludur, boş bırakılamaz."

        batch_no = (d.get("batch_no") or "").strip()
        customer_name = (d.get("customer_name") or "").strip()

        if batch_no:
            # 1) Aynı batch numarası farklı müşteride olamaz
            existing_batch = db.query(BatchEntry).filter(BatchEntry.batch_no == batch_no).first()
            if existing_batch and existing_batch.customer_name and existing_batch.customer_name.strip().lower() != customer_name.lower():
                return (f"Bu batch numarası ({batch_no}) başka bir müşteriye ({existing_batch.customer_name}) aittir. "
                        f"Aynı batch numarasıyla farklı müşteri kaydı oluşturulamaz.")

            # 2) Üretimde olan (çıkışı yapılmamış) batch numarasıyla yeni cihaz girilemez.
            #    Statü 100/101 = hâlâ giriş aşaması (izin), 102+ ve 128 değil = üretimde (kilit).
            in_production = db.query(BatchEntry).filter(
                BatchEntry.batch_no == batch_no,
                BatchEntry.statu_code.notin_([100, 101, 128])
            ).first()
            if in_production:
                return (f"'{batch_no}' numaralı batch üretim aşamasında (statü {in_production.statu_code}) ve henüz "
                        f"çıkışı yapılmadı. Çıkış (statü 128) tamamlanmadan bu batch numarasıyla yeni cihaz kaydı oluşturulamaz.")

        # 3) Aynı cihazın başka bir batch altında açık (statü 128 olmayan) servisi varsa engellenir.
        imei_val = (d.get("imei_number") or "").strip()
        serial_val = (d.get("serial_number") or "").strip()
        internal_val = (d.get("internal_id") or "").strip()
        active = self._find_active_service_for_device(db, imei_val, serial_val, internal_val)
        if active:
            # Aynı batch numarasına ait bir güncellemeyse izin ver
            active_batch = (active.get("batch_no") or "").strip().lower()
            current_batch = batch_no.strip().lower()
            if not (current_batch and active_batch and current_batch == active_batch):
                match_id = imei_val or serial_val or internal_val
                return (f"Cihaz ({match_id}) için farklı bir batch ({active.get('batch_no') or 'Tanımsız'}) altında zaten aktif servis var (Statü: {active['statu_code']}). "
                        f"Süreç tamamlanmadan (statü 128) aynı cihaz başka bir batch numarasıyla tekrar girilemez.")

        return None

    @Slot(str, result=str)
    def validate_batch_entry(self, data_json):
        """İçe aktarma öncesi dry-run: satırın kurallara uyup uymadığını KAYIT YAPMADAN
        döner. ok=True ise aktarılabilir; ok=False ise 'message' hata nedenidir."""
        db = SessionLocal()
        try:
            d = json.loads(data_json or "{}")
            err = self._validate_new_batch_entry(db, d)
            return json.dumps({"success": True, "ok": err is None, "message": err or ""}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "ok": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def import_defined_batch_entry(self, data_json):
        """Excel içe aktarma için: SADECE sistemde (batch_entries'te) zaten TANIMLI olan
        cihazları günceller. Cihaz (IMEI/Seri/Internal ID) tanımlı değilse ya da batch
        numarası sistemde yoksa satır reddedilir (ok=False + neden). Yeni/bilinmeyen
        cihaz veya batch OLUŞTURMAZ."""
        from models.batch_entry import BatchEntry
        from sqlalchemy import func, or_
        db = SessionLocal()
        try:
            d = json.loads(data_json or "{}")
            imei = (d.get("imei_number") or "").strip()
            serial = (d.get("serial_number") or "").strip()
            internal = (d.get("internal_id") or "").strip()
            batch_no = (d.get("batch_no") or "").strip()

            # 0) Cihaz Modeli Doğrulaması (Sistemde tanımlı olmayan modeller engellenir)
            model_val = (d.get("model") or "").strip()
            if model_val:
                is_valid_m, m_err_msg = self._validate_product_model(db, model_val)
                if not is_valid_m:
                    return json.dumps({"success": True, "ok": False, "message": m_err_msg}, ensure_ascii=False)

            # 1) Cihaz (IMEI/Seri/Internal ID) batch_entries'te tanımlı mı?
            conds = []
            if imei:
                conds.append(func.lower(func.trim(BatchEntry.imei_number)) == imei.lower())
            if serial:
                conds.append(func.lower(func.trim(BatchEntry.serial_number)) == serial.lower())
            if internal:
                conds.append(func.lower(func.trim(BatchEntry.internal_id)) == internal.lower())

            target = None
            if conds:
                target = db.query(BatchEntry).filter(or_(*conds)).order_by(BatchEntry.id.desc()).first()
            if not target:
                ident = imei or serial or internal or "-"
                return json.dumps({"success": True, "ok": False,
                                   "message": f"Cihaz ({ident}) sistemde tanımlı değil, içe aktarılmadı."}, ensure_ascii=False)

            # 1b) Model verildiyse sistemde tanımlı geçerli bir model olmalı (örn. 'iPhone 19' reddedilir)
            model_val = (d.get("model") or "").strip()
            if model_val:
                is_valid_m, m_err_msg = self._validate_product_model(db, model_val)
                if not is_valid_m:
                    return json.dumps({"success": True, "ok": False, "message": m_err_msg}, ensure_ascii=False)

            # 2) Batch numarası (verildiyse) sistemde tanımlı mı?
            if batch_no:
                batch_exists = db.query(BatchEntry).filter(
                    func.lower(func.trim(BatchEntry.batch_no)) == batch_no.lower()
                ).first()
                if not batch_exists:
                    return json.dumps({"success": True, "ok": False,
                                       "message": f"Batch numarası ({batch_no}) sistemde tanımlı değil, içe aktarılmadı."}, ensure_ascii=False)

            # 3) Var olan cihazı Excel verisiyle güncelle. Ama gelen tüm alanlar zaten
            #    mevcut kayıtla AYNIYSA (hiçbir değişiklik yok) mükerrer sayılıp reddedilir.
            update_fields = ["customer_no", "customer_name", "batch_no", "model", "gb", "color",
                             "defects", "screen_test", "power_test", "flow"]
            changed = False
            for field in update_fields:
                val = d.get(field)
                if val in (None, ""):
                    continue
                if (getattr(target, field) or "") != str(val).strip():
                    changed = True
                    break
            if not changed:
                ident = imei or serial or internal or "-"
                return json.dumps({"success": True, "ok": False,
                                   "message": f"Cihaz ({ident}) zaten aynı bilgilerle kayıtlı, değişiklik yok (mükerrer)."}, ensure_ascii=False)

            # Power Test her zaman zorunludur (satır power_test taşımasa bile, güncelleme
            # sonrasında kayıttaki nihai değer boş kalmamalı).
            final_power_test = str(d.get("power_test") or target.power_test or "").strip()
            if not final_power_test:
                return json.dumps({"success": True, "ok": False,
                                   "message": "Power Test alanı zorunludur, boş bırakılamaz."}, ensure_ascii=False)

            valid_flow_values = self._get_flow_values(db)
            flow_input = (d.get("flow", "") or "").strip()
            if flow_input:
                flow_map = {v.lower(): v for v in valid_flow_values}
                if flow_input.lower() in flow_map:
                    d["flow"] = flow_map[flow_input.lower()]
                else:
                    for v in valid_flow_values:
                        if flow_input.lower() in v.lower() or v.lower().endswith(flow_input.lower()):
                            d["flow"] = v
                            break

            for field in update_fields:
                val = d.get(field)
                if val not in (None, ""):
                    setattr(target, field, str(val).strip())
            db.commit()
            return json.dumps({"success": True, "ok": True, "id": target.id})
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] import_defined_batch_entry hatası: {e}")
            return json.dumps({"success": False, "ok": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def create_batch_entry(self, data_json):
        import uuid
        from models.batch_entry import BatchEntry
        db = SessionLocal()
        try:
            d = json.loads(data_json or "{}")

            # Model doğrulama (tanımsız model engelleme, örn. iPhone 19)
            model_val = d.get("model", "").strip()
            is_valid_m, m_err_msg = self._validate_product_model(db, model_val)
            if not is_valid_m:
                return json.dumps({"success": False, "message": m_err_msg})

            # Tüm giriş kuralları (farklı müşteri / üretimdeki batch / aktif cihaz) tek yerde.
            err = self._validate_new_batch_entry(db, d)
            if err:
                return json.dumps({"success": False, "message": err})

            imei_val = d.get("imei_number", "").strip()
            serial_val = d.get("serial_number", "").strip()

            valid_flow_values = self._get_flow_values(db)
            default_flow = "To refurbish" if "To refurbish" in valid_flow_values else (valid_flow_values[0] if valid_flow_values else "To refurbish")
            flow_input = (d.get("flow", "") or "").strip()
            
            # Esnek / Duyarsız eşleştirme (Örn: 'Refurbish' -> 'To refurbish', 'Repair' -> 'To repair')
            flow_value = None
            if flow_input:
                flow_map = {v.lower(): v for v in valid_flow_values}
                # Direkt (büyük/küçük harf duyarsız) eşleşme
                if flow_input.lower() in flow_map:
                    flow_value = flow_map[flow_input.lower()]
                else:
                    # Kısaltma/Takma ad eşleştirme
                    for v in valid_flow_values:
                        if flow_input.lower() in v.lower() or v.lower().endswith(flow_input.lower()):
                            flow_value = v
                            break

            if not flow_value:
                flow_value = default_flow if not flow_input else None

            if not flow_value:
                return json.dumps({"success": False, "message": f"Geçersiz Flow değeri: \"{flow_input}\". Geçerli değerler: {', '.join(valid_flow_values)}"})

            new_entry = BatchEntry(
                customer_no=d.get("customer_no", "").strip(),
                customer_name=d.get("customer_name", "").strip(),
                imei_number=imei_val,
                serial_number=serial_val,
                internal_id=d.get("internal_id", "").strip(),
                batch_no=d.get("batch_no", "").strip(),
                model=d.get("model", "").strip(),
                gb=d.get("gb", "").strip(),
                color=d.get("color", "").strip(),
                unit_price=float(d.get("unit_price") or 0.0),
                currency=d.get("currency", "EUR").strip() or "EUR",
                defects=d.get("defects", "").strip(),
                screen_test=d.get("screen_test", "").strip(),
                power_test=d.get("power_test", "").strip(),
                flow=flow_value,
                service_id=uuid.uuid4(),
            )
            db.add(new_entry)
            db.commit()
            return json.dumps({"success": True, "id": new_entry.id, "service_id": str(new_entry.service_id)})
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] create_batch_entry hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    def _try_update_defined_batch_entry(self, db, d):
        """import_defined_batch_entry'nin db.commit()/db.close() ÇAĞIRMAYAN gövdesi
        (paylaşılan session üzerinde toplu import için). Davranış birebir aynı."""
        from models.batch_entry import BatchEntry
        from sqlalchemy import func, or_

        imei = (d.get("imei_number") or "").strip()
        serial = (d.get("serial_number") or "").strip()
        internal = (d.get("internal_id") or "").strip()
        batch_no = (d.get("batch_no") or "").strip()

        model_val = (d.get("model") or "").strip()
        if model_val:
            is_valid_m, m_err_msg = self._validate_product_model(db, model_val)
            if not is_valid_m:
                return {"ok": False, "created": False, "message": m_err_msg}

        conds = []
        if imei:
            conds.append(func.lower(func.trim(BatchEntry.imei_number)) == imei.lower())
        if serial:
            conds.append(func.lower(func.trim(BatchEntry.serial_number)) == serial.lower())
        if internal:
            conds.append(func.lower(func.trim(BatchEntry.internal_id)) == internal.lower())

        target = None
        if conds:
            target = db.query(BatchEntry).filter(or_(*conds)).order_by(BatchEntry.id.desc()).first()
        if not target:
            ident = imei or serial or internal or "-"
            return {"ok": False, "created": False, "message": f"Cihaz ({ident}) sistemde tanımlı değil, içe aktarılmadı."}

        if batch_no:
            batch_exists = db.query(BatchEntry).filter(
                func.lower(func.trim(BatchEntry.batch_no)) == batch_no.lower()
            ).first()
            if not batch_exists:
                return {"ok": False, "created": False, "message": f"Batch numarası ({batch_no}) sistemde tanımlı değil, içe aktarılmadı."}

        update_fields = ["customer_no", "customer_name", "batch_no", "model", "gb", "color",
                         "defects", "screen_test", "power_test", "flow"]
        changed = False
        for field in update_fields:
            val = d.get(field)
            if val in (None, ""):
                continue
            if (getattr(target, field) or "") != str(val).strip():
                changed = True
                break
        if not changed:
            ident = imei or serial or internal or "-"
            return {"ok": False, "created": False, "message": f"Cihaz ({ident}) zaten aynı bilgilerle kayıtlı, değişiklik yok (mükerrer)."}

        # Power Test her zaman zorunludur. Bu satır power_test taşımasa bile (alan
        # değiştirilmediği için o zaman kayıttaki mevcut değer korunur), güncelleme
        # sonrasında oluşacak NİHAİ değeri mutasyondan ÖNCE hesaplayıp doğruluyoruz -
        # geçersizse target'a hiç dokunmadan erken dönüyoruz ki session kirlenmesin.
        final_power_test = str(d.get("power_test") or target.power_test or "").strip()
        if not final_power_test:
            return {"ok": False, "created": False, "message": "Power Test alanı zorunludur, boş bırakılamaz."}

        valid_flow_values = self._get_flow_values(db)
        flow_input = (d.get("flow", "") or "").strip()
        if flow_input:
            flow_map = {v.lower(): v for v in valid_flow_values}
            if flow_input.lower() in flow_map:
                d["flow"] = flow_map[flow_input.lower()]
            else:
                for v in valid_flow_values:
                    if flow_input.lower() in v.lower() or v.lower().endswith(flow_input.lower()):
                        d["flow"] = v
                        break

        for field in update_fields:
            val = d.get(field)
            if val not in (None, ""):
                setattr(target, field, str(val).strip())
        db.flush()
        return {"ok": True, "created": False, "message": "", "id": target.id}

    def _try_create_batch_entry(self, db, d):
        """create_batch_entry'nin db.commit()/db.close() ÇAĞIRMAYAN gövdesi (paylaşılan
        session üzerinde toplu import için). Davranış birebir aynı."""
        import uuid
        from models.batch_entry import BatchEntry

        model_val = d.get("model", "").strip()
        is_valid_m, m_err_msg = self._validate_product_model(db, model_val)
        if not is_valid_m:
            return {"ok": False, "created": False, "message": m_err_msg}

        err = self._validate_new_batch_entry(db, d)
        if err:
            return {"ok": False, "created": False, "message": err}

        imei_val = d.get("imei_number", "").strip()
        serial_val = d.get("serial_number", "").strip()

        valid_flow_values = self._get_flow_values(db)
        default_flow = "To refurbish" if "To refurbish" in valid_flow_values else (valid_flow_values[0] if valid_flow_values else "To refurbish")
        flow_input = (d.get("flow", "") or "").strip()

        flow_value = None
        if flow_input:
            flow_map = {v.lower(): v for v in valid_flow_values}
            if flow_input.lower() in flow_map:
                flow_value = flow_map[flow_input.lower()]
            else:
                for v in valid_flow_values:
                    if flow_input.lower() in v.lower() or v.lower().endswith(flow_input.lower()):
                        flow_value = v
                        break

        if not flow_value:
            flow_value = default_flow if not flow_input else None

        if not flow_value:
            return {"ok": False, "created": False, "message": f"Geçersiz Flow değeri: \"{flow_input}\". Geçerli değerler: {', '.join(valid_flow_values)}"}

        new_entry = BatchEntry(
            customer_no=d.get("customer_no", "").strip(),
            customer_name=d.get("customer_name", "").strip(),
            imei_number=imei_val,
            serial_number=serial_val,
            internal_id=d.get("internal_id", "").strip(),
            batch_no=d.get("batch_no", "").strip(),
            model=d.get("model", "").strip(),
            gb=d.get("gb", "").strip(),
            color=d.get("color", "").strip(),
            unit_price=float(d.get("unit_price") or 0.0),
            currency=d.get("currency", "EUR").strip() or "EUR",
            defects=d.get("defects", "").strip(),
            screen_test=d.get("screen_test", "").strip(),
            power_test=d.get("power_test", "").strip(),
            flow=flow_value,
            service_id=uuid.uuid4(),
        )
        db.add(new_entry)
        db.flush()
        return {"ok": True, "created": True, "message": "", "id": new_entry.id}

    @Slot(str, result=str)
    def bulk_process_batch_entries(self, rows_json):
        """Toplu (Excel) Batch Girişi içe aktarma. Her satır için önce mevcut cihazı
        güncellemeyi dener (_try_update_defined_batch_entry); mesajı "sistemde tanımlı
        değil" içeriyorsa (aynı frontend'deki eski substring kontrolüyle birebir aynı
        mantık) yeni kayıt oluşturmayı dener (_try_create_batch_entry). Eskiden bu N satır
        için N ayrı QWebChannel çağrısı + N ayrı DB commit gerekiyordu (uzak sunucuya N ayrı
        round-trip); şimdi TEK Slot çağrısında, satır başına SAVEPOINT (bir satırın hatası
        diğerlerini etkilemez) ile TEK connection üzerinden çalışıyor, sonda tek commit."""
        db = SessionLocal()
        try:
            try:
                rows = json.loads(rows_json or "[]")
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Geçersiz dosya verisi.", "results": []})

            if not rows:
                return json.dumps({"success": False, "message": "Dosyada içe aktarılacak satır bulunamadı.", "results": []})

            results = []
            for d in rows:
                d = dict(d or {})
                row_result = {"ok": False, "created": False, "message": ""}
                try:
                    with db.begin_nested():
                        update_result = self._try_update_defined_batch_entry(db, d)
                        if update_result["ok"] is False and update_result.get("message") and "sistemde tanımlı değil" in update_result["message"]:
                            row_result = self._try_create_batch_entry(db, d)
                        else:
                            row_result = update_result
                except Exception as row_ex:
                    row_result = {"ok": False, "created": False, "message": str(row_ex)}
                results.append(row_result)

            db.commit()
            return json.dumps({"success": True, "results": results}, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": f"İçe aktarma hatası: {str(e)}", "results": []})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def update_batch_entry(self, entry_id, data_json):
        from models.batch_entry import BatchEntry
        from datetime import datetime
        db = SessionLocal()
        try:
            entry = db.query(BatchEntry).filter(BatchEntry.id == int(entry_id)).first()
            if not entry:
                return json.dumps({"success": False, "message": "Kayıt bulunamadı."})

            d = json.loads(data_json or "{}")
            
            # Validation: Aynı batch numarasıyla farklı müşteri olamaz
            new_batch_no = d.get("batch_no", entry.batch_no).strip()
            new_customer_name = d.get("customer_name", entry.customer_name).strip()
            
            if new_batch_no:
                # Kendisi dışındaki kayıtları kontrol et
                existing_batch = db.query(BatchEntry).filter(
                    BatchEntry.batch_no == new_batch_no,
                    BatchEntry.id != entry.id
                ).first()
                
                if existing_batch and existing_batch.customer_name and existing_batch.customer_name.strip().lower() != new_customer_name.lower():
                    return json.dumps({
                        "success": False, 
                        "message": f"Bu batch numarası ({new_batch_no}) başka bir müşteriye ({existing_batch.customer_name}) aittir. Aynı batch numarasıyla farklı müşteri güncellenemez."
                    })
            entry.customer_no = d.get("customer_no", entry.customer_no).strip()
            entry.customer_name = d.get("customer_name", entry.customer_name).strip()
            entry.imei_number = d.get("imei_number", entry.imei_number).strip()
            entry.serial_number = d.get("serial_number", entry.serial_number).strip()
            entry.internal_id = d.get("internal_id", entry.internal_id).strip()
            if "model" in d:
                new_model = d.get("model", entry.model).strip()
                is_valid_m, m_err_msg = self._validate_product_model(db, new_model)
                if not is_valid_m:
                    return json.dumps({"success": False, "message": m_err_msg})
                entry.model = new_model
            entry.gb = d.get("gb", entry.gb).strip()
            entry.color = d.get("color", entry.color).strip()
            if "unit_price" in d:
                entry.unit_price = float(d.get("unit_price") or 0.0)
            if "currency" in d:
                entry.currency = d.get("currency", entry.currency).strip() or "EUR"
            entry.defects = d.get("defects", entry.defects).strip()
            entry.screen_test = d.get("screen_test", entry.screen_test).strip()
            entry.power_test = d.get("power_test", entry.power_test).strip()
            if not entry.power_test:
                return json.dumps({"success": False, "message": "Power Test alanı zorunludur, boş bırakılamaz."})
            if "flow" in d:
                new_flow = d.get("flow", entry.flow).strip()
                valid_flow_values = self._get_flow_values(db)
                if new_flow not in valid_flow_values:
                    return json.dumps({"success": False, "message": f"Geçersiz Flow değeri: \"{new_flow}\". Geçerli değerler: {', '.join(valid_flow_values)}"})
                entry.flow = new_flow
            entry.updated_at = datetime.now()

            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] update_batch_entry hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_batch_entry(self, entry_id):
        from models.batch_entry import BatchEntry
        db = SessionLocal()
        try:
            entry = db.query(BatchEntry).filter(BatchEntry.id == int(entry_id)).first()
            if not entry:
                return json.dumps({"success": False, "message": "Kayıt bulunamadı."})
            
            if entry.batch_no:
                db.query(BatchEntry).filter(BatchEntry.batch_no == entry.batch_no).delete(synchronize_session=False)
            else:
                db.delete(entry)
                
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] delete_batch_entry hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_batch_summary(self):
        from sqlalchemy import text
        db = SessionLocal()
        try:
            self._ensure_batch_entries_table()
            try:
                self.sync_customers_to_batch_entries()
            except Exception as sync_err:
                print(f"[WebBridge] sync_customers_to_batch_entries hatası: {sync_err}")

            sql = """
                SELECT
                    MAX(b.id) AS id,
                    COALESCE(NULLIF(b.batch_no, ''), 'Tanımsız Batch') AS document_number,
                    b.batch_no AS batch_no,
                    COALESCE(MAX(NULLIF(b.customer_name, '')), MAX(NULLIF(c.customer_name, '')), 'Tanımsız Müşteri') AS account_name,
                    COALESCE(MAX(NULLIF(b.customer_name, '')), MAX(NULLIF(c.customer_name, '')), 'Tanımsız Müşteri') AS customer_name,
                    COALESCE(MAX(NULLIF(b.customer_no, '')), MAX(NULLIF(c.code, '')), '-') AS customer_no,
                    COUNT(*) AS item_quantity,
                    COUNT(*) AS total_devices,
                    SUM(COALESCE(b.unit_price, 0)) AS total_price,
                    COALESCE(MAX(NULLIF(c.currency, '')), MAX(NULLIF(b.currency, '')), 'EUR') AS currency,
                    COALESCE(BOOL_AND(COALESCE(b.is_success, false)), false) AS is_success,
                    COALESCE(MAX(NULLIF(b.flow, '')), 'To refurbish') AS flow,
                    COALESCE(MAX(NULLIF(b.created_by, '')), 'io') AS create_by,
                    MAX(b.created_at) AS last_created
                FROM warehouse.batch_entries b
                LEFT JOIN LATERAL (
                    SELECT customer_name, code, currency
                    FROM warehouse.customers c
                    WHERE LOWER(b.customer_name) = LOWER(c.customer_name) OR b.customer_no = c.code
                    LIMIT 1
                ) c ON true
                GROUP BY
                    b.batch_no,
                    COALESCE(NULLIF(b.customer_no, ''), LOWER(NULLIF(b.customer_name, '')), 'tanimsiz')
                ORDER BY MAX(b.created_at) DESC;
            """
            rows = db.execute(text(sql)).mappings().all()

            batches = [{
                "id": r["id"],
                "document_date": r["last_created"].strftime("%d.%m.%Y") if r["last_created"] else "-",
                "document_number": r["document_number"],
                "account_name": r["account_name"] or "-",
                "is_success": bool(r["is_success"]),
                "item_quantity": int(r["item_quantity"]),
                "currency": r["currency"] or "EUR",
                "flow": r["flow"] or "To refurbish",
                "create_by": r["create_by"] or "io",

                # Legacy/compatibility fields
                "batch_no": r["batch_no"],
                "customer_name": r["customer_name"] or "-",
                "customer_no": r["customer_no"] or "-",
                "total_devices": int(r["total_devices"]),
                "total_price": float(r["total_price"] or 0.0),
                "last_created": r["last_created"].strftime("%d.%m.%Y %H:%M") if r["last_created"] else "-"
            } for r in rows]

            return json.dumps({"success": True, "batches": batches}, ensure_ascii=False)
        except Exception as e:
            print(f"[WebBridge] get_batch_summary hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def clear_all_batch_entries(self):
        from sqlalchemy import text
        db = SessionLocal()
        try:
            db.execute(text("TRUNCATE TABLE warehouse.batch_entries RESTART IDENTITY;"))
            db.commit()
            return json.dumps({"success": True, "message": "Tüm Batch kayıtları başarıyla temizlendi."})
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] clear_all_batch_entries hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def bulk_delete_batch_entries(self, ids_json):
        from models.batch_entry import BatchEntry
        db = SessionLocal()
        try:
            ids = json.loads(ids_json or "[]")
            if not ids:
                return json.dumps({"success": False, "message": "Silinecek kayıt seçilmedi."})
            int_ids = [int(i) for i in ids]
            
            # Find the batch numbers for the selected IDs to delete the entire batch
            target_batches = db.query(BatchEntry.batch_no).filter(BatchEntry.id.in_(int_ids)).all()
            batch_nos = [t[0] for t in target_batches if t[0]]
            
            db.query(BatchEntry).filter(
                (BatchEntry.batch_no.in_(batch_nos)) | (BatchEntry.id.in_(int_ids))
            ).delete(synchronize_session=False)
            
            db.commit()
            return json.dumps({"success": True, "count": len(int_ids)})
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] bulk_delete_batch_entries hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def bulk_update_batch_flow(self, ids_json, new_flow):
        from models.batch_entry import BatchEntry
        from datetime import datetime
        db = SessionLocal()
        try:
            ids = json.loads(ids_json or "[]")
            if not ids or not new_flow:
                return json.dumps({"success": False, "message": "Kayıt veya durum seçilmedi."})

            new_flow = str(new_flow).strip()
            valid_flow_values = self._get_flow_values(db)
            if new_flow not in valid_flow_values:
                return json.dumps({"success": False, "message": f"Geçersiz Flow değeri: \"{new_flow}\". Geçerli değerler: {', '.join(valid_flow_values)}"})

            int_ids = [int(i) for i in ids]

            # Find the batch numbers for the selected IDs to update the entire batch
            target_batches = db.query(BatchEntry.batch_no).filter(BatchEntry.id.in_(int_ids)).all()
            batch_nos = [t[0] for t in target_batches if t[0]]
            
            db.query(BatchEntry).filter(
                (BatchEntry.batch_no.in_(batch_nos)) | (BatchEntry.id.in_(int_ids))
            ).update(
                {BatchEntry.flow: str(new_flow).strip(), BatchEntry.updated_at: datetime.now()},
                synchronize_session=False
            )
            
            db.commit()
            return json.dumps({"success": True, "count": len(int_ids)})
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] bulk_update_batch_flow hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def lookup_batch_entry(self, search_term):
        from models.batch_entry import BatchEntry
        from sqlalchemy import func, text
        db = SessionLocal()
        try:
            term = (search_term or "").strip()
            if not term or len(term) < 2:
                return json.dumps({"success": False, "message": "Arama terimi çok kısa."})
            
            term_lower = term.lower()
            
            # 1. Search in warehouse.batch_entries (Exact Match)
            entry = db.query(BatchEntry).filter(
                (func.lower(func.trim(BatchEntry.imei_number)) == term_lower) |
                (func.lower(func.trim(BatchEntry.serial_number)) == term_lower) |
                (func.lower(func.trim(BatchEntry.internal_id)) == term_lower) |
                (func.lower(func.trim(BatchEntry.batch_no)) == term_lower) |
                (func.lower(func.trim(BatchEntry.customer_no)) == term_lower)
            ).order_by(BatchEntry.id.desc()).first()

            # 2. Search in warehouse.batch_entries (ILIKE Partial Match)
            if not entry and len(term) >= 3:
                entry = db.query(BatchEntry).filter(
                    (BatchEntry.imei_number.ilike(f"%{term}%")) |
                    (BatchEntry.serial_number.ilike(f"%{term}%")) |
                    (BatchEntry.internal_id.ilike(f"%{term}%")) |
                    (BatchEntry.batch_no.ilike(f"%{term}%")) |
                    (BatchEntry.customer_no.ilike(f"%{term}%"))
                ).order_by(BatchEntry.id.desc()).first()

            if entry:
                data = {
                    "customer_no": entry.customer_no or '',
                    "customer_name": entry.customer_name or '',
                    "imei_number": entry.imei_number or '',
                    "serial_number": entry.serial_number or '',
                    "internal_id": entry.internal_id or '',
                    "batch_no": entry.batch_no or '',
                    # Cihaz etiketindeki "Brand:" satırı için gerekli (bkz. EtiketYazdirModal).
                    "brand": entry.brand or '',
                    "model": entry.model or '',
                    "gb": entry.gb or '',
                    "color": entry.color or '',
                    "unit_price": float(entry.unit_price or 0.0),
                    "currency": entry.currency or 'EUR',
                    "defects": entry.defects or '',
                    "screen_test": entry.screen_test or '',
                    "power_test": entry.power_test or '',
                    "flow": entry.flow or 'To refurbish',
                    "statu_code": entry.statu_code,
                    "id": entry.id,
                    "service_id": str(entry.service_id) if entry.service_id else None
                }

                # Phonecheck test sonuclarini (battery cycle, battery health vb.) cek ve ekle
                from models.phonecheck_test_result import PhonecheckTestResult
                from services.phonecheck_service import PhonecheckService

                lookup_imei = entry.imei_number or entry.serial_number
                if lookup_imei:
                    pc = db.query(PhonecheckTestResult).filter(
                        PhonecheckTestResult.imei == lookup_imei
                    ).order_by(PhonecheckTestResult.fetched_at.desc()).first()

                    # Lokal veritabanında yoksa Phonecheck Cloud API'sinden canlı çek ve kaydet
                    if not pc:
                        try:
                            pc_svc = PhonecheckService(db)
                            fetched = pc_svc.fetch_device(lookup_imei)
                            if fetched.get("success") and fetched.get("device"):
                                pc = pc_svc.save_from_phonecheck(fetched["device"], test_stage="AUTO_LOOKUP", imei=lookup_imei)
                                db.commit()
                        except Exception as _e:
                            print(f"[Phonecheck Live Fetch Error]: {_e}")

                    if pc:
                        data["battery_cycle"] = pc.battery_cycle
                        data["battery_health_percentage"] = pc.battery_health_percentage
                        data["grade"] = pc.grade or data.get("grade", "")
                        data["defects"] = pc.failed or data.get("defects", "")

                return json.dumps({"success": True, "found": True, "data": data}, ensure_ascii=False)

            # 3. Search in warehouse.customers (MIO Create)
            c_row = db.execute(text("""
                SELECT id, customer_name, code, short_name, imei_number, serial_number, internal_id,
                       brand, model, flow, customer_reported_complaint, currency
                FROM warehouse.customers
                WHERE LOWER(TRIM(COALESCE(imei_number, ''))) = LOWER(:t)
                   OR LOWER(TRIM(COALESCE(serial_number, ''))) = LOWER(:t)
                   OR LOWER(TRIM(COALESCE(internal_id, ''))) = LOWER(:t)
                   OR LOWER(TRIM(COALESCE(code, ''))) = LOWER(:t)
                   OR LOWER(TRIM(CONCAT('BATCH-MIO-', id))) = LOWER(:t)
                   OR (LENGTH(:t) >= 3 AND (
                       COALESCE(imei_number, '') ILIKE :t_like OR
                       COALESCE(serial_number, '') ILIKE :t_like OR
                       COALESCE(internal_id, '') ILIKE :t_like OR
                       COALESCE(code, '') ILIKE :t_like
                   ))
                ORDER BY id DESC LIMIT 1
            """), {"t": term, "t_like": f"%{term}%"}).mappings().first()

            if c_row:
                data = {
                    "customer_no": c_row["code"] or '',
                    "customer_name": c_row["short_name"] or c_row["customer_name"] or '',
                    "imei_number": c_row["imei_number"] or '',
                    "serial_number": c_row["serial_number"] or '',
                    "internal_id": c_row["internal_id"] or '',
                    "batch_no": f"BATCH-MIO-{c_row['id']}",
                    "model": f"{c_row['brand'] or ''} {c_row['model'] or ''}".strip(),
                    "gb": '',
                    "color": '',
                    "unit_price": 0.0,
                    "currency": c_row["currency"] or 'EUR',
                    "defects": c_row["customer_reported_complaint"] or '',
                    "screen_test": '',
                    "power_test": '',
                    "flow": c_row["flow"] if c_row["flow"] in self._get_flow_values(db) else 'To refurbish',
                    "statu_code": None,
                    "id": None
                }
                return json.dumps({"success": True, "found": True, "data": data}, ensure_ascii=False)

            return json.dumps({"success": True, "found": False})
        except Exception as e:
            print(f"[WebBridge] lookup_batch_entry hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def sync_customers_to_batch_entries(self):
        """Müşteriler/MIO tablosundaki cihaz ve müşteri kayıtlarını Batch Girişi tablosuna aktarır.
        Her get_batch_summary çağrısında (Batch Girişi ekranı her açılışında) tetiklendiğinden,
        müşteri sayısı kadar ayrı 'var mı?'/'aktif servis mi?' sorgusu (N+1) çalıştırmak yerine,
        SADECE bu müşterilerin IMEI/seri/internal id/batch no'suyla eşleşen satırlar tek bir
        WHERE ... = ANY(:liste) sorgusuyla çekilir - warehouse.batch_entries'in TAMAMINI (7600+
        satır ve büyümeye devam ediyor) her senkronizasyonda indirmek yerine, sadece ilgili
        birkaç satır ağdan taşınır. Bu, get_batch_summary'nin en yavaş kısmıydı."""
        import uuid
        from sqlalchemy import text
        from datetime import datetime
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT id, customer_name, code, short_name, imei_number, serial_number, internal_id,
                       brand, model, flow, customer_reported_complaint, currency, created_at
                FROM warehouse.customers
            """)).mappings().all()

            valid_flow_values = self._get_flow_values(db)

            batch_nos = [f"BATCH-MIO-{r['id']}" for r in rows]
            imeis = list({(r["imei_number"] or "").strip() for r in rows if (r["imei_number"] or "").strip()})
            serials = list({(r["serial_number"] or "").strip() for r in rows if (r["serial_number"] or "").strip()})
            internals = list({(r["internal_id"] or "").strip() for r in rows if (r["internal_id"] or "").strip()})
            imeis_lower = [v.lower() for v in imeis]
            serials_lower = [v.lower() for v in serials]

            existing_rows = db.execute(text("""
                SELECT id, batch_no, imei_number, serial_number, internal_id, customer_name, customer_no, currency
                FROM warehouse.batch_entries
                WHERE batch_no = ANY(:batch_nos) OR imei_number = ANY(:imeis)
                   OR serial_number = ANY(:serials) OR internal_id = ANY(:internals)
            """), {"batch_nos": batch_nos, "imeis": imeis, "serials": serials, "internals": internals}).mappings().all()
            by_batch_no, by_imei, by_serial, by_internal = {}, {}, {}, {}
            for er in existing_rows:
                if er["batch_no"]: by_batch_no.setdefault(er["batch_no"], er)
                if er["imei_number"]: by_imei.setdefault(er["imei_number"], er)
                if er["serial_number"]: by_serial.setdefault(er["serial_number"], er)
                if er["internal_id"]: by_internal.setdefault(er["internal_id"], er)

            # _find_active_service_for_device ile aynı eşleşme mantığı (LOWER/TRIM, statü != 128),
            # ama tüm cihazlar için tek sorguda - idx_batch_entries_imei_lower/serial_lower
            # ifade indekslerini kullanır.
            active_rows = db.execute(text("""
                SELECT LOWER(TRIM(imei_number)) AS imei, LOWER(TRIM(serial_number)) AS serial
                FROM warehouse.batch_entries
                WHERE (LOWER(TRIM(imei_number)) = ANY(:imeis) OR LOWER(TRIM(serial_number)) = ANY(:serials))
                  AND COALESCE(statu_code, 100) != 128
            """), {"imeis": imeis_lower, "serials": serials_lower}).mappings().all()
            active_imeis = {r["imei"] for r in active_rows if r["imei"]}
            active_serials = {r["serial"] for r in active_rows if r["serial"]}

            added_count = 0
            skipped_active_count = 0
            insert_rows = []
            for r in rows:
                imei = (r["imei_number"] or "").strip()
                serial = (r["serial_number"] or "").strip()
                internal = (r["internal_id"] or "").strip()
                c_no = (r["code"] or "").strip()
                c_name = (r["short_name"] or r["customer_name"] or "").strip()
                mio_batch_no = f"BATCH-MIO-{r['id']}"

                existing = by_batch_no.get(mio_batch_no) or (imei and by_imei.get(imei)) or (serial and by_serial.get(serial)) or (internal and by_internal.get(internal))

                if existing:
                    changed_fields = {}
                    if c_name and existing["customer_name"] != c_name:
                        changed_fields["customer_name"] = c_name
                    if c_no and existing["customer_no"] != c_no:
                        changed_fields["customer_no"] = c_no
                    if r["currency"] and existing["currency"] != r["currency"].upper():
                        changed_fields["currency"] = r["currency"].upper()
                    if changed_fields:
                        changed_fields["updated_at"] = datetime.now()
                        set_clause = ", ".join(f"{k} = :{k}" for k in changed_fields)
                        db.execute(text(f"UPDATE warehouse.batch_entries SET {set_clause} WHERE id = :id"), {**changed_fields, "id": existing["id"]})
                else:
                    # Bu cihaz (IMEI/seri no) başka bir kaynaktan zaten aktif bir servis
                    # döngüsündeyse (statü 128 değilse) senkronizasyon bu satırı atlar -
                    # aynı cihaz için iki açık servis oluşturulamaz.
                    if (imei and imei.lower() in active_imeis) or (serial and serial.lower() in active_serials):
                        skipped_active_count += 1
                        continue

                    full_model = " ".join(filter(None, [r["brand"], r["model"]])).strip()
                    flow_val = (r["flow"] or "").strip()
                    if flow_val not in valid_flow_values:
                        flow_val = 'To refurbish' if 'To refurbish' in valid_flow_values else (valid_flow_values[0] if valid_flow_values else 'To refurbish')

                    insert_rows.append({
                        "customer_no": c_no or 'MIO-001',
                        "customer_name": c_name or 'MIO Müşterisi',
                        "imei_number": imei,
                        "serial_number": serial,
                        "internal_id": internal,
                        "batch_no": mio_batch_no,
                        "model": full_model,
                        "gb": '',
                        "color": '',
                        "unit_price": 0.0,
                        "currency": (r["currency"] or 'TRY').upper(),
                        "defects": r["customer_reported_complaint"] or '',
                        "screen_test": '',
                        "power_test": '',
                        "flow": flow_val,
                        "service_id": uuid.uuid4(),
                        "created_at": r["created_at"] or datetime.now(),
                        "updated_at": datetime.now(),
                    })
                    added_count += 1

            if insert_rows:
                db.execute(text("""
                    INSERT INTO warehouse.batch_entries
                        (customer_no, customer_name, imei_number, serial_number, internal_id, batch_no,
                         model, gb, color, unit_price, currency, defects, screen_test, power_test,
                         flow, service_id, created_at, updated_at)
                    VALUES
                        (:customer_no, :customer_name, :imei_number, :serial_number, :internal_id, :batch_no,
                         :model, :gb, :color, :unit_price, :currency, :defects, :screen_test, :power_test,
                         :flow, :service_id, :created_at, :updated_at)
                """), insert_rows)

            db.commit()
            return json.dumps({"success": True, "added_count": added_count, "skipped_active_count": skipped_active_count})
        except Exception as e:
            db.rollback()
            print(f"[WebBridge] sync_customers_to_batch_entries hatası: {e}")
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()






    # ---------------------------------------------------------
    # MODUL 5: STATE MACHINE VE DOA GUARDRAIL ENDPOINTLERI
    # ---------------------------------------------------------
    @Slot(str, result=str)
    def get_device_by_barcode(self, barcode):
        """Barkod/IMEI okutulduğunda ilgili iş emrini ve mevcut statü kodunu bulur."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            if not barcode or not barcode.strip():
                return json.dumps({"success": False, "message": "Barkod/IMEI boş olamaz"})

            term = barcode.strip()

            sr = db.execute(text("""
                SELECT id, customer_name, brand, model, imei_number, imei_serial
                FROM warehouse.service_records
                WHERE LOWER(TRIM(imei_number)) = LOWER(:term) OR LOWER(TRIM(imei_serial)) = LOWER(:term)
                ORDER BY id DESC LIMIT 1
            """), {"term": term}).mappings().first()

            if not sr:
                return json.dumps({"success": False, "message": f"'{term}' için kayıtlı bir cihaz bulunamadı."})

            wo = db.execute(text("""
                SELECT id, status
                FROM warehouse.work_orders
                WHERE service_record_id = :sr_id AND work_order_type = 'SERVICE'
                ORDER BY id DESC LIMIT 1
            """), {"sr_id": sr["id"]}).mappings().first()

            if not wo:
                return json.dumps({"success": False, "message": "Bu cihaza ait bir iş emri bulunamadı."})

            try:
                current_statu_code = int(wo["status"])
            except (TypeError, ValueError):
                current_statu_code = 100  # Henüz numaralı statü akışına alınmamış eski/legacy iş emri

            return json.dumps({
                "success": True,
                "work_order_id": wo["id"],
                "service_record_id": sr["id"],
                "imei": sr["imei_number"] or sr["imei_serial"] or term,
                "customer_name": sr["customer_name"],
                "model": " ".join(filter(None, [sr["brand"], sr["model"]])) or None,
                "current_statu_code": current_statu_code,
                "raw_status": wo["status"],
            })
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_all_statu_transitions(self):
        """Statü Geçiş Ekranı'nın buton menüsü için tüm (enabled) kaynak→hedef
        statü geçişlerini rol (to_dest) bilgisiyle birlikte döner."""
        from models.service_statu_map import ServiceStatuMap
        db = SessionLocal()
        try:
            rows = db.query(ServiceStatuMap).filter_by(enabled=True).order_by(
                ServiceStatuMap.to_dest, ServiceStatuMap.order_number
            ).all()
            transitions = [{
                "code": r.code,
                "parent_statu": r.parent_statu,
                "child_statu": r.child_statu,
                "is_positive": r.is_positive,
                "to_dest": r.to_dest,
                "short_name": r.short_name,
            } for r in rows]
            return json.dumps({"success": True, "transitions": transitions})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(int, result=str)
    def get_allowed_transitions(self, current_statu_code):
        from services.state_machine_service import StateMachineService
        db = SessionLocal()
        try:
            svc = StateMachineService(db)
            transitions = svc.get_allowed_transitions(current_statu_code)
            return json.dumps({"success": True, "transitions": transitions})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, int, int, str, str, result=str)
    def execute_statu_transition(self, service_record_id, current_statu_code, target_statu_code, request_type_code, test_result_code):
        from services.state_machine_service import StateMachineService
        from services.repair_service import RepairService
        
        db = SessionLocal()
        try:
            # 1. DOA Guardrail Kontrolü: Eger RMA'ya (1003, 134 vb veya Red durumu ise) gidiyorsa kontrol et
            is_rma = target_statu_code in [1003, 134, 135, 136] or (test_result_code and "Fail" in test_result_code)
            if is_rma:
                # Kullanılmıs parcalari kontrol et
                doa_check = self._internal_check_doa(db, service_record_id)
                if doa_check.get("has_consumed_parts"):
                    return json.dumps({
                        "success": False, 
                        "error_code": "DOA_TRANSFER_REQUIRED",
                        "message": "Bu cihaz iadeye yönlendirilmiştir! Üzerinde depodan çıkılmış parçalar var.",
                        "parts": doa_check.get("parts")
                    })
            
            # 2. State Machine Validasyonu
            svc = StateMachineService(db)
            # If empty string, convert to None
            req_type = request_type_code if request_type_code else None
            test_res = test_result_code if test_result_code else None
            
            result = svc.execute_transition(current_statu_code, target_statu_code, req_type, test_res)
            
            if not result.get("success"):
                return json.dumps(result)
                
            new_statu = result.get("new_statu_code")
            
            # 3. Ana kaydi guncelle
            # service_records tablosu (veya projede kullanilan tablo). Projede ana tablo "service_records" olmayabilir
            # Fakat web_bridge icinde service_records var oldugu farz ediliyor
            # execute_raw ile de yapabiliriz tablo adi net degilse
            from sqlalchemy import text
            db.execute(text("UPDATE warehouse.work_orders SET status = :st WHERE id = :id"), {"st": str(new_statu), "id": service_record_id})
            
            # 4. Eger yeni statu 109 ise Alt Onarim (RepairRecord) uret (Modul 4)
            if new_statu == 109:
                # Cihazin modelini ve kategorisini almaliyiz.
                # Demo amaciyla statik bir model ve kategori gonderiyoruz (Bunu gercekte veritabanindan ceker)
                rep_svc = RepairService(db)
                # Ozet olarak "Ti-Battery" gibi faultlari getirmeliyiz.
                # Bunu simdilik bos gecelim veya generic bir ti-battery verelim (test amacli)
                rep_svc.generate_concurrent_repairs(service_record_id, "iP11", ["Ti-Battery"])
                
            db.commit()
            return json.dumps({"success": True, "new_statu_code": new_statu, "message": result.get("message")})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # ---------------------------------------------------------
    # PARTI/BATCH STATU GECIS EKRANI (ortak kullanilan basit ekran)
    # Kod tablosu ve gecis kurallari service_statu / service_statu_map
    # uzerinden okunur; sadece hedef kayit batch_entries.statu_code'dur.
    # ---------------------------------------------------------
    def _find_batch_entry_by_term(self, db, term):
        from models.batch_entry import BatchEntry
        from sqlalchemy import func as sqlfunc
        term_clean = term.strip()
        return db.query(BatchEntry).filter(
            (sqlfunc.lower(sqlfunc.trim(BatchEntry.imei_number)) == term_clean.lower()) |
            (sqlfunc.lower(sqlfunc.trim(BatchEntry.serial_number)) == term_clean.lower()) |
            (sqlfunc.lower(sqlfunc.trim(BatchEntry.internal_id)) == term_clean.lower()) |
            (sqlfunc.lower(sqlfunc.trim(BatchEntry.batch_no)) == term_clean.lower())
        ).order_by(BatchEntry.id.desc()).first()

    @Slot(str, result=str)
    def scan_batch_entry_statu(self, term):
        """IMEI/Seri/Internal ID/Batch No okutulduğunda partiyi bulur, mevcut statüsünü
        ve (varsa) izinli bir sonraki statüsünü/statülerini döner. Geçişi uygulamaz."""
        from models.service_statu import ServiceStatu
        from services.state_machine_service import StateMachineService
        db = SessionLocal()
        try:
            if not term or not term.strip():
                return json.dumps({"success": False, "message": "IMEI/Seri/Internal ID/Batch No boş olamaz"})

            entry = self._find_batch_entry_by_term(db, term)
            if not entry:
                return json.dumps({"success": False, "message": f"'{term.strip()}' için kayıtlı bir parti/cihaz bulunamadı."})

            current_code = entry.statu_code if entry.statu_code is not None else 100
            current_statu = db.query(ServiceStatu).filter_by(code=current_code).first()
            current_name = current_statu.short_name if current_statu else str(current_code)

            svc = StateMachineService(db)
            transitions = svc.get_allowed_transitions(current_code)

            return json.dumps({
                "success": True,
                "entry_id": entry.id,
                "imei": entry.imei_number or entry.serial_number or entry.internal_id or "",
                "batch_no": entry.batch_no or "",
                "flow": entry.flow or "",
                # Etiket basımı için gerekli cihaz alanları (bkz. EtiketYazdirModal).
                "serial_number": entry.serial_number or "",
                "internal_id": entry.internal_id or "",
                "brand": entry.brand or "",
                "model": entry.model or "",
                "gb": entry.gb or "",
                "color": entry.color or "",
                "current_statu_code": current_code,
                "current_statu_name": current_name,
                "transitions": transitions,
            })
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_phonecheck_device_by_imei(self, term):
        """Batch girisi ekraninda IMEI/Seri okutuldugunda cihaz bilgilerini (model,
        hafiza, renk, seri, grade, test sonuclari) dogrudan Phonecheck'ten ceker.
        YAN ETKISIZDIR: hicbir statu gecisi yapmaz, kayit olusturmaz. Phonecheck'te
        batch numarasi ve musteri bilgisi bulunmadigindan bunlar donmez."""
        from services.phonecheck_service import PhonecheckService

        t = (term or "").strip()
        if not t:
            return json.dumps({"success": False, "message": "IMEI/Seri boş olamaz."})

        db = SessionLocal()
        try:
            pc = PhonecheckService(db)
            fetched = pc.fetch_device(t)
            if not fetched.get("success"):
                return json.dumps({
                    "success": False,
                    "needs_manual": fetched.get("needs_manual", False),
                    "message": fetched.get("message", "Phonecheck'te cihaz bulunamadı."),
                })

            d = fetched["device"]

            def _pick(*keys):
                for k in keys:
                    v = d.get(k)
                    if v not in (None, ""):
                        return str(v)
                return ""

            # Working/Failed alanindan basit ekran/guc testi cikarimi
            failed = _pick("Failed").lower()
            working = _pick("Working").lower()
            fail_all = (working == "no")
            screen_test = "BAŞARISIZ" if (fail_all or "lcd" in failed or "screen" in failed or "touch" in failed) else "BAŞARILI"
            power_test = "BAŞARISIZ" if (fail_all or "power" in failed or "boot" in failed) else "BAŞARILI"

            data = {
                "imei_number": _pick("IMEI"),
                "serial_number": _pick("Serial"),
                "model": _pick("Model"),
                "gb": _pick("Memory"),
                "color": _pick("Color"),
                "grade": _pick("Grade"),
                "defects": _pick("Failed"),
                "notes": _pick("Notes"),
                "screen_test": screen_test,
                "power_test": power_test,
                "battery_cycle": d.get("BatteryCycle"),
                "battery_health_percentage": d.get("BatteryHealthPercentage"),
            }
            return json.dumps({"success": True, "found": True, "data": data}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def fetch_phonecheck_and_transition(self, imei):
        """IMEI/Seri/Internal ID okutulduğunda partiyi bulur, Phonecheck'ten cihaz/test
        bilgisini ceker (IMEI varsa IMEI, yoksa Seri Numarasi ile sorgular), sonucu
        Pass1/Fail1'e cevirir ve batch_entries.statu_code'u Statumap akis semasina gore gunceller."""
        from models.batch_entry import BatchEntry
        from models.phonecheck_test_result import PhonecheckTestResult
        from services.phonecheck_service import PhonecheckService
        from services.state_machine_service import StateMachineService

        term = (imei or "").strip()
        if not term:
            return json.dumps({"success": False, "message": "IMEI/Seri/Internal ID boş olamaz."})

        db = SessionLocal()
        try:
            # IMEI, Seri Numarasi, Internal ID veya Batch No ile okutulmus olabilir.
            entry = self._find_batch_entry_by_term(db, term)
            if not entry:
                return json.dumps({"success": False, "message": f"'{term}' için kayıtlı bir batch_entries kaydı bulunamadı."})

            # Phonecheck'e her zaman gercek IMEI ile, o yoksa Seri Numarasi ile sorulur
            # (Internal ID Phonecheck'te taninmadigindan sorguda kullanilmaz).
            pc_term = (entry.imei_number or entry.serial_number or "").strip()
            if not pc_term:
                return json.dumps({"success": False, "message": "Bu cihaz için IMEI veya Seri Numarası tanımlı değil, Phonecheck sorgusu yapılamıyor."})

            pc = PhonecheckService(db)
            svc = StateMachineService(db)
            allowed = svc.get_allowed_transitions(entry.statu_code)
            positive = next((t for t in allowed if t["is_positive"]), None)

            # Test asamasi kodu service_statu_map.code formatindadir: "103_104"
            test_stage = (pc.build_stage(entry.statu_code, positive["target_statu_code"])
                          if positive else str(entry.statu_code))

            # Bu adimda basarisiz deneme hakki dolmus mu?
            if pc.failed_limit_reached(pc_term, test_stage):
                from services.phonecheck_service import MAX_FAILED_ATTEMPTS
                return json.dumps({
                    "success": False,
                    "message": f"Bu cihaz için bu test adımında en fazla {MAX_FAILED_ATTEMPTS} başarısız deneme hakkı var, hak doldu.",
                })

            fetched = pc.fetch_device(pc_term)
            if not fetched.get("success"):
                # Cihaz Phonecheck'te yok -> arayuz manuel doldurma formunu acmali
                fetched["test_stage"] = test_stage
                return json.dumps(fetched)

            device = fetched["device"]
            test_result_code = pc.to_test_result_code(device)

            # Her sorgu (sonuc ne olursa olsun) phonecheck_test_results'a kaydedilir.
            pc.save_from_phonecheck(device, test_stage, imei=pc_term)

            if test_result_code is None:
                return json.dumps({
                    "success": True,
                    "pending": True,
                    "message": "Phonecheck testi henüz tamamlanmadı, statü değiştirilmedi.",
                    "raw": device,
                })

            if not positive:
                return json.dumps({
                    "success": False,
                    "message": f"{entry.statu_code} statüsünden izinli bir sonraki geçiş tanımlı değil.",
                })

            old_statu_code = entry.statu_code
            result = svc.execute_transition(old_statu_code, positive["target_statu_code"], None, test_result_code)
            if not result.get("success"):
                return json.dumps(result)

            entry.statu_code = result["new_statu_code"]
            entry.is_success = (test_result_code == "Pass1")
            self._record_statu_change(
                db, entry.id, entry.imei_number, old_statu_code, entry.statu_code,
                staff=getattr(entry, "created_by", None),
                note=f"Test sonucu ({test_result_code}) — {result.get('message') or ''}".strip(),
            )
            db.commit()
            clear_api_cache()

            return json.dumps({
                "success": True,
                "imei": pc_term,
                "old_statu_code": old_statu_code,
                "new_statu_code": entry.statu_code,
                "test_result_code": test_result_code,
                "message": result.get("message"),
            })
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, int, int, str, result=str)
    def fetch_phonecheck_test(self, term, current_statu_code, target_statu_code, note=""):
        """Belirtilen statü geçişi için Phonecheck'ten test verisini çeker ve kaydeder.
        Statüyü DEĞİŞTİRMEZ - sadece test kaydı düşer. Geçişi de birlikte yapmak için
        fetch_phonecheck_and_transition kullanılır.

        test_stage her zaman service_statu_map.code formatındadır: "103_104", "125_109".

        note doluysa kaydedilen satırın notes alanına yazılır (ekrandan girilen not).

        Cihaz Phonecheck'te bulunamazsa needs_manual=True döner; bu durumda
        arayüz manuel doldurma formunu açmalı ve save_phonecheck_manual çağırmalıdır."""
        from services.phonecheck_service import PhonecheckService, MAX_FAILED_ATTEMPTS
        db = SessionLocal()
        try:
            svc = PhonecheckService(db)
            stage = svc.build_stage(current_statu_code, target_statu_code)

            if svc.failed_limit_reached(term, stage):
                return json.dumps({
                    "success": False,
                    "message": f"Bu cihaz için bu test adımında en fazla {MAX_FAILED_ATTEMPTS} başarısız deneme hakkı var, hak doldu.",
                    "test_stage": stage,
                })

            result = svc.fetch_device(term)
            if not result.get("success"):
                result["test_stage"] = stage
                return json.dumps(result)

            device = result["device"]
            record = svc.save_from_phonecheck(device, stage, imei=term)
            note = (note or "").strip()
            if note:
                record.notes = note
                db.commit()
            return json.dumps({
                "success": True,
                "test_stage": stage,
                "test_result_code": svc.to_test_result_code(device),
                "record_id": record.id,
                "attempt_no": record.attempt_no,
                "working": record.working,
                "grade": record.grade,
                "failed": record.failed,
            })
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, result=str)
    def save_phonecheck_manual(self, imei, test_stage, manual_reason, entered_by, fields_json):
        """Phonecheck'te bulunamayan cihaz için elle girilen test verisini kaydeder.
        manual_reason (açıklama) zorunludur."""
        from services.phonecheck_service import PhonecheckService
        db = SessionLocal()
        try:
            fields = json.loads(fields_json or "{}")
            svc = PhonecheckService(db)
            result = svc.save_manual(
                imei=imei,
                test_stage=test_stage,
                manual_reason=manual_reason,
                entered_by=entered_by or None,
                fields=fields,
            )
            return json.dumps(result)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_detected_parts_by_imei(self, term):
        """Depo > Servis ekranindaki Tespit Parca tablosu icin: test_result_faults
        tablosundaki hatali_parcaN/hataN ciftlerini duz satirlara cevirir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            term = (term or "").strip()
            if not term:
                return json.dumps({"success": False, "message": "IMEI boş olamaz."})

            entry = self._find_batch_entry_by_term(db, term)
            lookup_imei = (entry.imei_number if entry else None) or term
            lookup_internal = (entry.internal_id if entry else None) or ""

            rows = db.execute(text("""
                SELECT id, description,
                       hatali_parca1, hata1, hatali_parca2, hata2, hatali_parca3, hata3,
                       hatali_parca4, hata4, hatali_parca5, hata5, hatali_parca6, hata6,
                       hatali_parca7, hata7, hatali_parca8, hata8, hatali_parca9, hata9,
                       hatali_parca10, hata10
                FROM warehouse.test_result_faults
                WHERE imei_number = :imei OR (internal_id IS NOT NULL AND internal_id = :internal_id)
                ORDER BY created_at DESC
            """), {"imei": lookup_imei, "internal_id": lookup_internal}).fetchall()

            items = []
            for row in rows:
                record_id, description = row[0], row[1]
                pairs = [(row[2 + i * 2], row[3 + i * 2]) for i in range(10)]
                for idx, (part, fault) in enumerate(pairs, start=1):
                    if not part and not fault:
                        continue
                    items.append({
                        "id": f"{record_id}-{idx}",
                        "name": part or "",
                        "status": "",
                        "factorySerial": "",
                        "notice": description or "",
                        "currentSerial": "",
                        "test": fault or "",
                    })

            return json.dumps({"success": True, "items": items})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_repair_records_by_imei(self, term):
        """Depo > Servis ekranindaki Alt Onarimlar / Onarim Parca ve Iscilikleri
        tablolari icin: repair_records'tan bu cihaza ait kayitlari doner.
        QAC alani, o cihaza ait en guncel Phonecheck test sonucundan (grade/working) doldurulur —
        TEC/SupplyStatu/Labour Phonecheck'te karsiligi olmayan ic surec bilgileri oldugu icin bos kalir."""
        from sqlalchemy import text
        from models.phonecheck_test_result import PhonecheckTestResult
        db = SessionLocal()
        try:
            term = (term or "").strip()
            if not term:
                return json.dumps({"success": False, "message": "IMEI boş olamaz."})

            entry = self._find_batch_entry_by_term(db, term)
            lookup_imei = (entry.imei_number if entry else None) or term

            rows = db.execute(text("""
                SELECT rr.department_mission, rrt.short_name, rr.operation_type_code,
                       rr.item_category, rr.part_item_code, rr.item_fault_code, rr.notes,
                       rr.created_at, rr.updated_at
                FROM warehouse.repair_records rr
                LEFT JOIN warehouse.repair_result_type rrt ON rrt.code = rr.repair_result_type_code
                WHERE rr.service_record_id = :imei
                ORDER BY rr.created_at DESC
            """), {"imei": lookup_imei}).fetchall()

            latest_pc = db.query(PhonecheckTestResult).filter(
                PhonecheckTestResult.imei == lookup_imei
            ).order_by(PhonecheckTestResult.fetched_at.desc()).first()

            qac_value = ""
            battery_cycle = None
            battery_health = None
            if latest_pc:
                qac_value = latest_pc.grade or latest_pc.working or ""
                battery_cycle = latest_pc.battery_cycle
                battery_health = latest_pc.battery_health_percentage

            def fmt(dt):
                return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

            items = [{
                "missionGroup": r[0] or "",
                "repairStatu": r[1] or "",
                "tec": "",
                "repairStartTime": fmt(r[7]),
                "repairFinishTime": fmt(r[8]) if r[8] and r[8] != r[7] else "",
                "qac": qac_value,
                "testResult": r[5] or r[6] or "",
                "item": r[4] or "",
                "type": r[3] or r[2] or "",
                "supplyStatu": "",
                "labour": "",
                "fault": r[5] or "",
            } for r in rows]

            return json.dumps({
                "success": True,
                "items": items,
                "battery_cycle": battery_cycle,
                "battery_health": battery_health,
            })
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_status_history_by_imei(self, term):
        """Depo > Servis ekranindaki Durum sekmesi icin cihazin EKSIKSIZ statu gecmisini uretir.
        Uc kaynak birlestirilir:
          1) batch_entry_statu_history — her gecisin kalici logu (bundan sonraki gecisler tam).
          2) phonecheck_test_results — test_stage 'AAA_BBB' formatindan gecmis geriye donuk
             yeniden kurulur (log tablosu olmadan once yapilmis gecisler icin).
          3) Kayit (100) + guncel statu — her zaman iki uc nokta olarak eklenir.
        Ayni statu+dakika birden fazla kaynaktan gelirse tekillestirilir (log > phonecheck >
        sentetik oncelik sirasiyla). Sonuc en yeni ustte olacak sekilde siralanir."""
        from models.service_statu import ServiceStatu
        from models.batch_entry_statu_history import BatchEntryStatuHistory
        from models.phonecheck_test_result import PhonecheckTestResult
        db = SessionLocal()
        try:
            term = (term or "").strip()
            if not term:
                return json.dumps({"success": False, "message": "IMEI boş olamaz."})

            entry = self._find_batch_entry_by_term(db, term)
            if not entry:
                return json.dumps({"success": False, "message": f"'{term}' için kayıtlı bir cihaz bulunamadı."})

            # Statü etiketi (Türkçe) — tekrar sorguyu önlemek için cache'lenir.
            _label_cache = {}
            def label(code):
                if code is None:
                    return ""
                if code not in _label_cache:
                    s = db.query(ServiceStatu).filter_by(code=code).first()
                    _label_cache[code] = statu_label_tr(code, s.short_name if s else None)
                return _label_cache[code]

            def fmt(dt):
                return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

            # (öncelik, dt, kod, staff, text) topla; öncelik yüksek olan tekilleştirmede kazanır.
            events = []

            # 1) Kalıcı geçiş logu (en güvenilir kaynak)
            lookup_imei = (entry.imei_number or entry.serial_number or term).strip()
            log_rows = (db.query(BatchEntryStatuHistory)
                        .filter((BatchEntryStatuHistory.batch_entry_id == entry.id)
                                | (BatchEntryStatuHistory.imei == lookup_imei))
                        .all())
            for r in log_rows:
                events.append((3, r.created_at, r.new_statu_code, r.staff_name or "",
                               r.note or f"{label(r.old_statu_code)} → {label(r.new_statu_code)}"))

            # 2) Phonecheck test adımlarından geriye dönük yeniden kurma.
            #    test_stage 'AAA_BBB' => cihaz o anda BBB statüsüne geçmiştir.
            pc_rows = (db.query(PhonecheckTestResult)
                       .filter(PhonecheckTestResult.imei == lookup_imei)
                       .all())
            for r in pc_rows:
                stage = (r.test_stage or "").strip()
                parts = stage.split("_")
                if len(parts) == 2 and parts[1].isdigit():
                    new_code = int(parts[1])
                    src = label(int(parts[0])) if parts[0].isdigit() else parts[0]
                    events.append((2, r.fetched_at, new_code, getattr(r, "manual_entered_by", "") or "",
                                   f"Test adımı: {src} → {label(new_code)}"))

            # 3) Uç noktalar: güncel statü + kayıt (100). Log/phonecheck boşsa da her zaman görünür.
            last_update = entry.statu_update_time or entry.updated_at
            events.append((1, last_update, entry.statu_code, "",
                           f"Güncel durum: {label(entry.statu_code)}"))
            events.append((0, entry.created_at, 100, entry.created_by or "",
                           "Parti/cihaz sisteme kaydedildi."))

            # Tekilleştir: aynı (statü kodu, dakika) için en yüksek öncelikli olayı tut.
            best = {}
            for prio, dt, code, staff, text in events:
                if code is None:
                    continue
                key = (code, fmt(dt))
                if key not in best or prio > best[key][0]:
                    best[key] = (prio, dt, code, staff, text)

            # En yeni üstte (en son yapılan statü işlemi ilk sırada): dakikaya yuvarlanmış
            # string yerine SANİYE hassasiyetinde gerçek datetime'a göre azalan sıralanır —
            # böylece aynı dakika içinde yapılan ardışık geçişler de doğru sırada gelir.
            # Naive (tz'siz) değerler UTC kabul edilir ki karşılaştırma tutarlı olsun.
            def _sort_key(e):
                dt = e[1]
                if not dt:
                    return float("-inf")
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=_dt.timezone.utc)
                return dt.timestamp()
            ordered = sorted(best.values(), key=_sort_key, reverse=True)

            items = [{
                "date": fmt(dt),
                "staffName": staff,
                "statu": label(code),
                "text": text,
            } for (_prio, dt, code, staff, text) in ordered]

            return json.dumps({"success": True, "items": items})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_phonecheck_history_by_imei(self, term):
        """Depo > Servis ekranindaki Test sekmesi icin: bu cihaza ait tum
        phonecheck_test_results kayitlarini (en yeni once) doner."""
        from models.phonecheck_test_result import PhonecheckTestResult
        db = SessionLocal()
        try:
            term = (term or "").strip()
            if not term:
                return json.dumps({"success": False, "message": "IMEI boş olamaz."})

            entry = self._find_batch_entry_by_term(db, term)
            lookup_imei = (entry.imei_number or entry.serial_number) if entry else None
            lookup_imei = (lookup_imei or term).strip()

            rows = (db.query(PhonecheckTestResult)
                    .filter(PhonecheckTestResult.imei == lookup_imei)
                    .order_by(PhonecheckTestResult.fetched_at.desc())
                    .all())

            def fmt(dt):
                return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

            # Kritik parca orijinallik kontrolu (Ana Kamera / Batarya / Eski Pil).
            # Her kayit icin HER ZAMAN 3 satir doner; veri yoksa "unknown" (gri).
            from services.phonecheck_service import parse_all_parts, parse_critical_parts

            items = []
            for r in rows:
                raw_parts = getattr(r, "parts", None)
                critical, parts_remark = parse_critical_parts(raw_parts)
                all_parts, _ = parse_all_parts(raw_parts)
                items.append({
                    "deviceUpdatedD": fmt(r.fetched_at),
                    "grade": r.grade or "",
                    "partInfoRemark": r.notes or "",
                    # "Parts" artik gercekten Phonecheck'in Parts verisidir. Eskiden bu
                    # alan r.failed (basarisiz testler) tasiyordu - sutun basligi "Parts"
                    # oldugu icin yanlisti; basarisiz testler ayri "Failed" alanina alindi.
                    "parts": all_parts,
                    "partsRemark": parts_remark,
                    "criticalParts": critical,
                    "failed": r.failed or "",
                    "stationID": r.station_id or "",
                    "version": r.version or "",
                    "batteryCycle": r.battery_cycle if r.battery_cycle is not None else "",
                })

            return json.dumps({"success": True, "items": items})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_phonecheck_stored_by_imei(self, term):
        """Cihaza ait EN GÜNCEL phonecheck_test_results kaydını (yerel tablo) TÜM alanlarıyla
        döner. Yerel kayıt YOKSA canlı Phonecheck API'sine düşer (fallback) — böylece test
        aşamasından geçmemiş IMEI'ler için de veri gelir. Servis Onarımları / Teknisyen
        ekranlarındaki Müşteri Arıza Tespiti, Notes ve batarya bilgileri buradan doldurulur."""
        from models.phonecheck_test_result import PhonecheckTestResult
        db = SessionLocal()
        try:
            term = (term or "").strip()
            if not term:
                return json.dumps({"success": False, "message": "IMEI boş olamaz."})

            entry = self._find_batch_entry_by_term(db, term)
            lookup_imei = ((entry.imei_number or entry.serial_number) if entry else None) or term
            lookup_imei = lookup_imei.strip()

            r = (db.query(PhonecheckTestResult)
                 .filter(PhonecheckTestResult.imei == lookup_imei)
                 .order_by(PhonecheckTestResult.fetched_at.desc())
                 .first())
            if not r:
                # Yerel kayıt yoksa CANLI Phonecheck API'sinden çek — böylece test aşamasından
                # geçmemiş / tabloya yansımamış IMEI'ler için de Müşteri Arıza Tespiti ve Notes gelir.
                try:
                    from services.phonecheck_service import PhonecheckService
                    pc = PhonecheckService(db)
                    fetched = pc.fetch_device(lookup_imei)
                    if fetched.get("success"):
                        dv = fetched.get("device") or {}
                        def _g(*keys):
                            for k in keys:
                                v = dv.get(k)
                                if v not in (None, ""):
                                    return str(v)
                            return ""
                        def _gi(key):
                            try:
                                v = dv.get(key)
                                return int(float(v)) if v not in (None, "") else None
                            except (TypeError, ValueError):
                                return None
                        live = {
                            "imei": _g("IMEI") or lookup_imei,
                            "test_stage": "", "test_type": _g("Type"),
                            "test_start_time": _g("StartTime"), "test_end_time": _g("EndTime"),
                            "station_id": _g("StationID"), "working": _g("Working"),
                            "passed": _g("Passed"), "failed": _g("Failed"), "pending": _g("Pending"),
                            "model": _g("Model"), "memory": _g("Memory"), "serial": _g("Serial"),
                            "color": _g("Color"), "grade": _g("Grade"), "version": _g("Version"),
                            "notes": _g("Notes"),
                            "battery_cycle": _gi("BatteryCycle"),
                            "battery_health_percentage": _gi("BatteryHealthPercentage"),
                            "grading_results": _g("GradingResults"),
                            "fetched_at": "", "is_manual": False, "manual_reason": "",
                        }
                        return json.dumps({"success": True, "found": True, "source": "live", "data": live}, ensure_ascii=False)
                except Exception as live_err:
                    print(f"[WebBridge] get_phonecheck_stored_by_imei canlı fallback hatası: {live_err}")
                return json.dumps({"success": True, "found": False, "data": None})

            data = {
                "imei": r.imei or "",
                "test_stage": r.test_stage or "",
                "test_type": r.test_type or "",
                "test_start_time": r.test_start_time or "",
                "test_end_time": r.test_end_time or "",
                "station_id": r.station_id or "",
                "working": r.working or "",
                "passed": r.passed or "",
                "failed": r.failed or "",
                "pending": r.pending or "",
                "model": r.model or "",
                "memory": r.memory or "",
                "serial": r.serial or "",
                "color": r.color or "",
                "grade": r.grade or "",
                "version": r.version or "",
                "notes": r.notes or "",
                "battery_cycle": r.battery_cycle,
                "battery_health_percentage": r.battery_health_percentage,
                "grading_results": r.grading_results or "",
                "fetched_at": r.fetched_at.strftime("%Y-%m-%d %H:%M") if r.fetched_at else "",
                "is_manual": bool(r.is_manual),
                "manual_reason": r.manual_reason or "",
            }
            return json.dumps({"success": True, "found": True, "data": data}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_service_info_by_imei(self, term):
        """Depo > Servis ekranindaki IMEI aramasi icin: batch_entries'teki tum
        bilinen bilgileri INFO_FIELDS ile ayni anahtarlarla doner."""
        from models.service_statu import ServiceStatu
        db = SessionLocal()
        try:
            term = (term or "").strip()
            if not term:
                return json.dumps({"success": False, "message": "IMEI boş olamaz."})

            entry = self._find_batch_entry_by_term(db, term)
            if not entry:
                return json.dumps({"success": False, "message": f"'{term}' için kayıtlı bir cihaz bulunamadı."})

            statu = db.query(ServiceStatu).filter_by(code=entry.statu_code).first()

            def fmt(dt):
                return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

            return json.dumps({
                "success": True,
                "fields": {
                    "serviceNumber": str(entry.service_id) if entry.service_id else str(entry.id),
                    "productBrand": entry.brand or "",
                    "productFamily": entry.product_family or "",
                    "productCategory": entry.product_category or "",
                    "productModel": entry.model or "",
                    "product": entry.product_full_name or "",
                    "itemColor": entry.color or "",
                    "itemInternalId": entry.internal_id or "",
                    "itemSerialNo": entry.serial_number or "",
                    "itemImei": entry.imei_number or "",
                    "itemImei2": "",
                    "customer": entry.customer_name or "",
                    "requestType": entry.flow or "",
                    "rmaReason": "",
                    "receiveGrade": entry.receive_grade or "",
                    "createDate": fmt(entry.created_at),
                    "statuUpdateDate": fmt(entry.statu_update_time or entry.updated_at),
                    "updateDate": fmt(entry.updated_at),
                },
                "statu_code": entry.statu_code,
                "statu_name": statu.short_name if statu else "",
            })
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def find_device_by_term(self, term):
        """IMEI/Seri/Internal ID okutularak cihaz bilgisini (marka/model/hafıza/renk)
        batch_entries'ten getirir. Servis Kaydı ekranındaki IMEI okutma alanı için kullanılır."""
        db = SessionLocal()
        try:
            term = (term or "").strip()
            if not term:
                return json.dumps({"success": False, "message": "IMEI/Seri/Internal ID boş olamaz."})

            entry = self._find_batch_entry_by_term(db, term)
            if not entry:
                return json.dumps({"success": False, "message": f"'{term}' için kayıtlı bir cihaz bulunamadı."})

            return json.dumps({
                "success": True,
                "imei_number": entry.imei_number or "",
                "serial_number": entry.serial_number or "",
                "internal_id": entry.internal_id or "",
                "brand": entry.brand or "",
                "model": entry.model or "",
                "memory": entry.gb or "",
                "color": entry.color or "",
            })
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(int, result=str)
    def get_batch_entries_by_statu(self, statu_code):
        """Belirtilen statüdeki (örn. 106 - Müşteri onayına sunulacak) tüm parti/cihazları
        listeler. Müşteri onayı bekleyecek ekranındaki tik/çarpı listesi için kullanılır."""
        from models.batch_entry import BatchEntry
        db = SessionLocal()
        try:
            entries = db.query(BatchEntry).filter(BatchEntry.statu_code == statu_code).order_by(BatchEntry.id.desc()).all()
            items = [{
                "entry_id": e.id,
                "imei": e.imei_number or e.serial_number or e.internal_id or "",
                "batch_no": e.batch_no or "",
                "flow": e.flow or "",
            } for e in entries]
            return json.dumps({"success": True, "items": items})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, int, int, result=str)
    def execute_batch_entry_statu_transition(self, entry_id, current_statu_code, target_statu_code):
        """Partiyi belirtilen kaynak statüden hedef statüye taşır. Kaynak statü kaydın
        mevcut durumuyla eşleşmiyorsa veya geçiş kuralı tanımlı değilse hata döner."""
        from models.batch_entry import BatchEntry
        from models.service_statu import ServiceStatu
        from services.state_machine_service import StateMachineService
        db = SessionLocal()
        try:
            entry = db.query(BatchEntry).filter(BatchEntry.id == int(entry_id)).first()
            if not entry:
                return json.dumps({"success": False, "message": "Parti/cihaz bulunamadı."})

            actual_code = entry.statu_code if entry.statu_code is not None else 100

            def statu_name(code):
                s = db.query(ServiceStatu).filter_by(code=code).first()
                return s.short_name if s else str(code)

            device_label = " ".join(filter(None, [entry.imei_number, entry.batch_no, entry.flow]))

            if actual_code != current_statu_code:
                return json.dumps({
                    "success": False,
                    "message": f"{device_label} mevcut statüsü {statu_name(actual_code)} ({actual_code}) — bu okutmaya uygun statü değil (beklenen: {current_statu_code})."
                })

            svc = StateMachineService(db)
            if not svc.validate_transition(current_statu_code, target_statu_code):
                return json.dumps({
                    "success": False,
                    "message": f"{device_label} mevcut statüsü {statu_name(actual_code)} ({actual_code}) — bu okutmaya uygun statü değil."
                })

            old_name = statu_name(current_statu_code)
            new_name = statu_name(target_statu_code)

            entry.statu_code = target_statu_code
            self._record_statu_change(
                db, entry.id, entry.imei_number, current_statu_code, target_statu_code,
                staff=getattr(entry, "created_by", None),
                note=f"{old_name} ({current_statu_code}) → {new_name} ({target_statu_code})",
            )
            db.commit()

            return json.dumps({
                "success": True,
                "new_statu_code": target_statu_code,
                "message": f"{device_label} {old_name} ({current_statu_code}) statüsünden {new_name} ({target_statu_code}) statüsüne alındı."
            })
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, int, int, int, str, str, str, bool, result=str)
    def submit_test_result(self, entry_id, current_statu_code, success_statu_code, fail_statu_code, result, description, faults_json, log_exit_test=False):
        """Ara Test / Son Test sonucunu işler.
        result='success' ise cihazı success_statu_code'a aktarır. log_exit_test=True ise (Son Test ekranı)
        açıklama zorunludur ve phonecheck_test_results'a "Çıkış Testi" olarak manuel kayıt düşülür.
        result='fail' ise açıklama ve en az bir hatalı parça/hata kodu zorunludur, cihaz fail_statu_code'a geri döner."""
        from models.batch_entry import BatchEntry
        from models.service_statu import ServiceStatu
        from models.test_result_fault import TestResultFault
        from models.repair_record import RepairRecord
        from models.phonecheck_test_result import PhonecheckTestResult
        from services.state_machine_service import StateMachineService
        db = SessionLocal()
        try:
            entry = db.query(BatchEntry).filter(BatchEntry.id == int(entry_id)).first()
            if not entry:
                return json.dumps({"success": False, "message": "Parti/cihaz bulunamadı."})

            actual_code = entry.statu_code if entry.statu_code is not None else 100

            def statu_name(code):
                s = db.query(ServiceStatu).filter_by(code=code).first()
                return s.short_name if s else str(code)

            device_label = " ".join(filter(None, [entry.imei_number, entry.batch_no, entry.flow]))

            if actual_code != current_statu_code:
                return json.dumps({
                    "success": False,
                    "message": f"{device_label} mevcut statüsü {statu_name(actual_code)} ({actual_code}) — bu okutmaya uygun statü değil (beklenen: {current_statu_code})."
                })

            if result == "success":
                if log_exit_test:
                    from services.phonecheck_service import PhonecheckService
                    pc = PhonecheckService(db)
                    stage = pc.build_stage(current_statu_code, success_statu_code)
                    imei = entry.imei_number or ""
                    timestamp = __import__("datetime").datetime.now().strftime('%d.%m.%Y %H:%M')
                    auto_note = f"[{timestamp}] Test olumlu — {device_label}"
                    db.add(PhonecheckTestResult(
                        imei=imei,
                        test_stage=stage,
                        attempt_no=pc.attempt_count(imei, stage) + 1,
                        working="Yes",
                        notes=auto_note,
                        is_manual=True,
                        manual_reason=auto_note,
                        manual_entered_by=getattr(entry, "created_by", None)
                    ))
                target_statu_code = success_statu_code
            elif result == "fail":
                if not description or not description.strip():
                    return json.dumps({"success": False, "message": "Test başarısız için açıklama zorunludur."})
                try:
                    fault_lines = json.loads(faults_json) if faults_json else []
                except Exception:
                    fault_lines = []
                if not fault_lines:
                    return json.dumps({"success": False, "message": "En az bir hatalı parça / hata kodu seçmelisiniz."})
                if len(fault_lines) > 10:
                    return json.dumps({"success": False, "message": "En fazla 10 hatalı parça / hata kodu seçebilirsiniz."})

                if log_exit_test:
                    from services.phonecheck_service import PhonecheckService, MAX_FAILED_ATTEMPTS
                    pc = PhonecheckService(db)
                    stage = pc.build_stage(current_statu_code, fail_statu_code)
                    imei = entry.imei_number or ""

                    if pc.failed_limit_reached(imei, stage):
                        return json.dumps({"success": False, "message": f"Bu cihaz için bu test adımında en fazla {MAX_FAILED_ATTEMPTS} başarısız deneme hakkı var, hak doldu."})

                    timestamp = __import__("datetime").datetime.now().strftime('%d.%m.%Y %H:%M')
                    fail_note = f"[{timestamp}] Test Başarısız — {description.strip()}\nHatalı Parçalar: " + "; ".join(fault_lines)
                    db.add(PhonecheckTestResult(
                        imei=imei,
                        test_stage=stage,
                        attempt_no=pc.attempt_count(imei, stage) + 1,
                        working="No",
                        notes=fail_note,
                        is_manual=True,
                        manual_reason=fail_note,
                        manual_entered_by=getattr(entry, "created_by", None)
                    ))

                timestamp = __import__("datetime").datetime.now().strftime('%d.%m.%Y %H:%M')
                note = f"[{timestamp}] Test Başarısız — {description.strip()}\nHatalı Parçalar: " + "; ".join(fault_lines)
                entry.defects = (entry.defects + "\n\n" + note) if entry.defects else note

                device_ref = entry.imei_number or entry.batch_no or str(entry.id)
                # repair_records.service_record_id hem eski (IMEI ile yazılmış) hem yeni
                # (service_id ile yazılmış) kayıtlarla eşleşsin diye ikisi de aranır.
                repair_refs = [r for r in [device_ref, str(entry.service_id) if entry.service_id else None] if r]
                fault_row_kwargs = {}
                for idx, fault_line in enumerate(fault_lines, start=1):
                    if ": " in fault_line:
                        part_category, fault_text = fault_line.split(": ", 1)
                    else:
                        part_category, fault_text = None, fault_line
                    fault_row_kwargs[f"hatali_parca{idx}"] = part_category
                    fault_row_kwargs[f"hata{idx}"] = fault_text

                db.add(TestResultFault(
                    service_id=entry.id,
                    imei_number=entry.imei_number,
                    internal_id=entry.internal_id,
                    description=description.strip(),
                    created_by=getattr(entry, "created_by", None),
                    **fault_row_kwargs
                ))

                db.query(RepairRecord).filter(
                    RepairRecord.service_record_id.in_(repair_refs),
                    RepairRecord.repair_result_type_code == 1002
                ).update({"repair_result_type_code": 1001}, synchronize_session=False)

                target_statu_code = fail_statu_code
            else:
                return json.dumps({"success": False, "message": "Geçersiz sonuç türü."})

            svc = StateMachineService(db)
            if not svc.validate_transition(current_statu_code, target_statu_code):
                return json.dumps({
                    "success": False,
                    "message": f"{device_label} mevcut statüsü {statu_name(actual_code)} ({actual_code}) — bu okutmaya uygun statü değil."
                })

            old_name = statu_name(current_statu_code)
            new_name = statu_name(target_statu_code)

            entry.statu_code = target_statu_code
            self._record_statu_change(
                db, entry.id, entry.imei_number, current_statu_code, target_statu_code,
                staff=getattr(entry, "created_by", None),
                note=f"{old_name} ({current_statu_code}) → {new_name} ({target_statu_code})",
            )
            db.commit()

            return json.dumps({
                "success": True,
                "new_statu_code": target_statu_code,
                "message": f"{device_label} {old_name} ({current_statu_code}) statüsünden {new_name} ({target_statu_code}) statüsüne alındı."
            })
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    def _internal_check_doa(self, db, service_record_id):
        from sqlalchemy import text
        # Bu is emrine bagli ve is_issued=True olan parcalari getirir
        # Projede work_order_parts tablosu var
        # part_id de dönülür: warehouse.stock ve warehouse.stock_movements part_code ile
        # değil part_id ile çalışır (bkz. models/stock.py, models/stock_movement.py).
        sql = text("""
            SELECT wp.id, wp.part_code, wp.quantity, p.name, p.id AS part_id
            FROM warehouse.work_order_parts wp
            LEFT JOIN warehouse.parts p ON p.item_code = wp.part_code
            WHERE wp.work_order_id = :wo_id AND wp.is_issued = true
        """)
        parts = db.execute(sql, {"wo_id": service_record_id}).mappings().all()
        return {
            "has_consumed_parts": len(parts) > 0,
            "parts": [dict(p) for p in parts]
        }

    @Slot(str, result=str)
    def check_doa_status(self, service_record_id):
        db = SessionLocal()
        try:
            res = self._internal_check_doa(db, service_record_id)
            return json.dumps({"success": True, "data": res})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()
            
    @Slot(str, result=str)
    def transfer_to_doa(self, service_record_id):
        # Kullanılmıs (issued) parcalari DOA deposuna tasir
        db = SessionLocal()
        try:
            from sqlalchemy import text
            doa_loc_id = _get_system_location_id(db, "doa_stock")
            if not doa_loc_id:
                return json.dumps({"success": False, "message": "Sistem DOA deposu bulunamadi."})
                
            parts = self._internal_check_doa(db, service_record_id).get("parts", [])
            for p in parts:
                # Stoktan dus (Good -> DOA transferine gerek yok, zaten work_order_parts is_issued=true yapilirken good_stocktan dusuldu)
                # Sadece doa_stock artirilir. 
                from models.stock import Stock
                from models.stock_movement import StockMovement
                # NOT: Stock ve StockMovement modellerinde 'part_code' kolonu YOKTUR;
                # ikisi de part_id ile çalışır. Eskiden burada part_code/from_location_id/
                # to_location_id/movement_type/reference_document veriliyordu - bunların
                # hiçbiri model alanı olmadığı için çağrı TypeError fırlatıyor ve
                # transfer_to_doa hiç çalışmıyordu.
                part_id = p.get("part_id")
                if not part_id:
                    # Parça katalogda bulunamadıysa bu satırı atla; stok bozulmasın.
                    logging.warning(f"transfer_to_doa: '{p['part_code']}' parts tablosunda bulunamadı, atlandı.")
                    continue

                # DOA stok artir
                ds = db.query(Stock).filter_by(location_id=doa_loc_id, part_id=part_id).first()
                if not ds:
                    ds = Stock(location_id=doa_loc_id, part_id=part_id, quantity=0)
                    db.add(ds)
                ds.quantity = (ds.quantity or 0) + p["quantity"]

                # Hareketi kaydet
                mov = StockMovement(
                    part_id=part_id,
                    part_name_snapshot=p.get("name") or p["part_code"],
                    source_location_id=None,        # Cihazdan çıkıyor
                    target_location_id=doa_loc_id,
                    quantity=p["quantity"],
                    type="İade",
                    movement_kind="Return",
                    created_by="system",
                    description=f"DOA Transfer - Parça: {p['part_code']} - İş Emri: {service_record_id}"
                )
                db.add(mov)
                
                # work_order_parts tablosundaki kaydi iptal et veya sil (cihazdan sokuldu)
                db.execute(text("UPDATE warehouse.work_order_parts SET is_issued = false, issued_quantity = 0 WHERE id = :pid"), {"pid": p["id"]})
                
            db.commit()
            return json.dumps({"success": True, "message": "Parçalar başarıyla DOA deposuna aktarıldı."})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_repair_records(self, service_record_id):
        """service_record_id bir work_order_id, ham IMEI veya service_id (UUID) olabilir -
        çağıran taraf hangisini geçtiğini bilmeyebileceğinden, IMEI ise cihazın
        batch_entries.service_id'si de (varsa) eşleşme aranan değerlere eklenir (bkz.
        get_repair_operations_by_imei'deki aynı desen)."""
        db = SessionLocal()
        try:
            from sqlalchemy import text
            refs = [service_record_id]
            batch = self._resolve_batch_entry_by_ref(db, service_record_id)
            if batch and batch["service_id"] and str(batch["service_id"]) not in refs:
                refs.append(str(batch["service_id"]))
            sql = text("SELECT * FROM warehouse.repair_records WHERE service_record_id = ANY(:refs) ORDER BY created_at DESC")
            records = db.execute(sql, {"refs": refs}).mappings().all()
            
            # JSON serialization of UUID/Datetime
            out = []
            for r in records:
                d = dict(r)
                d["id"] = str(d["id"])
                d["created_at"] = str(d["created_at"])
                d["updated_at"] = str(d["updated_at"])
                out.append(d)
                
            return json.dumps({"success": True, "records": out})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # Qt slotu DEĞİLDİR: arayüzden çağrılmaz, get_repair_operations_by_imei'nin
    # açık DB oturumunu paylaşan dahili yardımcıdır.
    def _get_device_critical_parts(self, db, imei):
        """Cihazin EN YENI Phonecheck kaydindaki Parts verisinden 3 kritik parcayi
        (Ana Kamera / Batarya / Eski Pil) cikarir. (liste, remarks) doner.

        Phonecheck kaydi ya da Parts verisi yoksa ucu de 'unknown' doner - arayuzde
        gri gorunur, "orijinal degil" ile karistirilmaz.

        Sorgu patlarsa (ornegin 'parts' kolonu henuz olusmamis eski bir semada)
        ozellik sessizce devre disi kalir; cagiran metodun asil isi bozulmasin diye
        oturum rollback edilir - Postgres'te hatali sorgu sonrasi ayni transaction'da
        yapilan her sorgu da hata verir."""
        from sqlalchemy import text
        from services.phonecheck_service import parse_critical_parts

        raw = None
        term = (imei or "").strip()
        if term:
            try:
                row = db.execute(text("""
                    SELECT parts FROM warehouse.phonecheck_test_results
                    WHERE LOWER(TRIM(imei)) = LOWER(:t) AND parts IS NOT NULL
                    ORDER BY fetched_at DESC LIMIT 1
                """), {"t": term}).first()
                raw = row[0] if row else None
            except Exception as e:
                db.rollback()
                print(f"[WebBridge] kritik parca okunamadi ({term}): {e}")
                raw = None
        return parse_critical_parts(raw)

    # ---------------------------------------------------------
    # "İADE EDİLECEK" (Servis Onarımları -> Cihazı İadeye Al)
    # repair_records.service_record_id, generate_concurrent_repairs()'in
    # yazdığı gibi aslında work_orders.id'yi (string) tutar.
    # ---------------------------------------------------------
    @Slot(str, result=str)
    def get_repair_operations_by_imei(self, imei):
        """Servis Onarımları / İade ekranı için IMEI'ye ait iş emrini, parça
        listesini (depodan çıkmış olanlar dahil) ve onarım kayıtlarını döner.
        Bağlı bir Servis Kaydı/İş Emri yoksa (üretim verisinde sık görülen durum),
        onarım kayıtlarını doğrudan IMEI'ye bağlı olarak arar (work_order_id: null döner)
        ki 'Onarım Ekle' iş emri olmayan cihazlarda da çalışabilsin."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            if not imei or not imei.strip():
                return json.dumps({"success": False, "message": "IMEI boş olamaz"})
            term = imei.strip()

            sr = db.execute(text("""
                SELECT id, customer_name, brand, model, memory, color, imei_number, imei_serial,
                       customer_complaint, preliminary_diagnosis
                FROM warehouse.service_records
                WHERE LOWER(TRIM(imei_number)) = LOWER(:term) OR LOWER(TRIM(imei_serial)) = LOWER(:term)
                ORDER BY id DESC LIMIT 1
            """), {"term": term}).mappings().first()

            wo = None
            if sr:
                wo = db.execute(text("""
                    SELECT id, status, assigned_technician
                    FROM warehouse.work_orders
                    WHERE service_record_id = :sr_id AND work_order_type = 'SERVICE'
                    ORDER BY id DESC LIMIT 1
                """), {"sr_id": sr["id"]}).mappings().first()

            # ── Bağlı Servis Kaydı + İş Emri bulundu: tam veri (parçalar dahil) ──
            if sr and wo:
                try:
                    current_statu_code = int(wo["status"])
                except (TypeError, ValueError):
                    current_statu_code = None

                part_rows = db.execute(text("""
                    SELECT wop.id, wop.quantity, wop.status,
                           p.item_code, p.name, p.brand, p.model, p.color
                    FROM warehouse.work_order_parts wop
                    LEFT JOIN warehouse.parts p ON p.id = wop.part_id
                    WHERE wop.work_order_id = :wo_id
                    ORDER BY wop.id DESC
                """), {"wo_id": wo["id"]}).mappings().all()

                parts = [{
                    "id": str(r["id"]),
                    "itemCode": r["item_code"] or "",
                    "name": " ".join(filter(None, [r["brand"], r["model"], r["color"]])) or (r["name"] or "-"),
                    "qty": r["quantity"],
                    # DOAReturnModal (frontend) sadece "OUT" konumundaki parçaları güvenlik kontrolüne sokar.
                    "location": "OUT" if r["status"] == "Teslim Edildi" else "-",
                } for r in part_rows]

                repair_refs = [str(wo["id"]), str(sr["id"])]
                if sr["imei_number"]:
                    repair_refs.append(str(sr["imei_number"]).strip())
                if sr["imei_serial"]:
                    repair_refs.append(str(sr["imei_serial"]).strip())
                be_row_sr = db.execute(text("SELECT service_id FROM warehouse.batch_entries WHERE LOWER(TRIM(imei_number)) = LOWER(:t) LIMIT 1"), {"t": term}).mappings().first()
                if be_row_sr and be_row_sr["service_id"]:
                    repair_refs.append(str(be_row_sr["service_id"]))
                device_info = {
                    "imei": sr["imei_number"] or sr["imei_serial"] or term,
                    "productInfo": " ".join(filter(None, [sr["brand"], sr["model"], sr["color"], sr["memory"]])) or "-",
                    "customerRequest": sr["customer_complaint"] or "",
                    "customerDiagnosis": sr["preliminary_diagnosis"] or "",
                    "serviceStatus": current_statu_code,
                    # work_orders.status sayısal koda geçmediyse (ör. "Beklemede",
                    # "Devam Ediyor" gibi eski metin statüler) ham değer burada kalır.
                    "statusText": wo["status"] or "",
                }
                work_order_id_out = wo["id"]
                service_record_id_out = sr["id"]
                current_statu_code_out = current_statu_code
                found_batch_entry = True
            # ── Bağlı Servis Kaydı/İş Emri yok: onarım kayıtları doğrudan IMEI'ye bağlı aranır ──
            else:
                parts = []
                be_row = db.execute(text("""
                    SELECT customer_diagnosis, service_id, statu_code, brand, model, color, imei_number, serial_number FROM warehouse.batch_entries
                    WHERE LOWER(TRIM(imei_number)) = LOWER(:term) OR LOWER(TRIM(serial_number)) = LOWER(:term) OR LOWER(TRIM(internal_id)) = LOWER(:term)
                    ORDER BY id DESC LIMIT 1
                """), {"term": term}).mappings().first()

                product_info = ""
                if be_row:
                    product_info = " ".join(filter(None, [be_row["brand"], be_row["model"], be_row["color"]])) or (be_row["model"] or "")

                device_info = {
                    "imei": term,
                    "productInfo": product_info,
                    "customerRequest": "",
                    "customerDiagnosis": (be_row["customer_diagnosis"] if be_row else "") or "",
                    "serviceStatus": be_row["statu_code"] if be_row else None,
                    "statusText": str(be_row["statu_code"]) if be_row and be_row["statu_code"] is not None else "",
                }
                found_batch_entry = be_row is not None
                work_order_id_out = None
                service_record_id_out = str(be_row["service_id"]) if be_row and be_row["service_id"] else None
                current_statu_code_out = be_row["statu_code"] if be_row else None
                repair_refs = [term]
                if be_row and be_row["service_id"]:
                    repair_refs.append(str(be_row["service_id"]))

            repair_rows = db.execute(text("""
                SELECT rr.id, rr.department_mission, rr.notes, rr.repair_result_type_code, rr.warranty_code,
                       rr.part_item_code, rr.item_fault_code, rr.operation_type_code, rr.supply_status_code,
                       rr.assigned_technician, rr.assigned_by, rr.assigned_at, rr.created_at, rr.updated_at,
                       rrt.short_name AS result_name, rrt.is_cancelled, rrt.is_success,
                       -- mission_group_name SELECT'ten dusmustu (merge kaybi): mg JOIN'i duruyordu
                       -- ama kolon secilmiyordu, mapping r["mission_group_name"] okuyunca
                       -- NoSuchColumnError firlatiyor, metod success:false donuyor ve arayuzde
                       -- HER cihazda "Aktif onarim kaydi yok" gorunuyordu.
                       mg.short_name AS mission_group_name,
                       it.short_name AS part_name, pp.item_category AS item_category,
                       pp.stock_tracking_type, pp.id AS part_id,
                       fault.short_name AS fault_name,
                       opt.short_name AS operation_type_name,
                       sup.short_name AS supply_status_name,
                       COALESCE(NULLIF(TRIM(au.fullname), ''), rr.assigned_technician, rr.supply_requested_by) AS assigned_technician_name
                FROM warehouse.repair_records rr
                LEFT JOIN warehouse.repair_result_type rrt ON rrt.code = rr.repair_result_type_code
                LEFT JOIN organization.mission_groups mg ON mg.code = rr.department_mission
                LEFT JOIN warehouse.item it ON it.code = rr.part_item_code
                LEFT JOIN warehouse.parts pp ON pp.item_code = rr.part_item_code
                LEFT JOIN warehouse.item_fault fault ON fault.code = rr.item_fault_code
                LEFT JOIN warehouse.repair_item_operation_type opt ON opt.code = rr.operation_type_code
                LEFT JOIN warehouse.item_supply_status sup ON sup.code = rr.supply_status_code
                LEFT JOIN warehouse.users au ON (au.username = rr.assigned_technician OR au.username = rr.supply_requested_by)
                WHERE rr.service_record_id = ANY(:refs)
                   OR LOWER(TRIM(rr.service_record_id)) = LOWER(:term)
                   OR EXISTS (
                       SELECT 1 FROM warehouse.batch_entries be
                       WHERE (be.service_id IS NOT NULL AND strpos(rr.service_record_id, be.service_id::text) > 0)
                         AND (LOWER(TRIM(be.imei_number)) = LOWER(:term) OR LOWER(TRIM(be.serial_number)) = LOWER(:term) OR LOWER(TRIM(be.internal_id)) = LOWER(:term))
                   )
                ORDER BY rr.created_at DESC
            """), {"refs": repair_refs, "term": term}).mappings().all()

            repair_loc_id = _get_system_location_id(db, "repair_stock")
            repairs = []
            for r in repair_rows:
                tracking_type = (r["stock_tracking_type"] or "Stok Takipli").strip() if r["part_item_code"] else "Stoksuz"
                is_stoksuz = tracking_type in ("Stok Takipsiz", "Stoksuz", "Servis Hizmeti", "Yazılım/Hizmet") or not r["part_item_code"]
                
                is_delivered = is_stoksuz
                if not is_stoksuz:
                    # Depocu Parça Teslim ekranından parça çıkışını yapıp supply_status_code'u
                    # 'Stoktan Çıktı' yapmadığı sürece teslim edilmiş SAYILMAZ.
                    # NOT: 'stoktan çıktı' warehouse.item_supply_status'taki GEÇERLİ koddur.
                    # 'teslim edildi' o tabloda YOK - eski kayıtlar için geriye dönük kabul edilir.
                    if (r["supply_status_code"] or "").strip().lower() in ("stoktan çıktı", "teslim edildi", "teslim", "completed"):
                        is_delivered = True
                    else:
                        is_delivered = False

                repairs.append({
                    "id": str(r["id"]),
                    "missionGroupCode": r["department_mission"] or "",
                    "missionGroup": r["mission_group_name"] or r["department_mission"] or "-",
                    "statusCode": r["repair_result_type_code"],
                    "statusName": r["result_name"] or str(r["repair_result_type_code"]),
                    "isCancelled": bool(r["is_cancelled"]),
                    "chargeType": "FREE" if r["warranty_code"] == "IW" else "PAID",
                    "partItemCode": r["part_item_code"] or "",
                    "partName": r["part_name"] or "",
                    "itemCategory": r["item_category"] or "",
                    "stockTrackingType": tracking_type,
                    "isStoksuz": is_stoksuz,
                    "isDelivered": is_delivered,
                    "faultCode": r["item_fault_code"] or "",
                    "faultName": r["fault_name"] or r["item_fault_code"] or "",
                    "operationTypeCode": r["operation_type_code"] or "",
                    "operationTypeName": r["operation_type_name"] or "",
                    "supplyStatusCode": r["supply_status_code"] or "",
                    "supplyStatusName": r["supply_status_name"] or "",
                    "assignedTechnician": r["assigned_technician"] or "",
                    "assignedTechnicianName": r["assigned_technician_name"] or "",
                    "assignedBy": r["assigned_by"] or "",
                    "assignedAt": r["assigned_at"].isoformat() if r["assigned_at"] else "",
                    # Onarım Detay grid'indeki "Tarih" sütunu bunu okur (onarımın oluşturulma anı).
                    # Türkiye yerel saati (naive created_at UTC'dir - bkz. fmt_tr_datetime).
                    "createdAt": fmt_tr_datetime(r["created_at"]),
                    "updatedAt": fmt_tr_datetime(r["updated_at"]),
                    "notes": r["notes"] or "",
                })

            if not (sr and wo) and not repairs and not found_batch_entry:
                return json.dumps({"success": False, "message": "Bu cihaza ait bir iş emri veya onarım kaydı bulunamadı."})

            # ── Phonecheck test sonuclarini (battery cycle, battery health) ekle ──
            from models.phonecheck_test_result import PhonecheckTestResult
            from services.phonecheck_service import PhonecheckService

            pc_lookup_imei = term
            if sr and (sr["imei_number"] or sr["imei_serial"]):
                pc_lookup_imei = sr["imei_number"] or sr["imei_serial"]
            elif 'be_row' in locals() and be_row and (be_row["imei_number"] or be_row["serial_number"]):
                pc_lookup_imei = be_row["imei_number"] or be_row["serial_number"]

            pc = db.query(PhonecheckTestResult).filter(
                PhonecheckTestResult.imei == pc_lookup_imei
            ).order_by(PhonecheckTestResult.fetched_at.desc()).first()

            # Yerel DB'de test kaydi yoksa, canlı Phonecheck Cloud API'sine sor
            if not pc and pc_lookup_imei:
                try:
                    pc_svc = PhonecheckService(db)
                    fetched = pc_svc.fetch_device(pc_lookup_imei)
                    if fetched.get("success") and fetched.get("device"):
                        pc = pc_svc.save_from_phonecheck(fetched["device"], test_stage="AUTO_LOOKUP", imei=pc_lookup_imei)
                        db.commit()
                except Exception as _e:
                    print(f"[Phonecheck Live Fetch Error]: {_e}")

            battery_cycle = pc.battery_cycle if pc else None
            battery_health = pc.battery_health_percentage if pc else None

            # Kritik parça orijinallik kontrolü (Ana Kamera / Batarya / Eski Pil).
            # Kaynak: bu cihaza ait EN YENİ Phonecheck kaydının Parts alanı. Her iki
            # device_info dalı da aynı sonucu alsın diye tek noktada, dönüşten hemen
            # önce ekleniyor.
            device_info["criticalParts"], device_info["partsRemark"] = \
                self._get_device_critical_parts(db, device_info.get("imei") or term)

            return json.dumps({
                "success": True,
                "work_order_id": work_order_id_out,
                "service_record_id": service_record_id_out,
                "current_statu_code": current_statu_code_out,
                "device": device_info,
                "parts": parts,
                "repairs": repairs,
                "battery_cycle": battery_cycle,
                "battery_health": battery_health,
            })
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_department_technicians(self, department_code):
        """Belirtilen departmanda (BATTERY, DISPLAY, CAMERA, CASE, L1REPAIR, L2REPAIR, L3REPAIR)
        görevli teknisyenleri (warehouse.users) döndürür."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            dept = (department_code or "").strip().upper()
            if not dept:
                return json.dumps({"success": False, "message": "Departman koda boş olamaz."})

            rows = db.execute(text("""
                SELECT username, fullname, role, gorev
                FROM warehouse.users
                WHERE (
                    UPPER(COALESCE(gorev, '')) LIKE :pattern1 OR
                    UPPER(COALESCE(gorev, '')) LIKE :pattern2 OR
                    UPPER(COALESCE(role, '')) LIKE :pattern1
                )
                ORDER BY fullname ASC
            """), {
                "pattern1": f"%{dept}%",
                "pattern2": f"%TEC_{dept}%"
            }).mappings().all()

            items = [{
                "username": r["username"],
                "fullname": r["fullname"] or r["username"],
                "role": r["role"] or "",
                "gorev": r["gorev"] or "",
            } for r in rows]

            return json.dumps({"success": True, "technicians": items}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def assign_repair_to_technician(self, department_code, imei_or_term, technician_username):
        """Teknisyen adının altındaki okutma kutusundan IMEI/Seri okutulduğunda:
        O departmana ait en eski 'Teknisyene Atanacak' (1000) durumundaki onarım kaydını
        veya belirtilen IMEI'li kaydı bulur, teknisyene atar (1001 - Teknisyene Atandı)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            dept = (department_code or "").strip().upper()
            term = (imei_or_term or "").strip()
            tech = (technician_username or "").strip()

            if not dept or not term or not tech:
                return json.dumps({"success": False, "message": "Departman, IMEI ve Teknisyen bilgileri zorunludur."})

            # Teknisyen varlığını doğrula
            tech_user = db.execute(text("SELECT username, fullname FROM warehouse.users WHERE username = :u"), {"u": tech}).mappings().first()
            if not tech_user:
                return json.dumps({"success": False, "message": f"'{tech}' kullanıcısı bulunamadı."})

            # İlgili departmanda bu IMEI'ye ait TÜM açık onarım kayıtlarını bul ve hepsini TEK teknisyene ata.
            # (Aynı IMEI ve aynı departman için iki farklı teknisyen ataması oluşması engellenir).
            repairs = db.execute(text("""
                SELECT rr.id, rr.repair_result_type_code, rr.service_record_id, rr.assigned_technician, rr.supply_requested_by
                FROM warehouse.repair_records rr
                WHERE UPPER(TRIM(rr.department_mission)) = :dept
                  AND COALESCE(rr.repair_result_type_code, 0) NOT IN (1002, 1003)
                  AND (
                    LOWER(TRIM(rr.service_record_id)) = LOWER(:term) OR
                    EXISTS (
                        SELECT 1 FROM warehouse.batch_entries be
                        WHERE (be.service_id IS NOT NULL AND strpos(rr.service_record_id, be.service_id::text) > 0)
                          AND (LOWER(TRIM(be.imei_number)) = LOWER(:term) OR LOWER(TRIM(be.serial_number)) = LOWER(:term) OR LOWER(TRIM(be.internal_id)) = LOWER(:term))
                    )
                  )
            """), {"dept": dept, "term": term}).mappings().all()

            if not repairs:
                kapali = db.execute(text("""
                    SELECT COUNT(*) FROM warehouse.repair_records rr
                    WHERE UPPER(TRIM(rr.department_mission)) = :dept
                      AND rr.repair_result_type_code IN (1002, 1003)
                      AND (
                        LOWER(TRIM(rr.service_record_id)) = LOWER(:term) OR
                        EXISTS (
                            SELECT 1 FROM warehouse.batch_entries be
                            WHERE (be.service_id IS NOT NULL AND strpos(rr.service_record_id, be.service_id::text) > 0)
                              AND (LOWER(TRIM(be.imei_number)) = LOWER(:term) OR LOWER(TRIM(be.serial_number)) = LOWER(:term) OR LOWER(TRIM(be.internal_id)) = LOWER(:term))
                        )
                      )
                """), {"dept": dept, "term": term}).scalar() or 0
                if kapali:
                    return json.dumps({"success": False, "message":
                        f"'{term}' cihazının {dept} departmanındaki {kapali} onarım kaydının tamamı kapalı."},
                        ensure_ascii=False)
                return json.dumps({"success": False, "message": f"'{term}' IMEI/cihazı için {dept} departmanında onarım kaydı bulunamadı."}, ensure_ascii=False)

            r_ids = [r["id"] for r in repairs]

            # Statüyü Teknisyene Atandı (1001) yap ve tüm açık satırların teknisyen alanlarını
            # güncelle. KANONİK assigned_technician yazılır (Üretim Kaydını Görüntüle ekranı bunu
            # okur); geriye dönük uyum için supply_requested_by da set edilir.
            db.execute(text("""
                UPDATE warehouse.repair_records
                SET repair_result_type_code = 1001,
                    supply_requested_by = :tech,
                    assigned_technician = :tech,
                    assigned_by = :tech,
                    assigned_at = NOW(),
                    updated_at = NOW()
                WHERE id = ANY(:rids)
            """), {"tech": tech, "rids": r_ids})
            db.commit()

            tech_name = tech_user["fullname"] or tech_user["username"]
            return json.dumps({
                "success": True,
                "message": f"'{term}' cihazının {dept} onarımı {tech_name} üzerine atandı ({len(r_ids)} kayıt).",
                "assignedTo": tech,
                "assignedToName": tech_name
            }, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_repair_pool_by_department(self, department_code):
        """Onarım Havuzu ekranı için belirtilen departmana (department_code: BATTERY, DISPLAY vb.)
        ait tüm aktif ve tamamlanmış onarım kayıtlarını sisteme giriş tarihine göre (en eskiden en yeniye)
        cihaz/batch detaylarıyla döndürür."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            dept = (department_code or "").strip().upper()
            if not dept:
                return json.dumps({"success": False, "message": "Departman kodu boş olamaz."})

            rows = db.execute(text("""
                SELECT 
                    rr.id AS repair_id,
                    rr.service_record_id,
                    rr.department_mission,
                    rr.repair_result_type_code,
                    rrt.short_name AS status_name,
                    rrt.is_cancelled,
                    rrt.is_success,
                    rr.operation_type_code,
                    opt.short_name AS operation_type_name,
                    rr.warranty_code,
                    rr.item_category,
                    rr.part_item_code,
                    pp.name AS part_name,
                    rr.item_fault_code,
                    fault.short_name AS fault_name,
                    rr.supply_status_code,
                    sup.short_name AS supply_status_name,
                    COALESCE(rr.assigned_technician, rr.supply_requested_by) AS assigned_technician,
                    COALESCE(NULLIF(TRIM(u.fullname), ''), rr.assigned_technician, rr.supply_requested_by) AS assigned_technician_name,
                    rr.notes,
                    rr.created_at,
                    rr.updated_at,
                    be.imei_number,
                    be.serial_number,
                    be.internal_id,
                    be.batch_no,
                    be.model,
                    be.gb,
                    be.color,
                    be.customer_name,
                    be.statu_code AS batch_status_code
                FROM warehouse.repair_records rr
                LEFT JOIN warehouse.repair_result_type rrt ON rrt.code = rr.repair_result_type_code
                LEFT JOIN warehouse.parts pp ON pp.item_code = rr.part_item_code
                LEFT JOIN warehouse.item_fault fault ON fault.code = rr.item_fault_code
                LEFT JOIN warehouse.repair_item_operation_type opt ON opt.code = rr.operation_type_code
                LEFT JOIN warehouse.item_supply_status sup ON sup.code = rr.supply_status_code
                LEFT JOIN warehouse.users u ON (u.username = rr.assigned_technician OR u.username = rr.supply_requested_by)
                LEFT JOIN warehouse.batch_entries be ON LOWER(TRIM(be.imei_number)) = LOWER(TRIM(rr.service_record_id))
                    OR LOWER(TRIM(be.serial_number)) = LOWER(TRIM(rr.service_record_id))
                    OR LOWER(TRIM(be.internal_id)) = LOWER(TRIM(rr.service_record_id))
                    OR (be.service_id IS NOT NULL AND strpos(rr.service_record_id, be.service_id::text) > 0)
                WHERE UPPER(TRIM(rr.department_mission)) = :dept
                ORDER BY rr.created_at ASC
            """), {"dept": dept}).mappings().all()

            # Tarih/saat Türkiye yerel saatinde ve gg.aa.yyyy SS:DD formatında (bkz. fmt_tr_datetime).
            fmt = fmt_tr_datetime

            items = []
            for r in rows:
                product_info = " ".join(filter(None, [r["model"], r["gb"], r["color"]])) or "-"
                items.append({
                    "repairId": str(r["repair_id"]),
                    "serviceRecordId": r["service_record_id"] or "",
                    "departmentMission": r["department_mission"] or "",
                    "statusCode": r["repair_result_type_code"],
                    "statusName": r["status_name"] or str(r["repair_result_type_code"]),
                    "isCancelled": bool(r["is_cancelled"]),
                    "isSuccess": bool(r["is_success"]),
                    "operationTypeCode": r["operation_type_code"] or "",
                    "operationTypeName": r["operation_type_name"] or "",
                    "warrantyCode": r["warranty_code"] or "",
                    "itemCategory": r["item_category"] or "",
                    "partItemCode": r["part_item_code"] or "",
                    "partName": r["part_name"] or "",
                    "itemFaultCode": r["item_fault_code"] or "",
                    "faultName": r["fault_name"] or r["item_fault_code"] or "",
                    "supplyStatusCode": r["supply_status_code"] or "",
                    "supplyStatusName": r["supply_status_name"] or "",
                    "assignedTechnician": r["assigned_technician"] or "",
                    "assignedTechnicianName": r["assigned_technician_name"] or r["assigned_technician"] or "",
                    "notes": r["notes"] or "",
                    "createdAt": fmt(r["created_at"]),
                    "updatedAt": fmt(r["updated_at"]),
                    "imei": r["imei_number"] or r["service_record_id"] or "-",
                    "serialNo": r["serial_number"] or "",
                    "internalId": r["internal_id"] or "",
                    "batchNo": r["batch_no"] or "",
                    "productInfo": product_info,
                    "customerName": r["customer_name"] or "",
                    "batchStatusCode": r["batch_status_code"],
                })

            return json.dumps({"success": True, "items": items}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_service_statu_list(self):
        """warehouse.service_statu'daki tüm statü kodlarını (kısa ad + gerekli mission) getirir.
        Servis Onarımları ekranındaki statü-bazlı rol/yetki kontrolünün kaynağıdır.
        Çalışma zamanında pratikte hiç değişmediğinden 5 dakika önbelleklenir."""
        def _compute():
            from sqlalchemy import text
            db = SessionLocal()
            try:
                rows = db.execute(text("""
                    SELECT code, short_name, mission, is_closed
                    FROM warehouse.service_statu
                    ORDER BY code ASC
                """)).mappings().all()
                items = [{"code": r["code"], "short_name": r["short_name"] or "", "mission": r["mission"] or "", "is_closed": bool(r["is_closed"])} for r in rows]
                return json.dumps({"success": True, "service_statu": items}, ensure_ascii=False)
            except Exception as e:
                return json.dumps({"success": False, "message": str(e)})
            finally:
                db.close()
        return self._cached_json("service_statu_list", 300, _compute)

    @Slot(str, int, result=str)
    def admin_set_batch_entry_statu(self, imei, target_statu_code):
        """Genel Bakış > Statü Kontrol ekranından çağrılır. Normal iş akışındaki
        execute_batch_entry_statu_transition'ın aksine StateMachineService kurallarını
        (hangi statüden hangi statüye geçilebileceği) uygulamaz - IMEI'ye ait en güncel
        batch_entries satırının statu_code'unu doğrudan hedef statüye ayarlar. Manuel/idari
        düzeltme amaçlıdır (ör. yanlışlıkla yanlış statüde kalmış bir cihazı düzeltmek)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            imei = (imei or "").strip()
            if not imei:
                return json.dumps({"success": False, "message": "IMEI boş olamaz."})

            entry = db.execute(text("""
                SELECT id, statu_code FROM warehouse.batch_entries
                WHERE LOWER(TRIM(imei_number)) = LOWER(:imei)
                ORDER BY id DESC LIMIT 1
            """), {"imei": imei}).mappings().first()
            if not entry:
                return json.dumps({"success": False, "message": "Bu IMEI için Batch Girişi kaydı bulunamadı."})

            target_row = db.execute(text("SELECT code, short_name FROM warehouse.service_statu WHERE code = :c"), {"c": target_statu_code}).mappings().first()
            if not target_row:
                return json.dumps({"success": False, "message": f"Geçersiz statü kodu: {target_statu_code}"})

            old_code = entry["statu_code"] if entry["statu_code"] is not None else 100
            old_row = db.execute(text("SELECT short_name FROM warehouse.service_statu WHERE code = :c"), {"c": old_code}).mappings().first()
            old_name = old_row["short_name"] if old_row else str(old_code)

            db.execute(text("UPDATE warehouse.batch_entries SET statu_code = :c WHERE id = :id"), {"c": target_statu_code, "id": entry["id"]})
            self._record_statu_change(
                db, entry["id"], imei, old_code, target_statu_code,
                note=f"Manuel/idari düzeltme: {old_name} ({old_code}) → {target_row['short_name']} ({target_statu_code})",
            )
            db.commit()
            return json.dumps({
                "success": True,
                "new_statu_code": target_statu_code,
                "message": f"{imei}: {old_name} ({old_code}) statüsünden {target_row['short_name']} ({target_statu_code}) statüsüne alındı."
            }, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    def _get_user_missions(self, db, username):
        """(mission_kod_listesi, is_admin) döner. warehouse.users.gorev (virgülle ayrılmış
        organization.missions.code değerleri) ve role='admin'/'developer' muafiyetine bakar."""
        from sqlalchemy import text
        if not username:
            return [], False
        row = db.execute(text("SELECT role, gorev FROM warehouse.users WHERE username = :u"), {"u": username}).mappings().first()
        if not row:
            return [], False
        is_admin = (row["role"] or "").strip().lower() in ("admin", "developer")
        missions = [m.strip() for m in (row["gorev"] or "").split(",") if m.strip()]
        return missions, is_admin

    def _get_required_mission_for_work_order(self, db, work_order_id):
        """work_order -> service_record -> IMEI -> en son batch_entries.statu_code -> service_statu.mission
        zincirini çözer. Herhangi bir halka eksikse (kural uygulanamıyorsa) None döner."""
        from sqlalchemy import text
        try:
            wo_id = int(work_order_id)
        except (TypeError, ValueError):
            return None

        sr = db.execute(text("""
            SELECT sr.imei_number, sr.imei_serial
            FROM warehouse.work_orders wo
            JOIN warehouse.service_records sr ON sr.id = wo.service_record_id
            WHERE wo.id = :id
        """), {"id": wo_id}).mappings().first()
        if not sr:
            return None

        imei = (sr["imei_number"] or sr["imei_serial"] or "").strip()
        if not imei:
            return None

        batch = db.execute(text("""
            SELECT statu_code FROM warehouse.batch_entries
            WHERE LOWER(TRIM(imei_number)) = LOWER(:imei)
            ORDER BY id DESC LIMIT 1
        """), {"imei": imei}).mappings().first()
        if not batch or batch["statu_code"] is None:
            return None

        statu = db.execute(text("""
            SELECT mission FROM warehouse.service_statu WHERE code = :code
        """), {"code": batch["statu_code"]}).mappings().first()
        return (statu["mission"] or None) if statu else None

    def _get_required_mission_for_repair(self, db, repair_id):
        """Bir repair_records.id için gerekli mission'ı çözer (üzerindeki device_ref'e bakarak)."""
        from models.repair_record import RepairRecord
        rec = db.query(RepairRecord).filter(RepairRecord.id == repair_id).first()
        if not rec:
            return None
        return self._get_required_mission_for_ref(db, rec.service_record_id)

    def _resolve_batch_entry_by_ref(self, db, ref):
        """Verilen ref bir service_id (UUID), IMEI, Seri No veya Dahili ID
        olabilir - hepsini dener, en güncel eşleşen batch_entries satırını döner."""
        from sqlalchemy import text
        ref = (ref or "").strip()
        if not ref:
            return None
        return db.execute(text("""
            SELECT id, statu_code, service_id, imei_number FROM warehouse.batch_entries
            WHERE service_id::text = :ref 
               OR LOWER(TRIM(imei_number)) = LOWER(:ref)
               OR LOWER(TRIM(serial_number)) = LOWER(:ref)
               OR LOWER(TRIM(internal_id)) = LOWER(:ref)
            ORDER BY id DESC LIMIT 1
        """), {"ref": ref}).mappings().first()

    def _resolve_service_record_id_for_new_repair(self, db, device_ref):
        """Yeni bir repair_records satırı için service_record_id'yi çözer: device_ref gerçek
        bir work_order_id ise (warehouse.work_orders'da gerçekten var olan bir id - IMEI'ler de
        tamamen sayısal olduğundan sadece int() dönüşümünün başarılı olması yeterli değildir,
        gerçek varlığı kontrol edilir) değişmeden kullanılır; IMEI ise ve o cihazın
        batch_entries'te bir service_id'si varsa (artık her yeni girişte üretiliyor), yeni
        onarım kayıtları IMEI yerine service_id ile yazılır - böylece aynı cihazın farklı
        service_id'li (yeni giriş) dönemlerine ait onarımlar birbirine karışmaz. Batch entry
        hiç yoksa (yalnızca customers/MIO durumu) ham device_ref (IMEI) korunur - eski
        davranışla aynı."""
        from sqlalchemy import text
        device_ref = str(device_ref).strip()
        try:
            wo_id = int(device_ref)
        except (TypeError, ValueError):
            wo_id = None
        if wo_id is not None:
            wo_exists = db.execute(text("SELECT id FROM warehouse.work_orders WHERE id = :id"), {"id": wo_id}).first()
            if wo_exists:
                return device_ref
        batch = self._resolve_batch_entry_by_ref(db, device_ref)
        if batch and batch["service_id"]:
            return str(batch["service_id"])
        return device_ref

    def _find_active_service_for_device(self, db, imei, serial_number=None, internal_id=None):
        """Verilen IMEI, seri no veya dahili ID ile eşleşen (herhangi biri bile eşleşse), HENÜZ KAPANMAMIŞ
        (statü 128 'Çıkışı yapıldı' değil) en güncel batch_entries satırını döner - yoksa None."""
        from sqlalchemy import text
        imei = (imei or "").strip()
        serial_number = (serial_number or "").strip()
        internal_id = (internal_id or "").strip()
        if not imei and not serial_number and not internal_id:
            return None
        clauses = []
        params = {}
        if imei:
            clauses.append("LOWER(TRIM(imei_number)) = LOWER(:imei)")
            params["imei"] = imei
        if serial_number:
            clauses.append("LOWER(TRIM(serial_number)) = LOWER(:serial)")
            params["serial"] = serial_number
        if internal_id:
            clauses.append("LOWER(TRIM(internal_id)) = LOWER(:internal_id)")
            params["internal_id"] = internal_id
        return db.execute(text(f"""
            SELECT id, service_id, batch_no, statu_code FROM warehouse.batch_entries
            WHERE ({' OR '.join(clauses)}) AND COALESCE(statu_code, 100) != 128
            ORDER BY id DESC LIMIT 1
        """), params).mappings().first()

    def _get_required_mission_for_ref(self, db, device_ref):
        """work_order_id (sayısal, warehouse.work_orders.id) veya bağlı bir iş emri yoksa
        doğrudan IMEI/service_id referansından statü-mission zincirini çözer. add_repair_record
        vb. 'Onarım Ekle' cihaza bağlı bir servis iş emri olmadan da çalışabildiği için
        (repair_records.service_record_id o durumda service_id veya IMEI tutar) üç yolu da destekler."""
        from sqlalchemy import text
        if not device_ref:
            return None
        device_ref = str(device_ref).strip()

        try:
            wo_id = int(device_ref)
        except (TypeError, ValueError):
            wo_id = None

        if wo_id is not None:
            wo_exists = db.execute(text("SELECT id FROM warehouse.work_orders WHERE id = :id"), {"id": wo_id}).first()
            if wo_exists:
                return self._get_required_mission_for_work_order(db, wo_id)

        # work_order olarak çözülemedi -> device_ref'i service_id veya IMEI olarak dene.
        batch = self._resolve_batch_entry_by_ref(db, device_ref)
        if not batch or batch["statu_code"] is None:
            return None

        statu = db.execute(text("""
            SELECT mission FROM warehouse.service_statu WHERE code = :code
        """), {"code": batch["statu_code"]}).mappings().first()
        return (statu["mission"] or None) if statu else None

    @Slot(str, str, str, result=str)
    def update_customer_diagnosis(self, device_ref, diagnosis_text, username):
        """'Müşteri Arıza Tespiti'ni günceller. Sadece test teknisyenleri (QAC ailesi: QAC,
        QAC_TL, QAC_DISPLAY, QAC_CASE, QAC_L3, QAC_CAMERA) ve cihazın mevcut statüsüne göre
        zaten yetkili olan kullanıcılar düzenleyebilir.
        device_ref, bağlı bir servis iş emri varsa work_order_id'dir (bu durumda
        warehouse.service_records.preliminary_diagnosis'e yazılır); yoksa (üretim verisinde
        sık görülen durum) doğrudan cihazın IMEI'sidir (bu durumda warehouse.batch_entries.
        customer_diagnosis'e yazılır) - bkz. add_repair_record/get_repair_operations_by_imei'deki
        aynı device_ref deseni."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            diag = diagnosis_text.strip() if diagnosis_text else None
            ref = (device_ref or "").strip()
            if not ref:
                return json.dumps({"success": False, "message": "Cihaz referansı boş olamaz."})

            wo = None
            try:
                wo = db.execute(text("SELECT service_record_id FROM warehouse.work_orders WHERE id = :id"), {"id": int(ref)}).mappings().first()
            except ValueError:
                wo = None

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                is_test_technician = any(m == "QAC" or m.startswith("QAC_") for m in user_missions)
                if not is_test_technician:
                    return json.dumps({"success": False, "message": "Bu alanı sadece test teknisyenleri düzenleyebilir."})
                required = self._get_required_mission_for_ref(db, ref)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            if wo and wo["service_record_id"]:
                db.execute(text("""
                    UPDATE warehouse.service_records SET preliminary_diagnosis = :diag WHERE id = :id
                """), {"diag": diag, "id": wo["service_record_id"]})
            else:
                result = db.execute(text("""
                    UPDATE warehouse.batch_entries SET customer_diagnosis = :diag
                    WHERE LOWER(TRIM(imei_number)) = LOWER(:imei)
                """), {"diag": diag, "imei": ref})
                if result.rowcount == 0:
                    return json.dumps({"success": False, "message": "Cihaz bulunamadı (iş emri veya Batch Girişi kaydı yok)."})
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, result=str)
    def add_repair_record(self, device_ref, mission_group_code, warranty_code, notes, username, part_item_code="", item_fault_code="", operation_type_code=""):
        """Bir cihaza yeni bir alt onarım kaydı (warehouse.repair_records) ekler.
        Servis Onarımları ekranındaki 'Onarım Ekle' aksiyonunun kalıcı karşılığıdır.
        device_ref, bağlı bir servis iş emri varsa work_order_id'dir; yoksa cihazın IMEI'sidir.
        Kaydın service_record_id'si _resolve_service_record_id_for_new_repair ile çözülür:
        cihazın batch_entries.service_id'si varsa (artık her yeni girişte üretiliyor) onarım
        kaydı IMEI yerine service_id ile yazılır; yoksa (yalnızca customers/MIO gibi
        batch_entries dışı durumlarda) ham IMEI korunur — her iki durumda da kayıt sonradan
        get_repair_operations_by_imei ile hem IMEI hem service_id üzerinden bulunabilir.
        part_item_code/item_fault_code/operation_type_code opsiyoneldir (Demontaj ekranının
        'Parça'/'Arıza Tespiti'/'İşlem' seçimleri)."""
        import uuid
        from models.repair_record import RepairRecord
        db = SessionLocal()
        try:
            if not device_ref or not str(device_ref).strip():
                return json.dumps({"success": False, "message": "Cihaz bulunamadı."})
            if not mission_group_code or not mission_group_code.strip():
                return json.dumps({"success": False, "message": "Görev grubu zorunludur."})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                required = self._get_required_mission_for_ref(db, device_ref)
                # Yeni onarım ekleme işleminde yetki kontrolü esnetilebilir veya cihaz teknisyene atandıysa izin verilir
                if required and required not in user_missions and not any(m in ("TEC_DISMANTLE", "QAC") or m.startswith("TEC_") or m.startswith("QAC_") for m in user_missions):
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            service_ref = self._resolve_service_record_id_for_new_repair(db, device_ref)
            part_code = part_item_code.strip() if part_item_code else None

            # AYNI PARÇA İKİ KEZ EKLENEMEZ. Aynı cihaza aynı parça için ikinci bir onarım
            # kaydı açılması hem stoktan iki kez düşülmesine hem de teknisyene aynı işin iki
            # kez görünmesine yol açıyordu. İptal edilmiş (1003) kayıtlar sayılmaz; iptal
            # edilen bir parça yeniden eklenebilmeli.
            if part_code:
                mevcut = db.query(RepairRecord).filter(
                    RepairRecord.service_record_id == service_ref,
                    RepairRecord.part_item_code == part_code,
                    RepairRecord.repair_result_type_code != 1003,
                ).first()
                if mevcut:
                    return json.dumps({
                        "success": False,
                        "message": f"'{part_code}' bu cihaza zaten eklenmiş. Aynı parça ikinci kez eklenemez; "
                                   f"mevcut satırı düzenleyin veya iptal edip yeniden ekleyin."
                    }, ensure_ascii=False)

            # L1REPAIR ve L2REPAIR aynı cihazda birlikte olamaz: biri varsa diğeri eklenemez.
            # Yalnızca aktif (iptal edilmemiş, 1003 değil) kayıtlar sayılır - iptal edilmiş bir
            # L1/L2 onarımı bu kısıtlamayı artık tetiklemez.
            OPPOSING_REPAIR_TEAMS = {"L1REPAIR": ("L2REPAIR", "L1", "L2"), "L2REPAIR": ("L1REPAIR", "L2", "L1")}
            team_code = mission_group_code.strip().upper()
            opposing = OPPOSING_REPAIR_TEAMS.get(team_code)
            if opposing:
                opposing_code, team_label, opposing_label = opposing
                conflict = db.query(RepairRecord).filter(
                    RepairRecord.service_record_id == service_ref,
                    RepairRecord.department_mission == opposing_code,
                    RepairRecord.repair_result_type_code != 1003,
                ).first()
                if conflict:
                    return json.dumps({
                        "success": False,
                        "message": f"Bu cihazda zaten aktif bir {opposing_label} onarımı var. Aynı cihaza hem {team_label} hem {opposing_label} onarımı eklenemez."
                    }, ensure_ascii=False)

            rec = RepairRecord(
                id=uuid.uuid4(),
                service_record_id=service_ref,
                department_mission=mission_group_code.strip(),
                repair_result_type_code=1000,
                warranty_code=warranty_code.strip() if warranty_code else "OOW",
                notes=notes.strip() if notes else None,
                part_item_code=part_item_code.strip() if part_item_code else None,
                item_fault_code=item_fault_code.strip() if item_fault_code else None,
                operation_type_code=operation_type_code.strip() if operation_type_code else None,
            )
            db.add(rec)
            db.commit()

            # Karar sonrası yeniden onarım: Cihaz demontaj kararını (Müşteri Onayı / Üretime
            # Aktar) geçip Üretim aşamasına (109) alındıktan sonra teknisyen YENİ bir onarım
            # eklerse, bu onarımın da müşteri onayına veya üretime yönlendirilebilmesi için
            # cihazı Demontaj karar aşamasına (105 - Awaiting production planning acceptance)
            # geri çekeriz. Böylece Servis Onarım ekranında "Müşteri Onayına Gönder / Üretime
            # Aktar" kararı yeniden görünür ve submit_dismantle_decision tekrar çalıştırılabilir.
            # Sadece karar SONRASI statüden (109) geri çekilir; normal ilk demontaj (104/105)
            # sırasında eklenen onarımlar statüyü değiştirmez. Geçiş yalnızca state machine'in
            # izin verdiği durumda (109->105) uygulanır.
            REOPEN_DECISION_FROM = {109}
            RESET_TARGET_STATU = 105
            reopened = False
            try:
                from services.state_machine_service import StateMachineService
                from models.batch_entry import BatchEntry
                be_ref = self._resolve_batch_entry_by_ref(db, device_ref)
                if be_ref and be_ref["statu_code"] is not None:
                    cur = int(be_ref["statu_code"])
                    if cur in REOPEN_DECISION_FROM and StateMachineService(db).validate_transition(cur, RESET_TARGET_STATU):
                        be = db.query(BatchEntry).filter(BatchEntry.id == int(be_ref["id"])).first()
                        if be:
                            be.statu_code = RESET_TARGET_STATU
                            self._record_statu_change(
                                db, be.id, be.imei_number, cur, RESET_TARGET_STATU,
                                note="Yeni onarım eklendi — karar için demontaj aşamasına geri çekildi (109 → 105)",
                            )
                            db.commit()
                            reopened = True
            except Exception as reopen_err:
                db.rollback()
                print(f"[WebBridge] add_repair_record statü geri çekme hatası: {reopen_err}")

            return json.dumps({"success": True, "id": str(rec.id), "reopened_for_decision": reopened})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def open_device_for_dismantle(self, imei, username):
        """Demontaj ekranında bir cihaz açıldığında çağrılır. Cihazın batch_entries.flow
        değerine göre warehouse.flow_dgd_mapping'den ilgili DGD işçilik kodunu bulur ve
        (henüz eklenmemişse) cihaza otomatik bir onarım kaydı (warehouse.repair_records,
        part_item_code=<dgd_kodu>, department_mission='DISMANTLE') ekler. İdempotenttir -
        aynı cihaz tekrar açıldığında zaten aktif bir DGD satırı varsa tekrar eklemez.
        add_repair_record'daki 'karar sonrası statü geri çekme' side-effect'i BİLEREK
        kullanılmıyor - bu otomatik/sessiz bir atama, kullanıcının bilinçli bir 'yeni onarım
        ekledim' aksiyonu değil."""
        import uuid
        from sqlalchemy import text
        from models.repair_record import RepairRecord
        db = SessionLocal()
        try:
            imei = (imei or "").strip()
            if not imei:
                return json.dumps({"success": False, "message": "IMEI boş olamaz."})

            entry = self._resolve_batch_entry_by_ref(db, imei)
            if not entry:
                return json.dumps({"success": True, "attached": False})

            flow = (db.execute(text("SELECT flow FROM warehouse.batch_entries WHERE id = :id"), {"id": entry["id"]}).scalar() or "").strip()
            if not flow:
                return json.dumps({"success": True, "attached": False})

            mapping = db.execute(text("""
                SELECT dgd_item_code FROM warehouse.flow_dgd_mapping
                WHERE LOWER(TRIM(flow_code)) = LOWER(:flow) AND enabled = TRUE
                LIMIT 1
            """), {"flow": flow}).mappings().first()
            if not mapping:
                return json.dumps({"success": True, "attached": False, "message": f"'{flow}' akışı için tanımlı DGD kodu yok."})
            dgd_item_code = mapping["dgd_item_code"]

            repair_refs = [imei]
            if entry["service_id"]:
                repair_refs.append(str(entry["service_id"]))

            existing = db.execute(text("""
                SELECT rr.id FROM warehouse.repair_records rr
                LEFT JOIN warehouse.repair_result_type rrt ON rrt.code = rr.repair_result_type_code
                WHERE rr.service_record_id = ANY(:refs) AND rr.part_item_code = :code
                  AND COALESCE(rrt.is_cancelled, FALSE) = FALSE
                LIMIT 1
            """), {"refs": repair_refs, "code": dgd_item_code}).mappings().first()
            if existing:
                return json.dumps({"success": True, "attached": False, "already_present": True})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                required = self._get_required_mission_for_ref(db, imei)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            rec = RepairRecord(
                id=uuid.uuid4(),
                service_record_id=self._resolve_service_record_id_for_new_repair(db, imei),
                department_mission="DISMANTLE",
                repair_result_type_code=1000,
                warranty_code="OOW",
                part_item_code=dgd_item_code,
                notes="Flow otomatik DGD ataması",
            )
            db.add(rec)
            db.commit()
            return json.dumps({"success": True, "attached": True, "dgd_item_code": dgd_item_code})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def toggle_dgd_repair_team(self, repair_id, username):
        """DGD (otomatik demontaj işçiliği) satırının onarım takımını L1REPAIR ↔ L2REPAIR
        arasında tek tuşla değiştirir. Satır hâlâ eski varsayılan 'DISMANTLE' kodundaysa
        (ekranda 'L1 Onarımı' gösterilir) L1 sayılıp L2REPAIR'e geçirilir. add_repair_record'daki
        L1/L2 karşılıklı dışlama kontrolünden BİLEREK muaf tutulur - bu yeni bir onarım eklemez,
        var olan tek satırın takımını değiştirir, kullanıcı isteğiyle her zaman izin verilir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            row = db.execute(text("""
                SELECT rr.id, rr.department_mission, rr.repair_result_type_code, p.item_category
                FROM warehouse.repair_records rr
                LEFT JOIN warehouse.parts p ON p.item_code = rr.part_item_code
                WHERE rr.id = :id
            """), {"id": repair_id}).mappings().first()

            if not row:
                return json.dumps({"success": False, "message": "Onarım kaydı bulunamadı."})
            if row["repair_result_type_code"] == 1003:
                return json.dumps({"success": False, "message": "İptal edilmiş bir onarımın takımı değiştirilemez."})
            is_dgd = (row["item_category"] == "DGD") or str(row["department_mission"] or "").upper() == "DISMANTLE"
            if not is_dgd:
                return json.dumps({"success": False, "message": "Bu işlem sadece DGD (otomatik demontaj işçiliği) satırları için geçerlidir."})

            new_team = "L1REPAIR" if row["department_mission"] == "L2REPAIR" else "L2REPAIR"
            db.execute(text("""
                UPDATE warehouse.repair_records SET department_mission = :team WHERE id = :id
            """), {"team": new_team, "id": repair_id})
            db.commit()
            return json.dumps({"success": True, "new_team": new_team})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def apply_dgd_return(self, device_ref, username):
        """Demontaj ekranındaki 'İade Et' aksiyonu: cihazın aktif DGD işçilik satırlarını
        (part_item_code, warehouse.parts.item_category='DGD', 'DGDDEC' hariç) iptal eder
        (repair_result_type_code=1003, mevcut 'İptal Et' konvansiyonuyla aynı - satır asla
        silinmez/üzerine yazılmaz) ve yerine tek bir DGDDEC (iade işçiliği) satırı ekler."""
        import uuid
        from sqlalchemy import text
        from models.repair_record import RepairRecord
        db = SessionLocal()
        try:
            device_ref = (device_ref or "").strip()
            if not device_ref:
                return json.dumps({"success": False, "message": "Cihaz bulunamadı."})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                required = self._get_required_mission_for_ref(db, device_ref)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            repair_refs = [device_ref]
            entry = self._resolve_batch_entry_by_ref(db, device_ref)
            if entry and entry["service_id"]:
                repair_refs.append(str(entry["service_id"]))

            active_dgd_rows = db.execute(text("""
                SELECT rr.id, rr.warranty_code
                FROM warehouse.repair_records rr
                JOIN warehouse.parts p ON p.item_code = rr.part_item_code
                LEFT JOIN warehouse.repair_result_type rrt ON rrt.code = rr.repair_result_type_code
                WHERE rr.service_record_id = ANY(:refs) AND p.item_category = 'DGD'
                  AND rr.part_item_code != 'DGDDEC'
                  AND COALESCE(rrt.is_cancelled, FALSE) = FALSE
            """), {"refs": repair_refs}).mappings().all()
            if not active_dgd_rows:
                return json.dumps({"success": False, "message": "İade edilecek aktif bir DGD işçilik kaydı yok."})

            for row in active_dgd_rows:
                db.execute(text("""
                    UPDATE warehouse.repair_records SET repair_result_type_code = 1003 WHERE id = :id
                """), {"id": row["id"]})

            warranty_code = active_dgd_rows[0]["warranty_code"] or "OOW"
            rec = RepairRecord(
                id=uuid.uuid4(),
                service_record_id=self._resolve_service_record_id_for_new_repair(db, device_ref),
                department_mission="DISMANTLE",
                repair_result_type_code=1000,
                warranty_code=warranty_code,
                part_item_code="DGDDEC",
                notes="DGD → DGDDEC iade dönüşümü",
            )
            db.add(rec)
            db.commit()
            return json.dumps({"success": True, "converted_count": len(active_dgd_rows)})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # --- MÜŞTERİ FİYAT MATRİSİ ---

    @Slot(result=str)
    def get_price_matrix_customers(self):
        """Fiyat matrisinin sütunlarını oluşturan müşteri listesini döner. warehouse.customers,
        uygulamanın gerçek 'Müşteriler' ekranının okuyup yazdığı canlı tablodur (organization.customers
        yalnızca Excel'den beslenen, kullanıcı tarafından düzenlenmeyen bir referans kopyasıdır)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT code, short_name, currency FROM warehouse.customers
                WHERE code IS NOT NULL ORDER BY short_name
            """)).mappings().all()
            items = [{"code": r["code"], "short_name": r["short_name"] or r["code"], "currency": r["currency"] or ""} for r in rows]
            return json.dumps({"success": True, "customers": items}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_price_matrix_brands(self):
        """Fiyat matrisinde önce seçilecek marka listesini döner - matrisin varsayılan
        yükleme davranışı artık tüm katalog (30 bin+ satır) yerine tek bir markanın
        parçalarını göstermektir, bu yüzden marka seçimi ilk ve zorunlu adımdır.
        warehouse.parts.brand serbest metin olarak girildiğinden aynı marka birden
        çok yazımla var olabilir (Samsung/SAMSUNG, Oppo/OPPO gibi); bunlar büyük/küçük
        harf duyarsız tek bir seçenekte gruplanır, etiket için en sık kullanılan yazım
        seçilir. Marka taşımayan DGD işçilik kodları için ayrı bir sözde-marka eklenir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT brand, COUNT(*) AS cnt
                FROM warehouse.parts
                WHERE brand IS NOT NULL AND brand != '' AND COALESCE(item_category, '') != 'DGD'
                GROUP BY brand
            """)).mappings().all()

            grouped = {}
            for r in rows:
                raw = (r["brand"] or "").strip()
                if not raw:
                    continue
                key = raw.upper()
                g = grouped.setdefault(key, {"label": raw, "label_count": 0, "count": 0})
                g["count"] += r["cnt"]
                if r["cnt"] > g["label_count"]:
                    g["label_count"] = r["cnt"]
                    g["label"] = raw

            brands = sorted(
                [{"value": key, "label": g["label"], "count": g["count"]} for key, g in grouped.items()],
                key=lambda b: b["label"]
            )

            dgd_count = db.execute(text(
                "SELECT COUNT(*) FROM warehouse.parts WHERE item_category = 'DGD'"
            )).scalar() or 0
            if dgd_count:
                brands.append({"value": "__DGD__", "label": "İşçilik (DGD)", "count": dgd_count})

            return json.dumps({"success": True, "brands": brands}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_price_matrix_product_types(self, brand=""):
        """Fiyat matrisinde markanın yanındaki 'Ürün Tipi' (Akıllı Telefon/Tablet/Dizüstü
        Bilgisayar/Bluetooth Kulaklık/Akıllı Saat) filtresinin seçeneklerini döner - bkz.
        PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL (model metninden türetilir, warehouse.parts'ta
        ayrı bir DB kolonu yoktur). Ekran adı (label) uydurulmaz; CASE ifadesinin ürettiği
        değer (SMART PHONE/TABLET/...) warehouse.product_category.code ile birebir aynı
        olduğundan gerçek Türkçe adı (short_name) o var olan referans tablosundan JOIN ile
        çekilir - 'value' (filtreleme için kullanılan kod) sabit kalır, sadece 'label'
        (ekranda görünen) product_category'den gelir. Model listesiyle 'uyumlu' çalışır:
        bkz. get_price_matrix_models'in product_type parametresi. '__DGD__' için ürün tipi
        ayrımı yoktur (işçilik kodlarının cihaz modeli yoktur)."""
        from sqlalchemy import text
        brand = (brand or "").strip()
        if brand == "__DGD__":
            return json.dumps({"success": True, "product_types": []}, ensure_ascii=False)

        db = SessionLocal()
        try:
            params = {}
            clause = "WHERE COALESCE(item_category, '') != 'DGD'"
            if brand:
                clause += " AND UPPER(brand) = UPPER(:brand)"
                params["brand"] = brand
            rows = db.execute(text(f"""
                SELECT sub.product_type AS value, COALESCE(pc.short_name, sub.product_type) AS label, sub.cnt
                FROM (
                    SELECT {PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL} AS product_type, COUNT(*) AS cnt
                    FROM warehouse.parts
                    {clause}
                    GROUP BY product_type
                ) sub
                LEFT JOIN warehouse.product_category pc ON pc.code = sub.product_type
                ORDER BY sub.cnt DESC
            """), params).mappings().all()
            product_types = [{"value": r["value"], "label": r["label"], "count": r["cnt"]} for r in rows]
            return json.dumps({"success": True, "product_types": product_types}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def get_price_matrix_models(self, brand="", product_type=""):
        """Seçili marka (+ opsiyonel ürün tipi) için cihaz modeli listesini döner - fiyat
        matrisinde markadan sonraki daraltma adımlarından biridir, ürün tipi ve kategori ile
        birlikte 'uyumlu' çalışır (ürün tipi seçiliyse model listesi de SADECE o tipe göre
        daralır: bkz. PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL). Filtreleme değeri ('value') HER
        ZAMAN warehouse.parts.model'in ham hâlidir (parça sorguları bunu birebir kullanır) -
        ama bu kolon marka bazında tutarsız: bazı satırlarda okunaklı isim ('Galaxy S21'),
        bazılarında (özellikle Apple telefonlarında) kısa dahili kod var ('iP12PM'). Ekranda
        okunaklı isim ('label') göstermek için warehouse.product_family'de aynı marka altında
        code = model eşleşen bir kayıt varsa short_name'i ('iPhone 12 Pro Max') kullanılır,
        yoksa ham model metni aynen gösterilir (product_family her modeli kapsamıyor)."""
        from sqlalchemy import text
        brand = (brand or "").strip()
        product_type = (product_type or "").strip()
        if not brand or brand == "__DGD__":
            return json.dumps({"success": True, "models": []}, ensure_ascii=False)

        db = SessionLocal()
        try:
            params = {"brand": brand}
            clause = "WHERE UPPER(parts.brand) = UPPER(:brand) AND parts.model IS NOT NULL AND parts.model != ''"
            if product_type:
                clause += f" AND ({PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL}) = :product_type"
                params["product_type"] = product_type
            rows = db.execute(text(f"""
                SELECT parts.model AS value, COALESCE(pf.short_name, parts.model) AS label, COUNT(*) AS cnt
                FROM warehouse.parts
                LEFT JOIN warehouse.product_family pf
                    ON UPPER(pf.code) = UPPER(parts.model) AND UPPER(pf.brand) = UPPER(parts.brand)
                {clause}
                GROUP BY parts.model, pf.short_name
                ORDER BY label
            """), params).mappings().all()
            models = [{"value": r["value"], "label": r["label"], "count": r["cnt"]} for r in rows]
            return json.dumps({"success": True, "models": models}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def get_price_matrix_categories(self, brand="", model="", product_type=""):
        """Seçili marka (+ opsiyonel ürün tipi/model) için item_category (parça tipi:
        Middle Frame, Back Glass, LCD...) listesini döner - fiyat matrisinde markadan
        sonraki daraltma adımlarından biridir, ürün tipi ve model ile birlikte 'uyumlu'
        çalışır (ikisi de seçiliyse kategori listesi SADECE o alt kümeye göre daralır).
        brand boş verilirse tüm markalardaki kategoriler döner; '__DGD__' için kategori
        ayrımı yoktur (işçilik kodları zaten tek bir marka-eşleniğinde toplanmıştır)."""
        from sqlalchemy import text
        brand = (brand or "").strip()
        model = (model or "").strip()
        product_type = (product_type or "").strip()
        if brand == "__DGD__":
            return json.dumps({"success": True, "categories": []}, ensure_ascii=False)

        db = SessionLocal()
        try:
            params = {}
            clause = "WHERE item_category IS NOT NULL AND item_category != '' AND item_category != 'DGD'"
            if brand:
                clause += " AND UPPER(brand) = UPPER(:brand)"
                params["brand"] = brand
            if model:
                clause += " AND model = :model"
                params["model"] = model
            if product_type:
                clause += f" AND ({PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL}) = :product_type"
                params["product_type"] = product_type
            rows = db.execute(text(f"""
                SELECT item_category, COUNT(*) AS cnt
                FROM warehouse.parts
                {clause}
                GROUP BY item_category
                ORDER BY item_category
            """), params).mappings().all()
            categories = [{"value": r["item_category"], "label": r["item_category"], "count": r["cnt"]} for r in rows]
            return json.dumps({"success": True, "categories": categories}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, result=str)
    def get_price_matrix_items(self, search="", brand="", category="", model="", product_type=""):
        """Fiyat matrisinin satırlarını oluşturan item_code listesini döner (fiziksel parçalar
        + DGD işçilik kodları), her satırda türetilmiş bir 'İşçilik'/'Parça' etiketiyle.
        warehouse.parts on binlerce satır olabildiğinden (bkz. get_parts/get_stock_status),
        frontend artık markaya (ve opsiyonel ürün tipi/model/kategoriye) göre daraltılmış bir
        alt küme ister - bu yüzden herhangi biri verildiğinde doğrudan (küçük, hızlı)
        sorgulanır, önbelleklenmez. Hiçbir filtre verilmeyen (search='', brand='', category='',
        model='', product_type='') geriye dönük çağrı için tam katalog, diğer büyük listelerle
        aynı api_cache/*.json + fetch_url deseniyle döner."""
        from sqlalchemy import text
        search = (search or "").strip()
        brand = (brand or "").strip()
        category = (category or "").strip()
        model = (model or "").strip()
        product_type = (product_type or "").strip()
        unfiltered = not search and not brand and not category and not model and not product_type

        if unfiltered:
            filename = "price_matrix_items.json"
            path = os.path.join(get_cache_dirs()[0], filename)
            fetch_url = f"/api_cache/{filename}"
            if os.path.exists(path):
                return json.dumps({"success": True, "fetch_url": fetch_url})

        db = SessionLocal()
        try:
            clauses = []
            params = {}
            if search:
                clauses.append("(item_code ILIKE :s OR name ILIKE :s)")
                params["s"] = f"%{search}%"
            if brand == "__DGD__":
                clauses.append("item_category = 'DGD'")
            elif brand:
                clauses.append("UPPER(brand) = UPPER(:brand)")
                params["brand"] = brand
            if model:
                clauses.append("model = :model")
                params["model"] = model
            if product_type:
                clauses.append(f"({PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL}) = :product_type")
                params["product_type"] = product_type
            if category:
                clauses.append("item_category = :category")
                params["category"] = category
            clause = ("WHERE " + " AND ".join(clauses)) if clauses else ""

            rows = db.execute(text(f"""
                SELECT item_code, name,
                       CASE WHEN item_category = 'DGD' THEN 'İşçilik' ELSE 'Parça' END AS item_type
                FROM warehouse.parts
                {clause}
                ORDER BY item_type, item_code
            """), params).mappings().all()
            items = [{"item_code": r["item_code"], "name": r["name"] or "", "item_type": r["item_type"]} for r in rows if r["item_code"]]

            if unfiltered:
                json_data = json.dumps({"success": True, "items": items}, ensure_ascii=False)
                write_to_cache("price_matrix_items.json", json_data)
                fetch_url = f"/api_cache/price_matrix_items.json"
                return json.dumps({"success": True, "fetch_url": fetch_url})

            return json.dumps({"success": True, "items": items}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def get_price_matrix(self, brand="", category="", model="", product_type=""):
        """warehouse.customer_item_prices'taki (item_code, customer_code, price) satırlarını
        döner - frontend bunu {item_code: {customer_code: price}} pivot haritasına çevirir.
        Varsayılan görünüm (Müşteri Fiyat Matrisi'nde marka/model/kategori seçilmeden) TÜM
        katalog + işçilik kodlarıdır (bkz. get_price_matrix_items'ın unfiltered dalı) - bu
        yüzden hiç filtre verilmediğinde bu fonksiyon da diğer büyük listelerle aynı
        api_cache/*.json + fetch_url deseniyle döner: 30 bin parça x N müşteri onbinlerce/
        yüzbinlerce satıra çıkabildiğinden (test verisiyle 572K oldu) tüm tabloyu her
        seferinde QWebChannel'dan inline döndürmek get_price_matrix_items için çözdüğümüz
        sorunu geri getirir. Marka/model/kategori/ürün-tipi verildiğinde (kullanıcı daraltma
        filtrelerinden birini seçtiğinde) items ile AYNI filtrelerle, warehouse.parts'a JOIN
        edilerek doğrudan (küçük, hızlı) sorgulanır, önbelleklenmez."""
        from sqlalchemy import text
        brand = (brand or "").strip()
        category = (category or "").strip()
        model = (model or "").strip()
        product_type = (product_type or "").strip()
        unfiltered = not brand and not category and not model and not product_type

        if unfiltered:
            filename = "price_matrix_prices.json"
            path = os.path.join(get_cache_dirs()[0], filename)
            fetch_url = f"/api_cache/{filename}"
            if os.path.exists(path):
                return json.dumps({"success": True, "fetch_url": fetch_url})

        db = SessionLocal()
        try:
            if unfiltered:
                rows = db.execute(text(
                    "SELECT item_code, customer_code, price FROM warehouse.customer_item_prices"
                )).mappings().all()
            else:
                clauses = []
                params = {}
                if brand == "__DGD__":
                    clauses.append("item_category = 'DGD'")
                elif brand:
                    clauses.append("UPPER(brand) = UPPER(:brand)")
                    params["brand"] = brand
                if model:
                    clauses.append("model = :model")
                    params["model"] = model
                if product_type:
                    clauses.append(f"({PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL}) = :product_type")
                    params["product_type"] = product_type
                if category:
                    clauses.append("item_category = :category")
                    params["category"] = category
                clause = ("WHERE " + " AND ".join(clauses)) if clauses else ""

                rows = db.execute(text(f"""
                    SELECT cip.item_code, cip.customer_code, cip.price
                    FROM warehouse.customer_item_prices cip
                    JOIN warehouse.parts ON parts.item_code = cip.item_code
                    {clause}
                """), params).mappings().all()

            items = [{"item_code": r["item_code"], "customer_code": r["customer_code"], "price": float(r["price"])} for r in rows]

            if unfiltered:
                json_data = json.dumps({"success": True, "prices": items}, ensure_ascii=False)
                write_to_cache("price_matrix_prices.json", json_data)
                fetch_url = f"/api_cache/price_matrix_prices.json"
                return json.dumps({"success": True, "fetch_url": fetch_url})

            return json.dumps({"success": True, "prices": items}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def save_price_matrix_batch(self, rows_json, username):
        """Fiyat matrisi ızgarasındaki tüm 'kirli' (değiştirilmiş) hücreleri TEK bir transaction
        içinde kaydeder. rows_json: [{item_code, customer_code, price}] - price null/boşsa o
        hücre matristen SİLİNİR (warehouse.item.satis'e geri düşülür), doluysa upsert edilir."""
        import uuid
        from sqlalchemy import text
        db = SessionLocal()
        try:
            try:
                rows = json.loads(rows_json) if rows_json else []
            except (TypeError, ValueError):
                return json.dumps({"success": False, "message": "Geçersiz veri formatı."})

            updated_count = 0
            deleted_count = 0
            for row in rows:
                item_code = (row.get("item_code") or "").strip()
                customer_code = (row.get("customer_code") or "").strip()
                price = row.get("price", None)
                if not item_code or not customer_code:
                    continue
                if price is None or price == "":
                    result = db.execute(text("""
                        DELETE FROM warehouse.customer_item_prices
                        WHERE item_code = :item_code AND customer_code = :customer_code
                    """), {"item_code": item_code, "customer_code": customer_code})
                    deleted_count += result.rowcount
                else:
                    db.execute(text("""
                        INSERT INTO warehouse.customer_item_prices (id, item_code, customer_code, price, updated_by, updated_at)
                        VALUES (:id, :item_code, :customer_code, :price, :username, now())
                        ON CONFLICT (item_code, customer_code)
                        DO UPDATE SET price = EXCLUDED.price, updated_by = EXCLUDED.updated_by, updated_at = now()
                    """), {"id": str(uuid.uuid4()), "item_code": item_code, "customer_code": customer_code, "price": float(price), "username": username or None})
                    updated_count += 1

            db.commit()
            return json.dumps({"success": True, "updated_count": updated_count, "deleted_count": deleted_count})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def get_prices_for_items(self, item_codes_csv, customer_code):
        """get_effective_price'ın toplu (çoklu item_code, TEK sorgu) hali - Üretim Kaydını
        Görüntüle ve Üretime Aktar ekranlarındaki 'Fiyat' sütunu için, bir onarım grubundaki
        her parça için ayrı ayrı çağrı yapmak yerine tek seferde tüm fiyatları getirir. Aynı
        kural: önce customer_item_prices (müşteriye özel), yoksa warehouse.item.satis
        (genel varsayılan)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            item_codes = list({c.strip() for c in (item_codes_csv or "").split(",") if c.strip()})
            customer_code = (customer_code or "").strip()
            if not item_codes:
                return json.dumps({"success": True, "prices": {}})

            prices = {}
            if customer_code:
                rows = db.execute(text("""
                    SELECT item_code, price FROM warehouse.customer_item_prices
                    WHERE customer_code = :customer_code AND item_code = ANY(:codes)
                """), {"customer_code": customer_code, "codes": item_codes}).mappings().all()
                for r in rows:
                    prices[r["item_code"]] = float(r["price"])

            missing = [c for c in item_codes if c not in prices]
            if missing:
                rows2 = db.execute(text("""
                    SELECT code, satis FROM warehouse.item WHERE code = ANY(:codes)
                """), {"codes": missing}).mappings().all()
                for r in rows2:
                    if r["satis"] is not None:
                        prices[r["code"]] = float(r["satis"])

            return json.dumps({"success": True, "prices": prices})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def get_effective_price(self, item_code, customer_code):
        """item_code x customer_code için geçerli fiyatı döner: önce customer_item_prices'tan,
        yoksa warehouse.item.satis'ten (global varsayılan). Şu an sadece lookup Slot'u olarak
        var - herhangi bir fatura/maliyet hesaplamasına bağlanması kasıtlı olarak kapsam dışı."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            item_code = (item_code or "").strip()
            customer_code = (customer_code or "").strip()
            if not item_code:
                return json.dumps({"success": False, "message": "item_code zorunludur."})

            if customer_code:
                row = db.execute(text("""
                    SELECT price FROM warehouse.customer_item_prices
                    WHERE item_code = :item_code AND customer_code = :customer_code
                """), {"item_code": item_code, "customer_code": customer_code}).mappings().first()
                if row:
                    return json.dumps({"success": True, "price": float(row["price"]), "source": "matrix"})

            default_price = db.execute(text("SELECT satis FROM warehouse.item WHERE code = :code"), {"code": item_code}).scalar()
            if default_price is not None:
                return json.dumps({"success": True, "price": float(default_price), "source": "item_default"})

            return json.dumps({"success": True, "price": None, "source": "none"})
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    def _get_effective_price(self, db, item_code, customer_code):
        """get_effective_price Slot'unun DB oturumu paylaşan iç hali (aynı kural: önce
        customer_item_prices, yoksa item.satis) - submit_dismantle_decision'ın hedef fiyat
        limit kontrolü için eklenen parçaların toplam fiyatını hesaplarken kullanır."""
        from sqlalchemy import text
        item_code = (item_code or "").strip()
        if not item_code:
            return None
        customer_code = (customer_code or "").strip()
        if customer_code:
            row = db.execute(text("""
                SELECT price FROM warehouse.customer_item_prices
                WHERE item_code = :item_code AND customer_code = :customer_code
            """), {"item_code": item_code, "customer_code": customer_code}).mappings().first()
            if row:
                return float(row["price"])
        default_price = db.execute(text("SELECT satis FROM warehouse.item WHERE code = :code"), {"code": item_code}).scalar()
        return float(default_price) if default_price is not None else None

    # --- MÜŞTERİ HEDEF FİYAT MATRİSİ ---
    # Demontaj ekranında eklenen onarım parçalarının toplam fiyatı (bkz. _get_effective_price)
    # bu tablodaki (müşteri, model, screen test, power test) limitini aşarsa cihaz otomatik
    # Müşteri Onayına yönlendirilir (bkz. submit_dismantle_decision). brand/product_type,
    # product_family_code'dan otomatik türetilir - Müşteri Fiyat Matrisi'ndeki marka/ürün
    # tipi/model desenleriyle (get_price_matrix_brands/get_price_matrix_models,
    # PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL) TUTARLI kalması için aynı yardımcılar kullanılır.

    @Slot(result=str)
    def get_target_price_customers(self):
        """Hedef Fiyat Matrisi'nin müşteri seçeneklerini döner - get_price_matrix_customers
        ile aynı kaynak (warehouse.customers), ayrı bir Slot olarak tutulur ki iki modül
        birbirinden bağımsız evrimleşebilsin."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT code, short_name, currency FROM warehouse.customers
                WHERE code IS NOT NULL ORDER BY short_name
            """)).mappings().all()
            items = [{"code": r["code"], "short_name": r["short_name"] or r["code"], "currency": r["currency"] or ""} for r in rows]
            return json.dumps({"success": True, "customers": items}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_target_price_brands(self):
        """Marka listesini warehouse.product_family'den döner (parts değil - burada model
        seçimi product_family_code üzerinden yapılır). get_price_matrix_brands ile AYNI
        büyük/küçük harf normalizasyon deseni: aynı marka birden çok yazımla olabilir
        (Samsung/SAMSUNG), büyük/küçük harf duyarsız tek seçenekte gruplanır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT brand, COUNT(*) AS cnt
                FROM warehouse.product_family
                WHERE brand IS NOT NULL AND brand != ''
                GROUP BY brand
            """)).mappings().all()

            grouped = {}
            for r in rows:
                raw = (r["brand"] or "").strip()
                if not raw:
                    continue
                key = raw.upper()
                g = grouped.setdefault(key, {"label": raw, "label_count": 0, "count": 0})
                g["count"] += r["cnt"]
                if r["cnt"] > g["label_count"]:
                    g["label_count"] = r["cnt"]
                    g["label"] = raw

            brands = sorted(
                [{"value": key, "label": g["label"], "count": g["count"]} for key, g in grouped.items()],
                key=lambda b: b["label"]
            )
            return json.dumps({"success": True, "brands": brands}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_target_price_models(self, brand=""):
        """Seçili markaya ait cihaz modeli (warehouse.product_family: code + short_name)
        listesini döner - product_family_code, Hedef Fiyat Matrisi'nde TEK zorunlu cihaz
        seçim noktasıdır (bkz. models/customer_target_price.py). Her satırda ürün tipi
        (product_type) de PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL ile short_name'den türetilerek
        döner - frontend bunu salt-okunur gösterir, create_customer_target_price de aynı
        türetmeyi tekrar (sunucu tarafında, güvenilir) yapar."""
        from sqlalchemy import text
        brand = (brand or "").strip()
        if not brand:
            return json.dumps({"success": True, "models": []}, ensure_ascii=False)

        db = SessionLocal()
        try:
            case_sql = PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL.replace("model", "short_name")
            rows = db.execute(text(f"""
                SELECT code, short_name, ({case_sql}) AS product_type
                FROM warehouse.product_family
                WHERE UPPER(brand) = UPPER(:brand) AND code IS NOT NULL AND code != ''
                ORDER BY short_name
            """), {"brand": brand}).mappings().all()
            models = [{"value": r["code"], "label": r["short_name"] or r["code"], "productType": r["product_type"]} for r in rows]
            return json.dumps({"success": True, "models": models}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def get_customer_target_prices(self, customer_code=""):
        """Hedef Fiyat Matrisi tablosundaki kuralları döner - customer_code verilirse
        sadece o müşteriye ait olanlar (CRUD ekranındaki liste görünümü)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            customer_code = (customer_code or "").strip()
            params = {}
            clause = ""
            if customer_code:
                clause = "WHERE customer_code = :customer_code"
                params["customer_code"] = customer_code
            rows = db.execute(text(f"""
                SELECT id, customer_code, product_family_code, brand, product_type,
                       screen_test_result, power_test_result, target_price, currency
                FROM warehouse.customer_target_prices
                {clause}
                ORDER BY customer_code, brand, product_family_code
            """), params).mappings().all()
            items = [{
                "id": str(r["id"]),
                "customerCode": r["customer_code"],
                "productFamilyCode": r["product_family_code"],
                "brand": r["brand"] or "",
                "productType": r["product_type"] or "",
                "screenTestResult": r["screen_test_result"],
                "powerTestResult": r["power_test_result"],
                "targetPrice": float(r["target_price"]),
                "currency": r["currency"] or "",
            } for r in rows]
            return json.dumps({"success": True, "rules": items}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, result=str)
    def create_customer_target_price(self, customer_code, product_family_code, screen_test_result, power_test_result, target_price, username):
        """Yeni bir Hedef Fiyat Matrisi kuralı ekler. brand/product_type/currency BURADA,
        sunucu tarafında product_family_code ve customer'dan türetilir - istemciden gelen
        bir değer YOKTUR, bu yüzden marka/model/müşteri arasında çelişki oluşamaz (bkz.
        tasarım kararı: model her zaman zorunlu, brand/tip otomatik dolar)."""
        import uuid
        from sqlalchemy import text
        db = SessionLocal()
        try:
            customer_code = (customer_code or "").strip()
            product_family_code = (product_family_code or "").strip()
            screen_test_result = (screen_test_result or "").strip().upper()
            power_test_result = (power_test_result or "").strip().upper()

            if not customer_code:
                return json.dumps({"success": False, "message": "Müşteri zorunludur."})
            if not product_family_code:
                return json.dumps({"success": False, "message": "Model zorunludur."})
            if screen_test_result not in ("OK", "NOK", "BOŞ"):
                return json.dumps({"success": False, "message": "Screen Test sonucu OK/NOK/BOŞ olmalıdır."})
            # Power Test, Screen Test'ten daha üstün/öncelikli bir alan olduğu için (fiyatlandırma
            # kuralı buna dayanıyor) her zaman OK veya NOK olmalıdır, asla boş/BOŞ olamaz.
            if power_test_result not in ("OK", "NOK"):
                return json.dumps({"success": False, "message": "Power Test sonucu OK veya NOK olmalıdır, boş (BOŞ) bırakılamaz."})
            try:
                price_val = float(target_price)
            except (TypeError, ValueError):
                return json.dumps({"success": False, "message": "Hedef fiyat sayısal olmalıdır."})

            case_sql = PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL.replace("model", "short_name")
            fam = db.execute(text(f"""
                SELECT brand, ({case_sql}) AS product_type
                FROM warehouse.product_family WHERE code = :code
            """), {"code": product_family_code}).mappings().first()
            if not fam:
                return json.dumps({"success": False, "message": f"'{product_family_code}' kodlu model bulunamadı."})

            cust = db.execute(text("SELECT currency FROM warehouse.customers WHERE code = :code"),
                               {"code": customer_code}).mappings().first()

            db.execute(text("""
                INSERT INTO warehouse.customer_target_prices
                    (id, customer_code, product_family_code, brand, product_type,
                     screen_test_result, power_test_result, target_price, currency, updated_by, updated_at)
                VALUES (:id, :customer_code, :pfc, :brand, :ptype, :screen, :power, :price, :currency, :user, now())
            """), {
                "id": str(uuid.uuid4()), "customer_code": customer_code, "pfc": product_family_code,
                "brand": fam["brand"], "ptype": fam["product_type"], "screen": screen_test_result,
                "power": power_test_result, "price": price_val,
                "currency": (cust["currency"] if cust else None), "user": username or None,
            })
            db.commit()
            return json.dumps({"success": True, "message": "Hedef fiyat kuralı eklendi."})
        except Exception as e:
            db.rollback()
            msg = str(e)
            if "uq_customer_target_price" in msg:
                msg = "Bu müşteri/model/test kombinasyonu için zaten bir kural var."
            return json.dumps({"success": False, "message": msg})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def update_customer_target_price(self, id_str, target_price, username):
        """Var olan bir kuralın hedef fiyatını günceller. Müşteri/model/test kombinasyonu
        (kuralın kimliği) değiştirilemez - değişmesi gerekiyorsa kural silinip yeniden
        eklenir (delete_customer_target_price + create_customer_target_price)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            try:
                price_val = float(target_price)
            except (TypeError, ValueError):
                return json.dumps({"success": False, "message": "Hedef fiyat sayısal olmalıdır."})

            result = db.execute(text("""
                UPDATE warehouse.customer_target_prices
                SET target_price = :price, updated_by = :user, updated_at = now()
                WHERE id::text = :id
            """), {"price": price_val, "user": username or None, "id": id_str})
            if result.rowcount == 0:
                return json.dumps({"success": False, "message": "Kural bulunamadı."})
            db.commit()
            return json.dumps({"success": True, "message": "Hedef fiyat güncellendi."})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, result=str)
    def delete_customer_target_price(self, id_str):
        from sqlalchemy import text
        db = SessionLocal()
        try:
            result = db.execute(text("DELETE FROM warehouse.customer_target_prices WHERE id::text = :id"), {"id": id_str})
            if result.rowcount == 0:
                return json.dumps({"success": False, "message": "Kural bulunamadı."})
            db.commit()
            return json.dumps({"success": True, "message": "Hedef fiyat kuralı silindi."})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def bulk_import_customer_target_prices(self, rows_json, username):
        """Hedef Fiyat Matrisi için toplu (Excel) içe aktarma. bulk_import_customers'daki
        AYNI 'hepsi ya da hiçbiri' deseni: önce TÜM satırlar doğrulanır (müşteri/model
        kod VEYA okunaklı isimle çözülür, test sonucu OK/NOK/BOŞ olmalı, fiyat sayısal
        olmalı); herhangi bir satırda hata varsa HİÇBİR satır kaydedilmez, tüm hatalar
        satır numarasıyla döner. Doğrulama geçerse tüm satırlar UPSERT edilir (aynı
        müşteri/model/test kombinasyonu dosyada tekrar geçerse veya sistemde zaten
        varsa fiyatı GÜNCELLENİR, save_price_matrix_batch'teki ON CONFLICT deseniyle
        aynı - create_customer_target_price'ın aksine burada duplicate reddedilmez,
        çünkü Excel'in asıl kullanım amacı var olan bir listeyi güncellemektir)."""
        import uuid
        from sqlalchemy import text
        db = SessionLocal()
        try:
            try:
                rows = json.loads(rows_json or "[]")
            except (ValueError, TypeError):
                return json.dumps({"success": False, "message": "Geçersiz dosya verisi.", "errors": []})
            if not rows:
                return json.dumps({"success": False, "message": "Dosyada içe aktarılacak satır bulunamadı.", "errors": []})

            customers = db.execute(text("SELECT code, short_name, currency FROM warehouse.customers WHERE code IS NOT NULL")).mappings().all()
            customer_by_code = {c["code"].strip().lower(): c["code"] for c in customers}
            customer_by_name = {(c["short_name"] or "").strip().lower(): c["code"] for c in customers if c["short_name"]}
            currency_by_code = {c["code"]: c["currency"] for c in customers}

            families = db.execute(text("SELECT code, short_name, brand FROM warehouse.product_family WHERE code IS NOT NULL")).mappings().all()
            family_by_code = {f["code"].strip().lower(): f for f in families}
            family_by_name = {(f["short_name"] or "").strip().lower(): f for f in families if f["short_name"]}

            errors = []
            valid_rows = []
            seen_keys_in_file = {}

            for idx, row in enumerate(rows):
                row_num = idx + 2  # 1. satır başlık
                row = row or {}

                def get_val(key):
                    v = row.get(key)
                    return str(v).strip() if v is not None else ""

                musteri_raw = get_val("musteri")
                model_raw = get_val("model")
                screen_raw = get_val("screen_test").upper()
                power_raw = get_val("power_test").upper()
                price_raw = get_val("hedef_fiyat")

                if not musteri_raw:
                    errors.append({"row": row_num, "field": "Müşteri", "message": "Müşteri boş olamaz."})
                if not model_raw:
                    errors.append({"row": row_num, "field": "Model", "message": "Model boş olamaz."})
                if screen_raw not in ("OK", "NOK", "BOŞ", "BOS"):
                    errors.append({"row": row_num, "field": "Screen Test", "message": f"\"{screen_raw}\" geçersiz. OK, NOK veya BOŞ olmalıdır."})
                # Power Test her zaman OK/NOK olmalıdır - BOŞ kabul edilmez (Screen Test'ten
                # farklı olarak, fiyatlandırma kuralı Power Test'e dayandığından zorunludur).
                if power_raw not in ("OK", "NOK"):
                    errors.append({"row": row_num, "field": "Power Test", "message": f"\"{power_raw}\" geçersiz. Power Test OK veya NOK olmalıdır, boş bırakılamaz."})
                price_val = None
                if not price_raw:
                    errors.append({"row": row_num, "field": "Hedef Fiyat", "message": "Hedef Fiyat boş olamaz."})
                else:
                    try:
                        price_val = float(str(price_raw).replace(",", "."))
                    except ValueError:
                        errors.append({"row": row_num, "field": "Hedef Fiyat", "message": f"\"{price_raw}\" sayısal değil."})

                customer_code = customer_by_code.get(musteri_raw.lower()) or customer_by_name.get(musteri_raw.lower())
                if musteri_raw and not customer_code:
                    errors.append({"row": row_num, "field": "Müşteri", "message": f"\"{musteri_raw}\" sistemde tanımlı bir müşteri değil (kod veya ad ile eşleşmedi)."})

                fam = family_by_code.get(model_raw.lower()) or family_by_name.get(model_raw.lower())
                if model_raw and not fam:
                    errors.append({"row": row_num, "field": "Model", "message": f"\"{model_raw}\" sistemde tanımlı bir cihaz modeli değil (kod veya ad ile eşleşmedi)."})

                if customer_code and fam:
                    screen_norm = "BOŞ" if screen_raw in ("BOŞ", "BOS") else screen_raw
                    power_norm = power_raw  # buraya kadar gelindiyse zaten OK/NOK, BOŞ olamaz
                    dup_key = (customer_code, fam["code"], screen_norm, power_norm)
                    if dup_key in seen_keys_in_file:
                        errors.append({"row": row_num, "field": "Müşteri/Model/Test", "message": f"Bu kombinasyon dosyada satır {seen_keys_in_file[dup_key]} ile tekrarlanıyor - son değer kullanılacak."})
                    seen_keys_in_file[dup_key] = row_num
                    valid_rows.append({
                        "customer_code": customer_code, "product_family_code": fam["code"], "brand": fam["brand"],
                        "screen": screen_norm, "power": power_norm, "price": price_val,
                        "currency": currency_by_code.get(customer_code),
                    })

            blocking_errors = [e for e in errors if "tekrarlanıyor" not in e["message"]]
            if blocking_errors:
                return json.dumps({"success": False, "message": f"{len(blocking_errors)} satırda hata bulundu, hiçbir kayıt eklenmedi.", "errors": errors}, ensure_ascii=False)

            case_sql = PRICE_MATRIX_PRODUCT_TYPE_CASE_SQL.replace("model", "short_name")
            imported = 0
            for r in valid_rows:
                ptype = db.execute(text(f"SELECT ({case_sql}) FROM warehouse.product_family WHERE code = :c"), {"c": r["product_family_code"]}).scalar()
                db.execute(text("""
                    INSERT INTO warehouse.customer_target_prices
                        (id, customer_code, product_family_code, brand, product_type,
                         screen_test_result, power_test_result, target_price, currency, updated_by, updated_at)
                    VALUES (:id, :customer_code, :pfc, :brand, :ptype, :screen, :power, :price, :currency, :user, now())
                    ON CONFLICT (customer_code, product_family_code, screen_test_result, power_test_result)
                    DO UPDATE SET target_price = EXCLUDED.target_price, currency = EXCLUDED.currency,
                                  updated_by = EXCLUDED.updated_by, updated_at = now()
                """), {
                    "id": str(uuid.uuid4()), "customer_code": r["customer_code"], "pfc": r["product_family_code"],
                    "brand": r["brand"], "ptype": ptype, "screen": r["screen"], "power": r["power"],
                    "price": r["price"], "currency": r["currency"], "user": username or None,
                })
                imported += 1

            db.commit()
            return json.dumps({"success": True, "imported": imported, "message": f"{imported} kural içe aktarıldı."}, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e), "errors": []})
        finally:
            db.close()

    @Slot(str, str, str, str, str, str, str, str, result=str)
    def update_repair_record(self, repair_id, mission_group_code, warranty_code, notes, username, part_item_code="", item_fault_code="", operation_type_code=""):
        """Mevcut bir onarım kaydını (warehouse.repair_records) tüm alanlarıyla günceller.
        Demontaj ekranındaki 'Teklif Parçaları' listesindeki bir satırın Düzenle aksiyonu -
        aynı toolbar (Parça/Arıza Tespiti/Onarım Takımı/Ücret Tipi/Açıklama) tekrar kullanılarak
        değerler değiştirilip buraya gönderilir. add_repair_record ile aynı RBAC deseni."""
        db = SessionLocal()
        try:
            from models.repair_record import RepairRecord
            rec = db.query(RepairRecord).filter(RepairRecord.id == repair_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Onarım kaydı bulunamadı."})
            if not mission_group_code or not mission_group_code.strip():
                return json.dumps({"success": False, "message": "Görev grubu zorunludur."})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                required = self._get_required_mission_for_ref(db, rec.service_record_id)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            rec.department_mission = mission_group_code.strip()
            rec.warranty_code = warranty_code.strip() if warranty_code else "OOW"
            rec.notes = notes.strip() if notes else None
            rec.part_item_code = part_item_code.strip() if part_item_code else None
            rec.item_fault_code = item_fault_code.strip() if item_fault_code else None
            rec.operation_type_code = operation_type_code.strip() if operation_type_code else None
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def delete_repair_record(self, repair_id, username):
        """Bir onarım kaydını (warehouse.repair_records) siler. add_repair_record ile aynı
        RBAC deseni - sadece cihazın mevcut statüsünde yetkili olan (veya admin) kullanıcılar silebilir."""
        db = SessionLocal()
        try:
            from models.repair_record import RepairRecord
            rec = db.query(RepairRecord).filter(RepairRecord.id == repair_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Onarım kaydı bulunamadı."})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                required = self._get_required_mission_for_ref(db, rec.service_record_id)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            db.delete(rec)
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, result=str)
    def submit_dismantle_decision(self, imei, username):
        """Demontaj Teknisyeni'nin bir cihaza eklediği onarım kayıtlarını, cihazın Flow'una
        (Akış Durumu, warehouse.batch_entries.flow - warehouse.service_request_type.code ile
        aynı değer kümesi: 'To repair', 'To refurbish', 'Battery only', 'To RMA') göre
        warehouse.service_request_item_category'de tanımlı, o flow için önceden onaylanmış
        (is_customer_approved=TRUE) parça kategorileriyle karşılaştırır. Eklenen onarımların
        HEPSİ bu flow için onaylı kategorilerdeyse cihazı 109'a (Üretime Aktar) taşır; flow'un
        desteklemediği/onaylı olmayan bir kategoriden parça eklenmişse 106'ya (Müşteri Onayına
        Gönder) taşır. İstisna: 'To RMA' ve 'To refurbish' akışlarında müşteri onayı hiç
        aranmaz, kategoriden bağımsız her zaman doğrudan 109'a (Üretime Aktar) taşınır.
        Gerçek statü geçişi mevcut, doğrulanmış execute_batch_entry_statu_transition
        üzerinden yapılır.

        Ayrıca Müşteri Hedef Fiyat Matrisi limit kontrolü uygulanır (bkz. aşağıdaki blok):
        kategori onaylı olsa bile, eklenen parçaların toplam fiyatı müşterinin bu model +
        test sonucu için tanımladığı hedef fiyatı aşıyorsa karar zorla Müşteri Onayına
        çevrilir. Tanımlı bir hedef fiyat kuralı yoksa bu kontrol hiçbir şeyi değiştirmez."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            imei = (imei or "").strip()
            if not imei:
                return json.dumps({"success": False, "message": "IMEI boş olamaz."})

            entry = db.execute(text("""
                SELECT id, statu_code, flow, service_id, customer_no, model, screen_test, power_test
                FROM warehouse.batch_entries
                WHERE LOWER(TRIM(imei_number)) = LOWER(:imei)
                ORDER BY id DESC LIMIT 1
            """), {"imei": imei}).mappings().first()
            if not entry:
                return json.dumps({"success": False, "message": "Bu IMEI için Batch Girişi kaydı bulunamadı."})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                is_dismantle_technician = "TEC_DISMANTLE" in user_missions
                if not is_dismantle_technician:
                    return json.dumps({"success": False, "message": "Bu işlemi sadece Demontaj Teknisyeni yapabilir."})
                required = self._get_required_mission_for_ref(db, imei)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            # rr.service_record_id hem eski (IMEI ile yazılmış) hem yeni (service_id ile
            # yazılmış) kayıtlarla eşleşsin diye ikisi de aranır - bkz. get_repair_operations_by_imei.
            repair_refs = [imei]
            if entry["service_id"]:
                repair_refs.append(str(entry["service_id"]))

            repair_rows = db.execute(text("""
                SELECT rr.id, rr.part_item_code, p.item_category
                FROM warehouse.repair_records rr
                LEFT JOIN warehouse.parts p ON p.item_code = rr.part_item_code
                LEFT JOIN warehouse.repair_result_type rrt ON rrt.code = rr.repair_result_type_code
                WHERE rr.service_record_id = ANY(:refs) AND COALESCE(rrt.is_cancelled, FALSE) = FALSE
            """), {"refs": repair_refs}).mappings().all()
            if not repair_rows:
                return json.dumps({"success": False, "message": "Önce en az bir onarım eklemelisiniz."})

            # Ham flow KODA çevrilir: alanda 7644 cihaz kısa adı ("Refurbish") tutuyor,
            # kural ve kategori tablosu ise kodu ("To refurbish") kullanıyor. Ham değerle
            # karşılaştırılınca bu cihazların HEPSİ gereksiz yere müşteri onayına düşüyordu.
            flow = self._kanonik_flow(db, entry["flow"])
            # To RMA ve To refurbish akışlarında müşteri onayı hiç aranmaz, kategoriden
            # bağımsız olarak her zaman doğrudan Üretime Aktarılır.
            if flow.lower() in self.ONAY_GEREKTIRMEYEN_FLOWLAR:
                all_approved = True
            else:
                approved_categories = set()
                if flow:
                    cat_rows = db.execute(text("""
                        SELECT DISTINCT item_category FROM warehouse.service_request_item_category
                        WHERE LOWER(TRIM(service_request_type)) = LOWER(:flow) AND is_customer_approved = TRUE
                    """), {"flow": flow}).fetchall()
                    approved_categories = {c[0].strip().lower() for c in cat_rows if c[0]}

                all_approved = flow and all(
                    (r["item_category"] or "").strip().lower() in approved_categories
                    for r in repair_rows
                )

            # ── Müşteri Hedef Fiyat Matrisi limit kontrolü ──────────────────────
            # Eklenen onarımların (işçilik dahil, repair_rows zaten hepsini kapsıyor) toplam
            # parça fiyatı - _get_effective_price ile AYNI kural (önce customer_item_prices,
            # yoksa item.satis) - müşterinin bu model + test sonucu kombinasyonu için
            # tanımladığı hedef fiyatı aşarsa, kategori onaylı olsa BİLE cihaz zorla Müşteri
            # Onayına gönderilir. customer_target_prices'ta bu kombinasyon için TANIMLI bir
            # kural yoksa (batch_entries.customer_no boşsa, model warehouse.product_family'de
            # çözülemezse, ya da tam eşleşen bir satır yoksa) bu blok hiçbir şeyi DEĞİŞTİRMEZ -
            # sadece yukarıdaki kategori mantığı karar verir.
            price_limit_exceeded = False
            price_limit_info = None
            customer_code = (entry["customer_no"] or "").strip()
            model_text = (entry["model"] or "").strip()
            if customer_code and model_text:
                fam = db.execute(text("""
                    SELECT code FROM warehouse.product_family
                    WHERE LOWER(code) = LOWER(:m) OR LOWER(short_name) = LOWER(:m)
                    LIMIT 1
                """), {"m": model_text}).mappings().first()
                if fam:
                    def _map_test_result(raw):
                        v = (raw or "").strip().upper()
                        if v in ("BAŞARILI", "BASARILI"):
                            return "OK"
                        if v in ("BAŞARISIZ", "BASARISIZ"):
                            return "NOK"
                        return "BOŞ"

                    screen_result = _map_test_result(entry["screen_test"])
                    power_result = _map_test_result(entry["power_test"])

                    target_row = db.execute(text("""
                        SELECT target_price FROM warehouse.customer_target_prices
                        WHERE customer_code = :customer_code AND product_family_code = :pfc
                          AND screen_test_result = :screen AND power_test_result = :power
                    """), {
                        "customer_code": customer_code, "pfc": fam["code"],
                        "screen": screen_result, "power": power_result,
                    }).mappings().first()

                    if target_row:
                        total_price = 0.0
                        for r in repair_rows:
                            p = self._get_effective_price(db, r["part_item_code"], customer_code)
                            if p is not None:
                                total_price += p
                        target_price = float(target_row["target_price"])
                        if total_price > target_price:
                            price_limit_exceeded = True
                            price_limit_info = {"total_price": total_price, "target_price": target_price}

            all_approved = all_approved and not price_limit_exceeded
            target_statu_code = 109 if all_approved else 106

            # Cihaz zaten hedef statüdeyse (örn. Müşteri Onayına gönderilmiş 106'da bir cihaza
            # yeni onarım eklenip karar tekrar verilirse hedef yine 106 olur) idempotent kabul
            # edilir; aksi halde execute_batch_entry_statu_transition geçersiz 106->106 geçişi
            # deneyip "bu okutmaya uygun statü değil" hatası verirdi. Cihaz zaten doğru statüde,
            # eklenen yeni onarım da aynı müşteri onayı kapsamına dahil olur.
            price_limit_note = ""
            if price_limit_exceeded and price_limit_info:
                price_limit_note = (f" (Toplam parça fiyatı {price_limit_info['total_price']:.2f}, "
                                     f"hedef limit {price_limit_info['target_price']:.2f} aşıldı.)")

            if int(entry["statu_code"]) == int(target_statu_code):
                return json.dumps({
                    "success": True,
                    "new_statu_code": target_statu_code,
                    "decision": "URETIME_AKTAR" if all_approved else "MUSTERI_ONAYI",
                    "priceLimitExceeded": price_limit_exceeded,
                    "message": (("Cihaz zaten Üretim aşamasında; yeni onarım bu kapsama eklendi."
                                 if all_approved else
                                 "Cihaz zaten Müşteri Onayı kapsamında; yeni onarım da bu onaya dahil edildi.")
                                + price_limit_note)
                }, ensure_ascii=False)

            result_json = self.execute_batch_entry_statu_transition(str(entry["id"]), int(entry["statu_code"]), int(target_statu_code))
            result = json.loads(result_json)
            result["decision"] = "URETIME_AKTAR" if all_approved else "MUSTERI_ONAYI"
            result["priceLimitExceeded"] = price_limit_exceeded
            if price_limit_note and result.get("message"):
                result["message"] = result["message"] + price_limit_note
            return json.dumps(result, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    # ── ONARIM BİTİŞ TESTİ ───────────────────────────────────────────────
    # Bu görev gruplarında teknisyen onarımı "tamamladığında" kayıt DOĞRUDAN
    # 1002 (Onarım Tamamlandı) OLMAZ; önce 1006 (Onarım Testi Bekleniyor =
    # "Onarım Bitiş Testine Aktarıldı") statüsüne geçer. Ayrı bir "Onarım Bitiş
    # Testi" ekranında test edilir: başarılı → 1002, başarısız → 1001 (kayıt
    # teknisyende açık iş olarak kalır). Diğer departmanlarda akış değişmez.
    COMPLETION_TEST_DEPARTMENTS = {"CAMERA", "L3REPAIR", "DISPLAY", "CASE"}
    # 1006 kaydın bitiş testinde beklediğini, 1007 testten kaldığını gösterir
    # (bkz. warehouse.repair_result_type). Tamamlama işlemi bu statüdeki kayıtları
    # "zaten test aşamasında" kabul edip tekrar işlemez.
    COMPLETION_TEST_PENDING_CODE = 1006

    def _needs_completion_test(self, department_mission):
        """Verilen görev grubu (department_mission) bir onarım bitiş testi
        gerektiriyor mu? Kamera / L3 / Ekran / Kasa için True döner."""
        return (department_mission or "").strip().upper() in self.COMPLETION_TEST_DEPARTMENTS

    @Slot(str, str, str, result=str)
    def quick_complete_repair(self, device_ref, mission_group_code, username):
        """Hızlı Onarım Bitiş ekranı: bir cihaz okutulduğunda, verilen GÖREV GRUBUNDAKİ
        uygun onarımları tek seferde 1002 (Tamamlandı) yapar.

        KISMİ KAPATMA: kayıtlar TEK TEK değerlendirilir. Şartları sağlayanlar kapanır,
        sağlamayanlar OLDUĞU GİBİ bırakılır ve sebebi ayrı satır olarak döner. Hepsi
        ya da hiçbiri DEĞİLDİR - bir cihazda 3 kayıt varsa ve 2'si hazırsa o 2'si kapanır.

        Şartlar yeniden yazılmaz; update_repair_status ile AYNI yardımcı kullanılır
        (_repair_completion_blocker): teknisyen ataması + parça depodan çıkışı +
        müşteri onayı. Böylece bu ekran müşteri onayını atlayamaz.

        Zaten tamamlanmış (1002) ve iptal edilmiş (1003) kayıtlara dokunulmaz."""
        from models.repair_record import RepairRecord
        from sqlalchemy import text

        db = SessionLocal()
        try:
            ref = (device_ref or "").strip()
            grup = (mission_group_code or "").strip()
            if not ref:
                return json.dumps({"success": False, "message": "IMEI / Seri No / Internal ID boş olamaz."}, ensure_ascii=False)
            if not grup:
                return json.dumps({"success": False, "message": "Görev grubu belirtilmedi."}, ensure_ascii=False)

            # Cihazı bul. Onarım kayıtları service_id ya da ham IMEI ile yazılmış
            # olabiliyor (bkz. _resolve_service_record_id_for_new_repair), ikisi de aranır.
            be = db.execute(text("""
                SELECT id, imei_number, serial_number, internal_id, service_id, model
                FROM warehouse.batch_entries
                WHERE LOWER(TRIM(imei_number)) = LOWER(:t)
                   OR LOWER(TRIM(serial_number)) = LOWER(:t)
                   OR LOWER(TRIM(internal_id)) = LOWER(:t)
                ORDER BY id DESC LIMIT 1
            """), {"t": ref}).mappings().first()

            refs = [ref]
            imei = ref
            if be:
                imei = (be["imei_number"] or be["serial_number"] or ref).strip()
                if be["service_id"]:
                    refs.append(str(be["service_id"]))
                if be["imei_number"]:
                    refs.append(str(be["imei_number"]).strip())

            # Yetki: kullanıcının bu GÖREV GRUBUNA bağlı bir görevi olmalı. Yoksa bir
            # batarya teknisyeni kamera onarımlarını kapatabilirdi.
            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                yetkili = db.execute(text("""
                    SELECT 1 FROM organization.missions m
                    JOIN organization.mission_groups g ON g.id = m.mission_group_id
                    WHERE g.code = :grup AND m.code = ANY(:kodlar) LIMIT 1
                """), {"grup": grup, "kodlar": list(user_missions or [])}).first()
                if not yetkili:
                    return json.dumps({
                        "success": False,
                        "message": f"Bu işlem için '{grup}' görev grubuna bağlı bir yetkiniz yok."
                    }, ensure_ascii=False)

            # 1006 (bitiş testinde bekliyor) da hariç tutulur: bu görev grubunda
            # onarımı bitirilen kayıt teste alınmıştır, "hızlı bitiş" onu tekrar
            # işlememelidir - sonucu artık Onarım Bitiş Testi ekranı belirler.
            kayitlar = db.query(RepairRecord).filter(
                RepairRecord.service_record_id.in_(refs),
                RepairRecord.department_mission == grup,
                ~RepairRecord.repair_result_type_code.in_([1002, 1003, 1006]),
            ).all()

            if not kayitlar:
                return json.dumps({
                    "success": False,
                    "imei": imei,
                    "message": f"Bu cihazda '{grup}' için açık onarım yok."
                }, ensure_ascii=False)

            sonuclar = []
            kapanan = 0
            for rec in kayitlar:
                engel = self._repair_completion_blocker(db, rec)
                if engel:
                    sonuclar.append({
                        "repairId": str(rec.id),
                        "partItemCode": rec.part_item_code or "",
                        "completed": False,
                        "message": engel,
                    })
                    continue
                # Kamera / L3 / Ekran / Kasa: doğrudan tamamlanmaz, önce bitiş
                # testine (1006) aktarılır. Diğer departmanlar 1002 ile kapanır.
                if self._needs_completion_test(rec.department_mission):
                    rec.repair_result_type_code = self.COMPLETION_TEST_PENDING_CODE
                    kapanan += 1
                    sonuclar.append({
                        "repairId": str(rec.id),
                        "partItemCode": rec.part_item_code or "",
                        "technician": rec.assigned_technician or "",
                        "completed": True,
                        "toTest": True,
                        "message": "Onarım bitiş testine aktarıldı.",
                    })
                else:
                    rec.repair_result_type_code = 1002
                    kapanan += 1
                    sonuclar.append({
                        "repairId": str(rec.id),
                        "partItemCode": rec.part_item_code or "",
                        "technician": rec.assigned_technician or "",
                        "completed": True,
                        "toTest": False,
                        "message": "Onarım tamamlandı.",
                    })

            # Tek commit: kısmi kapatma yapiliyor ama yazma atomik olmali.
            db.commit()

            return json.dumps({
                "success": True,
                "imei": imei,
                "missionGroup": grup,
                "total": len(kayitlar),
                "completed": kapanan,
                "skipped": len(kayitlar) - kapanan,
                "results": sonuclar,
            }, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    # Qt slotu DEĞİLDİR - açık bir DB oturumu alan dahili yardımcı.
    def _repair_completion_blocker(self, db, rec):
        """Bir onarım kaydının 1002 (Tamamlandı) yapılmasını ENGELLEYEN sebebi döner;
        engel yoksa None.

        Üç kural birden aranır:
          1) Kayıt bir teknisyene ATANMIŞ olmalı,
          2) Parça stok takipliyse depodan çıkmış ('Stoktan Çıktı') olmalı - stok
             takipsiz parçada bu şart ARANMAZ (bkz. _is_part_stock_tracked, takip tipi
             parts'ta boşsa part_categories'e düşülür),
          3) 1001'den doğrudan 1002'ye geçiliyorsa ve cihazın Flow'u müşteri onayı
             gerektiriyorsa (To refurbish / To RMA hariç) önce onay adımından
             geçilmiş olmalı.

        Bu yardımcıyı hem update_repair_status hem quick_complete_repair çağırır.
        Kural iki yerde ayrı yazılırsa zamanla ayrışır ve aynı onarım bir ekrandan
        kapanıp diğerinden kapanmaz."""
        from sqlalchemy import text

        # 1) Teknisyen ataması
        if not (rec.assigned_technician or "").strip():
            return ("Bu onarım tamamlanamaz! Kayıt henüz bir teknisyene atanmamış. "
                    "Önce 'Teknisyene Ata' ile atama yapın.")

        # 2) Parça depodan çıkmış mı
        if rec.part_item_code:
            part_row = db.execute(
                text("SELECT id, name FROM warehouse.parts WHERE item_code = :c LIMIT 1"),
                {"c": rec.part_item_code.strip()},
            ).mappings().first()
            if part_row and self._is_part_stock_tracked(db, rec.part_item_code):
                # 'stoktan çıktı' warehouse.item_supply_status'taki GEÇERLİ koddur;
                # 'teslim edildi' o tabloda yok, eski kayıtlar için kabul edilir.
                is_delivered = (rec.supply_status_code or "").strip().lower() in (
                    "stoktan çıktı", "teslim edildi", "teslim", "1002", "completed")
                if not is_delivered:
                    return (f"Bu onarım tamamlanamaz! Eklenen '{part_row['name']}' "
                            f"({rec.part_item_code}) parçası henüz depodan teknisyene "
                            f"teslim edilmemiş (Good Stock'tan çıkışı yapılmamış).")

        # 3) Müşteri onayı
        if rec.repair_result_type_code == 1001:
            ref = str(rec.service_record_id or "")
            be_row = db.execute(text("""
                SELECT flow, statu_code FROM warehouse.batch_entries
                WHERE service_id::text = :ref
                   OR LOWER(TRIM(imei_number)) = LOWER(:ref)
                   OR LOWER(TRIM(serial_number)) = LOWER(:ref)
                   OR LOWER(TRIM(internal_id)) = LOWER(:ref)
                ORDER BY id DESC LIMIT 1
            """), {"ref": ref}).first()
            ham_flow = ((be_row[0] if be_row else "") or "").strip()
            device_statu = be_row[1] if be_row else None

            # batch_entries.flow bazen KODU ("To refurbish") bazen KISA ADI ("Refurbish")
            # tutuyor - canli veride 7644 kayit kisa ad, 33 kayit kod. Karsilastirma
            # kanonik KOD uzerinden yapilmali; ham metinle yapilirsa "Refurbish" akisindaki
            # cihazlar musteri onayi gerektiriyor sanilip HIC tamamlanamiyor.
            flow = ham_flow.lower()
            if ham_flow:
                srt = db.execute(text("""
                    SELECT code FROM warehouse.service_request_type
                    WHERE LOWER(TRIM(code)) = LOWER(:f) OR LOWER(TRIM(short_name)) = LOWER(:f)
                    LIMIT 1
                """), {"f": ham_flow}).first()
                if srt and srt[0]:
                    flow = srt[0].strip().lower()

            # Müşteri onayı GEREKTİRMEYEN akışlar. "Battery only" (batarya değişimi)
            # de buradadır: bu akış demontaj kararında doğrudan 109'a gider, müşteri
            # onayı (107→136) adımından HİÇ geçmez (bkz. state_machine_service
            # is_battery_only özel işleme). Kanonik kod ("Battery only ") TRIM+lower
            # ile "battery only"e indirgenir.
            NO_APPROVAL_FLOWS = {"to refurbish", "to rma", "battery only"}

            # Onay GEREKTİREN akışlarda (ör. "To repair") engel yalnızca cihaz HÂLÂ
            # onay bekleyen bir statüde PARK EDERKEN çıkmalı. Onay kararı verildiğinde
            # cihaz 109'a (Production in Progress) geçer - hangi yoldan gelirse gelsin:
            #   105→109 (doğrudan üretime), 106→109 / 136→109 (onay geldi).
            # Onay bekleyen cihaz 106/107/136'da durur, ASLA 109'da değildir. Bu yüzden
            # cihaz onay kapısını (109) geçmişse tekrar onay aranmaz - aksi halde onaydan
            # geçmiş "To repair" cihazları da hiç tamamlanamaz (statü geçmişine değil
            # yalnızca flow'a bakan eski davranışın hatası buydu).
            APPROVAL_PENDING_STATUS = {106, 107, 136}
            needs_approval = bool(flow) and flow not in NO_APPROVAL_FLOWS
            if needs_approval and device_statu in APPROVAL_PENDING_STATUS:
                return ("Bu onarım müşteri onayı alınmadan tamamlanamaz. Önce "
                        "'Müşteri Onayına Sun' adımından geçip onay alınmalı, ardından "
                        "'Onay Geldi - Tamamla' ile kapatılmalıdır.")

        return None

    def _repair_cancellation_blocker(self, db, rec):
        """Bir onarım kaydının 1003 (İptal Edildi) yapılmasını ENGELLEYEN sebebi döner;
        engel yoksa None.

        _repair_completion_blocker'ın TERSİ mantığı: tamamlama parçanın depodan çıkmış
        olmasını ŞART koşar, iptal ise tam tersine parçanın HÂLÂ depoda (henüz çıkmamış
        ya da geri alınmış) olmasını şart koşar - aksi halde fiilen depodan çıkıp
        teknisyende duran bir parça, sistemde 'iptal edildi' görünüp iz kaybeder. Parça
        stok takipsizse (bkz. _is_part_stock_tracked) bu şart aranmaz - Cihaz İade
        Prosedürü'ndeki aynı korumayla tutarlı (bkz. execute_device_return)."""
        from sqlalchemy import text
        if rec.part_item_code and self._is_part_stock_tracked(db, rec.part_item_code):
            is_delivered = (rec.supply_status_code or "").strip().lower() in (
                "stoktan çıktı", "teslim edildi", "teslim", "completed")
            if is_delivered:
                part_row = db.execute(
                    text("SELECT name FROM warehouse.parts WHERE item_code = :c LIMIT 1"),
                    {"c": rec.part_item_code.strip()},
                ).mappings().first()
                part_name = part_row["name"] if part_row else rec.part_item_code
                return (f"Bu onarım iptal edilemez! Eklenen '{part_name}' ({rec.part_item_code}) "
                        f"parçası hâlâ depoda görünmüyor (Stoktan Çıktı). Parçayı depocuya teslim "
                        f"edin — depocu 'Depo → Parça Teslim' ekranından geri alma işlemini "
                        f"tamamladıktan sonra tekrar deneyin.")
        return None

    @Slot(str, str, str, result=str)
    def update_repair_status(self, repair_id, new_status_code, username):
        """Bir alt onarım kaydının statüsünü (repair_result_type_code) günceller.
        Kodlar warehouse.repair_result_type tablosundaki gerçek anlamlarıyla kullanılır
        (1000 Teknisyene Atanacak ... 1003 Onarım İptal Edildi ... 1002 Onarım Tamamlandı vb.)."""
        from models.repair_record import RepairRecord
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rec = db.query(RepairRecord).filter(RepairRecord.id == repair_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Onarım kaydı bulunamadı."})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                required = self._get_required_mission_for_ref(db, rec.service_record_id)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            target_code = int(new_status_code)

            # Tamamlama sartlari TEK YERDE toplandi (_repair_completion_blocker).
            # Hizli Onarim Bitis ekrani da ayni yardimciyi cagirir - kural iki yerde
            # ayri yazilirsa zamanla ayrisir ve ayni onarim bir ekrandan kapanip
            # digerinden kapanmaz.
            applied_message = ""
            if target_code == 1002:
                engel = self._repair_completion_blocker(db, rec)
                if engel:
                    return json.dumps({"success": False, "message": engel}, ensure_ascii=False)
                # Kamera / L3 / Ekran / Kasa: "Onarımı Tamamla" kaydı 1002 yapmaz,
                # önce bitiş testine (1006) aktarır. Nihai 1002/1001 kararı Onarım
                # Bitiş Testi ekranından (submit_completion_test) verilir.
                if self._needs_completion_test(rec.department_mission):
                    target_code = self.COMPLETION_TEST_PENDING_CODE
                    applied_message = "Onarım bitiş testine aktarıldı."
                else:
                    applied_message = "Onarım tamamlandı."
            elif target_code == 1003:
                engel = self._repair_cancellation_blocker(db, rec)
                if engel:
                    return json.dumps({"success": False, "message": engel}, ensure_ascii=False)

            rec.repair_result_type_code = target_code
            db.commit()
            return json.dumps({"success": True, "appliedCode": target_code, "message": applied_message}, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    @Slot(str, result=str)
    def get_completion_test_pool(self, department_code):
        """Onarım Bitiş Testi ekranı: verilen departmanda (CAMERA / L3REPAIR /
        DISPLAY / CASE) BİTİŞ TESTİ BEKLEYEN (repair_result_type_code = 1006)
        kayıtları cihaz/batch detaylarıyla döndürür. Teknisyen onarımı bitirince
        kayıt bu havuza düşer; testçi başarılı/başarısız kararı verir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            dept = (department_code or "").strip().upper()
            if not dept:
                return json.dumps({"success": False, "message": "Departman kodu boş olamaz."})
            if dept not in self.COMPLETION_TEST_DEPARTMENTS:
                return json.dumps({
                    "success": False,
                    "message": f"'{dept}' bir onarım bitiş testi departmanı değil."
                }, ensure_ascii=False)

            rows = db.execute(text("""
                SELECT
                    rr.id AS repair_id,
                    rr.service_record_id,
                    rr.department_mission,
                    rr.repair_result_type_code,
                    rr.item_category,
                    rr.part_item_code,
                    pp.name AS part_name,
                    rr.item_fault_code,
                    fault.short_name AS fault_name,
                    rr.operation_type_code,
                    opt.short_name AS operation_type_name,
                    COALESCE(rr.assigned_technician, rr.supply_requested_by) AS assigned_technician,
                    COALESCE(NULLIF(TRIM(u.fullname), ''), rr.assigned_technician, rr.supply_requested_by) AS assigned_technician_name,
                    rr.notes,
                    rr.created_at,
                    rr.updated_at,
                    be.imei_number,
                    be.serial_number,
                    be.internal_id,
                    be.batch_no,
                    be.model,
                    be.gb,
                    be.color,
                    be.customer_name
                FROM warehouse.repair_records rr
                LEFT JOIN warehouse.parts pp ON pp.item_code = rr.part_item_code
                LEFT JOIN warehouse.item_fault fault ON fault.code = rr.item_fault_code
                LEFT JOIN warehouse.repair_item_operation_type opt ON opt.code = rr.operation_type_code
                LEFT JOIN warehouse.users u ON (u.username = rr.assigned_technician OR u.username = rr.supply_requested_by)
                LEFT JOIN warehouse.batch_entries be ON LOWER(TRIM(be.imei_number)) = LOWER(TRIM(rr.service_record_id))
                    OR LOWER(TRIM(be.serial_number)) = LOWER(TRIM(rr.service_record_id))
                    OR LOWER(TRIM(be.internal_id)) = LOWER(TRIM(rr.service_record_id))
                    OR (be.service_id IS NOT NULL AND strpos(rr.service_record_id, be.service_id::text) > 0)
                WHERE UPPER(TRIM(rr.department_mission)) = :dept
                  AND rr.repair_result_type_code = :pending
                ORDER BY rr.updated_at ASC, rr.created_at ASC
            """), {"dept": dept, "pending": self.COMPLETION_TEST_PENDING_CODE}).mappings().all()

            def fmt(dt):
                return dt.strftime("%Y-%m-%d %H:%M") if dt else ""

            items = []
            for r in rows:
                product_info = " ".join(filter(None, [r["model"], r["gb"], r["color"]])) or "-"
                items.append({
                    "repairId": str(r["repair_id"]),
                    "serviceRecordId": r["service_record_id"] or "",
                    "departmentMission": r["department_mission"] or "",
                    "itemCategory": r["item_category"] or "",
                    "partItemCode": r["part_item_code"] or "",
                    "partName": r["part_name"] or "",
                    "itemFaultCode": r["item_fault_code"] or "",
                    "faultName": r["fault_name"] or r["item_fault_code"] or "",
                    "operationTypeCode": r["operation_type_code"] or "",
                    "operationTypeName": r["operation_type_name"] or "",
                    "assignedTechnician": r["assigned_technician"] or "",
                    "assignedTechnicianName": r["assigned_technician_name"] or r["assigned_technician"] or "",
                    "notes": r["notes"] or "",
                    "createdAt": fmt(r["created_at"]),
                    "updatedAt": fmt(r["updated_at"]),
                    "imei": r["imei_number"] or r["service_record_id"] or "-",
                    "serialNo": r["serial_number"] or "",
                    "internalId": r["internal_id"] or "",
                    "batchNo": r["batch_no"] or "",
                    "productInfo": product_info,
                    "customerName": r["customer_name"] or "",
                })

            return json.dumps({"success": True, "items": items}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def submit_completion_test(self, repair_id, result, description, username):
        """Onarım Bitiş Testi kararı. Kayıt 1006 (bitiş testinde) iken çağrılır.
          result = 'pass' (başarılı)  -> 1002 (Onarım Tamamlandı)
          result = 'fail' (başarısız) -> 1001 (Teknisyene Atandı); kayıt atanmış
                     teknisyende AÇIK İŞ olarak kalır, açıklama ZORUNLUDUR.
        Karar her iki durumda da tarih/kullanıcı/sonuç ile notes'a eklenir."""
        from models.repair_record import RepairRecord
        from sqlalchemy import text
        import datetime

        db = SessionLocal()
        try:
            rec = db.query(RepairRecord).filter(RepairRecord.id == repair_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Onarım kaydı bulunamadı."}, ensure_ascii=False)

            raw = (result or "").strip().lower()
            is_pass = raw in ("pass", "success", "ok", "basarili", "başarılı", "1")
            is_fail = raw in ("fail", "failed", "nok", "basarisiz", "başarısız", "0")
            if not is_pass and not is_fail:
                return json.dumps({"success": False, "message": "Geçersiz test sonucu (pass/fail bekleniyor)."}, ensure_ascii=False)

            aciklama = (description or "").strip()
            # Karar yalnızca "bitiş testinde bekleyen" (1006) kayıtlar için verilir.
            if rec.repair_result_type_code != self.COMPLETION_TEST_PENDING_CODE:
                return json.dumps({
                    "success": False,
                    "message": "Bu kayıt onarım bitiş testinde değil (yalnızca 1006 statüsündeki kayıtlar test edilir)."
                }, ensure_ascii=False)

            # Başarısız testte arıza nedeni girilmesi zorunlu (karar #2).
            if is_fail and not aciklama:
                return json.dumps({
                    "success": False,
                    "message": "Test başarısız işaretlenirken açıklama (arıza nedeni) girilmesi zorunludur."
                }, ensure_ascii=False)

            # Yetki: kullanıcının kaydın GÖREV GRUBUNA bağlı bir görevi olmalı
            # (bkz. quick_complete_repair - aynı iki kod uzayı ayrımı geçerli).
            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                yetkili = db.execute(text("""
                    SELECT 1 FROM organization.missions m
                    JOIN organization.mission_groups g ON g.id = m.mission_group_id
                    WHERE g.code = :grup AND m.code = ANY(:kodlar) LIMIT 1
                """), {"grup": (rec.department_mission or "").strip(), "kodlar": list(user_missions or [])}).first()
                if not yetkili:
                    return json.dumps({
                        "success": False,
                        "message": f"Bu işlem için '{rec.department_mission}' görev grubuna bağlı bir yetkiniz yok."
                    }, ensure_ascii=False)

            if is_pass:
                # Başarılı testte de tamamlama şartları korunur (parça teslimi vb.).
                engel = self._repair_completion_blocker(db, rec)
                if engel:
                    return json.dumps({"success": False, "message": engel}, ensure_ascii=False)
                rec.repair_result_type_code = 1002
                sonuc_etiket = "BAŞARILI"
                mesaj = "Onarım bitiş testi başarılı — onarım tamamlandı."
            else:
                # Başarısız: kayıt teknisyene geri döner (atama korunur, açık iş olur).
                rec.repair_result_type_code = 1001
                sonuc_etiket = "BAŞARISIZ"
                mesaj = "Onarım bitiş testi başarısız — kayıt teknisyene geri gönderildi."

            zaman = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
            not_satiri = f"[Onarım Bitiş Testi · {sonuc_etiket} · {username or '?'} · {zaman}]"
            if aciklama:
                not_satiri += f" {aciklama}"
            rec.notes = (rec.notes + "\n" + not_satiri) if (rec.notes or "").strip() else not_satiri
            rec.updated_at = datetime.datetime.utcnow()

            db.commit()
            return json.dumps({
                "success": True,
                "result": "pass" if is_pass else "fail",
                "newStatusCode": rec.repair_result_type_code,
                "message": mesaj,
            }, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    @Slot(str, result=str)
    def get_technicians_for_mission(self, mission_code):
        """Teknisyene Atama modalinin kaynağı: verilen GÖREV GRUBUNA bağlı aktif kullanıcıları,
        üzerlerindeki açık iş sayısıyla birlikte döner.

        DİKKAT - iki ayrı kod uzayı var, doğrudan karşılaştırılamazlar:
          repair_records.department_mission -> organization.mission_groups.code  ("CASE", "BATTERY")
          warehouse.users.gorev             -> organization.missions.code        ("TEC_CASE", "QAC_CASE")
        Bağ, missions.mission_group_id üzerinden kurulur. Birebir string karşılaştırması
        yapılırsa liste HER ZAMAN boş döner (canlı veride doğrulandı).

        users.gorev virgülle ayrılmış mission kodları tutar; eşleşme TAM TOKEN üzerinden
        (unnest + TRIM) yapılır - LIKE '%kod%' olsaydı 'L1' araması 'L10'u da yakalardı.

        Bir görev grubunda hem TEC_* (teknisyen) hem QAC_* (son kontrol) görevleri olabilir.
        Hiçbiri elenmez; kullanıcının eşleşen görev adı (missionName) listede gösterilir ki
        atamayı yapan kimi seçtiğini görsün.

        mission_code boş gelirse tüm aktif kullanıcılar döner.
        Açık iş = o kullanıcıya atanmış, statüsü 1002 (Tamamlandı) / 1003 (İptal) OLMAYAN kayıtlar."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            code = (mission_code or "").strip()

            open_jobs_sql = """
                (SELECT COUNT(*) FROM warehouse.repair_records rr
                  WHERE rr.assigned_technician = u.username
                    AND COALESCE(rr.repair_result_type_code, 0) NOT IN (1002, 1003)) AS open_jobs
            """

            if not code:
                rows = db.execute(text(f"""
                    SELECT u.username,
                           COALESCE(NULLIF(TRIM(u.fullname), ''), u.username) AS fullname,
                           COALESCE(u.gorev, '') AS gorev,
                           COALESCE(u.role, '')  AS role,
                           '' AS mission_name,
                           {open_jobs_sql}
                    FROM warehouse.users u
                    WHERE COALESCE(u.account_enabled, TRUE) = TRUE
                    ORDER BY open_jobs ASC, fullname ASC
                """)).mappings().all()
            else:
                # grp_missions: verilen koda ait TÜM mission kodlari.
                # Kod bir mission_group kodu olabilir (normal durum) veya dogrudan bir
                # mission kodu olabilir (savunma amacli ikinci kosul).
                rows = db.execute(text(f"""
                    WITH grp_missions AS (
                        SELECT m.code, m.short_name
                        FROM organization.missions m
                        JOIN organization.mission_groups mg ON mg.id = m.mission_group_id
                        WHERE mg.code = :code
                        UNION
                        SELECT m2.code, m2.short_name
                        FROM organization.missions m2
                        WHERE m2.code = :code
                    )
                    SELECT u.username,
                           COALESCE(NULLIF(TRIM(u.fullname), ''), u.username) AS fullname,
                           COALESCE(u.gorev, '') AS gorev,
                           COALESCE(u.role, '')  AS role,
                           (SELECT STRING_AGG(DISTINCT gm.short_name, ', ')
                              FROM grp_missions gm
                             WHERE EXISTS (
                                   SELECT 1 FROM unnest(string_to_array(COALESCE(u.gorev, ''), ',')) g
                                    WHERE TRIM(g) = gm.code)) AS mission_name,
                           {open_jobs_sql}
                    FROM warehouse.users u
                    WHERE COALESCE(u.account_enabled, TRUE) = TRUE
                      AND EXISTS (
                          SELECT 1
                            FROM grp_missions gm
                            JOIN unnest(string_to_array(COALESCE(u.gorev, ''), ',')) g
                              ON TRIM(g) = gm.code)
                    ORDER BY open_jobs ASC, fullname ASC
                """), {"code": code}).mappings().all()

            techs = [{
                "username": r["username"],
                "fullname": r["fullname"],
                "gorev": r["gorev"],
                "role": r["role"],
                "missionName": r["mission_name"] or "",
                "openJobs": int(r["open_jobs"] or 0),
            } for r in rows]
            return json.dumps({"success": True, "technicians": techs}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def assign_technician_to_repair(self, repair_id, technician_username, username):
        """Bir ONARIMI (görev grubunu) bir teknisyene atar ve statüsünü 1001 (Teknisyene
        Atandı) yapar - ikisi TEK transaction içinde olur, yarım kalmış atama oluşmaz.

        ATAMA SEVİYESİ = ONARIM, PARÇA DEĞİL. repair_records tablosunda her satır bir
        parçayı taşır; atama parçaya göre yapılmaz. Verilen repair_id'nin ait olduğu
        onarımın (aynı service_record_id + aynı department_mission) TÜM aktif kayıtları
        aynı teknisyene yazılır. Böylece bir onarımın parçaları farklı teknisyenlere
        dağılmaz, arayüzde de teknisyen parça başına değil onarım başına gösterilir.

        Tamamlanmış (1002) ve iptal edilmiş (1003) kayıtlara DOKUNULMAZ - geçmiş
        atamaları bozmamak için.

        Sistemde henüz yetkilendirme kurulmadığı için atamayı KİMİN yaptığı kısıtlanmaz;
        teknisyen listesi zaten görev grubuna göre süzülerek geliyor (bkz.
        get_technicians_for_mission). Doğrulanan tek şey teknisyenin var ve aktif olduğudur.

        technician_username boş gönderilirse atama KALDIRILIR ve statü 1000
        (Teknisyene Atanacak) durumuna geri döner."""
        from models.repair_record import RepairRecord
        from sqlalchemy import text
        import datetime
        db = SessionLocal()
        try:
            rec = db.query(RepairRecord).filter(RepairRecord.id == repair_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Onarım kaydı bulunamadı."})

            tech = (technician_username or "").strip()
            actor = (username or "").strip() or None
            now = datetime.datetime.utcnow()

            # Onarımın kapsamı: aynı cihaz kaydı + aynı görev grubu, tamamlanmamış/iptal
            # edilmemiş tüm satırlar. Parça kodu ölçüt DEĞİLDİR.
            scope = """
                WHERE service_record_id = :sr
                  AND COALESCE(TRIM(department_mission), '') = COALESCE(TRIM(:dm), '')
                  AND COALESCE(repair_result_type_code, 0) NOT IN (1002, 1003)
            """
            scope_params = {"sr": rec.service_record_id, "dm": rec.department_mission}

            # Atamayı kaldırma
            if not tech:
                n = db.execute(text(f"""
                    UPDATE warehouse.repair_records
                    SET assigned_technician = NULL, supply_requested_by = NULL, assigned_by = :by, assigned_at = :now,
                        repair_result_type_code = 1000
                    {scope}
                """), {**scope_params, "by": actor, "now": now}).rowcount
                db.commit()
                return json.dumps({
                    "success": True, "assigned": False, "statusCode": 1000,
                    "affected": n,
                    "message": f"Atama kaldırıldı ({n} kayıt)."
                }, ensure_ascii=False)

            row = db.execute(text("""
                SELECT username,
                       COALESCE(NULLIF(TRIM(fullname), ''), username) AS fullname,
                       COALESCE(account_enabled, TRUE) AS enabled
                FROM warehouse.users WHERE username = :u
            """), {"u": tech}).mappings().first()
            if not row:
                return json.dumps({"success": False, "message": f"Kullanıcı bulunamadı: {tech}"}, ensure_ascii=False)
            if not row["enabled"]:
                return json.dumps({"success": False, "message": f"'{row['fullname']}' pasif durumda, atama yapılamaz."}, ensure_ascii=False)

            n = db.execute(text(f"""
                UPDATE warehouse.repair_records
                SET assigned_technician = :u, supply_requested_by = :u, assigned_by = :by, assigned_at = :now,
                    repair_result_type_code = 1001
                {scope}
            """), {**scope_params, "u": tech, "by": actor, "now": now}).rowcount
            db.commit()

            return json.dumps({
                "success": True,
                "assigned": True,
                "statusCode": 1001,
                "technician": tech,
                "technicianName": row["fullname"],
                "affected": n,
                "message": f"Onarım '{row['fullname']}' teknisyenine atandı ({n} kayıt)."
            }, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def update_repair_warranty(self, repair_id, warranty_code, username):
        """Bir alt onarım kaydının ücret tipini (warranty_code: IW=ücretsiz, OOW=ücretli) günceller."""
        from models.repair_record import RepairRecord
        db = SessionLocal()
        try:
            rec = db.query(RepairRecord).filter(RepairRecord.id == repair_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Onarım kaydı bulunamadı."})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                required = self._get_required_mission_for_ref(db, rec.service_record_id)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            rec.warranty_code = warranty_code
            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_item_supply_statuses(self):
        """warehouse.item_supply_status'daki tüm depo durumu kodlarını getirir. Onarım
        Parçaları ekranındaki 'Depo Durum' sütununun seçim kaynağıdır. Çalışma zamanında
        pratikte hiç değişmediğinden 5 dakika önbelleklenir."""
        def _compute():
            from sqlalchemy import text
            db = SessionLocal()
            try:
                rows = db.execute(text("""
                    SELECT code, short_name, is_success, is_cancelled
                    FROM warehouse.item_supply_status
                    ORDER BY order_number ASC NULLS LAST, short_name ASC
                """)).mappings().all()
                items = [{"code": r["code"], "short_name": r["short_name"] or r["code"], "is_success": bool(r["is_success"]), "is_cancelled": bool(r["is_cancelled"])} for r in rows]
                return json.dumps({"success": True, "supply_statuses": items}, ensure_ascii=False)
            except Exception as e:
                return json.dumps({"success": False, "message": str(e)})
            finally:
                db.close()
        return self._cached_json("item_supply_statuses", 300, _compute)

    # Depo teslimi yapabilen roller (yetkilendirme modülü kurulana kadar rol adına bakılır).
    WAREHOUSE_ROLES = ("depo", "depo sorumlusu", "depo müdürü", "depo muduru")

    def _is_warehouse_user(self, db, username):
        """Kullanıcının depo rollerinden birine sahip olup olmadığını döner.
        Admin/developer de depo işlemlerini yapabilir."""
        from sqlalchemy import text
        if not username:
            return False
        row = db.execute(text("SELECT role FROM warehouse.users WHERE username = :u"),
                         {"u": username}).mappings().first()
        if not row:
            return False
        role = (row["role"] or "").strip().lower()
        return role in self.WAREHOUSE_ROLES or role in ("admin", "developer")

    def _is_part_stock_tracked(self, db, item_code):
        """Parçanın stok takibine tabi olup olmadığını döner.
        DGD işçilik kalemleri ve stok takipsiz/servis hizmeti parçaları stok tutmaz (False döner)."""
        from sqlalchemy import text
        code = (item_code or "").strip()
        if not code:
            return False

        if code.upper().startswith("DGD"):
            return False

        dgd_row = db.execute(text("SELECT 1 FROM warehouse.flow_dgd_mapping WHERE LOWER(TRIM(dgd_item_code)) = LOWER(TRIM(:c)) LIMIT 1"), {"c": code}).first()
        if dgd_row:
            return False

        row = db.execute(text("""
            SELECT COALESCE(NULLIF(TRIM(p.stock_tracking_type), ''),
                            NULLIF(TRIM(pc.stock_tracking_type), '')) AS tt,
                   p.item_category,
                   p.part_category
            FROM warehouse.parts p
            LEFT JOIN warehouse.part_categories pc ON pc.id = p.part_category_id
            WHERE p.item_code = :c
            LIMIT 1
        """), {"c": code}).mappings().first()

        if row:
            cat = (row["item_category"] or "").strip().lower()
            pcat = (row["part_category"] or "").strip().lower()
            if cat in ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik') or pcat in ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik'):
                return False
            tt = (row["tt"] or "").strip().lower()
            if any(x in tt for x in ("takipsiz", "stoksuz", "hizmet", "dgd")):
                return False

        if not row or not row["tt"]:
            return True
        return "takipsiz" not in row["tt"].strip().lower()

    @Slot(str, str, result=str)
    def deliver_repair_part(self, repair_id, username):
        """DEPOCUNUN TESLİM İŞLEMİ - Servis Onarımları ekranındaki 'Teslim Et' butonu.

        Ön koşullar BACKEND'de doğrulanır (arayüzde butonu gizlemek güvenlik değildir):
          - kayıt statüsü 1001 (Teknisyene Atandı) olmalı
          - kayıtta eklenmiş bir parça (part_item_code) olmalı  <- 'eklenme parçası gerekli'
          - daha önce teslim edilmemiş olmalı
          - kullanıcı depo rollerinden birine (veya admin) sahip olmalı

        Sonuç: supply_status_code = 'Stoktan Çıktı'. Stok hareketi (Good Stock -1 /
        Repair Stock +1 + StockMovement) YALNIZCA parça 'Stok Takipli' ise oluşur;
        takipsiz parçada sadece durum değişir.

        Depocuda üretim mission'ı bulunmadığı için buradaki yetki kontrolü mission değil
        ROL tabanlıdır - update_repair_supply_status'un mission kontrolü depocuyu bloke ederdi."""
        from models.repair_record import RepairRecord
        from models.stock import Stock
        from sqlalchemy import text
        import datetime
        db = SessionLocal()
        try:
            if not self._is_warehouse_user(db, username):
                return json.dumps({"success": False, "message": "Parça teslimi için depo yetkisi gerekiyor."}, ensure_ascii=False)

            rec = db.query(RepairRecord).filter(RepairRecord.id == repair_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Onarım kaydı bulunamadı."}, ensure_ascii=False)

            if int(rec.repair_result_type_code or 0) != 1001:
                return json.dumps({"success": False, "message": "Yalnızca 'Teknisyene Atandı' (1001) statüsündeki kayıtlar teslim edilebilir."}, ensure_ascii=False)

            if not (rec.part_item_code or "").strip():
                return json.dumps({"success": False, "message": "Bu kayıtta eklenmiş bir parça yok. Parçayı teknisyen eklemelidir."}, ensure_ascii=False)

            if rec.supply_status_code in ("Stoktan Çıktı", "TESLIMEDILDI", "Teslim Edildi"):
                return json.dumps({"success": False, "message": "Bu parça zaten teslim edilmiş."}, ensure_ascii=False)

            code = rec.part_item_code.strip()
            tracked = self._is_part_stock_tracked(db, code)

            rec.supply_status_code = "Stoktan Çıktı"
            rec.supply_requested_by = (username or "").strip() or None
            rec.supply_requested_at = datetime.datetime.utcnow()

            moved = False
            if tracked:
                good_loc_id = _get_system_location_id(db, "good_stock")
                repair_loc_id = _get_system_location_id(db, "repair_stock")
                part_row = db.execute(text("SELECT id, name FROM warehouse.parts WHERE item_code = :c LIMIT 1"),
                                      {"c": code}).mappings().first()
                part_id = part_row["id"] if part_row else None

                if not part_id:
                    return json.dumps({"success": False, "message": f"Parça bulunamadı: {code}"}, ensure_ascii=False)

                g_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == good_loc_id).first()
                if not g_stock or (g_stock.quantity or 0) < 1:
                    return json.dumps({"success": False, "message": f"'{part_row['name']}' ({code}) Good Stock'ta tükenmiş, teslim edilemez."}, ensure_ascii=False)

                g_stock.quantity = max(0, (g_stock.quantity or 0) - 1)

                r_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == repair_loc_id).first()
                if r_stock:
                    r_stock.quantity = (r_stock.quantity or 0) + 1
                else:
                    db.add(Stock(part_id=part_id, location_id=repair_loc_id, quantity=1))

                from models.stock_movement import StockMovement
                db.add(StockMovement(
                    part_id=part_id,
                    part_name_snapshot=part_row["name"],
                    source_location_id=good_loc_id,
                    target_location_id=repair_loc_id,
                    quantity=1,
                    type="İç Transfer",
                    movement_kind="Transfer",
                    created_by=(username or "system"),
                    technician=rec.assigned_technician or None,
                    description=f"Depo Teslim - Parça: {code} - RepairRecord ID: {repair_id}"
                ))
                moved = True

            db.commit()
            clear_api_cache()
            return json.dumps({
                "success": True,
                "supplyStatusCode": "Stoktan Çıktı",
                "stockTracked": tracked,
                "stockMoved": moved,
                "message": ("Parça teslim edildi, stok Repair Stock'a aktarıldı."
                            if moved else
                            "Parça teslim edildi. (Stok takipsiz parça - stok hareketi oluşturulmadı.)")
            }, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def update_repair_supply_status(self, repair_id, supply_status_code, username):
        """Bir alt onarım kaydının Depo Durum'unu (supply_status_code, warehouse.item_supply_status.code)
        günceller - onarım için gereken parçanın depo/tedarik sürecindeki aşamasını izler.
        Kim ve ne zaman değiştirdiği supply_requested_by/supply_requested_at'e yazılır -
        Depo > Parça Teslim ekranının (get_repair_supply_requests) kaynağıdır."""
        from models.repair_record import RepairRecord
        from sqlalchemy import text
        import datetime
        db = SessionLocal()
        try:
            rec = db.query(RepairRecord).filter(RepairRecord.id == repair_id).first()
            if not rec:
                return json.dumps({"success": False, "message": "Onarım kaydı bulunamadı."})

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                required = self._get_required_mission_for_ref(db, rec.service_record_id)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            supply_status_code = (supply_status_code or "").strip() or None
            if supply_status_code:
                exists = db.execute(text("SELECT 1 FROM warehouse.item_supply_status WHERE code = :c"), {"c": supply_status_code}).first()
                if not exists:
                    return json.dumps({"success": False, "message": f"Geçersiz depo durum kodu: {supply_status_code}"})

            # Aynı onarım kaydı için parça talebi (Sipariş Edildi) yalnızca hiç Depo Durum
            # atanmamışken yapılabilir - Üretim Teknisyeni'nin "Talep Et" butonuna arka arkaya
            # basıp aynı parça için yeni talepler açması (ve dolayısıyla Parça Teslim'in Good
            # Stock'tan tekrar tekrar düşmesi) engellenir. Aynı parça için yeniden talep,
            # sadece Onarım Ekle'den açılan (supply_status_code'u boş başlayan) YENİ bir
            # onarım kaydıyla mümkündür.
            if supply_status_code == "Sipariş Edildi" and rec.supply_status_code:
                return json.dumps({"success": False, "message": "Bu parça için zaten bir depo talebi var. Aynı parça için yeniden talep, ancak Onarım Ekle'den yeni bir kayıt açılırsa yapılabilir."})

            previous_status = rec.supply_status_code
            rec.supply_status_code = supply_status_code
            rec.supply_requested_by = (username or "").strip() or None
            rec.supply_requested_at = datetime.datetime.utcnow()

            # Eger durum "Stoktan Çıktı" / "TESLIMEDILDI" yapıldıysa ve daha önce çıkarılmamışsa Good Stock'tan Repair Stock'a aktar.
            # STOK TAKİBİ KURALI: stok hareketi YALNIZCA "Stok Takipli" parçalarda oluşur.
            # Takipsiz parçada depo durumu yine değişir ama stok hiç düşmez (bkz. _is_part_stock_tracked).
            if supply_status_code in ["Stoktan Çıktı", "TESLIMEDILDI"] and previous_status not in ["Stoktan Çıktı", "TESLIMEDILDI"]:
                if rec.part_item_code and self._is_part_stock_tracked(db, rec.part_item_code):
                    # 1. Good Stock miktarını 1 düş
                    db.execute(text("""
                        UPDATE warehouse.stock s
                        SET quantity = GREATEST(0, s.quantity - 1)
                        FROM warehouse.parts p, warehouse.locations l
                        WHERE s.part_id = p.id
                          AND s.location_id = l.id
                          AND p.item_code = :code
                          AND l.kind = 'good_stock'
                    """), {"code": rec.part_item_code})

                    # 2. Repair Stock miktarını 1 artır (yoksa satır oluştur)
                    good_loc_id = _get_system_location_id(db, "good_stock")
                    repair_loc_id = _get_system_location_id(db, "repair_stock")
                    part_row = db.execute(text("SELECT id FROM warehouse.parts WHERE item_code = :c"), {"c": rec.part_item_code}).first()
                    part_id = part_row[0] if part_row else None

                    if part_id and repair_loc_id:
                        from models.stock import Stock
                        r_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == repair_loc_id).first()
                        if r_stock:
                            r_stock.quantity += 1
                        else:
                            db.add(Stock(part_id=part_id, location_id=repair_loc_id, quantity=1))

                    # 3. Stok Hareketi (audit log) kaydet
                    try:
                        if part_id and good_loc_id and repair_loc_id:
                            from models.stock_movement import StockMovement
                            # NOT: StockMovement'ta 'part_code' diye bir kolon yok (bkz.
                            # models/stock_movement.py). Onu vermek TypeError firlatiyor,
                            # asagidaki except bunu yutuyordu ve audit kaydi hic olusmuyordu.
                            # Parca kodu artik part_name_snapshot + description'a yaziliyor.
                            mov = StockMovement(
                                part_id=part_id,
                                part_name_snapshot=rec.part_item_code,
                                source_location_id=good_loc_id,
                                target_location_id=repair_loc_id,
                                quantity=1,
                                type="İç Transfer",
                                movement_kind="Transfer",
                                created_by=username or "system",
                                description=f"Good Stock'tan Repair Stock'a Parça Teslimi - Parça: {rec.part_item_code} - RepairRecord ID: {repair_id}"
                            )
                            db.add(mov)
                    except Exception as mov_err:
                        logging.warning(f"Stock movement log eklenirken hata: {mov_err}")

            db.commit()
            return json.dumps({"success": True})
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(result=str)
    def get_repair_supply_requests(self):
        """Depo > Parça Teslim ekranının kaynağı: Servis Onarımları / Onarım Parçaları
        ekranında Depo Durum'u (supply_status_code) set edilmiş TÜM alt onarım kayıtlarını
        (repair_records) - hangi teknisyenin ne zaman talep ettiği, hangi parça/kategori,
        cihaz bilgisi ve o parçanın Good Stock'taki GÜNCEL miktarıyla birlikte getirir.
        Parça Teslim, bu global listeyi sorgulanan IMEI'ye göre kendi tarafında filtreler.
        NOT: bu, work_order_parts tabanlı eski 'Tedarik İstekleri' (get_supply_requests)
        Slot'undan tamamen ayrı bir akıştır - repair_records.supply_status_code'a dayanır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rows = db.execute(text("""
                SELECT
                    rr.id, rr.department_mission, rr.notes, rr.part_item_code, rr.item_fault_code,
                    rr.supply_status_code, rr.supply_requested_by, rr.supply_requested_at,
                    rr.repair_result_type_code, rr.created_at, rr.item_category,
                    mg.short_name AS mission_group_name,
                    it.short_name AS part_name,
                    fault.short_name AS fault_name,
                    sup.short_name AS supply_status_name, sup.is_success AS supply_is_success, sup.is_cancelled AS supply_is_cancelled,
                    rrt.is_cancelled AS repair_is_cancelled,
                    be.imei_number, be.model AS device_model, be.batch_no, be.customer_name,
                    p.id AS part_id,
                    (
                        SELECT COALESCE(SUM(s.quantity), 0)
                        FROM warehouse.stock s
                        JOIN warehouse.locations l ON l.id = s.location_id
                        WHERE s.part_id = p.id AND l.kind = 'good_stock'
                    ) AS stock_qty
                FROM warehouse.repair_records rr
                LEFT JOIN organization.mission_groups mg ON mg.code = rr.department_mission
                LEFT JOIN warehouse.item it ON it.code = rr.part_item_code
                LEFT JOIN warehouse.parts p ON p.item_code = rr.part_item_code
                LEFT JOIN warehouse.item_fault fault ON fault.code = rr.item_fault_code
                LEFT JOIN warehouse.item_supply_status sup ON sup.code = rr.supply_status_code
                LEFT JOIN warehouse.repair_result_type rrt ON rrt.code = rr.repair_result_type_code
                LEFT JOIN LATERAL (
                    SELECT imei_number, model, batch_no, customer_name
                    FROM warehouse.batch_entries b
                    WHERE b.service_id::text = rr.service_record_id
                       OR LOWER(TRIM(b.imei_number)) = LOWER(TRIM(rr.service_record_id))
                    ORDER BY b.id DESC LIMIT 1
                ) be ON true
                WHERE rr.supply_status_code IS NOT NULL
                ORDER BY rr.supply_requested_at DESC NULLS LAST, rr.created_at DESC
                LIMIT 500
            """)).mappings().all()

            requests = [{
                "id": str(r["id"]),
                "missionGroupCode": r["department_mission"] or "",
                "missionGroup": r["mission_group_name"] or r["department_mission"] or "-",
                "partId": r["part_id"] or "",
                "partItemCode": r["part_item_code"] or "",
                "partName": r["part_name"] or "",
                "itemCategory": r["item_category"] or "",
                "faultName": r["fault_name"] or r["item_fault_code"] or "",
                "supplyStatusCode": r["supply_status_code"] or "",
                "supplyStatusName": r["supply_status_name"] or r["supply_status_code"] or "",
                "supplyIsSuccess": bool(r["supply_is_success"]),
                "supplyIsCancelled": bool(r["supply_is_cancelled"]),
                "requestedBy": r["supply_requested_by"] or "-",
                "requestedAt": r["supply_requested_at"].strftime("%d.%m.%Y %H:%M") if r["supply_requested_at"] else "-",
                "isCancelled": bool(r["repair_is_cancelled"]),
                "imei": r["imei_number"] or "",
                "deviceModel": r["device_model"] or "",
                "batchNo": r["batch_no"] or "",
                "customerName": r["customer_name"] or "",
                "stockQty": int(r["stock_qty"] or 0),
                "notes": r["notes"] or "",
            } for r in rows]
            return json.dumps({"success": True, "requests": requests}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def execute_device_return(self, device_ref, return_reason, username):
        """CİHAZ İADE PROSEDÜRÜ - cihazı 124 (Son Teste Teslim Edilecek) statüsüne alır.

        device_ref: add_repair_record ile AYNI desen - bağlı bir SERVICE iş emri varsa
        work_order_id'dir, yoksa cihazın IMEI'sidir (bkz. _get_required_mission_for_ref,
        _resolve_batch_entry_by_ref). Böylece SERVICE iş emri hiç oluşmamış (üretim
        verisinde sık görülen) cihazlar için de çalışır - eski sürüm SADECE work_order_id
        kabul ediyordu ve böyle cihazlarda hiç kullanılamıyordu.

        SERT ENGELLEME (tek koşul, ödün verilmez): cihaza bağlı stok takipli bir parça
        hâlâ 'Stoktan Çıktı' durumundaysa işlem TAMAMEN reddedilir - burada GOOD/DOA
        yönlendirmesi SORULMAZ (eski sürümün aksine). Parça önce fiziksel olarak
        depocuya teslim edilmeli, depocu 'Depo → Parça Teslim' ekranındaki 'Parçayı
        Geri Alma' ile (return_delivered_part) işlemi tamamlamalı, ancak ondan sonra
        bu prosedür tekrar denenebilir.

        Engel yoksa TEK transaction'da: (1) cihaza bağlı tüm AKTİF (1002/1003 dışı)
        onarım kayıtları 1003'e (İptal Edildi) çekilir, (2) bağlı bir SERVICE iş emri
        varsa onun statüsü 124'e alınır ve iade nedeni work_orders.return_reason'a
        yazılır; iş emri yoksa (üretim verisinin çoğunluğu) doğrudan
        batch_entries.statu_code 124'e çekilir ve iade nedeni customer_diagnosis'e
        not düşülür (bu tabloda ayrı bir 'iade nedeni' kolonu yok)."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            device_ref = (device_ref or "").strip()
            if not device_ref:
                return json.dumps({"success": False, "message": "Cihaz bulunamadı."})
            if not return_reason or not return_reason.strip():
                return json.dumps({"success": False, "message": "İade nedeni girilmelidir."})
            return_reason = return_reason.strip()

            user_missions, is_admin = self._get_user_missions(db, username)
            if not is_admin:
                required = self._get_required_mission_for_ref(db, device_ref)
                if required and required not in user_missions:
                    return json.dumps({"success": False, "message": f"Bu işlem için '{required}' yetkisi gerekiyor."})

            # device_ref gerçekten var olan bir work_order_id mi? (add_repair_record'daki
            # _resolve_service_record_id_for_new_repair ile aynı kontrol - IMEI'ler de tamamen
            # sayısal olduğundan sadece int() dönüşümü yeterli değildir, varlığı doğrulanır.)
            wo_id = None
            try:
                wo_candidate = int(device_ref)
            except (TypeError, ValueError):
                wo_candidate = None
            if wo_candidate is not None:
                if db.execute(text("SELECT id FROM warehouse.work_orders WHERE id = :id"), {"id": wo_candidate}).first():
                    wo_id = wo_candidate

            # Onarım kayıtlarına ulaşmak için olası TÜM referansları topla (apply_dgd_return
            # ile aynı desen): device_ref'in kendisi + varsa batch_entries.service_id.
            refs = [device_ref]
            entry = self._resolve_batch_entry_by_ref(db, device_ref)
            if entry and entry["service_id"]:
                refs.append(str(entry["service_id"]))
            if wo_id is not None and str(wo_id) not in refs:
                refs.append(str(wo_id))

            # 1. SERT ENGEL: stok takipli, hâlâ 'Stoktan Çıktı' olan parça var mı?
            issued_rows = db.execute(text("""
                SELECT rr.part_item_code, p.name AS part_name
                FROM warehouse.repair_records rr
                LEFT JOIN warehouse.parts p ON p.item_code = rr.part_item_code
                WHERE rr.service_record_id = ANY(:refs)
                  AND rr.repair_result_type_code NOT IN (1002, 1003)
                  AND rr.supply_status_code = 'Stoktan Çıktı'
            """), {"refs": refs}).mappings().all()

            blocking_names = []
            for row in issued_rows:
                if row["part_item_code"] and self._is_part_stock_tracked(db, row["part_item_code"]):
                    blocking_names.append(row["part_name"] or row["part_item_code"])

            if blocking_names:
                names = ", ".join(f"'{n}'" for n in dict.fromkeys(blocking_names))
                return json.dumps({
                    "success": False,
                    "message": f"{names} parçası/parçaları hâlâ depoda görünmüyor (Stoktan Çıktı). "
                               f"Parçaları depocuya teslim edin — depocu 'Depo → Parça Teslim' "
                               f"ekranından geri alma işlemini tamamladıktan sonra tekrar deneyin."
                }, ensure_ascii=False)

            # 2. Engel yok: cihaza bağlı tüm AKTİF onarımları iptal et.
            db.execute(text("""
                UPDATE warehouse.repair_records
                SET repair_result_type_code = 1003, updated_at = now()
                WHERE service_record_id = ANY(:refs) AND repair_result_type_code NOT IN (1002, 1003)
            """), {"refs": refs})

            # 3. Statüyü 124'e al (service_statu_map grafiğinden bağımsız, doğrudan).
            if wo_id is not None:
                db.execute(text("""
                    UPDATE warehouse.work_orders SET status = '124', return_reason = :reason WHERE id = :id
                """), {"reason": return_reason, "id": wo_id})
            elif entry:
                db.execute(text("""
                    UPDATE warehouse.batch_entries
                    SET statu_code = 124, updated_at = now(),
                        customer_diagnosis = COALESCE(customer_diagnosis || E'\\n', '') || :note
                    WHERE id = :id
                """), {"id": entry["id"], "note": f"[Cihaz İade Prosedürü] {return_reason}"})
            else:
                db.rollback()
                return json.dumps({"success": False, "message": "Cihaz kaydı (batch entry) bulunamadı."})

            db.commit()
            return json.dumps({
                "success": True,
                "new_statu_code": 124,
                "message": "Cihaz iade alındı, tüm onarımlar iptal edildi, 124 (Son Teste Teslim Edilecek) statüsüne yönlendirildi."
            })
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)})
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def get_deliverable_parts_for_device(self, brand, model, color="", imei_or_serial=""):
        """Parça Teslim ekranı için cihaza eklenmiş TÜM fiziki parçaları (teslim bekleyen ve teslim edilmiş) tek listede getirir. DGD işçilik kalemleri elenir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            brand_clean = (brand or "").strip()
            model_clean = (model or "").strip()
            color_clean = (color or "").strip()
            imei_clean = (imei_or_serial or "").strip()

            dgd_rows = db.execute(text("SELECT dgd_item_code FROM warehouse.flow_dgd_mapping WHERE dgd_item_code IS NOT NULL")).fetchall()
            dgd_mapped_codes = {r[0].strip().lower() for r in dgd_rows if r[0]}

            if not imei_clean:
                return json.dumps({"success": True, "parts": []}, ensure_ascii=False)

            be_row = db.execute(text("""
                SELECT service_id FROM warehouse.batch_entries
                WHERE LOWER(TRIM(imei_number)) = LOWER(:t) OR LOWER(TRIM(serial_number)) = LOWER(:t) OR LOWER(TRIM(internal_id)) = LOWER(:t)
                ORDER BY id DESC LIMIT 1
            """), {"t": imei_clean}).mappings().first()

            refs = [imei_clean]
            if be_row and be_row["service_id"]:
                refs.append(str(be_row["service_id"]))

            sr_row = db.execute(text("""
                SELECT id FROM warehouse.service_records
                WHERE LOWER(TRIM(imei_number)) = LOWER(:t) OR LOWER(TRIM(imei_serial)) = LOWER(:t)
                ORDER BY id DESC LIMIT 1
            """), {"t": imei_clean}).mappings().first()
            if sr_row and sr_row["id"]:
                refs.append(str(sr_row["id"]))

            # Cihazın eklenmiş tüm parçalarını çek
            repair_part_rows = db.execute(text("""
                SELECT 
                    rr.id AS repair_record_id,
                    rr.part_item_code,
                    rr.supply_status_code,
                    rr.supply_requested_by,
                    TO_CHAR(rr.supply_requested_at, 'YYYY-MM-DD HH24:MI') AS supply_requested_at
                FROM warehouse.repair_records rr
                LEFT JOIN warehouse.repair_result_type rrt ON rrt.code = rr.repair_result_type_code
                LEFT JOIN warehouse.parts p ON p.item_code = rr.part_item_code
                WHERE rr.service_record_id = ANY(:refs)
                  AND (rrt.is_cancelled IS NOT TRUE)
                  AND rr.part_item_code IS NOT NULL
                  AND TRIM(rr.part_item_code) <> ''
                  AND rr.repair_result_type_code = 1001
                  AND rr.assigned_technician IS NOT NULL
                  AND TRIM(rr.assigned_technician) <> ''
                  AND UPPER(TRIM(rr.part_item_code)) NOT LIKE 'DGD%'
                  AND LOWER(TRIM(COALESCE(p.item_category, ''))) NOT IN ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik')
                  AND LOWER(TRIM(COALESCE(p.part_category, ''))) NOT IN ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik')
                ORDER BY rr.created_at DESC
            """), {"refs": refs}).mappings().all()

            if not repair_part_rows:
                return json.dumps({"success": True, "parts": []}, ensure_ascii=False)

            part_codes = list({r["part_item_code"].strip() for r in repair_part_rows if r["part_item_code"]})

            parts_info_map = {}
            if part_codes:
                sql = text("""
                    SELECT
                        p.id,
                        p.item_code,
                        p.name,
                        p.brand,
                        p.model,
                        p.color,
                        p.item_category,
                        COALESCE(p.part_category, p.item_category) AS part_category,
                        p.stock_tracking_type,
                        mg.code AS repair_team_code,
                        mg.short_name AS repair_team_name,
                        (
                            SELECT COALESCE(SUM(s.quantity), 0)
                            FROM warehouse.stock s
                            JOIN warehouse.locations l ON l.id = s.location_id
                            WHERE s.part_id = p.id AND l.kind = 'good_stock'
                        ) AS good_stock_qty
                    FROM warehouse.parts p
                    LEFT JOIN LATERAL (
                        SELECT
                            CASE WHEN UPPER(LEFT(icm.mission, 4)) = 'TEC_' THEN SUBSTRING(icm.mission FROM 5) ELSE icm.mission END AS bare_code
                        FROM warehouse.item_category_mission icm
                        WHERE LOWER(TRIM(icm.item_category)) = LOWER(TRIM(p.item_category)) AND icm.enabled = TRUE
                        ORDER BY (CASE WHEN UPPER(icm.mission) IN ('TEC_L1REPAIR', 'TEC_L2REPAIR', 'TEC_L3REPAIR') THEN 1 ELSE 0 END) ASC
                        LIMIT 1
                    ) team_map ON true
                    LEFT JOIN organization.mission_groups mg ON mg.code = team_map.bare_code
                    WHERE p.item_code = ANY(:codes)
                """)
                p_rows = db.execute(sql, {"codes": part_codes}).mappings().all()
                for pr in p_rows:
                    code_key = (pr["item_code"] or "").strip().lower()
                    parts_info_map[code_key] = pr

            parts = []
            for rr in repair_part_rows:
                code_clean = (rr["part_item_code"] or "").strip()
                code_key = code_clean.lower()

                if code_key in dgd_mapped_codes or code_clean.upper().startswith("DGD"):
                    continue

                pr = parts_info_map.get(code_key) or {}
                item_cat = (pr.get("item_category") or "").strip().lower()
                part_cat = (pr.get("part_category") or "").strip().lower()

                if item_cat in ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik') or part_cat in ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik'):
                    continue

                qty = int(pr.get("good_stock_qty") or 0)
                tracking_type = (pr.get("stock_tracking_type") or "Stok Takipli").strip()

                status_clean = (rr["supply_status_code"] or "").strip().lower()
                is_delivered = status_clean in ('stoktan çıktı', 'teslim edildi', 'teslimedildi')

                parts.append({
                    "repairRecordId": str(rr["repair_record_id"]),
                    "id": str(pr.get("id")) if pr.get("id") else str(rr["repair_record_id"]),
                    "itemCode": code_clean,
                    "partName": pr.get("name") or code_clean,
                    "brand": pr.get("brand") or "",
                    "model": pr.get("model") or "",
                    "color": pr.get("color") or "",
                    "itemCategory": pr.get("item_category") or "",
                    "partCategory": pr.get("part_category") or pr.get("item_category") or "",
                    "repairTeamCode": pr.get("repair_team_code") or "",
                    "repairTeamName": pr.get("repair_team_name") or "Genel",
                    "goodStockQty": qty,
                    "stockTrackingType": tracking_type,
                    "isDelivered": is_delivered,
                    "deliveredBy": rr["supply_requested_by"] or "",
                    "deliveredAt": rr["supply_requested_at"] or "",
                    "isStoksuz": False,
                    "isAvailable": not is_delivered and (qty > 0)
                })

            return json.dumps({"success": True, "parts": parts}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    @Slot(str, str, str, result=str)
    def deliver_part_to_device(self, imei_or_serial, item_code, username=""):
        """Parça Teslim ekranında seçilen parçayı Good Stock'tan düşüp Repair Stock'a aktarır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            code = (item_code or "").strip()
            imei = (imei_or_serial or "").strip()
            user = (username or "").strip() or "Sistem"

            if not code:
                return json.dumps({"success": False, "message": "Parça kodu seçilmelidir."})

            part_row = db.execute(text("SELECT id, name FROM warehouse.parts WHERE item_code = :c LIMIT 1"), {"c": code}).mappings().first()
            if not part_row:
                return json.dumps({"success": False, "message": "Parça bulunamadı."})
            part_id = part_row["id"]

            # 1) Teknisyen Ataması Kontrolü (repair_result_type_code = 1001 olmalıdır)
            be_row = db.execute(text("""
                SELECT service_id FROM warehouse.batch_entries
                WHERE LOWER(TRIM(imei_number)) = LOWER(:t) OR LOWER(TRIM(serial_number)) = LOWER(:t) OR LOWER(TRIM(internal_id)) = LOWER(:t)
                ORDER BY id DESC LIMIT 1
            """), {"t": imei}).mappings().first()

            refs = [imei]
            if be_row and be_row["service_id"]:
                refs.append(str(be_row["service_id"]))

            sr_row = db.execute(text("""
                SELECT id FROM warehouse.service_records
                WHERE LOWER(TRIM(imei_number)) = LOWER(:t) OR LOWER(TRIM(imei_serial)) = LOWER(:t)
                ORDER BY id DESC LIMIT 1
            """), {"t": imei}).mappings().first()
            if sr_row and sr_row["id"]:
                refs.append(str(sr_row["id"]))

            pending = db.execute(text("""
                SELECT id FROM warehouse.repair_records
                WHERE service_record_id = ANY(:refs)
                  AND LOWER(TRIM(part_item_code)) = LOWER(TRIM(:code))
                  AND repair_result_type_code = 1001
                  AND assigned_technician IS NOT NULL AND TRIM(assigned_technician) <> ''
                  AND LOWER(TRIM(COALESCE(supply_status_code, '')))
                      NOT IN ('stoktan çıktı', 'teslim edildi', 'teslimedildi')
                ORDER BY created_at
                LIMIT 1
            """), {"refs": refs, "code": code}).mappings().first()

            if not pending:
                zaten = db.execute(text("""
                    SELECT 1 FROM warehouse.repair_records
                    WHERE service_record_id = ANY(:refs)
                      AND LOWER(TRIM(part_item_code)) = LOWER(TRIM(:code))
                      AND LOWER(TRIM(COALESCE(supply_status_code, '')))
                          IN ('stoktan çıktı', 'teslim edildi', 'teslimedildi')
                    LIMIT 1
                """), {"refs": refs, "code": code}).first()
                if zaten:
                    return json.dumps({
                        "success": False,
                        "message": "Bu parça zaten teslim edilmiş. Yeniden teslim için Servis Onarımları'ndan yeni bir onarım kaydı (parça) eklenmelidir."
                    }, ensure_ascii=False)
                return json.dumps({
                    "success": False,
                    "message": "Parça teslimatı yapılamaz! Bu parça/onarım henüz teknisyene atanmamış (Statü '1001 - Teknisyene Atandı' ve atanmış bir teknisyen olmalıdır)."
                }, ensure_ascii=False)

            pending_id = pending["id"]

            if not self._is_part_stock_tracked(db, code):
                db.execute(text("""
                    UPDATE warehouse.repair_records
                    SET supply_status_code = 'Stoktan Çıktı', supply_requested_by = :user, supply_requested_at = NOW()
                    WHERE id = :rid
                """), {"rid": pending_id, "user": user})
                db.commit()
                return json.dumps({
                    "success": True,
                    "stockTracked": False,
                    "stockMoved": False,
                    "message": f"'{part_row['name']}' ({code}) teslim edildi. (Stok takipsiz parça - stok hareketi oluşturulmadı.)"
                }, ensure_ascii=False)

            good_loc_id = _get_system_location_id(db, "good_stock")
            repair_loc_id = _get_system_location_id(db, "repair_stock")

            stock_row = db.execute(text("""
                SELECT id, quantity FROM warehouse.stock
                WHERE part_id = :pid AND location_id = :loc_id
                LIMIT 1
            """), {"pid": part_id, "loc_id": good_loc_id}).mappings().first()

            current_qty = stock_row["quantity"] if stock_row else 0
            if current_qty < 1:
                return json.dumps({"success": False, "message": f"'{part_row['name']}' ({code}) Good Stock'ta tükenmiş!"})

            db.execute(text("""
                UPDATE warehouse.stock SET quantity = GREATEST(0, quantity - 1) 
                WHERE id = :sid
            """), {"sid": stock_row["id"]})

            from models.stock import Stock
            r_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == repair_loc_id).first()
            if r_stock:
                r_stock.quantity += 1
            else:
                db.add(Stock(part_id=part_id, location_id=repair_loc_id, quantity=1))

            try:
                from models.stock_movement import StockMovement
                mov = StockMovement(
                    part_id=part_id,
                    part_name_snapshot=part_row["name"],
                    source_location_id=good_loc_id,
                    target_location_id=repair_loc_id,
                    quantity=1,
                    type="İç Transfer",
                    movement_kind="PARCA_TESLIM",
                    description=f"Parça Teslim - Parça: {code} - IMEI: {imei}",
                    created_by=user,
                    technician=user
                )
                db.add(mov)

                refs = [imei]
                be_row = db.execute(text("""
                    SELECT service_id FROM warehouse.batch_entries
                    WHERE LOWER(TRIM(imei_number)) = LOWER(:t) OR LOWER(TRIM(serial_number)) = LOWER(:t) OR LOWER(TRIM(internal_id)) = LOWER(:t)
                    ORDER BY id DESC LIMIT 1
                """), {"t": imei}).mappings().first()
                if be_row and be_row["service_id"]:
                    refs.append(str(be_row["service_id"]))

                sr_row = db.execute(text("""
                    SELECT id FROM warehouse.service_records
                    WHERE LOWER(TRIM(imei_number)) = LOWER(:t) OR LOWER(TRIM(imei_serial)) = LOWER(:t)
                    ORDER BY id DESC LIMIT 1
                """), {"t": imei}).mappings().first()
                if sr_row and sr_row["id"]:
                    refs.append(str(sr_row["id"]))
                    wo_row = db.execute(text("""
                        SELECT id FROM warehouse.work_orders
                        WHERE service_record_id = :sr_id
                        ORDER BY id DESC LIMIT 1
                    """), {"sr_id": sr_row["id"]}).mappings().first()
                    if wo_row and wo_row["id"]:
                        refs.append(str(wo_row["id"]))

                db.execute(text("""
                    UPDATE warehouse.repair_records
                    SET supply_status_code = 'Stoktan Çıktı', supply_requested_by = :user, supply_requested_at = NOW()
                    WHERE id = :rid
                """), {"rid": pending_id, "user": user})

            except Exception as mov_err:
                print(f"[WARN] StockMovement / RepairRecord güncelleme hatası: {mov_err}")

            db.commit()
            return json.dumps({
                "success": True, 
                "message": f"'{part_row['name']}' ({code}) başarıyla teslim edildi (Repair Stock'a aktarıldı)."
            }, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    @Slot(str, result=str)
    def get_delivered_parts_for_device(self, imei_or_serial):
        """Parça Teslim ekranında teslim edilmiş fiziki parçaları getirir. DGD ve stoksuz işçilik kalemleri elenir."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            imei_clean = (imei_or_serial or "").strip()
            if not imei_clean:
                return json.dumps({"success": True, "parts": []}, ensure_ascii=False)

            dgd_rows = db.execute(text("SELECT dgd_item_code FROM warehouse.flow_dgd_mapping WHERE dgd_item_code IS NOT NULL")).fetchall()
            dgd_mapped_codes = {r[0].strip().lower() for r in dgd_rows if r[0]}

            be_row = db.execute(text("""
                SELECT service_id FROM warehouse.batch_entries
                WHERE LOWER(TRIM(imei_number)) = LOWER(:t) OR LOWER(TRIM(serial_number)) = LOWER(:t) OR LOWER(TRIM(internal_id)) = LOWER(:t)
                ORDER BY id DESC LIMIT 1
            """), {"t": imei_clean}).mappings().first()

            refs = [imei_clean]
            if be_row and be_row["service_id"]:
                refs.append(str(be_row["service_id"]))

            sr_row = db.execute(text("""
                SELECT id FROM warehouse.service_records
                WHERE LOWER(TRIM(imei_number)) = LOWER(:t) OR LOWER(TRIM(imei_serial)) = LOWER(:t)
                ORDER BY id DESC LIMIT 1
            """), {"t": imei_clean}).mappings().first()
            if sr_row and sr_row["id"]:
                refs.append(str(sr_row["id"]))

            rows = db.execute(text("""
                SELECT 
                    rr.id AS repair_record_id,
                    rr.part_item_code,
                    COALESCE(p.name, rr.part_item_code) AS part_name,
                    p.brand,
                    p.model,
                    p.color,
                    p.item_category,
                    COALESCE(p.part_category, p.item_category) AS part_category,
                    rr.supply_requested_by,
                    TO_CHAR(rr.supply_requested_at, 'YYYY-MM-DD HH24:MI') AS supply_requested_at,
                    p.stock_tracking_type
                FROM warehouse.repair_records rr
                LEFT JOIN warehouse.parts p ON p.item_code = rr.part_item_code
                WHERE rr.service_record_id = ANY(:refs)
                  AND rr.part_item_code IS NOT NULL
                  AND TRIM(rr.part_item_code) <> ''
                  AND UPPER(TRIM(rr.part_item_code)) NOT LIKE 'DGD%'
                  AND LOWER(TRIM(COALESCE(p.item_category, ''))) NOT IN ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik')
                  AND LOWER(TRIM(COALESCE(p.part_category, ''))) NOT IN ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik')
                  AND LOWER(TRIM(COALESCE(rr.supply_status_code, ''))) IN ('stoktan çıktı', 'teslim edildi', 'teslimedildi')
                ORDER BY rr.supply_requested_at DESC NULLS LAST, rr.id DESC
            """), {"refs": refs}).mappings().all()

            parts = []
            for r in rows:
                code_clean = (r["part_item_code"] or "").strip()
                tracking_type = (r["stock_tracking_type"] or "Stok Takipli").strip()
                item_cat = (r["item_category"] or "").strip().lower()
                part_cat = (r.get("part_category") or "").strip().lower()

                is_dgd = code_clean.lower() in dgd_mapped_codes \
                    or code_clean.upper().startswith("DGD") \
                    or item_cat in ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik') \
                    or part_cat in ('dgd', 'dgd_labor', 'dgd labor', 'dgd_iscilik', 'dgd işçilik')

                is_stoksuz = tracking_type in ("Stok Takipsiz", "Stoksuz", "Servis Hizmeti", "Yazılım/Hizmet") or is_dgd

                if is_dgd or is_stoksuz:
                    continue

                parts.append({
                    "repairRecordId": str(r["repair_record_id"]),
                    "itemCode": code_clean,
                    "partName": r["part_name"] or "",
                    "brand": r["brand"] or "",
                    "model": r["model"] or "",
                    "color": r["color"] or "",
                    "deliveredBy": r["supply_requested_by"] or "",
                    "deliveredAt": r["supply_requested_at"] or "",
                    "stockTrackingType": tracking_type,
                })

            return json.dumps({"success": True, "parts": parts}, ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    @Slot(str, str, str, str, result=str)
    def return_delivered_part(self, repair_record_id, imei_or_serial, target_stock, username=""):
        """Parça Teslim ekranında teslim edilmiş bir parçayı geri alır ve seçilen stoğa (GOOD veya DOA) aktarır."""
        from sqlalchemy import text
        db = SessionLocal()
        try:
            rid_str = (repair_record_id or "").strip()
            imei = (imei_or_serial or "").strip()
            target = (target_stock or "").strip().upper()
            user = (username or "").strip() or "Sistem"

            if target not in ("GOOD", "DOA"):
                return json.dumps({"success": False, "message": "Hedef stok sadece 'GOOD' (Good Stock) veya 'DOA' (DOA Stock) olabilir."})

            rec = None
            if rid_str:
                rec = db.execute(text("""
                    SELECT id, part_item_code, service_record_id, supply_status_code
                    FROM warehouse.repair_records
                    WHERE id::text = :rid
                """), {"rid": rid_str}).mappings().first()

            if not rec:
                return json.dumps({"success": False, "message": "Teslim edilmiş onarım kaydı bulunamadı."})

            record_id = rec["id"]
            item_code = rec["part_item_code"]

            if not item_code:
                return json.dumps({"success": False, "message": "Kayda ait parça kodu bulunamadı."})

            # 1) supply_status_code sıfırla (Böylece parça teslim edilmiş durumdan çıkar)
            db.execute(text("""
                UPDATE warehouse.repair_records
                SET supply_status_code = NULL,
                    supply_requested_by = NULL,
                    supply_requested_at = NULL
                WHERE id = :rid
            """), {"rid": record_id})

            # 2) Stok Hareketi (Stok takipli parçalarda)
            part_row = db.execute(text("SELECT id, name FROM warehouse.parts WHERE item_code = :c LIMIT 1"), {"c": item_code}).mappings().first()

            if part_row and self._is_part_stock_tracked(db, item_code):
                part_id = part_row["id"]
                part_name = part_row["name"]

                repair_loc_id = _get_system_location_id(db, "repair_stock")
                
                if target == "GOOD":
                    target_loc_id = _get_system_location_id(db, "good_stock")
                    target_label = "Good Stock"
                    m_kind = "PARCA_IADE_GOOD"
                else:
                    target_loc_id = _get_system_location_id(db, "doa_stock")
                    target_label = "DOA Stock"
                    m_kind = "PARCA_IADE_DOA"

                # Repair Stock -1
                if repair_loc_id:
                    db.execute(text("""
                        UPDATE warehouse.stock
                        SET quantity = GREATEST(0, quantity - 1)
                        WHERE part_id = :pid AND location_id = :loc_id
                    """), {"pid": part_id, "loc_id": repair_loc_id})

                # Target Stock +1 (Good Stock veya DOA Stock)
                if target_loc_id:
                    from models.stock import Stock
                    t_stock = db.query(Stock).filter(Stock.part_id == part_id, Stock.location_id == target_loc_id).first()
                    if t_stock:
                        t_stock.quantity += 1
                    else:
                        db.add(Stock(part_id=part_id, location_id=target_loc_id, quantity=1))

                # Stock Movement Audit Log
                try:
                    from models.stock_movement import StockMovement
                    mov = StockMovement(
                        part_id=part_id,
                        part_name_snapshot=part_name,
                        source_location_id=repair_loc_id,
                        target_location_id=target_loc_id,
                        quantity=1,
                        type="İç Transfer",
                        movement_kind=m_kind,
                        description=f"Teslim Edilen Parça İadesi ({target_label}) - Parça: {item_code} - IMEI: {imei}",
                        created_by=user,
                        technician=user
                    )
                    db.add(mov)
                except Exception as mov_err:
                    print(f"[WARN] StockMovement kaydı eklenemedi: {mov_err}")

            db.commit()
            target_name = "Good Stock (Sağlam Depo)" if target == "GOOD" else "DOA Stock (Hasarlı/Arızalı Depo)"
            return json.dumps({
                "success": True,
                "message": f"'{part_row['name'] if part_row else item_code}' parçası teslimden geri alındı ve {target_name} alanına aktarıldı."
            }, ensure_ascii=False)
        except Exception as e:
            db.rollback()
            return json.dumps({"success": False, "message": str(e)}, ensure_ascii=False)
        finally:
            db.close()

    # Qt slotu DEĞİLDİR - get_db_status'un kullandığı, kısa zaman aşımlı ayrı motor.
    def _db_kontrol_motoru(self):
        """Yalnızca "ayakta mı" kontrolü için 2 saniye zaman aşımlı, havuzsuz motor.

        Uygulamanın ana motoru connect_timeout=5 ile kurulu ve 40'lık bir havuzu var.
        Bu kontrolü onun üzerinden yapmak iki sorun çıkarıyordu: (1) sunucu yanıt
        vermediğinde çağrı 5 saniye boyunca Qt ana iş parçacığını bloke ediyor, yani
        tüm arayüz donuyor; (2) sağlık kontrolü havuzdan bağlantı kapıp gerçek
        işlemleri bekletiyor. Ayrı ve havuzsuz bir motorla en kötü durum 2 saniye
        ve ana havuza hiç dokunulmuyor.
        """
        if getattr(self, "_db_kontrol_eng", None) is None:
            from sqlalchemy import create_engine
            from sqlalchemy.pool import NullPool
            from config.database import _build_database_url
            self._db_kontrol_eng = create_engine(
                _build_database_url(),
                poolclass=NullPool,
                connect_args={"connect_timeout": 2, "options": "-c statement_timeout=2000"},
            )
        return self._db_kontrol_eng

    @Slot(result=str)
    def get_print_support(self):
        """Yazdırmanın gerçekten mümkün olup olmadığını söyler.

        Asıl amaç: uygulamanın ESKİ bir süreci çalışıyorsa printRequested işleyicisi
        yoktur ve window.print() sessizce hiçbir şey yapmaz. Bu slot eski sürümde HİÇ
        BULUNMADIĞI için ekran durumu "eski sürüm" olarak ayırt edebilir.
        Yazıcı adı yalnızca bilgi amaçlıdır; çıktı PDF önizlemesi olarak açıldığından
        yazıcı kurulu olmasa da yazdırma çalışır.
        """
        try:
            from PySide6.QtPrintSupport import QPrinterInfo
            v = QPrinterInfo.defaultPrinter()
            # Yazıcı seçim penceresi açıldığı için varsayılan yazıcı olmasa da devam
            # edilir; ad yalnızca bilgi amaçlı döner.
            return json.dumps({"success": True, "supported": True, "reason": "",
                               "printer": "" if v.isNull() else v.printerName()},
                              ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": True, "supported": False, "reason": "hata",
                               "printer": "", "message": str(e)}, ensure_ascii=False)

    @Slot(result=str)
    def get_printer_forms(self):
        """Varsayılan yazıcının kağıt formlarını döner.

        Yazdır penceresindeki "Kağıt boyutu" listesiyle birebir aynıdır. Ölçüye göre
        otomatik eşleştirme her zaman doğru formu bulmuyor (DYMO'da 53.98x100.89 mm
        ölçüsünü üç ayrı form paylaşıyor) ve yanlış form seçilince sürücü etiketi
        kendi kenar boşluklarıyla basıp çerçeve/boş etiket üretiyor. Bu yüzden form
        Etiket Tasarımı ekranından elle seçilebiliyor.
        """
        try:
            from PySide6.QtGui import QPageSize
            from PySide6.QtPrintSupport import QPrinterInfo
            v = QPrinterInfo.defaultPrinter()
            if v.isNull():
                return json.dumps({"success": True, "printer": "", "forms": []})
            formlar = []
            for f in v.supportedPageSizes():
                b = f.size(QPageSize.Unit.Millimeter)
                formlar.append({"name": f.name(),
                                "width": round(b.width(), 2), "height": round(b.height(), 2)})
            return json.dumps({"success": True, "printer": v.printerName(),
                               "forms": formlar, "selected": self.etiket_form_adi},
                              ensure_ascii=False)
        except Exception as e:
            return json.dumps({"success": False, "message": str(e), "forms": []},
                              ensure_ascii=False)

    @Slot(str, result=str)
    def set_label_form(self, form_name):
        """Basımda kullanılacak kağıt formunun adını bildirir. Boş = otomatik seç."""
        self.etiket_form_adi = (form_name or "").strip()
        return json.dumps({"success": True, "selected": self.etiket_form_adi},
                          ensure_ascii=False)

    @Slot(bool, str, int, result=str)
    def set_print_preview(self, enabled, theme, label_count):
        """Yazdırma penceresini hazırlar: önizleme açık mı, tema, kaç etiket basılacak.

        Windows'un kendi yazdırma penceresindeki önizleme alanı Qt uygulamalarında
        "Bu uygulama yazdırma önizlemesini desteklemiyor" der ve doldurulamaz; bu
        yüzden yazıcı seçimi + önizleme uygulamanın kendi penceresinde gösterilir
        (bkz. main_window._baski_penceresi). TÜM basım noktalarında aynı davranır -
        Demontaj → "Üretime Aktar" dahil; kapatılırsa hiçbir pencere açılmaz.

        theme: "dark" | "light" — pencere uygulamanın temasıyla aynı görünsün diye.
        label_count: ekranın bastığı etiket sayısı. Pencere bunu üretilen SAYFA
        sayısıyla karşılaştırır; sayfa daha fazlaysa etiket sığmıyor demektir ve
        kullanıcı bunu kâğıt harcamadan görür.
        """
        self.baski_onizleme_istendi = bool(enabled)
        self.baski_temasi = "light" if str(theme).lower() == "light" else "dark"
        try:
            self.baski_etiket_sayisi = max(0, int(label_count))
        except (TypeError, ValueError):
            self.baski_etiket_sayisi = 0
        return json.dumps({"success": True, "enabled": self.baski_onizleme_istendi},
                          ensure_ascii=False)

    @Slot(result=str)
    def get_last_print_result(self):
        """Son yazdırma işinin sonucunu döner (ekran, basımdan sonra bunu sorar).

        Sessiz basımda operatör hiçbir şey görmüyordu: iş yazıcıya gitti mi, sürücü
        reddetti mi, kağıt ölçüsü tutmadı mı belli olmuyordu. Artık bu bilgi ekrana
        taşınıyor.
        """
        s = getattr(self, "son_yazdirma_sonucu", None)
        if not s:
            return json.dumps({"success": True, "durum": "yok"}, ensure_ascii=False)
        return json.dumps({"success": True, **s}, ensure_ascii=False)

    @Slot(float, float, result=str)
    def set_label_page_size(self, width_mm, height_mm):
        """Bir sonraki yazdırma işleminin kağıt ölçüsünü (mm) bildirir.

        CSS'teki @page kuralını sürücülerin çoğu yok sayıyor; kağıt boyutunu QPrinter
        üzerinde ayarlamak gerekiyor. Ekran window.print() çağırmadan HEMEN ÖNCE burayı
        çağırır, main_window._yazdirma_istegi de bu değeri okur (bkz. son_etiket_olcusu).
        Farklı etiket türleri farklı ölçüde olduğu için sabit bir değer kullanılamıyor.

        Ayrıca yazıcının GERÇEK sayfa tuvalini ölçüp döner; baskı CSS'i kutuyu bu
        ölçüye göre kurar. Etiket yazıcıları sıfır kenar boşluğunu kabul etmiyor
        (DYMO'da setPageMargins(0) False dönüyor), bu yüzden _kagit_ayarla artık
        setFullPage(True) ile tam sayfa modunu açıyor: tuval = medyanın tamamı.
        Dönen ölçü genelde sürücünün form ölçüsüdür (99014 için 53.98x100.89 mm) ve
        şablondaki yuvarlak 54x101'den birkaç yüzde mm küçüktür — kutu bu farkı
        bilmeden kurulursa kıl payı taşıyıp İKİNCİ ETİKETE düşüyordu.
        """
        try:
            g, y = float(width_mm), float(height_mm)
            if g <= 0 or y <= 0:
                return json.dumps({"success": False, "message": "Geçersiz ölçü."})
            self.son_etiket_olcusu = (g, y)
            alan = {"width": g, "height": y, "olculdu": False}
            try:
                from PySide6.QtWidgets import QApplication
                from PySide6.QtGui import QPageLayout
                from PySide6.QtPrintSupport import QPrinter, QPrinterInfo
                uygulama = QApplication.instance()
                # Headless sunucu (server.py) ve tarayıcıdan basım: QApplication YOK,
                # pencere de yok. Eskiden burada uygulama None iken topLevelWidgets()
                # çağrılıyor ve AttributeError atıyordu; ölçüm sessizce "ölçülemedi"ye
                # düşüyor, ekran kutuyu medya ölçüsüne göre kurup taşırıyordu.
                # Ayrıca o modda ölçülecek yazıcı SUNUCUNUN yazıcısıdır — kullanıcının
                # DYMO'suyla ilgisi yoktur, ölçmemek doğrusudur.
                pencere = None
                if uygulama is not None:
                    pencere = getattr(uygulama, "main_window", None)
                    if pencere is None or not hasattr(pencere, "_kagit_ayarla"):
                        # app.main_window'u main.py atıyor; test/gömülü kullanımda
                        # atanmamış olabiliyor. Pencereyi doğrudan da bulabilelim.
                        pencere = next((w for w in uygulama.topLevelWidgets()
                                        if hasattr(w, "_kagit_ayarla")), None)
                varsayilan = QPrinterInfo.defaultPrinter()
                if pencere is not None and not varsayilan.isNull():
                    # Gerçek işle aynı ayarlar uygulanır; bu kopya yalnızca ölçüm için.
                    deneme = QPrinter(varsayilan, QPrinter.PrinterMode.HighResolution)
                    pencere._kagit_ayarla(deneme, g, y)
                    r = deneme.pageLayout().paintRect(QPageLayout.Unit.Millimeter)
                    if r.width() > 1 and r.height() > 1:
                        alan = {"width": round(r.width(), 2),
                                "height": round(r.height(), 2), "olculdu": True,
                                "printer": varsayilan.printerName()}
            except Exception as e:
                print(f"[WARN] Basilabilir alan olculemedi: {e}")
            return json.dumps({"success": True, "printable": alan}, ensure_ascii=False)
        except Exception:
            pass
        return json.dumps({"success": False, "message": "Geçersiz ölçü."})

    @Slot(result=str)
    def get_db_status(self):
        """Veritabanına gerçekten ulaşılıp ulaşılamadığını söyler (giriş ekranındaki rozet).

        QWebChannel köprüsünün ayakta olması veritabanının da ayakta olduğu anlamına
        GELMEZ - köprü çalışırken sunucuya erişilemediği durumlar yaşandı. Bu yüzden
        burada gerçekten bir sorgu çalıştırılır.

        Sonuç 15 saniye önbelleklenir: ekran arka arkaya sorsa bile veritabanına tek
        bir sorgu gider, sonraki cevaplar anında döner.
        """
        import time as _t
        from sqlalchemy import text

        onbellek = getattr(self, "_db_durum_onbellek", None)
        if onbellek and (_t.time() - onbellek[0]) < 15:
            return onbellek[1]

        try:
            with self._db_kontrol_motoru().connect() as baglanti:
                baglanti.execute(text("SELECT 1"))
            cevap = json.dumps({"success": True, "connected": True,
                                "message": "Veritabanı bağlantısı aktif."}, ensure_ascii=False)
        except Exception as e:
            # Ham sürücü hatası kullanıcıya gösterilmez; sunucu adresi/parola sızdırabilir.
            print(f"[WARN] get_db_status: veritabanina erisilemiyor -> {type(e).__name__}")
            cevap = json.dumps({"success": True, "connected": False,
                                "message": "Veritabanına ulaşılamıyor."}, ensure_ascii=False)

        self._db_durum_onbellek = (_t.time(), cevap)
        return cevap

    @Slot(result=str)
    def get_app_version(self):
        """Mevcut uygulamanın sürüm bilgilerini (version.json) döndürür."""
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        vfile = os.path.join(base_dir, "version.json")
        if os.path.exists(vfile):
            try:
                with open(vfile, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    return json.dumps({"success": True, "data": data}, ensure_ascii=False)
            except Exception as e:
                return json.dumps({"success": False, "message": str(e)})
        return json.dumps({"success": True, "data": {"version": "v1.0.0", "app_name": "RemaLab WMS"}}, ensure_ascii=False)
