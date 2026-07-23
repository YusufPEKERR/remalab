"""MioCreate.xlsx -> warehouse.product_boms toplu içe aktarma scripti.

Excel'deki 'ProductBom' sayfası (id, code, item, productFamily, enabled, update)
ile veritabanındaki warehouse.product_boms tablosu (product_model, child_item_code,
quantity, status) farklı şemalara sahip. Bu script:
  - 'ProductFamily' sayfasından productFamily kodu (örn. 'iP11') -> okunabilir
    model adı (örn. 'iPhone 11') eşlemesini çıkarır (item_models lookup'ıyla aynı mantık),
  - 'ProductBom' sayfasındaki her satırı product_model/child_item_code/status olarak
    warehouse.product_boms tablosuna batch halinde upsert eder.

Kullanım: python import_product_bom.py
"""
import os
import sys

import openpyxl
from sqlalchemy import text

from config.database import engine

BATCH_SIZE = 500


def find_reference_excel_file():
    candidates = [
        f for f in os.listdir('.')
        if f.lower().endswith('.xlsx')
        and not f.startswith('~$')
        and ('dosya' in f.lower() or 'miocreate' in f.lower())
    ]
    return candidates[0] if candidates else None


def build_family_name_map(wb):
    ws = wb['ProductFamily']
    rows = list(ws.iter_rows(values_only=True))
    h_idx = next(i for i, r in enumerate(rows) if r and 'code' in [str(x).lower() for x in r if x is not None])
    headers = rows[h_idx]
    code_col = next(i for i, h in enumerate(headers) if h == 'code')
    shortname_col = next(i for i, h in enumerate(headers) if h == 'shortName')

    name_map = {}
    for r in rows[h_idx + 1:]:
        code = r[code_col]
        short_name = r[shortname_col]
        if code:
            name_map[str(code)] = str(short_name) if short_name else str(code)
    return name_map


def build_bom_rows(wb, family_name_map):
    ws = wb['ProductBom']
    rows = list(ws.iter_rows(values_only=True))
    h_idx = next(i for i, r in enumerate(rows) if r and 'item' in [str(x).lower() for x in r if x is not None])
    headers = rows[h_idx]
    item_col = next(i for i, h in enumerate(headers) if h == 'item')
    family_col = next(i for i, h in enumerate(headers) if h == 'productFamily')
    enabled_col = next(i for i, h in enumerate(headers) if h == 'enabled')

    bom_rows = []
    skipped = 0
    for r in rows[h_idx + 1:]:
        item_code = r[item_col]
        fam_code = r[family_col]
        if not item_code or not fam_code:
            skipped += 1
            continue
        product_model = family_name_map.get(str(fam_code), str(fam_code))
        status = 'Aktif' if r[enabled_col] else 'Pasif'
        bom_rows.append({
            'product_model': product_model,
            'child_item_code': str(item_code),
            'status': status,
        })
    return bom_rows, skipped


def ensure_unique_constraint(conn):
    conn.execute(text("""
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM pg_constraint WHERE conname = 'product_boms_model_item_uq'
            ) THEN
                ALTER TABLE warehouse.product_boms
                ADD CONSTRAINT product_boms_model_item_uq UNIQUE (product_model, child_item_code);
            END IF;
        END $$;
    """))


def insert_batch(conn, batch):
    values_sql = ", ".join(
        f"(:model{i}, :item{i}, 1, :status{i})" for i in range(len(batch))
    )
    params = {}
    for i, row in enumerate(batch):
        params[f"model{i}"] = row['product_model']
        params[f"item{i}"] = row['child_item_code']
        params[f"status{i}"] = row['status']
    conn.execute(text(f"""
        INSERT INTO warehouse.product_boms (product_model, child_item_code, quantity, status)
        VALUES {values_sql}
        ON CONFLICT (product_model, child_item_code)
        DO UPDATE SET status = EXCLUDED.status, updated_at = CURRENT_TIMESTAMP;
    """), params)


def main():
    limit = None
    if len(sys.argv) > 1:
        limit = int(sys.argv[1])

    fname = find_reference_excel_file()
    if not fname:
        print("MioCreate.xlsx bulunamadi.")
        sys.exit(1)

    print(f"Excel okunuyor: {fname}")
    wb = openpyxl.load_workbook(fname, data_only=True)
    family_name_map = build_family_name_map(wb)
    bom_rows, skipped = build_bom_rows(wb, family_name_map)
    wb.close()

    if limit:
        bom_rows = bom_rows[:limit]
        print(f"TEST MODU: sadece ilk {limit} satir islenecek.")

    print(f"Toplam gecerli satir: {len(bom_rows)} (atlanan: {skipped})")

    with engine.begin() as conn:
        ensure_unique_constraint(conn)

    inserted_total = 0
    with engine.begin() as conn:
        batch = []
        for row in bom_rows:
            batch.append(row)
            if len(batch) >= BATCH_SIZE:
                insert_batch(conn, batch)
                inserted_total += len(batch)
                batch = []
                print(f"  ... {inserted_total}/{len(bom_rows)}")
        if batch:
            insert_batch(conn, batch)
            inserted_total += len(batch)

    print(f"Tamamlandi. {inserted_total} satir warehouse.product_boms tablosuna aktarildi/guncellendi.")


if __name__ == '__main__':
    main()
