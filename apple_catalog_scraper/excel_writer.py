"""Katalog satırlarını .xlsx dosyasına yazar."""

from __future__ import annotations

import logging

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from . import config
from .catalog_builder import CatalogRow

logger = logging.getLogger(__name__)


def write_catalog_to_excel(rows: list[CatalogRow], output_path: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Apple Device Catalog"

    sheet.append(config.EXCEL_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"

    for row in rows:
        sheet.append(row.as_tuple())

    for col_index, header in enumerate(config.EXCEL_COLUMNS, start=1):
        max_len = len(header)
        for row in rows:
            value = row.as_tuple()[col_index - 1]
            max_len = max(max_len, len(str(value)))
        sheet.column_dimensions[get_column_letter(col_index)].width = min(max_len + 4, 40)

    workbook.save(output_path)
    logger.info("Excel dosyası yazıldı: %s (%d satır)", output_path, len(rows))
