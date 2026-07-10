"""
RemaLab WMS - Excel Utilities
Excel İçe/Dışa aktarma, önizleme ve dinamik kullanıcı dostu sütun eşleştirme logic'i.
"""

import pandas as pd
from PySide6.QtWidgets import (
    QDialog,
    QVBoxLayout,
    QTableWidget,
    QTableWidgetItem,
    QComboBox,
    QLabel,
    QDialogButtonBox,
    QFileDialog,
    QMessageBox,
    QHeaderView,
)
from PySide6.QtCore import Qt
from ui.translations import tr


class ExcelMappingDialog(QDialog):
    """Excel başlıkları ile Veritabanı sütunlarını eşleştirme ve Önizleme diyaloğu."""

    def __init__(
        self,
        excel_columns: list[str],
        db_columns: list[str],
        sample_df: pd.DataFrame,
        parent=None,
    ):
        super().__init__(parent)
        self.setWindowTitle(tr("excel.mapping_title"))
        self.setMinimumWidth(750)
        self.setMinimumHeight(550)
        self

        layout = QVBoxLayout(self)

        # Üst Bilgilendirme
        lbl = QLabel("1. Excel Sütun Başlıklarını Veritabanı Alanlarıyla Eşleştirin:")
        lbl
        layout.addWidget(lbl)

        # Eşleştirme tablosu
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(
            [tr("excel.db_column"), tr("excel.excel_column")]
        )
        self.table.verticalHeader().setVisible(False)
        self.table.setShowGrid(False)
        self.table.setMaximumHeight(200)
        self.table
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.table)

        # Orta Bilgilendirme (Önizleme Alanı)
        lbl_preview = QLabel("2. Seçilen Excel Dosyasından Örnek Veri Önizlemesi:")
        lbl_preview
        layout.addWidget(lbl_preview)

        # Excel Veri Önizleme Tablosu
        self.preview_table = QTableWidget()
        self.preview_table.verticalHeader().setVisible(False)
        self.preview_table.setShowGrid(True)
        self.preview_table
        layout.addWidget(self.preview_table)

        self.db_columns = db_columns
        self.excel_columns = excel_columns
        self.sample_df = sample_df
        self.combos = {}

        self._populate_mapping_table()
        self._populate_preview_table()

        # Ok / Cancel
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)

        buttons.button(QDialogButtonBox.Ok).setText(tr("db.save"))
        buttons.button(QDialogButtonBox.Ok)
        buttons.button(QDialogButtonBox.Cancel).setText(tr("db.cancel"))
        buttons.button(QDialogButtonBox.Cancel)

        layout.addWidget(buttons)

    def _populate_mapping_table(self):
        """Eşleştirme listesini hazırlar."""
        self.table.setRowCount(len(self.db_columns))

        friendly_names = {
            "part_id": f"{tr('parts.part_name')} (part_id)",
            "quantity": f"{tr('table.quantity')} (quantity)",
            "unit_price": f"{tr('inbound.unit_price')} (unit_price)",
            "total_cost": f"{tr('inbound.total_cost')} (total_cost)",
            "location_id": f"{tr('table.location_id')} (location_id)",
            "created_by": f"{tr('inbound.created_by')} (created_by)",
            "destination": f"{tr('outbound.destination')} (destination)",
            "part_category": f"{tr('table.part_category')} (part_category)",
            "brand_model": f"{tr('table.brand_model')} (brand_model)",
        }

        for idx, db_col in enumerate(self.db_columns):
            display_name = friendly_names.get(db_col, db_col)
            db_item = QTableWidgetItem(display_name)
            db_item.setFlags(db_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(idx, 0, db_item)

            combo = QComboBox()
            combo
            combo.addItem("[Eşleştirilmedi]", None)
            for excel_col in self.excel_columns:
                combo.addItem(excel_col, excel_col)

            # Otomatik eşleştirme denemeleri
            matched_idx = combo.findText(db_col, Qt.MatchContains)
            if matched_idx == -1:
                clean_friendly = display_name.split("(")[0].strip()
                matched_idx = combo.findText(clean_friendly, Qt.MatchContains)

            if matched_idx != -1:
                combo.setCurrentIndex(matched_idx)

            self.table.setCellWidget(idx, 1, combo)
            self.table.setRowHeight(idx, 36)
            self.combos[db_col] = combo

    def _populate_preview_table(self):
        """Seçilen Excel dosyasının ilk 5 satırını önizleme tablosuna yazar."""
        cols = self.sample_df.columns.tolist()
        self.preview_table.setColumnCount(len(cols))
        self.preview_table.setHorizontalHeaderLabels(cols)

        # İlk 5 satırı önizleme yap
        preview_rows = self.sample_df.head(5)
        self.preview_table.setRowCount(len(preview_rows))

        for r_idx, (_, row) in enumerate(preview_rows.iterrows()):
            for c_idx, col_name in enumerate(cols):
                val = str(row[col_name]) if not pd.isna(row[col_name]) else ""
                item = QTableWidgetItem(val)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.preview_table.setItem(r_idx, c_idx, item)

            self.preview_table.setRowHeight(r_idx, 32)

        self.preview_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

    def get_mappings(self) -> dict[str, str]:
        """Eşleştirilen alanları döndürür."""
        mapping = {}
        for db_col, combo in self.combos.items():
            val = combo.currentData()
            if val is not None:
                mapping[db_col] = val
        return mapping


def import_excel_flow(parent, db_columns: list[str], callback):
    """Excel kontrol etme ve interaktif eşleştirme akışı."""
    file_path, _ = QFileDialog.getOpenFileName(
        parent, tr("excel.select_file"), "", "Excel Files (*.xlsx *.xls)"
    )
    if not file_path:
        return

    try:
        # 1. Dosyayı oku
        df = pd.read_excel(file_path)
        excel_cols = df.columns.tolist()

        if len(df) == 0:
            QMessageBox.warning(
                parent, "Hata", "Seçilen Excel dosyasında hiç veri bulunamadı!"
            )
            return

        # 2. Önizleme ve eşleştirme modalını göster
        dialog = ExcelMappingDialog(excel_cols, db_columns, df, parent)
        if dialog.exec() == QDialog.Accepted:
            mapping = dialog.get_mappings()
            if not mapping:
                QMessageBox.warning(parent, "Uyarı", "Hiçbir sütun eşleştirilmedi!")
                return

            # Eşleştirilenleri yeniden adlandır ve işle
            inv_mapping = {v: k for k, v in mapping.items()}
            mapped_df = df[list(inv_mapping.keys())].rename(columns=inv_mapping)

            callback(mapped_df)
            QMessageBox.information(parent, "Başarılı", tr("excel.success"))

    except Exception as e:
        QMessageBox.critical(parent, "Hata", f"{tr('excel.error')} {e}")


def style_excel_file(filepath: str):
    """Excel dosyasını openpyxl kullanarak premium ve estetik bir tasarıma kavuşturur."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        wb = openpyxl.load_workbook(filepath)
        for sheet in wb.worksheets:
            # Renk Paleti ve Stiller
            header_fill = PatternFill(start_color="212B36", end_color="212B36", fill_type="solid") # Koyu şık gri/lacivert
            even_row_fill = PatternFill(start_color="F4F6F8", end_color="F4F6F8", fill_type="solid") # Açık gri alternatif satır
            odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Beyaz satır
            
            header_font = Font(name="Segoe UI", color="FFFFFF", bold=True, size=11)
            data_font = Font(name="Segoe UI", color="161C24", size=10)
            
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # Sadece altı çizili zarif kenarlık (Modern web tabloları gibi)
            light_gray_side = Side(style='thin', color='E2E8F0')
            modern_border = Border(bottom=light_gray_side, left=light_gray_side, right=light_gray_side)
            header_border = Border(bottom=Side(style='medium', color='1F6FEB')) # Başlığın altına mavi bir vurgu
            
            # Satır Yükseklikleri
            sheet.row_dimensions[1].height = 28 # Başlık daha ferah
            
            # Başlık satırını (Satır 1) biçimlendir
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = header_border
                
            # Tüm veri satırlarını biçimlendir ve sütun genişliklerini ayarla
            for col_idx, col in enumerate(sheet.columns, 1):
                max_length = 0
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for cell in col:
                    if cell.row > 1:
                        # Satır yüksekliği
                        sheet.row_dimensions[cell.row].height = 22
                        
                        # Alternatif arka plan rengi
                        if cell.row % 2 == 0:
                            cell.fill = even_row_fill
                        else:
                            cell.fill = odd_row_fill
                            
                        cell.font = data_font
                        cell.border = modern_border
                        cell.alignment = left_align
                        
                    try:
                        val_str = str(cell.value) if cell.value is not None else ""
                        if len(val_str) > max_length:
                            max_length = len(val_str)
                    except:
                        pass
                
                # Sütun genişliğini içeriğe göre ayarla (min 15, max 45)
                adjusted_width = min(max(max_length + 4, 15), 45)
                sheet.column_dimensions[col_letter].width = adjusted_width
                
            # İlk satırı dondur (sabit kalsın)
            sheet.freeze_panes = 'A2'
            
        wb.save(filepath)
    except Exception as e:
        print(f"Excel stili uygulanırken hata oluştu: {e}")


def export_excel_flow(parent, data: list[dict], default_filename: str):
    """Verileri Excel dosyası olarak dışa aktarır."""
    if not data:
        QMessageBox.warning(parent, "Uyarı", "Dışa aktarılacak veri bulunamadı!")
        return

    file_path, _ = QFileDialog.getSaveFileName(
        parent, "Dosyayı Kaydet", default_filename, "Excel Files (*.xlsx)"
    )
    if not file_path:
        return

    try:
        df = pd.DataFrame(data)
        df.to_excel(file_path, index=False)
        style_excel_file(file_path)
        QMessageBox.information(
            parent, "Başarılı", "Veriler başarıyla Excel dosyasına aktarıldı!"
        )
    except Exception as e:
        QMessageBox.critical(parent, "Hata", f"{tr('excel.error')} {e}")
